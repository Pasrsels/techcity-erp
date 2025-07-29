from django.utils import timezone
import json, datetime, openpyxl
from os import system 
import csv, base64
from django.core.files.base import ContentFile
from django.http import HttpResponse
from datetime import timedelta
from openpyxl.styles import Alignment, Font, PatternFill
from . models import *
import apps.inventory.tasks as tasks
from decimal import Decimal
from django.views import View
from django.db.models import Q, Sum, F, FloatField, ExpressionWrapper
from django.db import transaction
from django.contrib import messages
from utils.utils import generate_pdf
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse
from apps.finance.models import (
    StockTransaction,
    PurchaseOrderAccount, 
    PurchasesAccount,
    Expense,
    Currency,
    VATTransaction, 
    VATRate,
    Account,
    Cashbook,
    ExpenseCategory,
    AccountBalance,
    AccountTransaction
)
from . utils import (
    calculate_inventory_totals, 
    average_inventory_cost,
    generete_delivery_note
)
from . forms import (
    BatchForm,
    AddProductForm, 
    addCategoryForm, 
    addTransferForm, 
    DefectiveForm,
    RestockForm, 
    AddDefectiveForm, 
    ServiceForm, 
    AddSupplierForm,
    CreateOrderForm,
    noteStatusForm,
    PurchaseOrderStatusForm,
    ReorderSettingsForm,
    EditSupplierForm,
    StockTakeForm,
    AddShrinkageForm,
    AddWriteOffForm
)
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from channels.generic.websocket import  AsyncJsonWebsocketConsumer
from django.shortcuts import render, redirect, get_object_or_404, get_list_or_404
from permissions.permissions import (
    admin_required,
    # sales_required, 
    # accountant_required
)
from utils.account_name_identifier import account_identifier
from loguru import logger
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.core.files.uploadedfile import InMemoryUploadedFile
from apps.inventory.utils import best_price
from collections import defaultdict
from django.core.cache import cache
from django.core.paginator import Paginator
from typing import List, Dict, Any
from apps.finance.models import UserAccount
from django.template.loader import render_to_string

@login_required
def notifications_json(request):
    notifications = StockNotifications.objects.filter(inventory__branch=request.user.branch).select_related('inventory, inventory__branch').values(
        'inventory__product__name', 'type', 'notification', 'inventory__id'
    )
    return JsonResponse(list(notifications), safe=False)
    
@login_required
def batch_code(request):
    if request.method == 'GET':
        
        batch_codes = BatchCode.objects.all().values(
            'id',
            'code'
        )
        logger.info(batch_codes)
        return JsonResponse(list(batch_codes), safe=False)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            code = data.get('batch_code')

            BatchCode.objects.create(code=code)
            return JsonResponse({'success':True}, status=200)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)

@login_required
def temporary_purchase_order(request):
    logger.info('done')
    if request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name')
        product_id = data.get('product_id')
        quantity = data.get('quantity')
        price = data.get('price')
        supplier_id = data.get('supplier')
        
        temp_purchase_order = None
        
        with transaction.atomic():
        
            temp_purchase_order, created = TemporaryPurchaseOrder.objects.get_or_create(
            name=name,
                defaults={
                    'name': name,
                    'order_number': temp_purchase_order.generate_order_number(),
                    'branch': request.user.branch,
                    'user': request.user
                }
            )
            if created:
                temp_purchase_order.save()

            TemporaryPurchaseOrderItem.objects.create(
                temporary_purchase_order=temp_purchase_order,
                product_id=product_id,
                quantity=quantity,
                price=price,
                supplier_id=supplier_id, 
                unit_cost=price
            )
        
        return JsonResponse({'success':True, 'message':'Temporary purchase order created successfully'})

@login_required
def get_temporary_purchase_order_items(request, temp_po_id):
    if request.method == 'GET':
        temp_purchase_order_items = TemporaryPurchaseOrderItem.objects.filter(temporary_purchase_order=temp_po_id)
        return JsonResponse({'success':True, 'message':'Temporary purchase order items fetched successfully', 'temp_purchase_order_items':list(temp_purchase_order_items)})


@login_required
def product_list(request):
    """ for the pos """
    queryset = Inventory.objects.filter(branch=request.user.branch, status=True).select_related('branch')

    services = Service.objects.all().order_by('-name')

    search_query = request.GET.get('q', '') 
    product_id = request.GET.get('product', '')
    category_id = request.GET.get('category', '')

    if category_id:
        queryset = queryset.filter(category__id=category_id)
    if search_query:
        queryset = queryset.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query) 
        )
    if product_id:
        queryset = queryset.filter(id=product_id)

    inventory_data = list(queryset.values(
        'id', 
        'name', 
        'description', 
        'category__id', 
        'category__name',  
        'end_of_day',
        'price', 
        'quantity',
        'dealer_price',
        'image'
    ))
    
    # to be reviewed more 
    merged_data = [{
        'inventory_id': item['id'],
        'product_id': item['id'],
        'product_name':item['name'],
        'description': item['description'],
        'category': item['category__id'],
        'category_name': item['category__name'],
        'end_of_day':item['end_of_day'],
        'price': item['price'],
        'quantity': item['quantity'],
        'dealer_price':item['dealer_price'],
        'image':item['image']
    } for item in inventory_data]

    return JsonResponse(merged_data, safe=False)

@login_required
def branches_inventory(request):
    return render(request, 'branches_inventory.html')

@login_required
def branches_inventory_json(request):
    branches_inventory = Inventory.objects.filter(status=True).values(
        'product__name', 'price', 'quantity', 'branch__name'
    ).select_related('compay')
    return JsonResponse(list(branches_inventory), safe=False)
        
class AddProductView(LoginRequiredMixin, View):
    form_class = AddProductForm()
    initial = {'key':'value'}
    template_name = 'add_product.html'
    
    def get(self, request, *args, **kwargs):
        form = self.form_class
        return render(request, self.template_name, {'form': form})
    
    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = AddProductForm(request.POST)
            product_name = request.POST['name']
            
            try:
                product = Inventory.objects.get(name=product_name, branch=request.user.branch)                
                messages.warning(request, f'Product with name {product_name} exists.')
                return redirect('inventory:add_product')
            except Inventory.DoesNotExist:
                if form.is_valid():
                    product = form.save()
                    message = 'Product successfully created'
                    log_action = 'stock in'
                else:
                    messages.warning(request, 'Invalid form data')
                    return redirect('inventory:inventory')
                
            self.create_branch_inventory(product, log_action)
            
            messages.success(request, message)
        return redirect('inventory:inventory')


    def create_branch_inventory(self, product, log_action):
        try:
            inv = Inventory.objects.get(product__name=product.name)
            inv.quantity += product.quantity
            inv.price = product.price
            inv.cost = product.cost
            inv.save()
            self.activity_log(log_action, inv, inv)
        except Inventory.DoesNotExist:
            inventory = Inventory.objects.create(
                product=product,
                quantity=product.quantity,
                price=product.price,
                cost=product.cost,
                branch=self.request.user.branch,
                stock_level_threshold=product.min_stock_level
            )
            self.activity_log(log_action, inventory, inv=0)  
    
    def activity_log(self, action, inventory, inv):
        ActivityLog.objects.create(
            branch = self.request.user.branch,
            user=self.request.user,
            action= action,
            inventory=inventory,
            quantity=inventory.quantity,
            total_quantity=inventory.quantity + inv.quantity if action == 'update' else inventory.quantity
        )

from django.apps import apps


class ProcessTransferCartView(LoginRequiredMixin, View):
    """Handle product transfers between branches."""

    def post(self, request, *args, **kwargs) -> JsonResponse:
        """
        Process a transfer request or create a held transfer.
        Returns JSON response indicating success or failure.
        """
        try:
            data = json.loads(request.body)
            action = data['action']

            logger.info(f'data: {data}')

            if action == 'process':
                # Inline version of `process_transfer`
                Transfer = apps.get_model('inventory', 'Transfer')
                TransferItems = apps.get_model('inventory', 'TransferItems')
                Inventory = apps.get_model('inventory', 'Inventory')
                Branch = apps.get_model('company', 'Branch')
                ActivityLog = apps.get_model('inventory', 'ActivityLog')
                User = apps.get_model('users', 'User')

                with transaction.atomic():
                    user = request.user
                    user_branch = user.branch

                    branches_data = data['branches_to']
                    transfer_id = data.get('transfer_id', '')
                    cart = data['cart']

                    logger.info(f'cart data {cart}')

                    # Get branch objects
                    branch_objects = []
                    for branch in branches_data:
                        if branch.get('value'):
                            branch_obj = Branch.objects.get(id=branch['value'])
                        else:
                            branch_obj = Branch.objects.get(name=branch['name'])
                        branch_objects.append(branch_obj)

                    # Get products
                    products = Inventory.objects.filter(branch=user_branch).select_related('branch')
                    products_dict = {product.id: product for product in products}

                    # Validate quantities
                    product_quantities = defaultdict(int)
                    for item in cart:
                        product_quantities[item['product_id']] += item['quantity']

                    for product_id, total_quantity in product_quantities.items():
                        product = products_dict.get(int(product_id))
                        if total_quantity > product.quantity:
                            raise ValueError(f'Insufficient stock for product: {product.name}')

                    # Create or get transfer
                    branch_names = [branch['name'] for branch in branches_data]
                    if not transfer_id:
                        transfer = Transfer.objects.create(
                            branch=user_branch,
                            user=user,
                            transfer_ref=Transfer.generate_transfer_ref(user_branch.name, branch_names),
                            description='transfer'
                        )
                    else:
                        transfer = Transfer.objects.get(id=transfer_id)

                    transfer.transfer_to.set(branch_objects)

                    # Process transfer items
                    track_quantity = 0
                    transfer_items = []

                    for branch_obj in branch_objects:
                        for item in cart:
                            if item['branch_name'] == branch_obj.name:
                                product = products_dict.get(int(item['product_id']))
                                transfer_item = TransferItems(
                                    transfer=transfer,
                                    product=product,
                                    cost=item['cost'],
                                    price=item['price'],
                                    dealer_price=item['dealer_price'],
                                    quantity=item['quantity'],
                                    from_branch=user_branch,
                                    to_branch=branch_obj,
                                    description=f'from {user_branch} to {branch_obj}'
                                )
                                transfer_items.append(transfer_item)
                                track_quantity += item['quantity']

                    created_items = TransferItems.objects.bulk_create(transfer_items)

                    # Update inventory and create activity logs
                    for transfer_item in created_items:
                        inventory = Inventory.objects.select_for_update().get(
                            id=transfer_item.product.id,
                            branch__name=transfer_item.from_branch
                        )
                        inventory.quantity -= int(transfer_item.quantity)
                        inventory.save()

                        ActivityLog.objects.create(
                            invoice=None,
                            product_transfer=transfer_item,
                            branch=user_branch,
                            user=user,
                            action='transfer out',
                            dealer_price=transfer_item.dealer_price,
                            selling_price=transfer_item.price,
                            inventory=inventory,
                            system_quantity=inventory.quantity,
                            quantity=-transfer_item.quantity,
                            total_quantity=inventory.quantity,
                            description=f'to {transfer_item.to_branch}'
                        )

                        transfer.quantity += transfer_item.quantity

                    transfer.total_quantity_track = track_quantity
                    transfer.hold = False
                    transfer.date = datetime.datetime.now()
                    transfer.save()

                    return JsonResponse({'success': True, 'message': 'Transfer processed successfully'})

            else:
                return JsonResponse({'success': True, 'message': 'Held transfer saved'})

        except Exception as e:
            logger.error(f"Error processing transfer: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            })

    

    
@login_required
def delete_transfer(request, transfer_id):
    try:
        logger.info(f'transfer id: {transfer_id}')
        transfer = get_object_or_404(Transfer, id=transfer_id)
        logger.info(f'transfer: {transfer}')

        if transfer.receive_status:
            return JsonResponse({'success':False, 'message':f'Cancel failed the transfer is already received.'})

        transfer_items = TransferItems.objects.filter(transfer=transfer)
        logger.info(f'transfer: {transfer_items}')

        with transaction.atomic():
            inventory_updates = []
            activity_logs = []
            quantity_updates = {}

            for item in transfer_items:
                logger.info(f'From branch {item.from_branch}')

                product = Inventory.objects.get(branch=item.from_branch, id=item.product.id)
                logger.info(f'Product is: {product}')

                if product.id not in quantity_updates:
                    quantity_updates[product.id] = {'product': product, 'increment': 0}

                quantity_updates[product.id]['increment'] += item.quantity
            
            # to be further reviewed
            for update in quantity_updates.values():
                product = update['product']
                product.quantity += update['increment']
                inventory_updates.append(product)

                activity_logs.append(ActivityLog(
                    invoice=None,
                    product_transfer=None,  
                    branch=request.user.branch,
                    user=request.user,
                    action='transfer cancel',
                    inventory=product,
                    selling_price=None,  
                    dealer_price=None,  
                    quantity=update['increment'],
                    total_quantity=product.quantity,
                    description='Transfer cancelled'
                ))

            # Perform bulk updates
            Inventory.objects.bulk_update(inventory_updates, ['quantity'])
            ActivityLog.objects.bulk_create(activity_logs)

            transfer.delete = True
            transfer.save()

            logger.info('done')

        return JsonResponse({'success':True})
    except Exception as e:
        return JsonResponse({'success':False, 'message':f'{e}'})
        
        
@login_required       
def transfer_details(request, transfer_id):
    transfer = TransferItems.objects.filter(id=transfer_id).values(
        'product__name', 
        'transfer__transfer_ref', 
        'quantity', 
        'price', 
        'from_branch__name', 
        'to_branch__name'
    ).select_related(
        'transfer',
        'transfer_to',
        'transfer_from'
    )
    return JsonResponse(list(transfer), safe=False)

# requires a good name for the view
@login_required
def inventory(request):
    product_id = request.GET.get('id', '')

    if product_id:
        inventory_items = Inventory.objects.filter(
            id=product_id,
            status=True
        ).filter(
            branch=request.user.branch
        ).values()

        if inventory_items:
            return JsonResponse(list(inventory_items), safe=False)
        else:
            return JsonResponse({'error': 'Product not found in allowed branches'}, status=404)

    return JsonResponse({'error': 'Product ID is required'}, status=400)

@login_required
def add_inventory_view(request):
    try:
        name = request.POST.get('name')
        quantity = request.POST.get('quantity')
        category_name = request.POST.get('category')
        cost = request.POST.get('cost_price')
        price = request.POST.get('selling_price')
        dealer_price = request.POST.get('wholesale_price') or 0
        tax_type = request.POST.get('tax_type')
        min_stock = request.POST.get('min_stock_level')
        description = request.POST.get('description')
        end_of_day = request.POST.get('end_of_day') == 'on'
        service = request.POST.get('service') == 'on'
        image = request.FILES.get('image')

        category, _ = ProductCategory.objects.get_or_create(name=category_name)

        inventory = Inventory.objects.create(
            branch=request.user.branch,
            quantity=quantity,
            name=name,
            cost=cost,
            price=price,
            dealer_price=dealer_price,
            stock_level_threshold=min_stock,
            category=category,
            tax_type=tax_type,
            description=description,
            end_of_day=end_of_day,
            service=service,
            image=image
        )

        return JsonResponse({
            'status': 'success',
            'message': 'Product added successfully!',
            'product': {
                'id': inventory.id,
                'name': inventory.name
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
def inventory_index(request):
    form = ServiceForm()
    q = request.GET.get('q', '')  
    category = request.GET.get('category', '')    
    
    now = timezone.now() 
    today = now.date()  
    
    accessories = Accessory.objects.all()
    inventory = Inventory.objects.filter(
        branch=request.user.branch, 
        status=True, 
        disable=False,
        category__name=category
    ).select_related(
        'category',
        'branch'
    ).order_by('name')
    
    logs = ActivityLog.objects.filter(branch=request.user.branch).select_related('branch').order_by('-id')
    
    grouped_logs = {}
    ordered_grouped_logs = {}

    for log in logs:
        log_date = log.timestamp.date()  

        if log_date == today:
            date_key = 'Today'
        elif log_date == today - timedelta(days=1):
            date_key = 'Yesterday'
        else:
            date_key = log_date.strftime('%A, %d %B %Y')

        if not log.invoice:
            if date_key not in grouped_logs:
                grouped_logs[date_key] = {'logs': []}
            grouped_logs[date_key]['logs'].append(log)

    for special_day in ['Today', 'Yesterday']:
        if special_day in grouped_logs:
            ordered_grouped_logs[special_day] = grouped_logs[special_day]

    remaining_dates = sorted(
        [(k, v) for k, v in grouped_logs.items() if k not in ['Today', 'Yesterday']],
        key=lambda x: datetime.datetime.strptime(x[0], '%A, %d %B %Y') if not x[0] in ['Today', 'Yesterday'] else today
    )

    for date_key, data in remaining_dates:
        ordered_grouped_logs[date_key] = data
    

    if 'download' and 'excel' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={request.user.branch.name} stock.xlsx'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Add products data
        products = Inventory.objects.all()
        branches = Branch.objects.all().values_list('name', flat=True).distinct()
        row_offset = 0
        for branch in branches:
            worksheet['A' + str(row_offset + 1)] = f'{branch} Products'
            worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
            cell = worksheet['A' + str(row_offset + 1)]
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(size=16, bold=True)
            cell.fill = PatternFill(fgColor='AAAAAA', fill_type='solid')

            row_offset += 1 
            
            category_headers = ['Name', 'Cost', 'Price', 'Quantity']
            for col_num, header_title in enumerate(category_headers, start=1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = header_title
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            categories = Inventory.objects.filter(branch=request.user.branch).values_list('category__name', flat=True).distinct()
            for category in categories:
                products_in_category = products.filter(branch__name=branch, category__name=category)
                if products_in_category.exists():
                    worksheet['A' + str(row_offset + 1)] = category
                    cell = worksheet['A' + str(row_offset + 1)]
                    cell.font = Font(color='FFFFFF')
                    cell.fill = PatternFill(fgColor='0066CC', fill_type='solid')
                    worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
                    row_offset += 2

                for product in products.filter(branch__name=branch):
                    if product.category:
                        if category == product.category.name:
                            worksheet.append([product.name, product.cost, product.price, product.quantity])
                            row_offset += 1

        workbook.save(response)
        return response

    context = {
        'form': form,
        'total_cost': inventory.aggregate(total_cost=Sum(F('quantity') * F('cost')))['total_cost'] or 0,
        'total_price': inventory.aggregate(total_price=Sum(F('quantity') * F('price')))['total_price'] or 0,
        'search_query': q,
        'category': category,
        'accessories': accessories,
        'grouped_logs':ordered_grouped_logs
    }

    return render(request, 'inventory.html', context)

def logs_page(request):
    page_number = request.GET.get('page', 1)
    logs = ActivityLog.objects.select_related('inventory').order_by('-timestamp')
    paginator = Paginator(logs, 10) 
    page_obj = paginator.get_page(page_number)

    return render(request, 'partials/logs_page.html', {
        'logs': page_obj.object_list,
        'has_next': page_obj.has_next(),
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None
    })

@login_required
def edit_service(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == 'GET':
        form = ServiceForm(instance=service)
        return render(request, 'edit_service.html', {'form': form, 'service': service})
    
    if request.method == 'POST':
        form = ServiceForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, f'{service.name} successfully edited')
            return redirect('inventory:inventory')
        else:
            messages.warning(request, 'Please correct the errors below')
    
    messages.warning(request, 'Invalid request')
    return redirect('inventory:inventory')

        
@login_required   
def inventory_index_json(request):
    inventory = Inventory.objects.filter(branch=request.user.branch, status=True).values(
        'id', 
        'product__name', 
        'product__quantity', 
        'product__id',
        'price', 
        'cost', 
        'quantity', 
        'reorder'
    ).select_related('branch').order_by('product__name')
    return JsonResponse(list(inventory), safe=False)

@login_required 
@transaction.atomic
def activate_inventory(request, product_id):
    with transaction.atomic():
        product = Inventory.objects.select_for_update(id=product_id)
        product.status=True
        product.save()
        
    ActivityLog.objects.create(
        invoice = None,
        product_transfer = None,
        branch = request.user.branch,
        user=request.user,
        action= 'activated',
        inventory=product,
        quantity=product.quantity,
        total_quantity=product.quantity
    )
    
    messages.success(request, 'Product succefully activated')
    return redirect('inventory:inventory')

@login_required
def edit_inventory(request, product_id):
    """
    Edit inventory product details with proper concurrency handling and validation.
    Uses select_for_update to prevent race conditions during updates.
    """
    try:
        with transaction.atomic():
            inv_product = (
                Inventory.objects
                .select_for_update()
                .get(id=product_id, branch=request.user.branch)
            )

            logger.info(inv_product.category)

            if request.method == 'POST':
                original_quantity = inv_product.quantity
                
                try:
                    selling_price = Decimal(request.POST.get('price', 0))
                    dealer_price = Decimal(request.POST.get('dealer_price', 0))
                    cost = Decimal(request.POST.get('cost', 0))
                    quantity = int(request.POST.get('quantity', 0))
                    stock_level_threshold = int(request.POST.get('min_stock_level', 0))

                    category = ProductCategory.objects.get(id=int(request.POST.get('category')))

                    if selling_price < cost:
                        messages.warning(request, "Selling price cannot be less than cost")
                    if quantity < 0:
                        messages.warning(request, "Quantity cannot be negative")

                    inv_product.name = request.POST.get('name')
                    inv_product.description = request.POST.get('description')
                    inv_product.price = selling_price
                    inv_product.cost = cost
                    inv_product.dealer_price = dealer_price
                    inv_product.stock_level_threshold = stock_level_threshold
                    inv_product.quantity = quantity
                    inv_product.category = category
                    inv_product.end_of_day = request.POST.get('end_of_day') == 'on'

                    inv_product.save()

                    ActivityLog.objects.create(
                        branch=request.user.branch,
                        user=request.user,
                        action='Edit',
                        inventory=inv_product,
                        quantity=original_quantity,
                        total_quantity=quantity,
                        dealer_price=dealer_price,
                        selling_price=selling_price
                    )

                    logger.info(inv_product.category)

                    messages.success(request, f'{inv_product.name} updated successfully')
                    return redirect('inventory:inventory')

                except Exception as e:
                    messages.error(request, f'Invalid input: {str(e)}')
                    return render(request, 'inventory_form.html', {
                        'product': inv_product,
                        'title': f'Edit >>> {inv_product.name}',
                        'error': str(e)
                    })

    except Inventory.DoesNotExist:
        messages.error(request, 'Inventory item not found')
        return redirect('inventory:inventory')

    return render(request, 'inventory_form.html', {
        'product': inv_product,
        'title': f'Edit >>> {inv_product.name}'
    })

@login_required
def inventory_detail(request, id):
    purchase_order_items = PurchaseOrderItem.objects.all()
    
    inventory = Inventory.objects.get(id=id, branch=request.user.branch)
    
    logs = ActivityLog.objects.filter(
        inventory=inventory,
        branch=request.user.branch
    ).order_by('-timestamp__date', '-timestamp__time')
    

    # stock account data and totals (costs and quantities)
    stock_account_data = get_stock_account_data(logs)
    total_debits = sum(entry['cost'] for entry in stock_account_data if entry['type'] == 'debits')
    total_credits = sum(entry['cost'] for entry in stock_account_data if entry['type'] == 'credits')

    total_debits_quantity = sum(entry['quantity'] for entry in stock_account_data if entry['type'] == 'debits')
    total_credits_quantity = sum(entry['quantity'] for entry in stock_account_data if entry['type'] == 'credits')

    logger.info(f'debits {total_debits_quantity}')
    logger.info(f'debits {total_credits_quantity}')

    remaining_stock_quantity = total_debits_quantity - total_credits_quantity

    """ get inventory value based on the currencies in the system """
    inventory_value = []
    inventory_sold_value = []
    inventory_total_cost = inventory.cost * inventory.quantity
    
    for currency in Currency.objects.all():
        inventory_value.append(
            {
                'name': f'{currency.name}',
                'value': float(inventory_total_cost * currency.exchange_rate )if currency.exchange_rate == 1\
                        else float(inventory_total_cost * currency.exchange_rate)
            }
        )
        inventory_sold_value.append(
            {
                'name': f'{currency.name}',
                'value': logs.filter(invoice__invoice_return = False, invoice__currency=currency).\
                        aggregate(Sum('invoice__amount'))['invoice__amount__sum'] or 0
            }
        )
    
    logger.info(f'inventory value: {inventory_sold_value}')

    # logs = ActivityLog.objects.annotate(hour=Extract('timestamp', 'hour')).order_by('-hour')

    """ create log data structure for the activity log graph """
    sales_data = {}
    stock_in_data = {}
    transfer_data = {}
    labels = []

    for log in logs:
        month_name = log.timestamp.strftime('%B')  
        year = log.timestamp.strftime('%Y')
        month_year = f"{month_name} {year}"  

        if log.action == 'Sale':
            if month_year in sales_data:
                sales_data[month_year] += log.quantity
            else:
                sales_data[month_year] = log.quantity
        elif log.action in ('stock in', 'Update'):
            if month_year in stock_in_data:
                stock_in_data[month_year] += log.quantity
            else:
                stock_in_data[month_year] = log.quantity
        elif log.action == 'Transfer':
            if month_year in transfer_data:
                transfer_data[month_year] += log.quantity
            else:
                transfer_data[month_year] = log.quantity
        elif log.action == 'sale return':
            if month_year in transfer_data:
                transfer_data[month_year] += abs(log.quantity)
            else:
                transfer_data[month_year] = abs(log.quantity)

        if month_year not in labels:
            labels.append(month_year)

    """ download a log or stock account pdf """

    # if request.GET.get('logs'):
    #     download_stock_logs_account('logs', logs, inventory)
    # elif request.GET.get('account'):
    #     download_stock_logs_account('account', logs, inventory)
    
    logger.info(stock_account_data)
    
    return render(request, 'inventory_detail.html', {
        'inventory': inventory,
        'remaining_stock_quantity':remaining_stock_quantity,
        'stock_account_data':stock_account_data,
        'inventory_value':inventory_value,
        'inventory_sold_value':inventory_sold_value,
        'total_debits':total_debits,
        'total_credits':total_credits,
        'logs': logs,
        'items':purchase_order_items,
        'sales_data': list(sales_data.values()), 
        'stock_in_data': list(stock_in_data.values()),
        'transfer_data': list(transfer_data.values()),
        'labels': labels,
    })

def get_stock_account_data(logs):
    """ 
    logs data structure with both debits and credits, using default currency 
    debits => stock in,  transfer_in, sales_returns, positive_adjustments
    credits => transfer_out, supplier_returns, negative_adjustments, sale
"""
    stock_account = []
    for log in logs:
        if log.action in ['stock in', 'transfer_in', 'sale return', 'purchase edit +', 'stock adjustment']:
            entry_type = 'debits'
        elif log.action in ['transfer', 'returns', 'Sale', 'purchase edit -', 'write off']:
            entry_type = 'credits'
        else:
            continue  
        
        if log.action == 'sale return':
            inventory_cost = getattr(log.inventory, 'cost', Decimal('0.00'))
            cost = abs(log.quantity) * inventory_cost
            
            logger.info(f'action : {log.action}')
            logger.info(f'quantity : {abs(log.quantity)}')
            logger.info(f'inventory cost : {inventory_cost}')
            logger.info(f'cost : {cost}')

            stock_account.append({
                'type': entry_type,
                'description': log.action,
                'quantity': abs(log.quantity),
                'cost': cost,
                # 'currency': 'USD',
                'timestamp': log.timestamp,
                'user': log.user.username if log.user else 'Unknown',
                'branch': log.branch.name,
            })
            logger.info(stock_account)
        else:
            inventory_cost = getattr(log.inventory, 'cost', Decimal('0.00'))
            cost = abs(log.quantity) * inventory_cost
            
            logger.info(f'action : {log.action}')
            logger.info(f'quantity : {log.quantity}')
            logger.info(f'inventory cost : {inventory_cost}')
            logger.info(f'cost : {cost}')

            stock_account.append({
                'type': entry_type,
                'description': log.action,
                'quantity': log.quantity,
                'cost': cost,
                # 'currency': 'USD',
                'timestamp': log.timestamp,
                'user': log.user.username if log.user else 'Unknown',
                'branch': log.branch.name,
            })
            logger.info(stock_account)
    return stock_account

@login_required
def inventory_transfer_index(request):
    q = request.GET.get('q', '') 
    branch_id = request.GET.get('branch', '')
    page_number = request.GET.get('page', 1)  

    transfers = Transfer.objects.filter(
        Q(branch=request.user.branch) | Q(transfer_to=request.user.branch),
        delete=False
    ).annotate(
        total_quantity=Sum('transferitems__quantity'),
        total_received_qnt=Sum('transferitems__received_quantity'),
        total_r_difference=(F('transferitems__quantity')) - Sum(F('transferitems__received_quantity')),
        total_received_amount=Sum(F('transferitems__received_quantity') * F('transferitems__cost')),
        total_amount=ExpressionWrapper(
            Sum(F('transferitems__quantity') * F('transferitems__cost')),
            output_field=FloatField()
        ),
        check_all_received=Sum(F('transferitems__quantity') - F('transferitems__received_quantity')),
    ).order_by('-time')

    if q:
        transfers = transfers.filter(Q(transfer_ref__icontains=q) | Q(date__icontains=q))
    if branch_id:
        transfers = transfers.filter(transfer_to__id=branch_id)

    transfers_list = list(transfers)
    unique_transfers = list({t.id: t for t in transfers_list}.values())
    unique_transfers.sort(key=lambda x: x.time, reverse=True)  

    paginator = Paginator(unique_transfers, 20)
    paginated_transfers = paginator.get_page(page_number)

    if paginated_transfers:
        transfers_data = [
            {
                'id': transfer.id,
                'transfer_ref': transfer.transfer_ref,
                'branch': transfer.branch.name,
                'transfer_to': [branch.id for branch in transfer.transfer_to.all()],
                'total_quantity': transfer.total_quantity,
                'total_received_qnt': transfer.total_received_qnt,
                'total_r_difference': transfer.total_r_difference,
                'total_received_amount': transfer.total_received_amount,
                'total_amount': transfer.total_amount,
                'check_all_received': transfer.check_all_received,
                'time': transfer.time.strftime('%Y-%m-%d %H:%M:%S'),
                'username': transfer.user.username,
                'description': transfer.description 
            }
            for transfer in paginated_transfers
        ]
    else:
        transfers_data = []

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'transfers': transfers_data,
            'has_next': paginated_transfers.has_next(),
        })

    return render(request, 'transfers.html', {
        'transfers': paginated_transfers,
        'search_query': q,
    })

@login_required
def inventory_transfer_item_data(request, id):
    """
    Transfer items of the parent transfer
    """
    transfer_items = TransferItems.objects.filter(
        Q(to_branch=request.user.branch) | Q(from_branch=request.user.branch),
        transfer__id=id,
        transfer__delete=False,
    ).select_related(
        'product', 'from_branch', 'to_branch', 'action_by', 'received_by', 'transfer'
    ).annotate(
        total_amount=F('quantity') * F('product__cost')
    ).values(
        'id',
        'quantity', 
        'over_less_quantity', 
        'price', 
        'dealer_price', 
        'received', 
        'declined', 
        'over_less', 
        'quantity_track', 
        'description', 
        'over_less_description', 
        'received_quantity', 
        'cost', 
        'date', 
        'date_received', 
        'transfer__id',
        'from_branch__name',
        'product__name', 
        'to_branch__name', 
        'action_by__username', 
        'received_by__username',
        'received_back_quantity'
    )

    return JsonResponse(list(transfer_items), safe=False)

@login_required
def add_transfer_item(request, transfer_id):
    if request.method == 'GET':
        try:
            transfer = Transfer.objects.get(id=transfer_id)
            return render(request, 'add_transfer_item.html', {
                'transfer': transfer,
            })
        except Transfer.DoesNotExist:
            messages.warning(request, f'Transfer with ID {transfer_id} does not exist.')
            return redirect('inventory:transfers')

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            items = data.get('items', [])
            branches_data = data['branches_to']
            transfer_id = data.get('transfer_id', '')

            products = ProcessTransferCartView._get_products(request.user.branch)
            branch_obj_list = ProcessTransferCartView._get_branch_objects(branches_data)

            transfer = Transfer.objects.get(id=transfer_id)

            ProcessTransferCartView._process_transfer(data['cart'], products, branch_obj_list, transfer, request)
            
            return JsonResponse({'success': True, 'message': 'Items added to transfer successfully'}, status=200)

        except Transfer.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Transfer not found'}, status=404)
        except Exception as e:
            logger.error(f"Error adding items to transfer: {e}")
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        
@login_required
def held_transfer_json(request, transfer_id):
    transfer_items = Holdtransfer.objects.filter(transfer__id=transfer_id).values(
        'product__name',
        'from_branch__name',
        'to_branch__name',
        'quantity',
        'price',
        'cost',
        'dealer_price'
    )
    return JsonResponse(list(transfer_items), safe=False)

@login_required
def held_transfers(request):
    transfers = Transfer.objects.filter(
        Q(branch=request.user.branch),
        delete=False, 
        hold=True
    )
    logger.info(f'held transfers: {transfers}')
    return render(request, 'held_transfers.html',{
        'transfers':transfers
    })

@login_required
def process_held_transfer(request, transfer_id):
    try:
        transfer = Transfer.objects.get(
            id=transfer_id,
            branch=request.user.branch,
            delete=False, 
            hold=True
        )
        return render(request, 'process_held_transfers.html',{
            'transfer':transfer
        })
    except:
        messages.warning(request, f'Transfer with id {transfer_id} not found.')
        return redirect('inventory:transfers')

@login_required
def print_transfer(request, transfer_id):
    if request.method == 'GET':
        try:
            transfer = Transfer.objects.get(id=transfer_id)
            transfer_items = TransferItems.objects.filter(transfer=transfer)
        
            return render(request, 'components/ibt.html', {
                'date':datetime.datetime.now(),
                'transfer':transfer, 
                'transfer_items':transfer_items
            })
        except:
            messages.warning(request, 'Transfer doesnt exists')
            return redirect('inventory:transfers')
        
    if request.method == 'POST':
        try:
            transfer = Transfer.objects.get(id=transfer_id)
            transfer_items = TransferItems.objects.filter(transfer=transfer).values(
                'product__name',
                'product__price',
                'quantity',
                'to_branch__name', 
                'product__cost'
            )
        
            logger.info(transfer_items)
            return JsonResponse({'success':True, 'data':list(transfer_items)})
        except Exception as e:
            return JsonResponse({'success':False, 'meesage':f'{e}'})
    
@login_required
@transaction.atomic
def receive_inventory(request):
    if request.method == 'POST':
        try: 
            data = json.loads(request.body)
            logger.info(data)

            transfer_id = data.get('item_id')
            quantity_received = int(data.get('received_quantity'))
            serial_numbers = data.get('serial_numbers', [])
            received = True

            branch_transfer = get_object_or_404(TransferItems, id=transfer_id)
            logger.info(branch_transfer)
            transfer_obj = get_object_or_404(Transfer, id=branch_transfer.transfer.id)
            logger.info(transfer_obj)

            # if quantity_received > branch_transfer.quantity:
            #     return JsonResponse({'success': False, 'message': 'Quantity received cannot be more than quantity transferred'}, status=400)

            logger.info(branch_transfer)

            if received:
                logger.info(received)
                if quantity_received != branch_transfer.quantity:
                    branch_transfer.over_less_quantity = branch_transfer.quantity - quantity_received
                    branch_transfer.over_less = True
                    branch_transfer.save()
                
                logger.info(branch_transfer)

                # # validation for more quantity received
                # if quantity_received > branch_transfer.quantity:
                #     return JsonResponse({'success': False, 'message': 'Quantity received cannot be more than quantity transferred'}, status=400)

                
                with transaction.atomic():
                    product, created = Inventory.objects.get_or_create(
                        name=branch_transfer.product.name,
                        branch=request.user.branch,
                        defaults={
                            'cost': branch_transfer.cost,
                            'price': branch_transfer.price,
                            'quantity': quantity_received,
                            'dealer_price': branch_transfer.dealer_price,
                            'description': branch_transfer.product.description or '',
                            'category': branch_transfer.product.category,
                            'tax_type': branch_transfer.product.tax_type,
                            'stock_level_threshold': branch_transfer.product.stock_level_threshold,
                        }
                    )

                    if not created:
                        product.quantity += quantity_received
                        product.price = branch_transfer.price
                        product.cost = branch_transfer.cost
                        product.dealer_price = branch_transfer.dealer_price
                        product.save()

                    logger.info('created')

                    # for serial_number in serial_numbers:
                    #     print(serial_number)
                    #     serial_obj, _ = SerialNumber.objects.get_or_create(
                    #         serial_number=serial_number,
                    #         defaults={
                    #             'status':True
                    #         }
                    #     )
                    #     print(serial_obj, _)
                    #     product.serial_numbers.add(serial_obj)
                    #     logger.info('saved')

                    ActivityLog.objects.create(
                        branch=request.user.branch,
                        user=request.user,
                        action='stock in',
                        inventory=product,
                        dealer_price=product.dealer_price,
                        selling_price=product.price,
                        system_quantity=product.quantity,
                        quantity=quantity_received,
                        total_quantity=product.quantity,
                        product_transfer=branch_transfer,
                        description=f'From {branch_transfer.from_branch}, received: {branch_transfer.received_quantity} x {branch_transfer.quantity}'
                    )

                    product.batch += f'{branch_transfer.product.batch}, '
                    product.save()
                    logger.info('product received')

            branch_transfer.quantity_track = branch_transfer.quantity - quantity_received
            branch_transfer.received_quantity += quantity_received
            branch_transfer.received_by = request.user
            branch_transfer.received = True
            branch_transfer.description = f'received {quantity_received} out of {branch_transfer.quantity}'
            branch_transfer.save()

            logger.info(transfer_obj.total_quantity_track)

            transfer_obj.total_quantity_track -= quantity_received
            transfer_obj.save()

            if not transfer_obj.receive_status:
                transfer_obj.receive_status = True
                transfer_obj.save()
            
            logger.info('done')

            return JsonResponse({'success': True, 'message': 'Product received successfully'}, status=200)
        except TransferItems.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid transfer ID'}, status=400)
        except Transfer.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Transfer object not found'}, status=400)
        except Exception as e:
            logger.info(e)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
def edit_transfer_item(request, transfer_item_id):
    try:
        transfer_item = TransferItems.objects.get(id=transfer_item_id)
    except TransferItems.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Transfer item not found'}, status=404)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            logger.info(data)
            new_quantity = int(data.get('edit_quantity'))

            logger.info(new_quantity)

            if transfer_item.received:
                return JsonResponse({'success':False, 'message':f'{transfer_item.product.name}, is already received.'})

            if new_quantity is None:
                return JsonResponse({'success': False, 'message': 'Quantity is required'}, status=400)

            # if new_quantity < 0:
            #     return JsonResponse({'success': False, 'message': 'Quantity cannot be negative'}, status=400)

            with transaction.atomic():
                old_quantity = transfer_item.quantity
                transfer_item.quantity = new_quantity
                transfer_item.save()

                logger.info(transfer_item)

                # Update inventory
                inventory = Inventory.objects.select_for_update().get(
                    id=transfer_item.product.id,
                    branch__name=transfer_item.from_branch
                )
                
                if inventory.quantity < new_quantity:
                    return JsonResponse({'success': False, 'message': 'Insufficient stock to update transfer item quantity'}, status=400)
                
                inventory.quantity += old_quantity - new_quantity
                inventory.save()

                logger.info(inventory)

                # Log the activity
                ActivityLog.objects.create(
                    branch=request.user.branch,
                    user=request.user,
                    action='edit transfer item',
                    inventory=inventory,
                    quantity=inventory.quantity,
                    total_quantity=inventory.quantity,
                    description=f'Edited transfer item quantity from {old_quantity} to {new_quantity}'
                )

            return JsonResponse({'success': True, 'message': 'Transfer item updated successfully'}, status=200)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

@login_required
def receive_inventory_json(request):
    transfers =  TransferItems.objects.filter(to_branch=request.user.branch).order_by('-date')
    if request.method ==  'GET':
        transfers = transfers.values(
            'id',
            'date', 
            'quantity',
            'received', 
            'description',
            'date_received',
            'product__name', 
            'from_branch__name',
            'received_by__username'
        )
        return JsonResponse(list(transfers), safe=False)
    
    
@login_required
def update_notification_settings(request):
    settings, created = InventoryNotificationSettings.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        settings.low_stock = 'low_stock' in request.POST
        settings.out_of_stock = 'out_of_stock' in request.POST
        settings.movement_create = 'movement_create' in request.POST
        settings.movement_update = 'movement_update' in request.POST
        settings.movement_delete = 'movement_delete' in request.POST
        settings.movement_transfer = 'movement_transfer' in request.POST
        settings.save()
        messages.success(request, "Notification settings updated successfully.")
        return redirect('inventory:settings')

    return render(request, 'inventory/settings.html', {'settings': settings})

@login_required
def settings(request):
    return render(request, 'settings.html')
    
@login_required
@transaction.atomic
def over_less_list_stock(request):
    form = DefectiveForm()
    search_query = request.GET.get('search_query', '')
    
    transfers =  TransferItems.objects.filter(to_branch=request.user.branch).prefetch_related('product', 'transfer').order_by('-date')

    if search_query:
        transfers = transfers.filter(
            Q(transfer__product__name__icontains=search_query)|
            Q(transfer__transfer_ref__icontains=search_query)|
            Q(date__icontains=search_query)
        )
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            action = data.get('action')
            transfer_id = data.get('transfer_id')
            quantity = int(data.get('quantity', 0))
            branch = data.get('branch')
            reason = data.get('reason', '')
            action_taken = data.get('action_taken')
            
            branch_transfer = TransferItems.objects.filter(id=transfer_id).select_related('transfer').first()
            transfer = branch_transfer.transfer

            with transaction.atomic():
                if int(branch_transfer.received_quantity) != int(branch_transfer.quantity):

                    if action == 's_receive':
                        if quantity > branch_transfer.quantity:
                            return JsonResponse({'success': False, 'message': 'Quantity cannot be more than the transfer quantity'}, status=400)

                        product = Inventory.objects.get(name=branch_transfer.product.name, branch=request.user.branch)

                        product.quantity += quantity
                        product.save()

                        branch_transfer.description = f'Received {branch_transfer.received_quantity} from {branch_transfer.from_branch}'
                        branch_transfer.received_quantity += quantity
                        branch_transfer.save()

                        ActivityLog.objects.create(
                            branch=request.user.branch,
                            user=request.user,
                            action='stock in',
                            inventory=product,
                            quantity=quantity,
                            total_quantity=product.quantity,
                            product_transfer=branch_transfer,
                            description=f'Received {quantity} from {branch_transfer.from_branch}'
                        )

                        return JsonResponse({'success': True, 'message': 'Quantity successfully transferred to another branch'}, status=200)

                    if action == 'send_to':
                        if quantity > branch_transfer.quantity:
                            return JsonResponse({'success': False, 'message': 'Quantity cannot be more than the transfer quantity'}, status=400)

                        new_branch = Branch.objects.get(id=branch)
                        
                        try:
                            product = Inventory.objects.get(name=branch_transfer.product.name, branch=new_branch)
                        except Exception as e:
                            product = Inventory.objects.create(
                                name=branch_transfer.product.name,
                                branch=new_branch,
                                cost=branch_transfer.cost,
                                price=branch_transfer.price,
                                quantity=quantity,
                                dealer_price=branch_transfer.dealer_price,
                                description=branch_transfer.product.description or '',
                                category=branch_transfer.product.category,
                                tax_type=branch_transfer.product.tax_type,
                                stock_level_threshold=branch_transfer.product.stock_level_threshold,
                            )
                        
                        product.quantity += quantity
                        product.save()

                        ActivityLog.objects.create(
                            branch=request.user.branch,
                            user=request.user,
                            action='stock in',
                            inventory=product,
                            quantity=quantity,
                            total_quantity=product.quantity,
                            product_transfer=branch_transfer,
                            description=f'Received {quantity} from {branch_transfer.from_branch}'
                        )

                        branch_transfer.quantity -= quantity
                        branch_transfer.save()

                        branch_transfer.transfer.transfer_to.add(new_branch.id)
                        
                        try:
                            send_to_branch_transfer_item = TransferItems.objects.get(transfer__id=transfer.id, product__name=product.name, to_branch=new_branch)
                        except Exception as e:
                            send_to_branch_transfer_item = TransferItems.objects.create(
                                transfer=branch_transfer.transfer,
                                product=product,
                                cost=branch_transfer.cost,
                                price=branch_transfer.price,
                                dealer_price=branch_transfer.dealer_price,
                                quantity=quantity,
                                from_branch=branch_transfer.from_branch,
                                to_branch=new_branch,
                                description=f'Transferred {quantity} to {new_branch} from {branch_transfer.from_branch}'
                            )

                        send_to_branch_transfer_item.quantity += quantity
                        send_to_branch_transfer_item.save()
    
                        product.quantity -= quantity
                        product.save()

                        ActivityLog.objects.create(
                            branch=request.user.branch,
                            user=request.user,
                            action='transfer out',
                            inventory=product,
                            quantity=-quantity,
                            total_quantity=product.quantity,
                            product_transfer=branch_transfer,
                            description=f'Transferred {quantity} to {new_branch}'
                        )

                        return JsonResponse({'success': True, 'message': 'Quantity successfully transferred to another branch'}, status=200)
                    
                    if action in ['write-off', 'defective', 'shrinkage']:

                        if request.user.branch == branch_transfer.to_branch:
                            return JsonResponse({'success': False, 'message': 'Sorry, action not available'}, status=400)

                        product = Inventory.objects.get(        
                            id=branch_transfer.product.id,
                            branch=request.user.branch
                        )

                        if quantity > branch_transfer.quantity:
                            return JsonResponse({'success': False, 'message': 'Quantity cannot be more than the transfer quantity'}, status=400)

                        if action == 'write-off':
                            WriteOff.objects.create(
                                inventory_item=product,
                                quantity=quantity,
                                reason=f'{reason} (transfer number: {branch_transfer.transfer.transfer_ref})',
                                created_by=request.user
                            )
                        elif action == 'defective':
                            DefectiveItem.objects.create(
                                inventory_item=product,
                                quantity=quantity,
                                defect_description=f'{reason} (transfer number: {branch_transfer.transfer.transfer_ref})',
                                action_taken=action_taken,
                                created_by=request.user
                            )
                        elif action == 'shrinkage':
                            InventoryShrinkage.objects.create(
                                inventory_item=product,
                                quantity=quantity,
                                reason=action_taken,
                                created_by=request.user
                            )

                        product.quantity += quantity
                        product.save()

                        ActivityLog.objects.create(
                            branch = request.user.branch,
                            user=request.user,
                            action= 'stock in',
                            inventory=product,
                            quantity=quantity,
                            total_quantity=product.quantity,
                            product_transfer=branch_transfer,
                            description=f'Received from {transfer.transfer_ref}'
                        )

                        product.quantity -= quantity
                        product.save()
                        
                        branch_transfer.defect_quantity =+ quantity
                        description = f'{branch_transfer.defect_quantity} item(s) transferred to {action} account.'
                        branch_transfer.quantity -= int(quantity)
                        branch_transfer.description = description
                        branch_transfer.over_less_description = description
                        branch_transfer.action_by = request.user
                        branch_transfer.over_less = False
                        branch_transfer.save()

                        transfer.defective_status = True
                        transfer.save()

                        ActivityLog.objects.create(
                            branch=request.user.branch,
                            user=request.user,
                            action=action,
                            inventory=product,
                            quantity=-quantity,
                            total_quantity=product.quantity,
                            product_transfer=branch_transfer,
                            description=description
                        )

                        return JsonResponse({'success': True}, status=200)

                    if action == 'receive':
                        """
                            receiving the transfer on the current branch
                        """

                        if request.user.branch != branch_transfer.from_branch:
                            return JsonResponse({'success': False, 'message': 'Sorry, action not available'}, status=400)

                        product = Inventory.objects.get(
                            Q(name__icontains = branch_transfer.product.name), 
                            branch=request.user.branch
                        )

                        logger.info(f'Processing receiving for product: {product.name}')

                        if (branch_transfer.quantity - branch_transfer.received_quantity) < int(quantity):
                            return JsonResponse({'success':False, 'message':f'Cant receive more quantity'})
                            
                        product.quantity += int(quantity)
                        product.save()
                        
                        branch_transfer.received_back_quantity += quantity
                        branch_transfer.quantity -= quantity
                        branch_transfer.description = f'Received {branch_transfer.received_quantity} x {branch_transfer.quantity}'
                        branch_transfer.over_less = False
                        branch_transfer.over_less_description = 'received'
                        branch_transfer.save()

                        ActivityLog.objects.create(
                            branch = request.user.branch,
                            user=request.user,
                            action= 'stock in',
                            inventory=product,
                            quantity=quantity,
                            total_quantity=product.quantity,
                            product_transfer=branch_transfer,
                            description=f'Received from {transfer.transfer_ref}'
                        )

                        logger.info(f'Product received: {product.name}')

                        return JsonResponse({'success':True}, status=200)
                    
                    if action == 'send_back':

                        """
                            creation of a new transfer to the source branch with a product returned and
                            deduction of a the quantity transfered on the current transfer
                        """

                        # _adjust the transfer quantity of the currrent branch

                        branch_transfer.quantity -= int(quantity)
                        branch_transfer.description = f'Received {branch_transfer.quantity} and returned back {quantity}'
                        branch_transfer.received = True

                        receive_poduct = Inventory.objects.get(id=branch_transfer.product.id, branch=request.user.branch)

                        receive_poduct.quantity += int(quantity)

                        receive_poduct.save()
                        branch_transfer.save()

                        data = [
                            {
                                'transfer_id':branch_transfer.id,
                                'description':branch_transfer.description,
                                'status':branch_transfer.received 
                            }
                        ]
                    
                        return JsonResponse({'success':True, 'data':data}, status=200)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)

    return render(request, 'over_less_transfers.html', {'over_less_transfers':transfers, 'form':form})

@login_required
@transaction.atomic
def defective_product_list(request):
    form = RestockForm()
    defective_products = DefectiveProduct.objects.filter(branch=request.user.branch)
    
    # loss calculation
    if request.method == 'POST':
        data = json.loads(request.body)
        
        defective_id = data['product_id']
        quantity = data['quantity']
        
        try:
            d_product = DefectiveProduct.objects.get(id=defective_id, branch=request.user.branch)
            product = Inventory.objects.get(product__id=d_product.product.id, branch=request.user.branch)
        except:
            return JsonResponse({'success': False, 'message':'Product doesnt exists'}, status=400)
    
        product.quantity += quantity
        product.status = True if product.status == False else product.status
        product.save()
        
        d_product.quantity -= quantity
        d_product.save()
        
        ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= 'stock in',
            inventory=product,
            quantity=quantity,
            total_quantity=product.quantity,
            description = 'from defective products'
        )
        return JsonResponse({'success':True}, status=200)
    
    quantity = defective_products.aggregate(Sum('quantity'))['quantity__sum'] or 0
    price = defective_products.aggregate(Sum('product__cost'))['product__cost__sum'] or 0
    
    return render(request, 'defective_products.html', 
        {
            'total_cost': quantity * price,
            'defective_products':defective_products,
            'form':form,
        }
    )
    
@login_required
@transaction.atomic
def create_defective_product(request):
    form = AddDefectiveForm()
    if request.method == 'POST':
        form = AddDefectiveForm(request.POST)

        if form.is_valid():
            branch = request.user.branch
            product = form.cleaned_data['product']
            quantity = form.cleaned_data['quantity']
            
            # validation
            if quantity > product.quantity:
                messages.warning(request, 'Defective quantity cannot more than the products quantity')
                return redirect('inventory:create_defective_product')
            elif quantity == 0:
                messages.warning(request, 'Defective quantity cannot be less than zero')
                return redirect('inventory:create_defective_product')
            
            product.quantity -= quantity
            product.save()
        
            d_obj = form.save(commit=False)
            d_obj.branch = branch
            d_obj.branch_loss = branch
            
            d_obj.save()
            
            ActivityLog.objects.create(
                branch = branch,
                user=request.user,
                action= 'defective',
                inventory=product,
                quantity=quantity,
                total_quantity=product.quantity,
                description = ''
            )
            messages.success(request, 'Product successfuly saved')
            return redirect('inventory:defective_product_list')
        else:
            messages.success(request, 'Invalid form data')
    return render(request, 'add_defective_product.html', {'form':form})
        

@login_required
def add_inventory_transfer(request):
    form = addTransferForm()
    inventory = Inventory.objects.filter(branch=request.user.branch).order_by('-quantity')
    return render(request, 'add_transfer.html', {'fornm':form, 'inventory':inventory})

@login_required
@admin_required
def delete_inventory(request):
    product_id = request.GET.get('product_id', '')
    
    if product_id:
        inv = Inventory.objects.get(id=product_id, branch=request.user.branch)
        inv.status = False
        inv.save()
    
    ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= 'deactivated',
            inventory=inv,
            quantity=inv.quantity,
            total_quantity=inv.quantity
        )
    
    messages.success(request, 'Product successfully deleted')
    return redirect('inventory:inventory')

@login_required
# @admin_required
def add_product_category(request):
    
    if request.method == 'GET': 
        categories = ProductCategory.objects.all().values()
        logger.info(categories)
        return JsonResponse(list(categories), safe=False)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        category_name = data['name']
        
        if ProductCategory.objects.filter(name=category_name).exists():
            return JsonResponse({'success':False, 'message':'Category Exists'})
        
        category = ProductCategory.objects.create(
            name=category_name
        )
        return JsonResponse({'success':True, 'id': category.id, 'name':category.name})        

@login_required
def reorder_list(request):
    reorder_list = ReorderList.objects.filter(branch=request.user.branch)
        
    if request.method == 'GET':
        if 'download' in request.GET:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={request.user.branch.name} order.xlsx'
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            row_offset = 0
            
            worksheet['A' + str(row_offset + 1)] = f'Order List'
            worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
            cell = worksheet['A' + str(row_offset + 1)]
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(size=14, bold=True)
            cell.fill = PatternFill(fgColor='AAAAAA', fill_type='solid')

            row_offset += 1 
                
            category_headers = ['Name', 'Quantity']
            for col_num, header_title in enumerate(category_headers, start=1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = header_title
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            categories = reorder_list.values_list('product__product__category__name', flat=True).distinct()
            for category in categories:
                products_in_category = reorder_list.filter(product__product__category__name=category)
                if products_in_category.exists():
                    worksheet['A' + str(row_offset + 1)] = category
                    cell = worksheet['A' + str(row_offset + 1)]
                    cell.font = Font(color='FFFFFF')
                    cell.fill = PatternFill(fgColor='0066CC', fill_type='solid')
                    worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
                    row_offset += 2

                for product in reorder_list:
                    if category == product.product.product.category.name:
                        worksheet.append([product.product.product.name])
                        row_offset += 1

            workbook.save(response)
            return response
        return render(request, 'reorder_list.html', {'form':ReorderSettingsForm()})
        
@login_required
def reorder_list_json(request):
    order_list = ReorderList.objects.filter(branch=request.user.branch).values(
        'id', 
        'product__product__name',  
        'product__quantity',
        'quantity'
    )
    return JsonResponse(list(order_list), safe=False)

@login_required
@transaction.atomic
def clear_reorder_list(request):
    if request.method == 'GET':
        reorders = ReorderList.objects.filter(branch=request.user.branch)
        
        for item in reorders:
            inventory_items = Inventory.objects.filter(id=item.product.id)
            for product in inventory_items:
                product.reorder = False
                product.save()
            
        reorders.delete()
        
        messages.success(request, 'Reoder list success fully cleared')
        return redirect('inventory:reorder_list')

    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data['product_id']
    
        product = ReorderList.objects.get(id=product_id, branch=request.user.branch)
     
        inventory = Inventory.objects.get(id=product.product.id)
    
        product.delete()
        
        inventory.reorder=False
        inventory.save()
        
        return JsonResponse({'success':True}, status=200)
    
        
@login_required
@transaction.atomic
def create_order_list(request):
    if request.method == 'GET':
        products_below_five = Inventory.objects.filter(branch=request.user.branch, quantity__lte = 5).values(
            'id', 
            'product__id', 
            'product__name',
            'product__description',
            'quantity', 
            'reorder',
            'product__category__id',
            'product__category__name'
        )
        return JsonResponse(list(products_below_five), safe=False)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data['id']
        
        product = get_object_or_404(Inventory, id=product_id, branch=request.user.branch)
        ReorderList.objects.create(product=product, branch=request.user.branch)
        
        product.reorder = True
        product.save()
        return JsonResponse({'success': True}, status=201)
    return JsonResponse({'success': False, 'message':'Failed to add the product'}, status=400)
        
    

@login_required
def inventory_pdf(request):
    template_name = 'reports/inventory_pdf.html'
    category = request.GET.get('category', '')

    inventory = get_list_or_404(Inventory, branch=request.user.branch.id, product__category__name=category) if category else get_list_or_404(Inventory, branch=request.user.branch.id)
    title = f'{category} Inventory' if category else 'All Inventory'
    
    totals = calculate_inventory_totals(inventory)
    
    return generate_pdf(
        template_name, {
            'inventory':inventory,
            'title': f'{request.user.branch.name}: {title}',
            'date':datetime.date.today(),
            'total_cost':totals[0],
            'total_price':totals[1],
            'pdf_name':'Inventory'
        }
    )

@login_required
def transfers_report(request):
    
    template_name = 'reports/transfers.html'

    view = request.GET.get('view', '')
    choice = request.GET.get('type', '') 
    time_frame = request.GET.get('timeFrame', '')
    branch_id = request.GET.get('branch', '')
    product_id = request.GET.get('product', '')
    transfer_id = request.GET.get('transfer_id', '')
    
    transfers = TransferItems.objects.filter().order_by('-date') 
    
    today = datetime.date.today()
    
    if choice in ['All', '', 'Over/Less']:
        transfers = transfers
    
    if product_id:
        transfers = transfers.filter(product__id=product_id)
    if branch_id:
        transfers = transfers.filter(to_branch_id=branch_id)
    
    def filter_by_date_range(start_date, end_date):
        return transfers.filter(date__range=[start_date, end_date])
    
    date_filters = {
        'All': lambda: transfers, 
        'today': lambda: filter_by_date_range(today, today),
        'yesterday': lambda: filter_by_date_range(today - timedelta(days=1), today - timedelta(days=1)),
        'this week': lambda: filter_by_date_range(today - timedelta(days=today.weekday()), today),
        'this month': lambda: transfers.filter(date__month=today.month, issue_date__year=today.year),
        'this year': lambda: transfers.filter(date__year=today.year),
    }
    
    
    if time_frame in date_filters:
        transfers = date_filters[time_frame]()
         
    if view:
        return JsonResponse(list(transfers.values(
                'date',
                'product__name', 
                'price',
                'quantity', 
                'from_branch__name',
                'from_branch__id',
                'to_branch__id',
                'to_branch__name',
                'received_by__username',
                'date_received',
                'description',
                'received',
                'declined'
            )), 
            safe=False
        )
    
    if transfer_id:
        
        return JsonResponse(list(transfers.filter(id=transfer_id).values(
                'date',
                'product__name', 
                'price',
                'quantity', 
                'from_branch__name',
                'from_branch__id',
                'to_branch__id',
                'to_branch__name',
                'received_by__username',
                'date_received',
                'description',
                'received',
                'declined'
            )), 
            safe=False
        )
    
    return generate_pdf(
        template_name,
        {
            'title': 'Transfers', 
            'date_range': time_frame if time_frame else 'All',
            'report_date': datetime.date.today(),
            'transfers':transfers
        },
    )


@login_required
def reorder_from_notifications(request):
    if request.method == 'GET':
        notifications = StockNotifications.objects.filter(inventory__branch=request.user.branch, inventory__reorder=False, inventory__alert_notification=False).values(
            'quantity',
            'inventory__product__name', 
            'inventory__id', 
            'inventory__quantity' 
        )
        return JsonResponse(list(notifications), safe=False)
    
    if request.method == 'POST':
        # payload
        """
            inventory_id
        """
        
        data = json.loads(request.body)
        
        inventory_id = data['inventory_id']
        action_type = data['action_type']
        
        try:
            inventory = Inventory.objects.get(id=inventory_id)
            stock_notis = StockNotifications.objects.get(inventory=inventory)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        
        if action_type == 'add':
            inventory.reorder=True
            inventory.save()
            ReorderList.objects.create(
                quantity=0,
                product=inventory, 
                branch=request.user.branch
            )
            
        elif action_type == 'remove':
            inventory.alert_notification=True
            inventory.save()
    
        return JsonResponse({'success':True}, status=200)
    
    return JsonResponse({'success':False, 'message': 'Invalid request'}, status=400)

@login_required
def add_reorder_quantity(request):
    """
    Handles adding a quantity to a reorder item.

    Payload:
    - reorder_id: ID of the reorder item
    - quantity: Quantity to add
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

    reorder_id = data.get('reorder_id')
    reorder_quantity = data.get('quantity')

    if not reorder_id:
        return JsonResponse({'success': False, 'message': 'Reorder ID is required'}, status=400)

    if not reorder_quantity:
        return JsonResponse({'success': False, 'message': 'Reorder quantity is required'}, status=400)

    try:
        reorder_quantity = int(reorder_quantity)
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid reorder quantity'}, status=400)

    try:
        reorder = ReorderList.objects.get(id=reorder_id)
        reorder.quantity = reorder_quantity
        reorder.save()
        logger.info(reorder.quantity)
        return JsonResponse({'success': True, 'message': 'Reorder quantity updated successfully'}, status=200)
    except ReorderList.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Reorder item not found'}, status=404)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return JsonResponse({'success': False, 'message': 'An error occurred'}, status=500)


@login_required
def purchase_orders(request):
    form = CreateOrderForm()
    status_form = PurchaseOrderStatusForm()
    orders = PurchaseOrder.objects.filter(branch = request.user.branch).order_by('-order_date')
    orders_received = orders.filter(status='received')
    orders_on_hold = orders.filter(hold=True)
    orders_not_received = orders.exclude(status='received').exclude(hold=True)

    context = {
        'orders': orders,
        'orders_received': orders_received,
        'orders_on_hold': orders_on_hold,
        'orders_not_received': orders_not_received,
    }
   
    return render(request, 'purchase_orders.html', context)



@login_required
def create_purchase_order(request):
    if request.method == 'GET':
        supplier_form = AddSupplierForm()
        product_form = AddProductForm()
        suppliers = Supplier.objects.all()
        note_form = noteStatusForm()
        batch_form = BatchForm()

        products = Inventory.objects.filter(
            branch=request.user.branch,
            status=True,
            disable=False
        ).order_by('name')

        batch_codes = BatchCode.objects.all()

        return render(request, 'create_purchase_order.html', {
            'product_form': product_form,
            'supplier_form': supplier_form,
            'suppliers': suppliers,
            'note_form': note_form,
            'batch_form': batch_form,
            'batch_codes': batch_codes,
            'products': products
        })

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            purchase_order_data = data.get('purchase_order', {})
            purchase_order_items_data = data.get('po_items', [])
            expenses = data.get('expenses', [])
            cost_allocations = data.get('cost_allocations', [])
            hold = data.get('hold', False)
            supplier_payment_data = data.get('supplier_data')
            overide = data.get('overide')

            unique_expenses = []
            seen = set()
            for expense in expenses:
                expense_tuple = (expense['name'], expense['amount'])
                if expense_tuple not in seen:
                    seen.add(expense_tuple)
                    unique_expenses.append(expense)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

        batch = purchase_order_data['batch']
        delivery_date = purchase_order_data['delivery_date']
        status = purchase_order_data['status']
        notes = purchase_order_data['notes']
        total_cost = Decimal(purchase_order_data['total_cost'])
        discount = Decimal(purchase_order_data['discount'])
        tax_amount = Decimal(purchase_order_data['tax_amount'])
        other_amount = Decimal(purchase_order_data['other_amount'])
        payment_method = purchase_order_data.get('payment_method')

        if not all([batch, delivery_date, status, total_cost, payment_method]):
            return JsonResponse({'success': False, 'message': 'Missing required fields'}, status=400)

        try:
            with transaction.atomic():
                purchase_order = PurchaseOrder.objects.create(
                    order_number=PurchaseOrder.generate_order_number(),
                    batch=batch,
                    delivery_date=delivery_date,
                    status=status,
                    notes=notes,
                    total_cost=total_cost,
                    discount=discount,
                    tax_amount=tax_amount,
                    other_amount=other_amount,
                    branch=request.user.branch,
                    is_partial=False,
                    received=False,
                    hold=hold
                )

                purchase_order_items = []
                for item_data in purchase_order_items_data:
                    product_id = item_data['product_id']
                    product_name = item_data['product']
                    quantity = int(item_data['quantity'])
                    unit_cost = Decimal(item_data['price'])
                    actual_unit_cost = Decimal(item_data['price'])
                    supplier_id = item_data.get('supplier', [])

                    if not all([product_name, quantity, unit_cost, product_id]):
                        transaction.set_rollback(True)
                        return JsonResponse({'success': False, 'message': 'Missing fields in item data'}, status=400)

                    try:
                        from_other_branch = Inventory.objects.get(id=product_id)

                        existing_product = Inventory.objects.filter(
                            name=from_other_branch.name,
                            branch=request.user.branch,
                            category=from_other_branch.category,
                            disable=False
                        ).first()

                        if existing_product:
                            product = existing_product
                        else:
                            # Clone product
                            cloned_product = Inventory.objects.create(
                                branch=request.user.branch,
                                name=from_other_branch.name,
                                cost=from_other_branch.cost,
                                price=from_other_branch.price,
                                dealer_price=from_other_branch.dealer_price,
                                quantity=0,
                                status=from_other_branch.status,
                                stock_level_threshold=from_other_branch.stock_level_threshold,
                                reorder=from_other_branch.reorder,
                                alert_notification=from_other_branch.alert_notification,
                                batch=from_other_branch.batch,
                                category=from_other_branch.category,
                                tax_type=from_other_branch.tax_type,
                                description=from_other_branch.description,
                                end_of_day=from_other_branch.end_of_day,
                                service=from_other_branch.service,
                                image=from_other_branch.image,
                                disable=from_other_branch.disable,
                            )

                            cloned_product.suppliers.set(from_other_branch.suppliers.all())
                            product = cloned_product

                        # Use supplier logic
                        supplier = Supplier.objects.get(name__icontains="local")

                        # Create PO item
                        po_item = PurchaseOrderItem.objects.create(
                            purchase_order=purchase_order,
                            product=product,
                            quantity=quantity,
                            unit_cost=actual_unit_cost if overide == 'manual' else unit_cost,
                            actual_unit_cost=actual_unit_cost,
                            received_quantity=0,
                            received=False,
                            supplier=supplier,
                            price=0,
                            wholesale_price=0
                        )

                        purchase_order_items.append(po_item)
                        product.suppliers.add(po_item.supplier.id)
                        product.batch += f'{batch}, '
                        product.price = 0
                        product.save()

                    except Inventory.DoesNotExist:
                        transaction.set_rollback(True)
                        return JsonResponse({'success': False, 'message': f'Product with ID {product_id} not found.'}, status=404)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

        return JsonResponse({'success': True, 'message': 'Purchase order created successfully'})

@login_required
def confirm_purchase_order_items(request, po_id):
    try:
        order_items = PurchaseOrderItem.objects.filter(purchase_order__id=po_id).select_related('product')

        logs = []
        inventory_updates = []

        with transaction.atomic():
            for item in order_items:
                print(item.received_quantity)
                print(item.product)
                inventory_updates.append(
                    Inventory(
                        id=item.product.id,
                        quantity=item.product.quantity + item.received_quantity,
                        price=item.price,
                        dealer_price=item.wholesale_price,
                        cost=item.actual_unit_cost,
                        status=True,
                        disable=False,
                    )
                )

                existing_quantity = item.product.quantity 
                new_quantity = item.received_quantity + item.product.quantity

                quantity_change = abs(new_quantity -  existing_quantity)

                print(f' bvbproduct quantity {item.product.quantity}')
                print(f'update product quantity {item.product.quantity + item.received_quantity}')

                logs.append(
                    ActivityLog(
                        purchase_order=item.purchase_order,
                        branch=request.user.branch,
                        user=request.user,
                        action='stock in',
                        dealer_price=item.wholesale_price,
                        selling_price=item.price,
                        inventory=item.product,
                        quantity=quantity_change,
                        system_quantity=item.received_quantity,  
                        description=f'Stock in from batch {item.purchase_order.batch}',
                        total_quantity=item.product.quantity + item.received_quantity,
                    )
                )

            Inventory.objects.bulk_update(
                inventory_updates, ['quantity', 'price', 'dealer_price', 'cost']
            )

            ActivityLog.objects.bulk_create(logs)

        return JsonResponse({'success': True, 'message': 'All purchase order items processed'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

       
@login_required    
def if_purchase_order_is_received(request, purchase_order, tax_amount, payment_method):
    try:
        currency = Currency.objects.get(default=True)
        rate = VATRate.objects.get(status=True)

        # get account 
        account_details = account_identifier(request, currency, payment_method)
        account_name = account_details['account_name']
        account_type = account_details['account_type']

        account, _ = Account.objects.get_or_create(
            name=account_name,
            type=account_type
        )
        
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,

            defaults={
                'branch':request.user.branch,
                'currency':currency,
                'balance':0
            }
        )

        account_balance.balance -= purchase_order.total_cost
        
        account_balance.save()

        # get or create purchase order category
        category, _ = ExpenseCategory.objects.get_or_create(name='Purchase orders')
        logger.info(category)

        # create an expense and exclude the vat amount
        expense = Expense.objects.create(
            amount = purchase_order.total_cost - purchase_order.tax_amount,
            payment_method = payment_method,
            currency = currency, 
            category = category,
            user = request.user,
            branch = request.user.branch,
            status = False,
            purchase_order = purchase_order,
            description = f'Purchase order: {purchase_order.order_number}',
        )

        # create a cashbook entry
        Cashbook.objects.create(
            expense = expense,
            description = f'Purchase order: {purchase_order.order_number}',
            credit = True,
            amount = purchase_order.total_cost,
            currency = currency,
            branch = request.user.branch
        )

        # create account transaction log
        AccountTransaction.objects.create(
            account = account,
            expense = expense
        )

        # revisit
        # PurchaseOrderAccount.objects.create(
        #     purchase_order = purchase_order,
        #     amount = purchase_order.total_cost - purchase_order.tax_amount,
        #     balance = 0,
        #     expensed = False
        # )

        # create a vat entry
        VATTransaction.objects.create(
            purchase_order = purchase_order,
            vat_type='Input',
            vat_rate = rate.rate,
            tax_amount = tax_amount
        )

    except Exception as e:
        logger.info(e)
        return JsonResponse({'success':False, 'message':f'currency doesnt exists'})
    except VATRate.DoesNotExist:
        return JsonResponse({'success':False, 'message':f'Make sure you have a stipulated vat rate in the system'})

@login_required
def delete_purchase_order(request, purchase_order_id):
    try:
        # Retrieve the purchase order
        purchase_order = PurchaseOrder.objects.get(id=purchase_order_id)

        if purchase_order.received:
            return JsonResponse({'success': False, 'message': 'Cannot delete a received purchase order'}, status=400)

        with transaction.atomic():
            
            # Reverse related account transactions
            currency = Currency.objects.get(default=True)

            account_transaction = AccountTransaction.objects.filter(expense__purchase_order=purchase_order).first()
            if account_transaction:
                account_balance = AccountBalance.objects.get(account=account_transaction.account)

                # Reverse account balance adjustments
                account_balance.balance += purchase_order.total_cost
                account_balance.save()

                # Delete account transaction log
                account_transaction.delete()

            # Reverse Cashbook entry
            cashbook_entry = Cashbook.objects.filter(expense__purchase_order=purchase_order).first()
            if cashbook_entry:
                cashbook_entry.delete()

            # Reverse the VAT transaction
            vat_transaction = VATTransaction.objects.filter(purchase_order=purchase_order).first()
            if vat_transaction:
                vat_transaction.delete()

            # Reverse the Expense record
            expense = Expense.objects.filter(purchase_order=purchase_order).first()
            if expense:
                expense.delete()

            # Reverse other expenses related to the purchase order
            other_expenses = otherExpenses.objects.filter(purchase_order=purchase_order)
            if other_expenses.exists():
                other_expenses.delete()


            items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
            products = Inventory.objects.all()

            for item in items:
                for prod in products:
                    if item.product == prod.product:
                        prod.quantity -= item.received_quantity

                        # eliminate negative stock
                        if prod.quantity < 0:
                            prod.quantity = 0

                        prod.save()
                        
                        ActivityLog.objects.create(
                            branch = request.user.branch,
                            user= request.user,
                            action= 'delete',
                            inventory=prod,
                            quantity=item.received_quantity,
                            total_quantity=prod.quantity
                        )

            # Remove PurchaseOrderItems
            PurchaseOrderItem.objects.filter(purchase_order=purchase_order).delete()

            # Finally, delete the purchase order itself
            purchase_order.delete()

        return JsonResponse({'success': True, 'message': 'Purchase order deleted successfully'})
    except PurchaseOrder.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Purchase order not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
@login_required    
def download_delivery_note(request, po_id):
    delivery_note = get_object_or_404(DeliveryNote, purchase_order_id=po_id)

    if delivery_note.pdf:
        response = HttpResponse(delivery_note.pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Delivery_Note_{po_id}.pdf"'
        return response
    else:
        return HttpResponse("No PDF found for this delivery note.", status=404)

@login_required
@transaction.atomic
def change_purchase_order_status(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
    except PurchaseOrder.DoesNotExist:
        return JsonResponse({'error': f'Purchase order with ID: {order_id} does not exist'}, status=404)

    try:
        data = json.loads(request.body)
        status = data['status']
        
        if status:
            purchase_order.status=status

            with transaction.atomic():
                if purchase_order.status == 'received':
                    purchase_order.save()

                    tax_amount = purchase_order.tax_amount
                    payment_method = purchase_order.payment_method

                    if_purchase_order_is_received(
                        request, 
                        purchase_order, 
                        tax_amount,
                        payment_method
                    )
            
            return JsonResponse({'success':True}, status=200)
        else:
            return JsonResponse({'success':False, 'message':'Status is required'}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

@login_required
def print_purchase_order(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
    except PurchaseOrder.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    try:
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
    except PurchaseOrderItem.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    return render(request, 'print_purchase_order.html', 
        {
            'orders':purchase_order_items,
            'purchase_order':purchase_order
        }
    )

@login_required
def purchase_order_detail(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
    except PurchaseOrder.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exist')
        return redirect('inventory:purchase_orders')

    items = costAllocationPurchaseOrder.objects.filter(purchase_order=purchase_order)
    expenses = otherExpenses.objects.filter(purchase_order=purchase_order)
    purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)

    total_received_quantity = purchase_order_items.aggregate(Sum('received_quantity'))['received_quantity__sum'] or 0
    total_expected_profit = purchase_order_items.aggregate(Sum('expected_profit'))['expected_profit__sum'] or 0
    total_expected_dealer_profit = purchase_order_items.aggregate(Sum('dealer_expected_profit'))['dealer_expected_profit__sum'] or 0
    total_quantity = items.aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_expense_sum = expenses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0
    
    logger.info(f'total quantity = { total_received_quantity}')
    

    if request.GET.get('download') == 'csv':
        return generate_csv_response(items, purchase_order_items)

    return render(request, 'purchase_order_detail.html', {
        'items': items,
        'expenses': expenses,
        'order_items': purchase_order_items,  
        'total_quantity': total_received_quantity,# to be changed to total quantity,
        'total_received_quantity': total_received_quantity,
        'total_retail_profit': total_expected_profit,
        'total_expenses': total_expense_sum,
        'purchase_order': purchase_order,
        'total_wholesale_profit':total_expected_dealer_profit
    })

def generate_csv_response(items, po_items):
    """Generate CSV response for the purchase order"""
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="purchase_order.csv"'

    writer = csv.writer(response)

    writer.writerow(['Product', 'Quantity', 'Quantity Received', 'Selling Price', 'Dealer Price'])

    for item in items:
        received_quantity = 0
        if po_items.filter(product__name=item.product).exists():
           received_quantity = po_items.filter(product__name=item.product).first().received_quantity

        writer.writerow([
            item.product,
            item.quantity,
            received_quantity,
            f"${item.selling_price:.2f}",
            f"${item.dealer_price:.2f}",
        ])

    return response

@login_required
def sales_price_list_pdf(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
        
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
            
        context = {'items': purchase_order_items}

        template = get_template('pdf_templates/price_list.html')
        html = template.render(context)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="price_list.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)

        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=400)
        
        return response
    except PurchaseOrder.DoesNotExist:
            messages.warning(request, f'Purchase order with ID: {order_id} does not exist')
            return redirect('inventory:purchase_orders')

@login_required
def delete_purchase_order(request, purchase_order_id):
    if request.method != "DELETE":
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

    try:
        purchase_order = PurchaseOrder.objects.get(id=purchase_order_id)
    except PurchaseOrder.DoesNotExist:
        return JsonResponse({'success': False, 'message': f'Purchase order with ID {purchase_order_id} not found'}, status=404)

    try:
        purchase_order.delete()
        return JsonResponse({'success': True, 'message': 'Purchase order deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
def receive_order(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
    except PurchaseOrder.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    try:
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
    except PurchaseOrderItem.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    logger.info(purchase_order_items.all().values('product'))

    products = Inventory.objects.filter(branch=request.user.branch).values(
        'dealer_price', 
        'price', 
        'name',
        'id',
        'branch__name'
    )
    logger.info(f'branch: {request.user.branch}')
    # Convert products queryset to a dictionary for easy lookup by product ID
    product_prices = {product['id']: product for product in products}

    new_po_items =  []
    for item in purchase_order_items:
        if(item.product):
            logger.info(item.product.branch)
            product_name = item.product.id

            product_data = product_prices.get(product_name)
            logger.info(f'product prices {product_data}')
            if product_data:
                item.dealer_price = product_data['dealer_price']
                item.selling_price = product_data['price']
            else:
                item.dealer_price = 0  
                item.selling_price = 0 
            new_po_items.append(item)

    logger.info(purchase_order_items)

    
    return render(request, 'receive_order.html', 
        {
            'orders':purchase_order_items,
            'purchase_order':purchase_order
        }
    )

@transaction.atomic
def process_received_order(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            edit = data.get('edit')
            order_item_id = data.get('id')
            quantity = data.get('quantity', 0)
            selling_price = data.get('selling_price', 0)
            dealer_price = data.get('dealer_price', 0)
            expected_profit = data.get('expected_profit', 0)
            dealer_expected_profit = data.get('dealer_expected_profit', 0)

            logger.info('Processing order item')
            logger.info(data)
            
            order_item = PurchaseOrderItem.objects.get(id=order_item_id)
            cost = order_item.actual_unit_cost

            if edit:
                return edit_purchase_order_item(order_item_id, selling_price, dealer_price, expected_profit, dealer_expected_profit, quantity, cost, request)
            
            logger.info(order_item)
            order = PurchaseOrder.objects.get(id=order_item.purchase_order.id)

            # Update the order item with received quantity
            order_item.receive_items(quantity)
            order_item.received_quantity = quantity
            order_item.expected_profit = expected_profit
            order_item.dealer_expected_profit = dealer_expected_profit
            order_item.price = selling_price
            order_item.wholesale_pice = dealer_price
            order_item.received = True
            
            # Update or create inventory
            system_quantity = 0 # if new product
            try:
                
                with transaction.atomic():
                    inventory = Inventory.objects.get(id = order_item.product.id)
                    logger.info(inventory)
                    
                    system_quantity = inventory.quantity
                    
                    inventory.price = selling_price
                    inventory.dealer_price = dealer_price or 0
                    inventory.quantity += quantity

                    if inventory.batch:
                        inventory.batch += f'{order.batch}, '
                    else:
                        inventory.batch = f'{order.batch}, '      

                    cost = average_inventory_cost(inventory.id, order_item.actual_unit_cost, quantity, request.user.branch.id)
                    logger.info(order_item.actual_unit_cost)

                    inventory.cost = Decimal(round(cost, 2))
                    
                    logger.info(f'Inventory cost: {inventory.cost}')
                    
                    inventory.save()
                    
                    log = ActivityLog(
                        purchase_order=order_item.purchase_order,
                        branch=request.user.branch,
                        user=request.user,
                        action='stock in',
                        dealer_price = dealer_price,
                        selling_price = selling_price,
                        inventory=inventory,
                        quantity=quantity,
                        system_quantity=system_quantity,
                        description=f'Stock in from {order_item.purchase_order.batch}',
                        total_quantity=inventory.quantity
                    )
                    log.save()
                    
              
                    order_item.save()

                    logger.info(f'process order item received: {order_item.product}')

                    return JsonResponse({'success': True, 'message': 'Inventory updated successfully'}, status=200)
            except Product.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'Product with ID: {order_item.product.id} does not exist'}, status=404)
        except Exception as e:
            logger.info(e)
            return JsonResponse({'success': False, 'message': f'{e}'}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)



def edit_purchase_order_item(order_item_id, selling_price, dealer_price, expected_profit, dealer_expected_profit, quantity, cost, request):
    try:
        po_item = PurchaseOrderItem.objects.get(id=order_item_id)
        product = po_item.product.id
        try:
            inventory = Inventory.objects.get(id=product, branch=po_item.purchase_order.branch)
            system_quantity = inventory.quantity
            quantity_adjustment = 0
            description = ''
            if po_item.quantity == quantity:
                if inventory.quantity != quantity:
                    # adjust quantity
                    if inventory.quantity < quantity:
                        quantity_adjustment = quantity - inventory.quantity 
                        inventory.quantity += quantity_adjustment
                        action = 'purchase edit +'
                        description = f'Stock adjustment ({po_item.purchase_order.batch})'
                        logger.info(f'{quantity_adjustment}: quantity adjusted')
                    elif inventory.quantity > quantity:
                        quantity_adjustment = inventory.quantity - quantity 
                        inventory.quantity = quantity
                        # quantity_adjustment = quantity
                        action = 'purchase edit -'
                        description = f'Stock adjustment ({po_item.purchase_order.batch})'
                        logger.info(f'{quantity_adjustment}: quantity adjusted')
                    else:
                        action = 'price edit'
                        logger.info(f'{quantity_adjustment}: quantity adjusted')
                        inventory.quantity = quantity
                        description = f'Price adjustment ({po_item.purchase_order.batch})'
                else:
                    action='price edit'
                    description=f'Price adjustment ({po_item.purchase_order.batch})'
            else:
                # adjust quantity
                if inventory.quantity < quantity:
                    quantity_adjustment = quantity - inventory.quantity 
                    inventory.quantity += quantity_adjustment
                    action = 'purchase edit +'
                    description = f'Stock adjustment ({po_item.purchase_order.batch})'
                    logger.info(f'{quantity_adjustment}: quantity adjusted')
                elif inventory.quantity > quantity:
                    quantity_adjustment = inventory.quantity - quantity 
                    inventory.quantity = quantity
                    # quantity_adjustment = quantity
                    action = 'purchase edit -'
                    description = f'Stock adjustment ({po_item.purchase_order.batch})'
                    logger.info(f'{quantity_adjustment}: quantity adjusted')
                else:
                    action = 'price edit'
                    logger.info(f'{quantity_adjustment}: quantity adjusted')
                    inventory.quantity = quantity
                    description = f'Price adjustment ({po_item.purchase_order.batch})'
            
            # Update fields in the PurchaseOrderItem
            with transaction.atomic():
                po_item.selling_price = selling_price
                po_item.dealer_price = dealer_price
                po_item.expected_profit = expected_profit
                po_item.dealer_expected_profit = dealer_expected_profit
                po_item.received_quantity = quantity
                po_item.save()

                inventory.price = selling_price
                inventory.dealer_price = dealer_price
                
                # cost = average_inventory_cost(inventory.id, po_item.unit_cost, quantity, request.user.branch.id)
                # logger.info(po_item.actual_unit_cost)
                
                inventory.cost = cost
                inventory.save()

                ActivityLog.objects.create(
                    purchase_order=po_item.purchase_order,
                    branch=request.user.branch,
                    user=request.user,
                    action=action,
                    inventory=inventory,
                    quantity=quantity_adjustment,
                    system_quantity = system_quantity,
                    description=description,
                    total_quantity=inventory.quantity,
                    dealer_price = dealer_price,
                    selling_price = selling_price
                )

        except Inventory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Inventory not found for the product'}, status=404)

        logger.info('done')
        return JsonResponse({'success': True, 'message': 'Purchase Order Item updated successfully'}, status=200)
    
    except PurchaseOrderItem.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Purchase Order Item not found'}, status=404)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Product not found'}, status=404)
    
@login_required
def mark_purchase_order_done(request, po_id):
    if po_id:
        try:
            with transaction.atomic():
                purchase_order = PurchaseOrder.objects.select_for_update().get(id=po_id)
                purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
                
                if purchase_order.received:
                    return JsonResponse({'succes':False, 'message':'Purchase order already marked as received'})
                
                flag = True
                for item in purchase_order_items:
                    if not item.received:
                        flag = False
                        break
                    
                purchase_order.received = flag
                purchase_order.save()
                
                if flag:
                    return JsonResponse({'success':True, 'messages':'Purchase order has been successfully confirmed!'})
                else:
                    return JsonResponse({'success':False, 'message':'Purchase order not confirmed, Please receive all items.'})
        
        except PurchaseOrder.DoesNotExist:
            return JsonResponse({'success':False, 'message': f'Purchase order with ID {po_id} not found'})
        except Exception as e:
            logger.error(f'Error processing purchase order {po_id}: {str(e)}')
            return JsonResponse({'success':False, 'message':'An error occurred while processing the purchase order'})
    else:
        return JsonResponse({'success':False, 'message':'Purchase order ID is required'})

@login_required
def edit_purchase_order(request, po_id):
    if request.method == 'GET':
        purchase_order = PurchaseOrder.objects.get(id=po_id)
        supplier_form = AddSupplierForm()
        product_form = AddProductForm()
        suppliers = Supplier.objects.all()
        note_form = noteStatusForm()
        batch_form = BatchForm()

        batch_codes = BatchCode.objects.all()

        return render(request, 'edit_purchase_order.html', {
            'purchase_order':purchase_order,
            'product_form':product_form,
            'supplier_form':supplier_form,
            'suppliers':suppliers,
            'note_form':note_form,
            'batch_form':batch_form,
            'batch_codes':batch_codes,
         })
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            purchase_order_data = data.get('purchase_order', {})
            purchase_order_items_data = data.get('po_items', [])
            expenses = data.get('expenses', [])
            cost_allocations = data.get('cost_allocations', [])
            overide = data.get('overide')
            hold = data.get('hold')
            purchase_order_id = data.get('purchase_order_id')

            logger.info(purchase_order_items_data)

            # remove duplicates
            unique_expenses = []
            seen = set()
            for expense in expenses:
                expense_tuple = (expense['name'], expense['amount'])
                if expense_tuple not in seen:
                    seen.add(expense_tuple)
                    unique_expenses.append(expense)

            # get previous purchase_order
            last_purchase_order = PurchaseOrder.objects.get(id=po_id)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

        batch = purchase_order_data['batch']
        delivery_date = purchase_order_data['delivery_date']
        status = purchase_order_data['status']
        notes = purchase_order_data['notes']
        total_cost = Decimal(purchase_order_data['total_cost'])
        discount = Decimal(purchase_order_data['discount'])
        tax_amount = Decimal(purchase_order_data['tax_amount'])
        other_amount = Decimal(purchase_order_data['other_amount'])
        payment_method = purchase_order_data.get('payment_method')
        
        
        logger.info(f'hold: {hold}')
    
        if not all([delivery_date, status, total_cost, payment_method]):
            return JsonResponse({'success': False, 'message': 'Missing required fields'}, status=400)

        try:
            purchase_order = PurchaseOrder.objects.get(pk=purchase_order_id)
            with transaction.atomic():
                purchase_order.batch = batch
                purchase_order.delivery_date = delivery_date
                purchase_order.status = status
                purchase_order.order_date = purchase_order.order_date 
                purchase_order.notes = notes
                purchase_order.total_cost = total_cost
                purchase_order.discount = discount
                purchase_order.tax_amount = tax_amount
                purchase_order.other_amount = other_amount
                purchase_order.branch = request.user.branch
                purchase_order.is_partial = False
                purchase_order.received = False
                purchase_order.hold = hold
                purchase_order.save()
                
                logger.info(f'purchase order item: {purchase_order.total_cost}')
                

                products = Inventory.objects.filter(branch=request.user.branch)
                suppliers = Supplier.objects.all()
                logs = ActivityLog.objects.filter(purchase_order=last_purchase_order).select_related('inventory')

                products_dict = {product.id: product for product in products}
                suppliers_dict = {supplier.id: supplier for supplier in suppliers}
                logs_dict = {log.inventory_id: log.quantity for log in logs}
                
                logger.info(products_dict)

                supplier = Supplier.objects.get(id=1)
                
                logger.info(supplier)

                purchase_order_items_bulk = []
                existing_items = {item.product_id: item for item in PurchaseOrderItem.objects.filter(purchase_order=purchase_order)}

                for item_data in purchase_order_items_data:
                    product_id = int(item_data['product_id'])
                    product = products_dict.get(product_id)
                    log_quantity = logs_dict.get(product_id, 0)

                    if not product:
                        return JsonResponse({'success': False, 'message': 'Invalid product'}, status=400)

                    if product_id in existing_items:
                        # Update the existing item
                        existing_item = existing_items[product_id]
                        existing_item.quantity = item_data['quantity']
                        existing_item.unit_cost = item_data['price']
                        existing_item.actual_unit_cost = item_data['price']
                        existing_item.received_quantity = log_quantity
                        existing_item.supplier = supplier
                        existing_item.wholesale_price = 0
                        existing_item.received = False
                        existing_item.price = 0
                        existing_item.save()
                    else:
                        # Create a new item for bulk creation
                        purchase_order_items_bulk.append(
                            PurchaseOrderItem(
                                purchase_order=purchase_order,
                                product=product,
                                quantity=item_data['quantity'],
                                unit_cost=item_data['price'],
                                actual_unit_cost=item_data['price'],
                                received_quantity=log_quantity,
                                supplier=supplier,
                                wholesale_price=0,
                                received=False,
                                price=0,
                            )
                        )

                    product.price = 0
                    product.save()

                PurchaseOrderItem.objects.bulk_create(purchase_order_items_bulk)

                expense_bulk = []
                for expense in unique_expenses:
                    name = expense['name'] 
                    amount = expense['amount']
                    expense_bulk.append(
                        otherExpenses(
                            purchase_order=purchase_order,
                            name=name,
                            amount=amount
                        )
                    )
                otherExpenses.objects.bulk_create(expense_bulk)

                costs_list = []
                
                for cost in cost_allocations:
                    costs_list.append(
                        costAllocationPurchaseOrder(
                            purchase_order = purchase_order,
                            allocated = cost['allocated'],
                            allocationRate = cost['allocationRate'],
                            expense_cost = cost['expCost'],
                            price = cost['price'],
                            quantity = int(cost['quantity']),
                            product = cost['product'],
                            total = cost['total'],
                            total_buying = cost['totalBuying']
                        )
                    )
                costAllocationPurchaseOrder.objects.bulk_create(costs_list)
                    
                # if purchase_order.status in ['Received', 'received']:
                #     if_purchase_order_is_received(
                #         request, 
                #         purchase_order, 
                #         tax_amount, 
                #         payment_method
                #     )
                
                remove_purchase_order(po_id, request)
                
        except Exception as e:
            logger.info(e)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

        return JsonResponse({'success': True, 'message': 'Purchase order created successfully'})

def remove_purchase_order(purchase_order_id, request):
    try:
        # Retrieve the purchase order
        purchase_order = PurchaseOrder.objects.get(id=purchase_order_id)

        #logger.info(f'purchase_order: {purchase_order}')

        with transaction.atomic():
            
            # Reverse related account transactions
            currency = Currency.objects.filter(default=True).first()

            account_transaction = AccountTransaction.objects.filter(expense__purchase_order=purchase_order).first()
            if account_transaction:
                account_balance = AccountBalance.objects.get(account=account_transaction.account)

                # Reverse account balance adjustments
                account_balance.balance += purchase_order.total_cost
                account_balance.save()

                # Delete account transaction log
                account_transaction.delete()

            # Reverse Cashbook entry
            cashbook_entry = Cashbook.objects.filter(expense__purchase_order=purchase_order).first()
            if cashbook_entry:
                cashbook_entry.delete()

            # Reverse the VAT transaction
            vat_transaction = VATTransaction.objects.filter(purchase_order=purchase_order).first()
            if vat_transaction:
                vat_transaction.delete()

            # Reverse the Expense record
            expense = Expense.objects.filter(purchase_order=purchase_order).first()
            if expense:
                expense.delete()

            # Reverse other expenses related to the purchase order
            other_expenses = otherExpenses.objects.filter(purchase_order=purchase_order)
            if other_expenses.exists():
                other_expenses.delete()
            
            #deduct product quantity
            items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order, received=True)
            products = Inventory.objects.filter(branch=request.user.branch)

            logger.info(f'products in the branch {products}')

            for item in items:
                for prod in products:
                    if item.product == prod:
                        system_quantity = prod.quantity
                        prod.quantity -= item.received_quantity
                        
                        # eliminate negative stock
                        if prod.quantity < 0:
                            prod.quantity = 0

                        prod.save()
                        
                        ActivityLog.objects.create(
                            branch = request.user.branch,
                            user= request.user,
                            action= 'purchase edit -',
                            inventory=prod,
                            system_quantity=system_quantity,
                            quantity=item.received_quantity,
                            total_quantity=prod.quantity,
                            description=f'Purchase order: {item.purchase_order.batch} edit'
                        )
    except Exception as e:
        logger.info(e)

@login_required
def edit_purchase_order_data(request, po_id):
    try:
        expenses = otherExpenses.objects.filter(purchase_order__id=po_id).values()
    
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order__id=po_id).values(
            'purchase_order__id',
            'product__name',
            'product__id',
            'quantity',
            'unit_cost',
            'actual_unit_cost',
            'expected_profit',
            'supplier__name',
            'supplier'
        )

        return JsonResponse({'success':True, 'po_items':list(purchase_order_items), 'expenses':list(expenses)})

    except Exception as e:
        return JsonResponse({"success":False, 'message':f'{e}'})

#testing delete
@login_required
def supplier_delete(request, supplier_id):
    '''
        name,contact_person,email,product,address
    '''
    if request.method == 'GET':
        supplier = Supplier.objects.filter(id=supplier_id).values()
        logger.info(supplier)
        return JsonResponse({'success':True, 'data':list(supplier)})
         

    if request.method == "DELETE":
        try:
            supplier = Supplier.objects.get(id=supplier_id)
            supplier.delete = True
            supplier.save()

            return JsonResponse({'success':True}, status = 200)
        except Exception as e:
            logger.info(e)
            return JsonResponse({"success":False, "message":f"{e}"})
    return JsonResponse({"success":False, "message":"Invalid Request"})


#testing edit view
@login_required
def supplier_edit(request, supplier_id):
    if request.method == 'GET':
        supplier = Supplier.objects.filter(id=supplier_id).values()
        logger.info(supplier)
        return JsonResponse({'success':True, 'data':list(supplier)})
         
    if request.method == "POST":
         
        try:
            data = json.loads(request.body)
            logger.info(data)

            name = data.get('name')
            contact_person = data.get('contact_person')
            email = data.get('email')
            address = data.get('address')
            phone = data.get('phone')

            supplier = Supplier.objects.get(phone=phone)

            supplier.name=name
            supplier.contact_person=contact_person
            supplier.email=email
            supplier.phone=phone
            supplier.address=address
            
            supplier.save()
            logger.info(f'{supplier} saved')

            return JsonResponse({'success':True}, status = 200)
        except Exception as e:
            logger.info(e)
            return JsonResponse({"success":False, "message":f"{e}"})
    return JsonResponse({"success":False, "message":"Invalid Request"})

#payments
def supplier_payments(po, payment_data):
    #add payment
    data = PurchaseOrderItem.objects.filter(purchase_order=po).values('id', 'quantity', 'unit_cost', 'purchase_order__id', 'product__name', 'supplier_id')
    list_entries= []
    for item in data:
        item_id = item['id']
        quantity = item['quantity']
        unit_cost = item['unit_cost']

        for items in list_entries:
            t_amount = 0
            if items['id'] == item_id:
                amount = quantity * unit_cost
                t_amount = items['amount'] + amount
                items['amount'] =  t_amount
            else:
                amount = quantity * unit_cost
                items.append({'id': item_id, 'amount': amount})

    
        """
            [list_entries we have supplier id and total amount of goods he/she provided]
            [payment data supplier id, amount paid to the supplier, currency, payment method]   
            1. we want to create a payment for each supplier
            2. we want calculate the balance for each supplier and update the balance
            3. if the account the supplier doesnt exist we need to create it
        """

        for payment_info in payment_data:
            currency = Currency.objects.get(id=payment_info['currency'])
            supplier = Supplier.objects.get(id=payment_info['id'])

            account, _ = SupplierAccount.objects.get_or_create(
                supplier = supplier,
                defaults={
                    'currency':currency,
                    'balance':0
                }
            )

            SupplierAccountsPayments.objets.create(
                account=account,
                currency=currency,
                amount=payment_info['amount'],
                payment_method=payment_info['payment_method']
            ) 

            calucalateSupplierBalance(account, currency, payment_info['amount'])


        def calucalateSupplierBalance(account, currency, paid_amount):
            for supplier in list_entries:
                if account.supplier.id == supplier:
                    use_account = SupplierAccount.objets.get(currency=currency, account=account)
                    use_account.balance = supplier['amount'] - paid_amount
                

#Payment history
@login_required
def PaymentHistory(request, supplier_id):
    """
        order name
        purchase order amount
    """
    if request.method == 'GET':
        supplier_history = SupplierAccountsPayments.objects.filter(account__supplier_id = supplier_id).\
        values(
            'timestamp', 
            'amount',
            'account__balance',
            'user__username',
            'currency__name'
        )
        supplier_purchase_order_details = PurchaseOrderItem.objects.filter(supplier__id = supplier_id)
        list_details = {}
        for items in supplier_purchase_order_details:
            if items.purchase_order.order_number == list_details.get('order_number'):
                list_details['amount'] = items.quantity * items.unit_cost
            else:
                list_details['order_number'] = items.purchase_order.order_number
                list_details['amount']= items.quantity * items.unit_cost
        logger.info(list_details)
        logger.info(list(supplier_history))
        return JsonResponse({'success':True, 'history':list(supplier_history), 'pOrder': list_details}, status=200)
    return JsonResponse({'success':False, 'message':'Invalid request'}, status=500)


#individual supplier details
@login_required
def supplier_details_view(request,supplierId):
    if request.method == 'GET':
        try:
            supplier_details = Supplier.objects.get(id = supplierId)
            supplier_data = {
                'name': supplier_details.name,
                'contact_person': supplier_details.contact_person,
                'phone': supplier_details.phone,
                'email': supplier_details.email,
                'address': supplier_details.address 
            }

            logger.info(supplier_data)
            return JsonResponse({'success': True, 'data': supplier_data}, status = 200)
        except Exception as e:
            return JsonResponse({'success': False, 'response': f'{e}'}, status = 400)
    return JsonResponse({'success': False, 'response': 'invalid request'}, status = 500)

#life time details
@login_required
def view_LifeTimeOrders(request, supplier_id):
    #count of orders
    # total cost of all orders
    if request.method == 'GET':
        purchaseOrderDetails = PurchaseOrderItem.objects.get(supplier__id = supplier_id).values(
            'purchase_order__order_number', 'unit_cost', 'quantity'
        )
        list_entries = [{'id': supplier_id, 'number of orders': 0, 'total cost': 0}]
        count = 0
        for items in purchaseOrderDetails:
            if items['id'] == list_entries['id']:
                count += 1
                amount = (items['unit cost'] * items['quantity']) + list_entries['total cost']
                list_entries['total cost'] = amount
                list_entries['number of orders'] = count        
        return JsonResponse({'success': True, 'Data': list_entries}, status = 200)
    return JsonResponse({'success': True,  'response': 'invalid request'}, status = 500)

@login_required
@transaction.atomic
def supplier_view(request):
    if request.method == 'GET':
        supplier_products = Product.objects.all()
        supplier_balances = SupplierAccount.objects.all().values('supplier__id', 'balance', 'date')
        purchase_orders = PurchaseOrderItem.objects.all()

        list_orders = {}

        for items in purchase_orders:
            if list_orders.get(items.supplier.id):
                
                supplier = list_orders.get(items.supplier.id)
                # logger.info(f'quantity: {supplier}')
                supplier['quantity'] += items.quantity
                supplier['received_quantity'] += items.received_quantity
                supplier['returned'] += (items.quantity - items.received_quantity)
                supplier['amount'] += (items.unit_cost * items.received_quantity)

                if items.purchase_order.id == supplier['order_id']:
                    supplier['count'] = supplier['count']
                else:
                    supplier['count'] += 1
                    supplier['order_id'] = items.purchase_order.id
            else:
                list_orders[items.supplier.id] = {
                    'order_id': items.purchase_order.id,
                    'quantity' : items.quantity,
                    'received_quantity' : items.received_quantity,
                    'returned' : (items.quantity - items.received_quantity),
                    'amount' : (items.unit_cost * items.received_quantity),
                    'count' : 1
                }
                

        logger.info([list_orders])
        logger.info(supplier_balances)
        form = AddSupplierForm()
        suppliers = Supplier.objects.filter(delete = False)
        logger.info(suppliers)
        return render(request, 'Supplier/Suppliers.html', {
            'form':form,
            'products':supplier_products,
            'balances':supplier_balances,
            'life_time': [list_orders],
            'suppliers':suppliers
        })
        # except Exception as e:
        #     # logger.info(e)
        #     messages.error(request, f'{e}')
        #     return redirect('inventory:suppliers')
            
    if request.method == 'POST':
        """
        payload = {
            name,
            contact_person,
            email,
            product,
            address
        }
        """
        try:
            data = json.loads(request.body)
            logger.info(data)

            name = data.get('name')
            contact_person = data.get('contact_person')
            email = data.get('email')
            phone = data.get('phone')
            address = data.get('address')

            logger.info(name)

            # # check if all data exists
            # if not name or not contact_person or not email or not phone or not address:
            #     return JsonResponse({'success': False, 'message':'Fill in all the form data'}, status=400)

            # # check is supplier exists
            # if Supplier.objects.filter(email=email).exists() and Supplier.objects.filter(delete = True).exists():
            #     bring_back =  Supplier.objects.filter(email = email)
            #     bring_back.delete = False
            #     bring_back.update()
            #     logger.info(bring_back.delete)
            #     return JsonResponse({'success': True, 'response':f'Supplier{name} brought back'}, status=200)
            # elif Supplier.objects.filter(email=email).exists():
            #     return JsonResponse({'success': False, 'response':f'Supplier{name} already exists'}, status=400)
           
            logger.info('here')
            supplier = Supplier.objects.create(
                name = name,
                contact_person = contact_person,
                email = email,
                phone = phone,
                address = address,
                delete = False
            )
            logger.info('done')
                
                # SupplierAccount.objects.create(
                #     supplier = supplier,
                #     currency = Currency.objects.get(default = True),
                #     balance = 0,
                # )

                # SupplierAccount.objects.create(
                #     supplier = supplier,
                #     currency = Currency.objects.get(default = False),
                #     balance = 0,
                # )
            return JsonResponse({'success': True}, status=200)
        except Exception as e:
            logger.info(e)
            return JsonResponse({'success': False}, status=400)
    return JsonResponse({'success': False, 'message':'Invalid request'}, status=200)


@login_required
def supplier_list_json(request):
    suppliers = Supplier.objects.all().values(
        'id',
        'name'
    )
    return JsonResponse(list(suppliers), safe=False)

@login_required
def supplier_account(request, supplier_id):
    try:
        supplier_account = SupplierAccount.objects.get(id=supplier_id)
        return render(request, 'suppliers/supplier_account.html')
    except Exception as e:
        messages.warning(request, 'Supplier account doesnt exists')
        return redirect('inventory:suppliers')


@login_required
def supplier_prices(request, product_id):
    """
        {
            product_id: id
        }
    """
    try:
        best_three_prices = best_price(product_id)
        return JsonResponse({'success': True, 'suppliers': best_three_prices})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@transaction.atomic
def product(request):
    if request.method == 'POST':
        # payload
        """
            id,
            name,
            cost: float,
            quantity: int,
            category,
            tax_type,
            min_stock_level,
            description
        """
        try:
            data = json.loads(request.body)
            product_id = data.get('id', '')

            logger.info(f'product ID: {product_id}')
            
        except Exception as e:
            return JsonResponse({'success':False, 'message':'Invalid data'})

        image_data = data.get('image')

        # if image_data: revisit
        #     try:
        #         format, imgstr = image_data.split(';base64,') 
        #         ext = format.split('/')[-1]
        #         image = ContentFile(base64.b64decode(imgstr), name=f'{data['name']}.{ext}')
        #     except Exception as e:
        #         logger.error(f'Error decoding image: {e}')
        #         return JsonResponse({'success': False, 'message': 'Invalid image data'})

        try:
            category = ProductCategory.objects.get(id=data['category'])
        except ProductCategory.DoesNotExist:
            return JsonResponse({'success':False, 'message':f'Category Doesnt Exists'})
        
        if product_id:
            """editing the product"""
            with transaction.atomic():
                product = Inventory.objects.select_for_update().get(id=product_id, branch=request.user.branch)
                logger.info(f'Editing product: {product.name} ')
                product.name = data['name']
                product.price = data.get('price', 0)
                product.cost = data.get('cost', 0)    
                product.quantity = data.get('quantity', 0)  
                product.category = category  
                product.tax_type = data['tax_type']
                product.stock_level_threshold = data['min_stock_level']
                product.description = data['description']
                product.end_of_day = True if data.get('end_of_day') else False
                product.service = True if data.get('service') else False
                product.image=product.image
                product.batch = product.batch
                product.save()

        else:
            """creating a new product"""
            
            # validation for existance
            if Inventory.objects.filter(name__exact=data['name'], branch=request.user.branch).exists():
                return JsonResponse({'success':False, 'message':f'Product {data['name']} exists'})
            
            logger.info(f'Creating product: {data['name']}: {request.user.branch}')

            product = Inventory.objects.create(
                batch = '',
                name = data['name'],
                price = 0,
                cost = 0,
                quantity = 0,
                category = category,
                tax_type = data['tax_type'],
                stock_level_threshold = data['min_stock_level'],
                description = data['description'], 
                end_of_day = True if data['end_of_day'] else False,
                service = True if data['service'] else False,
                branch = request.user.branch,
                # image = image,
                status = True
            )
        
        return JsonResponse({'success':True}, status=200)

    if request.method == 'GET': # to be dynamic
        products = Inventory.objects.filter(
            Q(branch=request.user.branch),
            status=True,
            disable=False
        ).values(
            'id',
            'name',
            'quantity'
        ).order_by('name')     

        return JsonResponse(list(products), safe=False)
    
    return JsonResponse({'success':False, 'message':'Invalid request'}, status=400)

@login_required
def delete_product(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_id = data.get('id', '')

            product = Inventory.objects.get(id=product_id, branch=request.user.branch)

            logger.info(product)
            if product.quantity > 0:
                product.disable = True
                return JsonResponse({'False': True, 'message': 'Product cannot be deleted it have quantity more than zero.'})
            else:
                product.disable = True
            product.save()

            return JsonResponse({'success': True, 'message': 'Product deleted successfully.'})

        except Inventory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Product not found.'}, status=404)
        except Exception as e:
            logger.info(e)
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

#reorder settings
@login_required
def reorder_settings(request):
    """ method to set reorder settings"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            quantity_suggestion = data.get('suggestion')
            order_enough = data.get('enough')
            supplier = data.get('supplier', 'all')
            days_from = data.get('from')
            days_to = data.get('to')

            if not quantity_suggestion or not order_enough or not supplier:
                return JsonResponse({'success':False, 'message':'Please fill all required data.'})

            settings = reorderSettings.objects.get_or_create(id=1)

            settings.supplier=supplier,
            settings.quantity_suggestion = True if quantity_suggestion else False,
            settings.order_enough_stock = True if order_enough else False
            
            if order_enough:
                if not days_from or not days_to:
                    return JsonResponse({'success':False, 'message':'Please fill in the days.'})
                settings.number_of_days_from = days_from
                settings.number_of_days_to = days_to
                settings.save()
            
            return JsonResponse({'success':True, 'message':'Reorder Settings Succefully Saved.'}, status=200)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
    
    if request.method == 'GET':
        try:
            settings = reorderSettings.objects.filter(id=1).values()
            JsonResponse({'success':True, 'data':settings}, status=200)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        return JsonResponse({'success':False, 'message':'Invalid request'}, status=500)

# #stocktake
@login_required
def process_stock_take_item(request):
     if request.method == 'POST':
       try:
           data = json.loads(request.body)
           phy_quantity = data.get('quantity')
           stocktake_id =data.get('stocktake_id')
           
           s_item = StocktakeItem.objects.get(id=stocktake_id)
           s_item.quantity = int(phy_quantity)
           
           difference = abs(s_item.now_quantity - int(phy_quantity)) 
           
           logger.info(f'difference: {difference}')
           
           s_item.quantity_difference = difference
           s_item.cost = s_item.product.cost * s_item.quantity_difference

           s_item.stocktake.negative += difference
           s_item.stocktake.positive += int(phy_quantity)
           
           s_item.stocktake.negative_cost += difference * s_item.product.cost
           s_item.stocktake.positive_cost += int(phy_quantity)  * s_item.product.cost 
           s_item.recorded = True
           s_item.stocktake.save()        
           s_item.save()
           
           descripancy_value =  s_item.quantity_difference
           details_inventory= {'item_id': s_item.id, 'difference': descripancy_value}
           return JsonResponse({'success': True, 'data': details_inventory }, status = 200)
       except Exception as e:
           return JsonResponse({'success': False, 'response': e}, status = 400)
       
@login_required
def stocktake_pdf(request):
    try:
        data = json.loads(request.body)
        type = data.get('type')
        stocktake_id = data.get('stocktake_id')
        template_name = 'reports/stocktake.html'
        
        logger.info(f'Id: {stocktake_id}')
        
        stock_take = StockTake.objects.get(id=stocktake_id)
        
        stock_items = None
      
        if type == 'negative':
            stock_items = StocktakeItem.objects.filter(stocktake=stock_take, quantity_difference__lt=0)
        elif type == 'positve': 
            stock_items = StocktakeItem.objects.filter(stocktake=stock_take, quantity_difference__gt=0)
        elif type == 'all':
            stock_items = StocktakeItem.objects.all()

        total_cost = stock_items.aggregate(total=Sum('product__cost'))['total'] or 0
       
        return generate_pdf(
            template_name, {
                'stocktake_items':stock_items,
                'stock_take':stock_take,
                'total_cost':total_cost
            }
        )
        
    except Exception as e:
        logger.error(f'error processing pdf: {e}')
        return JsonResponse({'success':False, 'message':str(e)}, status=400)
    

@login_required
def stock_take_index(request):
    if request.method == 'GET':
        form = StockTakeForm()
        stock_takes = StockTake.objects.filter(branch=request.user.branch)
        
        negative = stock_takes.aggregate(total=Sum('negative'))['total'] or 0
        positive = stock_takes.aggregate(total=Sum('positive'))['total'] or 0
        negative_value = stock_takes.aggregate(total=Sum('negative_cost'))['total'] or 0
        positive_value = stock_takes.aggregate(total=Sum('positive_cost'))['total'] or 0
        
        return render(request, 'stocktake/stocktake.html', {
            'negative':negative,
            'positive':positive,
            'negative_value':negative_value,
            'positive_value':positive_value,
            'stocktakes': stock_takes,
            'form': form
        })

    if request.method == 'POST':
        form = StockTakeForm(request.POST)
        if form.is_valid():
            try:
                stock_take = form.save(commit=False)
                branch = request.user.branch
                stock_take.branch = branch
                stock_take.status = False
                stock_take.s_t_number = StockTake().stocktake_number(branch.name)
                stock_take.result = 0  
                stock_take.save()

                form.save_m2m()

                inventory = Inventory.objects.filter(branch=branch, disable=False)
                stocktake_items = [
                    StocktakeItem(
                        stocktake=stock_take,
                        product=product,
                        quantity=0,
                        quantity_difference=0
                    )
                    for product in inventory
                ]
                
                StocktakeItem.objects.bulk_create(stocktake_items)

                stock_takes = StockTake.objects.all()
                return render(request, 'stocktake/stocktake.html', {
                    'stocktakes': stock_takes,
                    'form': StockTakeForm(), 
                    'success': 'Stock take created successfully!'
                })

            except Exception as e:
                logger.error(f"Error in stock take creation: {e}")
                return render(request, 'stocktake/stocktake.html', {
                    'form': form,
                    'error': str(e),
                    'stocktakes': StockTake.objects.all()
                })
        else:
            return render(request, 'stocktake/stocktake.html', {
                'form': form,
                'error': 'Please correct the errors below.',
                'stocktakes': StockTake.objects.all()
            })
            
@login_required
def confirm_stocktake(request, stocktake_id):
    try:
        stocktake = StockTake.objects.get(id=stocktake_id)
        stocktakes = StocktakeItem.objects.filter(stocktake__id=stocktake_id)
        
        flag = True
        for stock in stocktakes:
            if not stock.recorded:
                if  not stock.product.quantity == 0:
                    flag = False
                    break
            
        if flag:
            stocktake.status = True
            stocktake.save()
            return JsonResponse({'success':True}, status=200)
        else:
            return JsonResponse({'success':False, 'message':'Please record all stock take items.'}, status=400)
    except Exception as e:
        logger.error(f'Error saving stocktake, {e}')
        return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        
@login_required
def stock_take_detail(request, stocktake_id):
    if request.method == 'GET':
   
        products = StocktakeItem.objects.filter(stocktake__id=stocktake_id)
        stocktake  = StockTake.objects.get(id=stocktake_id)
        return render(request, 'stocktake/stocktake_detail.html', {
            'products':products,
            'stocktake':stocktake
        })
    
        return redirect('inventory:stocktake_detail', stocktake_id)
        
    if request.method == 'POST':
        
        data = json.loads(request.body)
        id = data.get('inventory_id')
        
        if id:
            inventory = Inventory.objects.get(id=id, branch=request.user.branch)
            logs = ActivityLog.objects.filter(
                inventory=inventory,
                branch=request.user.branch
            ).select_related(
                'branch', 'inventory'
            ).order_by(
                '-timestamp__date', '-timestamp__time'
            )
            
            html = render_to_string('stocktake/partials/inventory_detail_partials.html', {
                'inventory': inventory,
                'logs': logs,
            })
            
            return JsonResponse({'success':True, 'html':html}, status=200)
        return JsonResponse({'success':False, 'message':'Error processing your request.'}, status=400)
    

def undo_accept_stocktake_item(request):
    from apps.finance.models import UserAccount, UserTransaction
    try:
        data = json.loads(request.body)
        stocktake_id = data.get('product_id')

        with transaction.atomic():
            stocktake = StocktakeItem.objects.select_related('product').get(id=stocktake_id)

            # if not stocktake.accepted:
            #     return JsonResponse({'success': False, 'message': 'Stocktake item not accepted yet.'}, status=400)

            original_diff = stocktake.quantity - stocktake.now_quantity or 0
            stocktake.product.quantity += abs(stocktake.quantity_difference)
            stocktake.now_quantity += abs(stocktake.quantity_difference)
            
            stocktake.product.quantity -= original_diff
            stocktake.product.save()

            UserTransaction.objects.filter(description=stocktake.note, amount=stocktake.cost).delete()

            expense = Expense.objects.filter(
                description=stocktake.note,
                amount=stocktake.cost,
                user=request.user
            ).first()
            if expense:
                Cashbook.objects.filter(expense=expense).delete()
                expense.delete()
                
            stocktake.recorded = False
            stocktake.has_diff = False
            stocktake.accepted = False
            stocktake.note = ''
            stocktake.company_loss = False
            stocktake.save()

            logger.success('Stocktake acceptance successfully undone.')

            return JsonResponse({'success': True, 'message': 'Undo successful', 'quantity': stocktake.now_quantity, 'product_id':stocktake.id}, status=200)

    except Exception as e:
        logger.error(f'Error undoing stocktake item: {e}')
        return JsonResponse({'success': False, 'message': str(e)}, status=400)
        

@login_required
def accept_stocktake_item(request):
    from apps.finance.models import UserAccount, UserTransaction
    try:
        data = json.loads(request.body)
        stocktake_id = data.get('product_id')
        users = data.get('users')
        note = data.get('note')
        company = data.get('company')

        logger.success(f'data: {data}')
        
        with transaction.atomic():
            stocktake = StocktakeItem.objects.get(id=stocktake_id)
            logger.info(f'quantity: {stocktake.quantity} : now_quantity {stocktake.now_quantity}')
            logger.info(f'diff: {stocktake.now_quantity} :{stocktake.quantity - stocktake.now_quantity }')
            
            stocktake.product.quantity += (stocktake.quantity - stocktake.now_quantity )
            stocktake.now_quantity = stocktake.product.quantity 
            stocktake.has_diff = True
            
            stocktake.save()
            
            logger.info(f'quantity: {stocktake.now_quantity}')
        
            stocktake.note = note
            stocktake.accepted = True
            
            if users:
                user_accounts = UserAccount.objects.all()
                user_accounts_transactons = []
                for user in users:
                    account = user_accounts.filter(user__id=int(user)).first()
                    account = create_user_account(int(user)) if not account else account
                    user_accounts_transactons.append(
                        UserTransaction(
                            account=account,
                            transaction_type='Stock loss',
                            amount = stocktake.cost,
                            description = note,
                            debit = stocktake.cost,
                            credit = 0,
                            received_by = request.user
                        )
                    )
                UserTransaction.objects.bulk_create(user_accounts_transactons)
                
            if company:
                stocktake.company_loss = True
            
            stocktake.save()
            
            credit = False
            debit = False    
                    
            if stocktake.quantity_difference < 0:
                credit=True
            else:
                debit=True
            
            currency=Currency.objects.get(default=True)
            category, _ = ExpenseCategory.objects.get_or_create(name='Stock Loss')
            
            logger.info(f'Category: {category}, {_}')
            
            if stocktake.cost >= 0:
                expense = Expense.objects.create(
                    amount = stocktake.cost,
                    payment_method = 'cash',
                    currency=currency,
                    description=stocktake.note,
                    user=request.user,
                    branch=request.user.branch,
                    is_recurring=True,
                    category=category
                )
            
            Cashbook.objects.create(
                expense=expense,
                description=note,
                credit=debit,
                debit=credit,
                branch=request.user.branch,
                created_by=request.user,
                updated_by=request.user,
                currency=currency,
                amount = stocktake.cost,
            )
                
            logger.success(f'Stocktake loss successfully debited to user account(s)')   
            return JsonResponse({'success':True, 'message':'success', 'quantity':stocktake.now_quantity}, status=201) 
        
    except Exception as e:
        logger.error(f'Error recording stocktake item: {e}')
        return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        
def create_user_account(id):
    account = UserAccount.objects.create(
        user_id=id,
        account_type='',
        balance=0,
        total_credits=0,
        total_debits=0,
    )
    logger.info(f'Account for user {account.user.username}')
    return account
        
#supplier payments
@login_required
def payments(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            supplier_id = data.get('id')
            supplier_amount = data.get('amount')
            supplier_currency_used = data.get('currency')
            supplier_method = data.get('payment_method')

            supplier_details = Supplier.objects.get(id = supplier_id)
            supplier_currency = Currency.objects.get(name = supplier_currency_used)
            supplier_payment = SupplierAccountsPayments.objects.filter(account__supplier__id = supplier_id)\
            .values('user', 'timestamp', 'amount', 'account__balance', 'payment_method')

            supplier_balance = supplier_payment['account__balance']
            # supplier_timestamp = supplier_payment['timestamp']
            # supplier_user = supplier_payment['amount']
            #supplier_pay_method = supplier_payment['payment_method']

            if supplier_balance <= 0:
                return JsonResponse({'success': True, 'response': 'We donot owe this supplier'})
            else:
                if supplier_method == 'USD':
                    new_balance = supplier_balance - supplier_amount
                else:
                    exchange_rate = Currency.objects.filter(name = supplier_method)
                    new_balance_zig = (supplier_balance * exchange_rate['exchange_rate'] ) - supplier_amount
                    new_balance = new_balance_zig/exchange_rate['exchange_rate']
            
            with transaction.Atomic():
                supplier_acc = SupplierAccount.objects.update(
                    suppliers = supplier_details,
                    currency = supplier_currency,
                    balance = new_balance,
                )

                SupplierAccountsPayments.objects.create(
                    account = supplier_acc,
                    payment_method = supplier_method,
                    currency = supplier_currency,
                    amount = supplier_amount,
                )
                return JsonResponse({'success': True, 'response': 'Data saved'})
        except Exception as e:
            return JsonResponse({'success': False, 'response': f'{e}'}, status = 400)

@login_required
def accessory_view(request, product_id):
    if request.method == 'POST':
       
        try:
            data = json.loads(request.body)
            logger.info(f'Accessories: {data}')
            product_id = data.get('product_id')
            accessory_data = data.get('accessories', [])  

            logger.info(accessory_data)

            product = Inventory.objects.get(id=product_id)
            logger.info(product)
            current_accessories = Accessory.objects.filter(main_product=product).values('accessory_product', 'quantity')
            current_ids = set(current_accessories.values_list('accessory_product', flat=True))

            logger.info(f'current ids: {current_ids}')

            input_data = {acc['id']: acc['quantity'] for acc in accessory_data}
            input_ids = set(input_data.keys())

            accessories_to_add = input_ids - current_ids
            accessories_to_remove = current_ids - input_ids

            logger.info(f'accessories to remove: {accessories_to_remove}')

            if Accessory.objects.filter(main_product=product).exists():
                acc = Accessory.objects.get(main_product=product)
            else:
                acc = Accessory.objects.create(main_product=product, quantity=0)

            if accessories_to_add:
                accessories_to_add_objs = Inventory.objects.filter(id__in=accessories_to_add)

                logger.info(f'Accessories to add: {accessories_to_add_objs}')

                for accessory in accessories_to_add_objs:
                    quantity = input_data[f'{accessory.id}']  
                    acc.accessory_product.add(int(accessory.id))
                    acc.quantity = quantity  
                    acc.save()

            if accessories_to_remove:
                accessories_to_remove_objs = Inventory.objects.filter(id__in=accessories_to_remove)

                for accessory in accessories_to_remove_objs:
                    acc.accessory_product.remove(accessory)
                    acc.save()
                
                logger.info('done')

            for accessory_id in input_ids.intersection(current_ids):
                quantity = input_data[accessory_id]
                acc_instance = Accessory.objects.filter(main_product=product, accessory_product_id=accessory_id).first()
                if acc_instance:
                    acc_instance.quantity = quantity
                    acc_instance.save()

            updated_accessories = Accessory.objects.filter(main_product=product).values(
                'id', 'main_product__name', 'accessory_product__name', 'quantity'
            )

            return JsonResponse({'success': True, 'data': list(updated_accessories)}, status=200)

        except Inventory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Product not found.'}, status=404)
        except Exception as e:
            logger.info(e)
            return JsonResponse({'success': False, 'message': str(e)}, status=400)       

@login_required
def get_accessory(request, product_id):
    """
    Get accessory data including related product information for a given main product.
    """
    if product_id:
        accessory = Accessory.objects.filter(
            main_product__id=product_id
        ).prefetch_related(
            'accessory_product'
        ).values(
            'id',
            'main_product__name',
            'quantity',
            'accessory_product__id',
            'accessory_product__name',
            'accessory_product__price',
        )
        
        return JsonResponse({
            'success': True,
            'data': list(accessory)
        })
    
    return JsonResponse({
        'success': False,
        'message': 'Product ID is required'
    }, status=400)

def vue_view(request):
    return render(request, 'vue.html')

# Loss management
@login_required
def loss_management(request):
    """
        to put caching later
    """
    write_off_totals = WriteOff.objects.aggregate(total=Sum(F('quantity') * F('inventory_item__cost')))['total'] or 0
    defective_totals = DefectiveItem.objects.aggregate(total=Sum(F('quantity') * F('inventory_item__cost')))['total'] or 0
    shrinkage_totals = InventoryShrinkage.objects.aggregate(total=Sum(F('quantity') * F('inventory_item__cost')))['total'] or 0

    context = {
        'defective_form':AddDefectiveForm(),
        'shrinkage_form':AddShrinkageForm(),
        'write_off_form':AddWriteOffForm(),
        'write_off':write_off_totals,
        'defective':defective_totals,
        'shrinkage':shrinkage_totals
    }

    return render(request, 'loss_management/loss_management.html', context)

@login_required
def loss_management_accounts(request, account_name):
    try:
        if account_name == 'shrinkage':
            shrinkage = InventoryShrinkage.objects.all().select_related(
                'inventory_item', 
                'created_by'
            ).values(
                'id',
                'inventory_item__id',
                'inventory_item__name',
                'inventory_item__cost',
                'quantity',
                'reason',
                'recorded_by',
                'created_at'
            ).annotate(
                total_cost = Sum(F('quantity') * F('inventory_item__cost'))
            )
            return JsonResponse(list(shrinkage), safe=False)

        if account_name == 'defective':
            defective = DefectiveItem.objects.all().select_related(
                'inventory_item', 
                'created_by'
            ).values(
                'id',
                'inventory_item__id',
                'inventory_item__name',
                'inventory_item__cost',
                'quantity',
                'defect_description',
                'action_taken',
                'created_by',
                'created_at'
            ).annotate(
                total_cost = Sum(F('quantity') * F('inventory_item__cost'))
            )

            return JsonResponse(list(defective), safe=False)
        
        if account_name == 'write-off':
            write_off_data = WriteOff.objects.all().select_related(
                'inventory_item', 
                'created_by'
            ).values(
                'id',
                'inventory_item__id',
                'inventory_item__name',
                'inventory_item__cost',
                'quantity',
                'reason',
                'created_by',
                'created_at'
            ).annotate(
                total_cost = Sum(F('quantity') * F('inventory_item__cost'))
            )

            return JsonResponse(list(write_off_data), safe=False)

    except Exception as e:
        logger.info(e)
        return JsonResponse({'success':False, 'message':f'{e}'}, status=200)
    
@login_required
@transaction.atomic
def create_defective(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_id = data.get('product_id')
            quantity = data.get('quantity')
            defect_description = data.get('defect_description')
            action_taken = data.get('action_taken')

            product = Inventory.objects.get(id=product_id, branch=request.user.branch)

            if quantity > product.quantity:
                return JsonResponse({'success': False, 'message': 'Defective quantity cannot be more than the product quantity'}, status=400)

            defective_item = DefectiveItem.objects.create(
                inventory_item=product,
                quantity=quantity,
                defect_description=defect_description,
                action_taken=action_taken,
                created_by=request.user
            )

            product.quantity -= quantity
            product.save()

            ActivityLog.objects.create(
                branch=request.user.branch,
                user=request.user,
                action='defective',
                inventory=product,
                quantity=quantity,
                total_quantity=product.quantity,
                description=f'Defective: {defect_description}'
            )

            return JsonResponse({'success': True, 'message': 'Defective item created successfully'}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@transaction.atomic
def create_shrinkage(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_id = data.get('product_id')
            quantity = data.get('quantity')
            reason = data.get('reason')

            product = Inventory.objects.get(id=product_id, branch=request.user.branch)

            if quantity > product.quantity:
                return JsonResponse({'success': False, 'message': 'Shrinkage quantity cannot be more than the product quantity'}, status=400)

            shrinkage_item = InventoryShrinkage.objects.create(
                inventory_item=product,
                quantity=quantity,
                reason=reason,
                created_by=request.user
            )

            product.quantity -= quantity
            product.save()

            ActivityLog.objects.create(
                branch=request.user.branch,
                user=request.user,
                action='shrinkage',
                inventory=product,
                quantity=quantity,
                total_quantity=product.quantity,
                description=f'Shrinkage: {reason}'
            )

            return JsonResponse({'success': True, 'message': 'Shrinkage item created successfully'}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@transaction.atomic
def create_write_off(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_id = data.get('product_id')
            quantity = data.get('quantity')
            reason = data.get('reason')

            product = Inventory.objects.get(id=product_id, branch=request.user.branch)

            if quantity > product.quantity:
                return JsonResponse({'success': False, 'message': 'Write-off quantity cannot be more than the product quantity'}, status=400)

            write_off_item = WriteOff.objects.create(
                inventory_item=product,
                quantity=quantity,
                reason=reason,
                created_by=request.user
            )

            product.quantity -= quantity
            product.save()

            ActivityLog.objects.create(
                branch=request.user.branch,
                user=request.user,
                action='write-off',
                inventory=product,
                quantity=quantity,
                total_quantity=product.quantity,
                description=f'Write-off: {reason}'
            )

            return JsonResponse({'success': True, 'message': 'Write-off item created successfully'}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
def get_cart_items(request):
    try:
        items = TemporaryPurchaseOrderItem.objects.filter(
            temporary_purchase_order__user=request.user,
            temporary_purchase_order__branch=request.user.branch
        ).select_related('product', 'supplier')

        items_data = [{
            'id': item.id,
            'product': item.product.name,
            'product_id': item.product.id,
            'supplier': item.supplier.name,
            'quantity': item.quantity,
            'price': float(item.price) if item.price else 0.0
        } for item in items]

        return JsonResponse({
            'success': True,
            'items': items_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)