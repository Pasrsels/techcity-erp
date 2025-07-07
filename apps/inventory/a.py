from rest_framework import views, status, viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from .serializers import *
from rest_framework.viewsets import ModelViewSet
from .serializers import *
from django.db import transaction
from django.db.models import Q, Sum, F, FloatField, ExpressionWrapper
from django.core.files.base import ContentFile
import base64
from decimal import Decimal
from .forms import *
from .views import get_stock_account_data
from django.shortcuts import render, redirect, get_object_or_404, get_list_or_404
from openpyxl.styles import Alignment, Font, PatternFill
import  json
from django.http.response import JsonResponse

class AddInventoryView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = InventorySerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            inventory = serializer.save()
            return Response({
                'status': 'success',
                'message': 'Product added successfully!',
                'product': {
                    'id': inventory.id,
                    'name': inventory.name
                }
            })
        return Response({'status': 'error', 'message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class InventoryViewset(viewsets.ModelViewSet):
    serializer_class = InventorySerializer

    def get_queryset(self):
        user = self.request.user
        inventory_details = Inventory.objects.filter(disable=False).values()
        return Response(inventory_details, status.HTTP_200_OK)

class CategoriesList(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request): 
        categories = ProductCategory.objects.all().values()
        logger.info(categories)
        return Response(categories, status.HTTP_200_OK)

class AddCategories(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        data = request.data
        category_name = data.get('name').upper()
        
        if ProductCategory.objects.filter(name=category_name).exists():
            return Response({'message':'Category Exists'}, status.HTTP_400_BAD_REQUEST)
        
        category = ProductCategory.objects.create(
            name=category_name
        )
        return Response({'id': category.id, 'name':category.name},status.HTTP_201_CREATED)        

class Products(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        products = Inventory.objects.filter(branch = request.user.branch, status=True, disable=False).values(
            'id',
            'name',
            'quantity',
            'price',
            'dealer_price',
            'category__name'
        ).order_by('name')  
      
        return Response(products, status.HTTP_200_OK)

class AddProducts(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        # payload
        """
            id,
            name,
            cost: float,
            quantity: int,
            category,
            tax_type,
            min_stock_level,
            description
        """
        try:
            data = request.data
            product_id = data.get('id', '')
            logger.info(data)   
        except Exception as e:
            return Response({'message':'Invalid data'}, status.HTTP_400_BAD_REQUEST)

        image_data = data.get('image', '')
        if image_data:
            try:
                format, imgstr = image_data.split(';base64,') 
                ext = format.split('/')[-1]
                image = ContentFile(base64.b64decode(imgstr), name=f'{data.get('name')}.{ext}')
            except Exception as e:
                logger.error(f'Error decoding image: {e}')
                return Response({'message': 'Invalid image data'}, status.HTTP_400_BAD_REQUEST)

        try:
            category = ProductCategory.objects.get(id=data.get('category'))
        except ProductCategory.DoesNotExist:
            return Response({'message':'Category Doesnt Exists'}, status.HTTP_400_BAD_REQUEST)
        
        if product_id:
            """editing the product"""
            logger.info(f'Editing product ')
            product = Inventory.objects.get(id=product_id, branch=request.user.branch)
            product.name = data.get('name')
            product.price = data.get('price', 0)
            product.cost = data.get('cost', 0)    
            product.quantity = data.get('quantity', 0)  
            product.category = category  
            product.tax_type = data.get('tax_type')
            product.stock_level_threshold = data.get('min_stocklevel')
            product.description = data.get('description')
            product.end_of_day = True if data.get('end_of_day') else False
            product.service = True if data.get('service') else False
            product.image=product.image
            product.batch = product.batch
        else:
            """creating a new product"""
            # validation for existance
            if Inventory.objects.filter(name=data.get('name')).exists():
                return Response({'message':f'Product {data.get('name')} exists'}, status.HTTP_400_BAD_REQUEST)
            logger.info(f'Creating ')
            product = Inventory.objects.create(
                batch = '',
                name = data.get('name'),
                price = 0,
                cost = 0,
                quantity = 0,
                category = category,
                tax_type = data.get('tax_type'),
                stock_level_threshold = data.get('min_stock_level'),
                description = data.get('description'), 
                end_of_day = True if data.get('end_of_day') else False,
                service = True if data.get('service') else False,
                branch = request.user.branch,
                # image = image,
                status = True
            )
        product.save()
        return Response(status.HTTP_201_CREATED)

class DeleteProducts(views.APIView):
    permission_classes = [IsAuthenticated]
    def delete(self, request):
        try:
            data = request.data
            product_id = data.get('id', '')
            product = Inventory.objects.get(id=product_id, branch=request.user.branch)

            logger.info(product)
            if product.quantity > 0:
                product.disable = True
                return Response({'False': True, 'message': 'Product cannot be deleted it have quantity more than zero.'})
            else:
                product.disable = True
            product.save()
            return Response({'message': 'Product deleted successfully.'}, status.HTTP_200_OK)
        except Inventory.DoesNotExist:
            return Response({'message': 'Product not found.'}, status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.info(e)
            return Response({'message': str(e)}, status.HTTP_406_NOT_ACCEPTABLE)

class InventoryList(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        
        logger.info(list(Inventory.objects.\
        filter(branch=request.user.branch, status=True).values()))
        inventory_data = Inventory.objects.filter(branch=request.user.branch, status=True).values()
        accessory_data = Accessory.objects.filter(main_product__branch=request.user.branch, main_product__status=True).values()

        logger.info(inventory_data)
        inventory = Inventory.objects.filter(branch=request.user.branch, status=True, disable=False).select_related(
        'category',
        'branch'
        ).order_by('name').values()
        logger.info(inventory_data)
        return Response(
            {
                'inventory':inventory_data,
                'accessory': accessory_data,
                'total_cost':inventory.aggregate(total_cost=Sum(F('quantity') * F('cost')))['total_cost'] or 0,
                'total_price':inventory.aggregate(total_price=Sum(F('quantity') * F('price')))['total_price'] or 0,
            }, 
            status.HTTP_200_OK
        )
    
class DeleteInventory(views.APIView):
    permission_classes = [IsAuthenticated]
    def delete(self, request, id):
        product_id = id

        if product_id:
            inv = Inventory.objects.get(id=product_id, branch=request.user.branch)
            inv.status = False
            inv.save()
        
            ActivityLog.objects.create(
                    branch = request.user.branch,
                    user=request.user,
                    action= 'deactivated',
                    inventory=inv,
                    quantity=inv.quantity,
                    total_quantity=inv.quantity
                )
            return Response(status.HTTP_202_ACCEPTED)
        return Response(status.HTTP_406_NOT_ACCEPTABLE)
    
class EditInventory(views.APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, product_id):
        inv_product = Inventory.objects.get(id=product_id, branch=request.user.branch)
        end_of_day = request.data.get('end_of_day')

        if end_of_day:
            inv_product.end_of_day = True
        
        selling_price = Decimal(request.data.get('price'))
        dealer_price = Decimal(request.data.get('dealer_price'))
        
        # think through
        quantity = inv_product.quantity
        inv_product.name=request.data.get('name')
        # inv_product.batch=request.data['batch_code']
        inv_product.description=request.data.get('description')
        inv_product.price = Decimal(request.data.get('price'))
        inv_product.cost = Decimal(request.data.get('cost'))
        inv_product.dealer_price = Decimal(request.data.get('dealer_price'))
        inv_product.stock_level_threshold = request.data.get('min_stock_level')
        inv_product.dealer_price = dealer_price
        inv_product.quantity = request.data.get('quantity')
        inv_product.save()
        
        ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= 'Edit',
            inventory=inv_product,
            quantity=quantity,
            total_quantity=inv_product.quantity,
            dealer_price = dealer_price,
            selling_price = selling_price
        )
        return Response({'product':inv_product, 'title':f'Edit >>> {inv_product.name}'}, status.HTTP_202_ACCEPTED)

class InventoryDetail(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, id):
        purchase_order_items = PurchaseOrderItem.objects.all()
        inventory = Inventory.objects.get(id=id, branch=request.user.branch)
        logs = ActivityLog.objects.filter(
            inventory=inventory,
            branch=request.user.branch
        ).order_by('-timestamp__date', '-timestamp__time')

        # stock account data and totals (costs and quantities)
        stock_account_data = get_stock_account_data(logs)
        total_debits = sum(entry['cost'] for entry in stock_account_data if entry['type'] == 'debits')
        total_credits = sum(entry['cost'] for entry in stock_account_data if entry['type'] == 'credits')

        total_debits_quantity = sum(entry['quantity'] for entry in stock_account_data if entry['type'] == 'debits')
        total_credits_quantity = sum(entry['quantity'] for entry in stock_account_data if entry['type'] == 'credits')

        logger.info(f'debits {total_debits_quantity}')

        remaining_stock_quantity = total_debits_quantity - total_credits_quantity

        """ get inventory value based on the currencies in the system """
        inventory_value = []
        inventory_sold_value = []
        inventory_total_cost = inventory.cost * inventory.quantity

        for currency in Currency.objects.all():
            inventory_value.append(
                {
                    'name': f'{currency.name}',
                    'value': float(inventory_total_cost * currency.exchange_rate )if currency.exchange_rate == 1\
                            else float(inventory_total_cost * currency.exchange_rate)
                }
            )
            inventory_sold_value.append(
                {
                    'name': f'{currency.name}',
                    'value': logs.filter(invoice__invoice_return = False, invoice__currency=currency).\
                            aggregate(Sum('invoice__amount'))['invoice__amount__sum'] or 0
                }
            )
        
        logger.info(f'inventory value: {inventory_sold_value}')

        # logs = ActivityLog.objects.annotate(hour=Extract('timestamp', 'hour')).order_by('-hour')

        """ create log data structure for the activity log graph """
        sales_data = {}
        stock_in_data = {}
        transfer_data = {}
        labels = []

        for log in logs:
            month_name = log.timestamp.strftime('%B')  
            year = log.timestamp.strftime('%Y')
            month_year = f"{month_name} {year}"  

            if log.action == 'Sale':
                if month_year in sales_data:
                    sales_data[month_year] += log.quantity
                else:
                    sales_data[month_year] = log.quantity
            elif log.action in ('stock in', 'Update'):
                if month_year in stock_in_data:
                    stock_in_data[month_year] += log.quantity
                else:
                    stock_in_data[month_year] = log.quantity
            elif log.action == 'Transfer':
                if month_year in transfer_data:
                    transfer_data[month_year] += log.quantity
                else:
                    transfer_data[month_year] = log.quantity

            if month_year not in labels:
                labels.append(month_year)

        """ download a log or stock account pdf """

        # if request.data.get('logs'):
        #     print(request.data.get('logs'))
        #     download_stock_logs_account('logs', logs, inventory)
        # elif request.data.get('account'):
        #     download_stock_logs_account('account', logs, inventory)
        
        inventory_serializer = InventorySerializer(inventory)
        return Response({
            'inventory': inventory_serializer,
            'remaining_stock_quantity':remaining_stock_quantity,
            'stock_account_data':stock_account_data,
            'inventory_value':inventory_value,
            'inventory_sold_value':inventory_sold_value,
            'total_debits':total_debits,
            'total_credits':total_credits,
            'logs': logs,
            'items':purchase_order_items,
            'sales_data': sales_data.values(), 
            'stock_in_data': stock_in_data.values(),
            'transfer_data': transfer_data.values(),
            'labels': labels,
        }, status.HTTP_200_OK)
    
class InventoryIndexJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        inventory = Inventory.objects.filter(branch=request.user.branch, status=True).values(
            'id', 'name', 'price', 'cost', 'quantity', 'reorder'
        ).order_by('name')
        return Response(inventory, status.HTTP_200_OK)

class ActivateInventory(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, product_id):
        product = get_object_or_404(Inventory, id=product_id)
        product.status=True
        product.save()
        
        ActivityLog.objects.create(
            invoice = None,
            product_transfer = None,
            branch = request.user.branch,
            user=request.user,
            action= 'activated',
            inventory=product,
            quantity=product.quantity,
            total_quantity=product.quantity
        )
        serializer = InventorySerializer(product)
        return Response(serializer.data, status.HTTP_200_OK)

class DefectiveProductList(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        defective_products = DefectiveProduct.objects.filter(branch=request.user.branch)
        
        # loss calculation
        data = request.data
        defective_id = data.get('product_id')
        quantity = data.get('quantity')       
        try:
            d_product = DefectiveProduct.objects.get(id=defective_id, branch=request.user.branch)
            product = Inventory.objects.get(product__id=d_product.product.id, branch=request.user.branch)
        except:
            return Response({'message':'Product doesnt exists'}, status.HTTP_400_BAD_REQUEST)
    
        product.quantity += quantity
        product.status = True if product.status == False else product.status
        product.save()
        
        d_product.quantity -= quantity
        d_product.save()
        
        ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= 'stock in',
            inventory=product,
            quantity=quantity,
            total_quantity=product.quantity,
            description = 'from defective products'
        )
        return Response(status.HTTP_200_OK)
    
        quantity = defective_products.aggregate(Sum('quantity'))['quantity__sum'] or 0
        price = defective_products.aggregate(Sum('product__cost'))['product__cost__sum'] or 0
        
        return render(request, 'defective_products.html', 
            {
                'total_cost': quantity * price,
                'defective_products':defective_products,
            }
        )

class BranchesInventory(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        branches_inventory = Inventory.objects.filter(status=True).values(
            'name', 'price', 'quantity', 'branch__name'
        )
        return Response(branches_inventory, status.HTTP_200_OK)

class DefectiveProductViewAdd(views.APIView):
    def get(self, request):
        defective_products = DefectiveProduct.objects.filter(branch=request.user.branch)
        quantity = defective_products.aggregate(Sum('quantity'))['quantity__sum'] or 0
        price = defective_products.aggregate(Sum('product__cost'))['product__cost__sum'] or 0
        
        return Response( 
            {
                'total_cost': quantity * price,
                'defective_products':defective_products,
            },
            status.HTTP_200_OK
        )
    # loss calculation
    def post(self, request):
        data = request.data
        
        defective_id = data['product_id']
        quantity = data['quantity']
        
        try:
            d_product = DefectiveProduct.objects.get(id=defective_id, branch=request.user.branch)
            product = Inventory.objects.get(product__id=d_product.product.id, branch=request.user.branch)
        except:
            return Response({'success': False, 'message':'Product doesnt exists'}, status.HTTP_400_BAD_REQUEST)
    
        product.quantity += quantity
        product.status = True if product.status == False else product.status
        product.save()
        
        d_product.quantity -= quantity
        d_product.save()
        
        ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= 'stock in',
            inventory=product,
            quantity=quantity,
            total_quantity=product.quantity,
            description = 'from defective products'
        )
        return Response({'success':True}, status.HTTP_200_OK)


class InventoryDetail(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, id):
        purchase_order_items = PurchaseOrderItem.objects.all().values()
        inventory = Inventory.objects.get(id=id, branch=request.user.branch)
        logs = ActivityLog.objects.filter(
            inventory=inventory,
            branch=request.user.branch
        ).order_by('-timestamp__date', '-timestamp__time')
        logger.info(logs)


        logs_serialized = ActivityLog.objects.filter(
            inventory=inventory,
            branch=request.user.branch
        ).order_by('-timestamp__date', '-timestamp__time').values()

        inventory_list = []
        inventory_list.append(
            {
                'ID': inventory.id,
                'Name': inventory.name,
            }
        )
        # stock account data and totals (costs and quantities)
        stock_account_data = get_stock_account_data(logs)
        total_debits = sum(entry['cost'] for entry in stock_account_data if entry['type'] == 'debits')
        total_credits = sum(entry['cost'] for entry in stock_account_data if entry['type'] == 'credits')

        total_debits_quantity = sum(entry['quantity'] for entry in stock_account_data if entry['type'] == 'debits')
        total_credits_quantity = sum(entry['quantity'] for entry in stock_account_data if entry['type'] == 'credits')

        logger.info(f'debits {total_debits_quantity}')
        logger.info(f'debits {total_credits_quantity}')

        remaining_stock_quantity = total_debits_quantity - total_credits_quantity

        """ get inventory value based on the currencies in the system """
        inventory_value = []
        inventory_sold_value = []
        inventory_total_cost = inventory.cost * inventory.quantity
        
        for currency in Currency.objects.all():
            inventory_value.append(
                {
                    'name': f'{currency.name}',
                    'value': float(inventory_total_cost * currency.exchange_rate )if currency.exchange_rate == 1\
                            else float(inventory_total_cost * currency.exchange_rate)
                }
            )
            inventory_sold_value.append(
                {
                    'name': f'{currency.name}',
                    'value': logs.filter(invoice__invoice_return = False, invoice__currency=currency).\
                            aggregate(Sum('invoice__amount'))['invoice__amount__sum'] or 0
                }
            )
        
        logger.info(f'inventory value: {inventory_sold_value}')

        # logs = ActivityLog.objects.annotate(hour=Extract('timestamp', 'hour')).order_by('-hour')

        """ create log data structure for the activity log graph """
        sales_data = {}
        stock_in_data = {}
        transfer_data = {}
        labels = []

        for log in logs:
            month_name = log.timestamp.strftime('%B')  
            year = log.timestamp.strftime('%Y')
            month_year = f"{month_name} {year}"  

            if log.action == 'Sale':
                if month_year in sales_data:
                    sales_data[month_year] += log.quantity
                else:
                    sales_data[month_year] = log.quantity
            elif log.action in ('stock in', 'Update'):
                if month_year in stock_in_data:
                    stock_in_data[month_year] += log.quantity
                else:
                    stock_in_data[month_year] = log.quantity
            elif log.action == 'Transfer':
                if month_year in transfer_data:
                    transfer_data[month_year] += log.quantity
                else:
                    transfer_data[month_year] = log.quantity
            elif log.action == 'sale return':
                if month_year in transfer_data:
                    transfer_data[month_year] += abs(log.quantity)
                else:
                    transfer_data[month_year] = abs(log.quantity)

            if month_year not in labels:
                labels.append(month_year)

        """ download a log or stock account pdf """

        # if request.GET.get('logs'):
        #     download_stock_logs_account('logs', logs, inventory)
        # elif request.GET.get('account'):
        #     download_stock_logs_account('account', logs, inventory)
        
        logger.info(stock_account_data)
        
        return Response({
            'inventory': inventory_list,
            'remaining_stock_quantity':remaining_stock_quantity,
            'stock_account_data':stock_account_data,
            'inventory_value':inventory_value,
            'inventory_sold_value':inventory_sold_value,
            'total_debits':total_debits,
            'total_credits':total_credits,
            'logs': logs_serialized,
            'items':purchase_order_items,
            'sales_data': list(sales_data.values()), 
            'stock_in_data': list(stock_in_data.values()),
            'transfer_data': list(transfer_data.values()),
            'labels': labels,
        }, status.HTTP_200_OK)

class NotificationJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        notifications = StockNotifications.objects.filter(inventory__branch=request.user.branch).values(
            'inventory__name', 'type', 'notification', 'inventory__id'
        )
        return Response(notifications, status.HTTP_200_OK)

class StockTakeViewEdit(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        products = Inventory.objects.filter(branch=request.user.branch).values()
        return Response({'products':products}, status.HTTP_200_OK)
    
    def post(self, request):
        """
            payload = {
                product_id:int
                pyhsical_quantity:int
            }
        """
        data = request.data
        prod_id = data.get('product_id')
        phy_quantity = data.get('physical_quantity')

        try:
            """
                1. get the product
                2. get the quantity
                3. condition to check between physical_quantity and quantity of the product
                4. json to the front {id:inventory.id, different:difference}
            """
            
            inventory_details = Inventory.objects.filter(id = prod_id).values('name', 'quantity','id')

            quantity = inventory_details['quantity']
            inventory_id = inventory_details['id']

            if quantity >= 0:
                descripancy_value =  quantity - phy_quantity
                details_inventory= {'inventory_id': inventory_id, 'difference': descripancy_value}
                return Response({'data': details_inventory }, status.HTTP_200_OK)
            return Response(status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            return Response({'response': e}, status.HTTP_400_BAD_REQUEST)
        
class ProcessStockTakeItem(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
       """
         payload = {
            pyhsical_quantity:int
            stocktake_id:int
        }
       """

       try:
           """
            1. get the product
            2. get the quantity
            3. condition to check between physical_quantity and quantity of the product
            4. json to the front {id:inventory.id, different:difference}
           """

           data = request.data
           phy_quantity = data.get('quantity')
           stocktake_id =data.get('stocktake_id')

           logger.info(phy_quantity)
           logger.info(type(phy_quantity))
           
           # update the stock item object with the quantity
           s_item = StocktakeItem.objects.get(id=stocktake_id)
           s_item.quantity = int(phy_quantity)
           
           s_item.quantity_difference = int(phy_quantity) - abs(s_item.product.quantity )
           logger.info(s_item.product.quantity)

           s_item.save()

           logger.info(f'{s_item.quantity_difference}')

           descripancy_value =  s_item.quantity_difference
           details_inventory= {'item_id': s_item.id, 'difference': descripancy_value}
           return Response({'data': details_inventory }, status.HTTP_200_OK)
       except Exception as e:
           return Response({'response': e}, status.HTTP_400_BAD_REQUEST)

class StockTakeDetail(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, stocktake_id):
        try:
            products = StocktakeItem.objects.filter(stocktake__id=stocktake_id).values()
            stocktake  = StockTake.objects.get(id=stocktake_id)
            serializer_data = StockTakeSerializer(stocktake)
            logger.info(stocktake)
            return Response({
                'products':products,
                'stocktake':serializer_data.data
            }, status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': f'{e}'}, status.HTTP_400_BAD_REQUEST)

class BranchCode(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        batch_codes = BatchCode.objects.all().values(
            'id',
            'code'
        )
        logger.info(batch_codes)
        return Response(batch_codes, status.HTTP_200_OK)
    
    def post(self, request):
        try:
            data = request.data
            code = data.get('batch_code')

            BatchCode.objects.create(code=code)
            return Response(status.HTTP_200_OK)
        except Exception as e:
            logger.info(e)
            return Response({'message':f'{e}'}, status.HTTP_406_NOT_ACCEPTABLE)

class SupplierViewAdd(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        try:
            supplier_products = Product.objects.all().values()
            supplier_balances = SupplierAccount.objects.all().values('supplier__id', 'balance', 'date')
            purchase_orders = PurchaseOrderItem.objects.all()

            list_orders = {}

            for items in purchase_orders:
                if list_orders.get(items.supplier.id):
                    
                    supplier = list_orders.get(items.supplier.id)
                    logger.info(f'quantity: {supplier}')
                    supplier['quantity'] += items.quantity
                    supplier['received_quantity'] += items.received_quantity
                    supplier['returned'] += (items.quantity - items.received_quantity)
                    supplier['amount'] += (items.unit_cost * items.received_quantity)

                    if items.purchase_order.id == supplier['order_id']:
                        supplier['count'] = supplier['count']
                    else:
                        supplier['count'] += 1
                        supplier['order_id'] = items.purchase_order.id
                else:
                    list_orders[items.supplier.id] = {
                        'order_id': items.purchase_order.id,
                        'quantity' : items.quantity,
                        'received_quantity' : items.received_quantity,
                        'returned' : (items.quantity - items.received_quantity),
                        'amount' : (items.unit_cost * items.received_quantity),
                        'count' : 1
                    }
                    

            logger.info([list_orders])
            logger.info(supplier_balances)
            form = AddSupplierForm()
            suppliers = Supplier.objects.filter(delete = False).values()
            logger.info(suppliers)

            return Response({
                'products':supplier_products,
                'balances':supplier_balances,
                'life_time': list_orders,
                'suppliers':suppliers
            }, status.HTTP_200_OK) 
        except Exception as e:
            logger.info(e)
            return Response({f'{e}'}, status.HTTP_400_BAD_REQUEST)
    def post(self, request):
        """
        payload = {
            name,
            contact_person,
            email,
            product,
            address
        }
        """
        try:
            data = request.data
            logger.info(data)

            name = data.get('name')
            contact_person = data.get('contact_person')
            email = data.get('email')
            phone = data.get('phone')
            address = data.get('address')

            logger.info(name)
            
            # check if all data exists
            if not name or not contact_person or not email or not phone or not address:
                return Response({'message':'Fill in all the form data'}, status.HTTP_400_BAD_REQUEST)

            # check is supplier exists
            if Supplier.objects.filter(email=email).exists() and Supplier.objects.filter(delete = True).exists():
                bring_back =  Supplier.objects.filter(email = email)
                bring_back.delete = False
                bring_back.update()
                logger.info(bring_back.delete)
                return Response({'response':f'Supplier{name} brought back'}, status.HTTP_202_ACCEPTED)
            elif Supplier.objects.filter(email=email).exists():
                return JsonResponse({'response':f'Supplier{name} already exists'}, status.HTTP_400_BAD_REQUEST)
           
            with transaction.atomic():
                if not Currency.objects.filter(name = 'USD').exists() and not Currency.objects.filter(name = 'ZIG').exists():
                    Currency.objects.create(
                        code = '001',
                        name = 'USD',
                        symbol = '$',
                        exchange_rate = 1,
                        default = True
                    )
                    Currency.objects.create(
                        code = '002',
                        name = 'ZIG',
                        symbol = 'Z',
                        exchange_rate = 26.78,
                        default =  False
                    )

                supplier = Supplier.objects.create(
                    name = name,
                    contact_person = contact_person,
                    email = email,
                    phone = phone,
                    address = address,
                    delete = False
                )
                SupplierAccount.objects.create(
                    supplier = supplier,
                    currency = Currency.objects.get(default = True),
                    balance = 0,
                )
                SupplierAccount.objects.create(
                    supplier = supplier,
                    currency = Currency.objects.get(name = 'ZIG'),
                    balance = 0,
                )
            return Response(status.HTTP_201_CREATED)
        except Exception as e:
            logger.info(e)
            return Response({f'{e}'}, status.HTTP_406_NOT_ACCEPTABLE)
    
class SupplierListJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        suppliers = Supplier.objects.all().values(
            'id',
            'name'
        )
        return Response(suppliers, status.HTTP_200_OK)

class SupplierDelete(views.APIView):
    permission_classes = [IsAuthenticated]
    def delete(self, request, supplier_id):
        try:
            supplier = Supplier.objects.get(id=supplier_id)
            supplier.delete = True
            supplier.save()
            return Response(status.HTTP_202_ACCEPTED)
        except Exception as e:
            logger.info(e)
            return Response({"message":f"{e}"}, status.HTTP_406_NOT_ACCEPTABLE)
        
    def put(self, request, supplier_id):
        try:
            data = request.data
            logger.info(data)

            name = data.get('name')
            contact_person = data.get('contact_person')
            email = data.get('email')
            address = data.get('address')
            phone = data.get('phone')

            supplier = Supplier.objects.get(phone=phone)

            supplier.name=name
            supplier.contact_person=contact_person
            supplier.email=email
            supplier.phone=phone
            supplier.address=address
            
            supplier.save()
            logger.info(f'{supplier} saved')

            return Response(status.HTTP_200_OK)
        except Exception as e:
            logger.info(e)
            return Response({"message":f"{e}"}, status.HTTP_406_NOT_ACCEPTABLE)
    
class SupplierPrices(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, product_id):
        """
            {
                product_id: id
            }
        """
        try:
            best_three_prices = best_price(product_id)
            logger.info(best_three_prices)
            return Response({'suppliers': best_three_prices}, status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': str(e)}, status.HTTP_406_NOT_ACCEPTABLE)

class SupplierPaymentHistory(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, supplier_id):
        """
            order name
            purchase order amount
        """
        supplier_history = SupplierAccountsPayments.objects.filter(account__supplier_id = supplier_id).\
        values(
            'timestamp', 
            'amount',
            'account__balance',
            'user__username',
            'currency__name'
        )
        supplier_purchase_order_details = PurchaseOrderItem.objects.filter(supplier__id = supplier_id)
        list_details = {}
        for items in supplier_purchase_order_details:
            if items.purchase_order.order_number == list_details.get('order_number'):
                list_details['amount'] = items.quantity * items.unit_cost
            else:
                list_details['order_number'] = items.purchase_order.order_number
                list_details['amount']= items.quantity * items.unit_cost
        logger.info(list_details)
        logger.info(list(supplier_history))
        return Response({'history':supplier_history, 'pOrder': list_details}, status.HTTP_200_OK)

class SupplierView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request,supplier_id):
        try:
            supplier_details = Supplier.objects.get(id = supplier_id)
            supplier_data = {
                'name': supplier_details.name,
                'contact_person': supplier_details.contact_person,
                'phone': supplier_details.phone,
                'email': supplier_details.email,
                'address': supplier_details.address 
            }
            logger.info(supplier_data)

            return Response({'data': supplier_data}, status.HTTP_200_OK)
        except Exception as e:
            return Response({'response': f'{e}'}, status.HTTP_406_NOT_ACCEPTABLE)

class CreateDefectiveProduct(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        form = DefectiveProductSerializer(data = request.data)
    
        if form.is_valid():
            branch = request.user.branch
            product = request.data.get('product')
            quantity = request.data.get('quantity')
            
            # validation
            if quantity > product.quantity:
                return Response({'Defective quantity cannot more than the products quantity'}, status.HTTP_400_BAD_REQUEST)
            elif quantity == 0:
                return Response({'Defective quantity cannot be less than zero'}, status.HTTP_400_BAD_REQUEST)
            
            product.quantity -= quantity
            product.save()
        
            d_obj = form.save(commit=False)
            d_obj.branch = branch
            d_obj.branch_loss = branch
            
            d_obj.save()
            
            ActivityLog.objects.create(
                branch = branch,
                user=request.user,
                action= 'defective',
                inventory=product,
                quantity=quantity,
                total_quantity=product.quantity,
                description = ''
            )
            #messages.success(request, 'Product successfuly saved')
            return Response(status.HTTP_200_OK)
        else:
            return Response({'message': 'Invalid form data'},status.HTTP_200_OK)

class CreateService(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(request):
        form = ServiceForm(request.POST)
        if form.is_valid():
            service_obj = form.save(commit=False)
            service_obj.cost = 0
            service_obj.branch = request.user.branch
            service_obj.save()
            messages.success(request, 'Service successfully created')
            return redirect('inventory:inventory')
        messages.warning(request, 'Invalid form data')

class EditService(views.APIView):
    def edit_service(request, service):
        form = ServiceForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, f'{service.name} successfully edited')
            return redirect('inventory:inventory')
        else:
            messages.warning(request, 'Please correct the errors below')

class  ReorderLists(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        reorder_list = ReorderList.objects.filter(branch=request.user.branch)
        if 'download' in request:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={request.user.branch.name} order.xlsx'
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            row_offset = 0
            
            worksheet['A' + str(row_offset + 1)] = f'Order List'
            worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
            cell = worksheet['A' + str(row_offset + 1)]
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(size=14, bold=True)
            cell.fill = PatternFill(fgColor='AAAAAA', fill_type='solid')

            row_offset += 1 
                
            category_headers = ['Name', 'Quantity']
            for col_num, header_title in enumerate(category_headers, start=1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = header_title
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            categories = reorder_list.values_list('product__product__category__name', flat=True).distinct()
            for category in categories:
                products_in_category = reorder_list.filter(product__product__category__name=category)
                if products_in_category.exists():
                    worksheet['A' + str(row_offset + 1)] = category
                    cell = worksheet['A' + str(row_offset + 1)]
                    cell.font = Font(color='FFFFFF')
                    cell.fill = PatternFill(fgColor='0066CC', fill_type='solid')
                    worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
                    row_offset += 2

                for product in reorder_list:
                    if category == product.product.product.category.name:
                        worksheet.append([product.product.product.name])
                        row_offset += 1

            workbook.save(response)
            return Response(response, status.HTTP_200_OK)
        return Response(status.HTTP_406_NOT_ACCEPTABLE)

class CreateandGetOrder(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        products_below_five = Inventory.objects.filter(branch=request.user.branch, quantity__lte = 5).values(
            'id', 
            'name',
            'description',
            'quantity', 
            'reorder',
            'category__id',
            'category__name'
        )
        return Response(products_below_five, status.HTTP_200_OK)
    def post(self, request):
        data = request.data
        product_id = data.get('id')
        
        product = get_object_or_404(Inventory, id=product_id, branch=request.user.branch)
        ReorderList.objects.create(product=product, branch=request.user.branch)
        
        product.reorder = True
        product.save()
        return Response(status.HTTP_201_CREATED)

class ReorderListJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        order_list = ReorderList.objects.filter(branch=request.user.branch).values(
            'id', 
            'product__name',  
            'product__quantity',
            'quantity'
        )
        return Response(order_list, status.HTTP_200_OK)

class ClearReorderList(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        reorders = ReorderList.objects.filter(branch=request.user.branch)
        
        for item in reorders:
            inventory_items = Inventory.objects.filter(id=item.product.id)
            for product in inventory_items:
                product.reorder = False
                product.save()
            
        reorders.delete()
        return Response({'Reoder list success fully cleared'}, status.HTTP_200_OK)
    def delete(self, request):
            data = request.data
            product_id = data.get('product_id')
        
            product = ReorderList.objects.get(id=product_id, branch=request.user.branch)
            inventory = Inventory.objects.get(id=product.product.id)
        
            product.delete() 
            inventory.reorder=False
            inventory.save()
            
            return Response(status.HTTP_200_OK)        
class ReorderFromNotification(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        notifications = StockNotifications.objects.filter(inventory__branch=request.user.branch, inventory__reorder=False, inventory__alert_notification=False).values(
            'quantity',
            'inventory__name', 
            'inventory__id', 
            'inventory__quantity' 
        )
        return Response(notifications, status.HTTP_200_OK)
    def post(self, request): 
        # payload
        """
            inventory_id
        """
        
        data = request.data
        inventory_id = data.get('inventory_id')
        action_type = data.get('action_type')        
        try:
            inventory = Inventory.objects.get(id=inventory_id)
            stock_notis = StockNotifications.objects.get(inventory=inventory)
        except Exception as e:
            return Response({'message':f'{e}'}, status.HTTP_406_NOT_ACCEPTABLE)
        
        if action_type == 'add':
            inventory.reorder=True
            inventory.save()
            ReorderList.objects.create(
                quantity=0,
                product=inventory, 
                branch=request.user.branch
            )   
        elif action_type == 'remove':
            inventory.alert_notification=True
            inventory.save()
    
        return Response(status.HTTP_201_CREATED)
    
class AddReorderQuantity(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        """
        Handles adding a quantity to a reorder item.

        Payload:
        - reorder_id: ID of the reorder item
        - quantity: Quantity to add
        """
        try:
            data = request.data
        except json.JSONDecodeError:
            return Response({'message': 'Invalid JSON payload'}, status.HTTP_406_NOT_ACCEPTABLE)

        reorder_id = data.get('reorder_id')
        reorder_quantity = data.get('quantity')

        if not reorder_id:
            return Response({'message': 'Reorder ID is required'}, status.HTTP_406_NOT_ACCEPTABLE)

        if not reorder_quantity:
            return Response({'message': 'Reorder quantity is required'}, status.HTTP_406_NOT_ACCEPTABLE)

        try:
            reorder_quantity = int(reorder_quantity)
        except ValueError:
            return Response({'message': 'Invalid reorder quantity'}, status.HTTP_406_NOT_ACCEPTABLE)

        try:
            reorder = ReorderList.objects.get(id=reorder_id)
            reorder.quantity = reorder_quantity
            reorder.save()
            logger.info(reorder.quantity)
            return Response({'message': 'Reorder quantity updated successfully'}, status.HTTP_201_CREATED)
        except ReorderList.DoesNotExist:
            return Response({'message': 'Reorder item not found'}, status.HTTP_404_NOT_FOUND)
        
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return Response({'message': f'{e}'}, status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class ReorderSettings(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        """ method to set reorder settings"""
        try:
            data = request.data
            quantity_suggestion = data.get('suggestion')
            order_enough = data.get('enough')
            supplier = data.get('supplier', 'all')
            days_from = data.get('from')
            days_to = data.get('to')

            if not quantity_suggestion or not order_enough or not supplier:
                return Response({'message':'Please fill all required data.'}, status.HTTP_406_NOT_ACCEPTABLE)

            settings = reorderSettings.objects.get_or_create(id=1)

            settings.supplier=supplier,
            settings.quantity_suggestion = True if quantity_suggestion else False,
            settings.order_enough_stock = True if order_enough else False
            
            if order_enough:
                # validate days 
                if not days_from or not days_to:
                    return Response({'message':'Please fill in the days.'}, status.HTTP_406_NOT_ACCEPTABLE)
                settings.number_of_days_from = days_from
                settings.number_of_days_to = days_to
                settings.save()
            
            return Response({'message':'Reorder Settings Succefully Saved.'}, status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'message':f'{e}'}, status.HTTP_400_BAD_REQUEST)
    
    def get(self, request):
        try:
            settings = reorderSettings.objects.filter(id=1).values()
            Response({'data':settings}, status.HTTP_200_OK)
        except Exception as e:
            return Response({'message':f'{e}'}, status.HTTP_400_BAD_REQUEST)

class PurchaseOrderList(views.APIView):#*
    permission_classes = [IsAuthenticated]
    def get(self, request):
        orders = PurchaseOrder.objects.filter(branch = request.user.branch).values().order_by('-order_date')

        items = PurchaseOrderItem.objects.filter(purchase_order__id=5)

        # Update the 'received' field for each item
        for item in items:
            item.expected_profit
            item.received_quantity
            item.save()
            
        # Perform a bulk update on the 'received' field
        PurchaseOrderItem.objects.bulk_update(items, ['expected_profit', 'received_quantity'])
        return Response( {'orders':orders}, status.HTTP_200_OK)

class PrintPurchaseOrder(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, order_id):
        try:
            purchase_order = PurchaseOrder.objects.get(id=order_id)
            purchase_order_serializer = PurchaseOrderSerializer(purchase_order)
        except PurchaseOrder.DoesNotExist:
            return Response({f'Purchase order with ID: {order_id} does not exists'}, status.HTTP_400_BAD_REQUEST)
        
        try:
            purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order).values()
        except PurchaseOrderItem.DoesNotExist:
            return Response({f'Purchase order with ID: {order_id} does not exists'}, status.HTTP_400_BAD_REQUEST)
        
        return Response( 
            {
                'orders':purchase_order_items,
                'purchase_order':purchase_order_serializer.data
            },
            status.HTTP_200_OK
        )
    
class PurchaseOrderListandCreate(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        orders = PurchaseOrder.objects.filter(branch = request.user.branch).values().order_by('-order_date')
        items = PurchaseOrderItem.objects.filter(purchase_order__id=5)

        # Update the 'received' field for each item
        for item in items:
            item.expected_profit
            item.received_quantity
            item.save()
            

        # Perform a bulk update on the 'received' field
        PurchaseOrderItem.objects.bulk_update(items, ['expected_profit', 'received_quantity'])
        return Response({'orders':orders}, status.HTTP_200_OK)

    def post(self, request): 
        try:
            data = request.data
            purchase_order_data = data.get('purchase_order', {})
            purchase_order_items_data = data.get('po_items', [])
            expenses = data.get('expenses', [])
            cost_allocations = data.get('cost_allocations', [])
            hold = data.get('hold', False)
            supplier_payment_data = data.get('supplier_data')

            unique_expenses = []

            seen = set()
            for expense in expenses:
                expense_tuple = (expense['name'], expense['amount'])
                if expense_tuple not in seen:
                    seen.add(expense_tuple)
                    unique_expenses.append(expense)

        except json.JSONDecodeError:
            return Response({'message': 'Invalid JSON payload'}, status.HTTP_400_BAD_REQUEST)

        # Extract data from purchase_order_data
        batch = purchase_order_data['batch']
        delivery_date = purchase_order_data['delivery_date']
        p_status = purchase_order_data['status']
        notes = purchase_order_data['notes']
        total_cost = Decimal(purchase_order_data['total_cost'])
        discount = Decimal(purchase_order_data['discount'])
        tax_amount = Decimal(purchase_order_data['tax_amount'])
        other_amount = Decimal(purchase_order_data['other_amount'])
        payment_method = purchase_order_data.get('payment_method')

        if not all([batch, delivery_date, p_status, total_cost, payment_method]):
            return Response({'message': 'Missing required fields'}, status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                purchase_order = PurchaseOrder(
                    order_number=PurchaseOrder.generate_order_number(),
                    batch=batch,
                    delivery_date=delivery_date,
                    status=p_status,
                    notes=notes,
                    total_cost=total_cost,
                    discount=discount,
                    tax_amount=tax_amount,
                    other_amount=other_amount,
                    branch=request.user.branch,
                    is_partial=False,
                    received=False,
                    hold=hold
                )
                purchase_order.save()

                purchase_order_items_bulk = []
                for item_data in purchase_order_items_data:
                    product_id = item_data['product_id']
                    product_name = item_data['product']
                    quantity = int(item_data['quantity'])
                    unit_cost = Decimal(item_data['price'])
                    actual_unit_cost = Decimal(item_data['actualPrice'])
                    supplier_id = item_data.get('supplier', [])

                    logger.info(f'Supplier id {{ supplier_id }}')

                    if not all([product_name, quantity, unit_cost, product_id]):
                        transaction.set_rollback(True)
                        return Response({'message': f'Missing fields in item data:{[product_name, quantity, unit_cost, product_id]}'}, status.HTTP_400_BAD_REQUEST)

                    try:
                        # Try to get the product in the user's branch
                        product = Inventory.objects.get(id=product_id, branch=request.user.branch)

                    except Inventory.DoesNotExist:
                        try:
                            # Try to get the product from branch 31
                            from_other_branch = Inventory.objects.get(id=product_id, branch__id=31)

                            # Clone the product into the user's branch
                            cloned_product = Inventory.objects.create(
                                branch=request.user.branch,
                                name=from_other_branch.name,
                                cost=from_other_branch.cost,
                                price=from_other_branch.price,
                                dealer_price=from_other_branch.dealer_price,
                                quantity=0, 
                                status=from_other_branch.status,
                                stock_level_threshold=from_other_branch.stock_level_threshold,
                                reorder=from_other_branch.reorder,
                                alert_notification=from_other_branch.alert_notification,
                                batch=from_other_branch.batch,
                                category=from_other_branch.category,
                                tax_type=from_other_branch.tax_type,
                                description=from_other_branch.description,
                                end_of_day=from_other_branch.end_of_day,
                                service=from_other_branch.service,
                                image=from_other_branch.image,
                                disable=from_other_branch.disable,
                            )

                            cloned_product.suppliers.set(from_other_branch.suppliers.all())

                            product = cloned_product

                        except Inventory.DoesNotExist:
                            transaction.set_rollback(True)
                            return Response({'message': f'Product with ID {product_id} not found in user branch or branch 31.'}, status=status.HTTP_404_NOT_FOUND)


                    supplier = Supplier.objects.get(id=supplier_id)

                    po_item = PurchaseOrderItem.objects.create(
                        purchase_order=purchase_order,
                        product=product,
                        quantity=quantity,
                        unit_cost=unit_cost,
                        actual_unit_cost=actual_unit_cost,
                        received_quantity=0,
                        received=False,
                        supplier = supplier,
                        # price=0,
                        # wholesale_price=0
                    )
                    product.suppliers.add(po_item.supplier.id)
                    product.batch += f'{batch}, '
                    product.price = 0
                    product.save()

                # Handle expenses
                expense_bulk = []
                for expense in unique_expenses:
                    name = expense['name']
                    amount = expense['amount']
                    expense_bulk.append(
                        otherExpenses(
                            purchase_order=purchase_order,
                            name=name,
                            amount=amount
                        )
                    )
                otherExpenses.objects.bulk_create(expense_bulk)

                # Cost allocations
                costs_list = []
                for cost in cost_allocations:
                    costs_list.append(
                        costAllocationPurchaseOrder(
                            purchase_order=purchase_order,
                            allocated=cost['allocated'],
                            allocationRate=cost['allocationRate'],
                            expense_cost=cost['expCost'],
                            price=cost['price'],
                            quantity=float(cost['price']),
                            product=cost['product'],
                            total=cost['total'],
                            total_buying=cost['totalBuying']
                        )
                    )
                costAllocationPurchaseOrder.objects.bulk_create(costs_list)

                # # Process finance updates
                # if not purchase_order.hold:
                #     if purchase_order.status.lower() == 'received':
                #         # if_purchase_order_is_received(
                #         #     request, 
                #         #     purchase_order, 
                #         #     tax_amount, 
                #         #     payment_method
                #         # ) 
                #         #
                #         # supplier_payments(purchase_order, supplier_payment_data, request)
                          
        except Exception as e:
            return Response({'message': str(e)}, status.HTTP_400_BAD_REQUEST)
        return Response({'message': 'Purchase order created successfully'}, status.HTTP_201_CREATED)

class PurchaseOrderInformation(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        suppliers = Supplier.objects.all().values()
        products = Inventory.objects.filter(branch=request.user.branch, status=True, disable=False).order_by('name').values()
        batch_codes = BatchCode.objects.all().values()
        return Response(
            {
                'suppliers':suppliers,
                'batch_codes':batch_codes,
                'products':products
            },
            status.HTTP_200_OK
        )

class PurchaseOrderDeleteandEdit(views.APIView):
    permission_classes = [IsAuthenticated]
    def delete(self, request, purchase_order_id):
        try:
            purchase_order = PurchaseOrder.objects.get(id=purchase_order_id)
        except PurchaseOrder.DoesNotExist:
            return Response({'message': f'Purchase order with ID {purchase_order_id} not found'}, status.HTTP_404_NOT_FOUND)

        try:
            purchase_order.delete()
            return Response({'message': 'Purchase order deleted successfully'}, status.HTTP_202_ACCEPTED)
        except Exception as e:
            return Response({'message': str(e)}, status.HTTP_500_INTERNAL_SERVER_ERROR)

    def put(self, request, purchase_order_id):
        try:
            expenses = otherExpenses.objects.filter(purchase_order__id=purchase_order_id).values()
        
            purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order__id=purchase_order_id).values(
                'purchase_order__id',
                'product__name',
                'product__id',
                'quantity',
                'unit_cost',
                'actual_unit_cost',
                'expected_profit',
                'supplier__name',
                'supplier'
            )

            return Response({'po_items':purchase_order_items, 'expenses':expenses}, status.HTTP_202_ACCEPTED)
        except Exception as e:
            return Response({'message':f'{e}'}, status.HTTP_500_INTERNAL_SERVER_ERROR)

class ReceiveOrder(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, order_id):
        try:
            purchase_order = PurchaseOrder.objects.get(id=order_id)
            purchase_order_serializer = PurchaseOrderSerializer(purchase_order)
        except PurchaseOrder.DoesNotExist:
            return Response({f'Purchase order with ID: {order_id} does not exists'}, status.HTTP_400_BAD_REQUEST)
        
        try:
            purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order).values()
        except PurchaseOrderItem.DoesNotExist:
            #messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
            return Response({f'Purchase order with ID: {order_id} does not exists'}, status.HTTP_400_BAD_REQUEST)
        
        logger.info(purchase_order_items.all().values('product'))

        products = Inventory.objects.filter(branch=request.user.branch).values(
            'dealer_price', 
            'price', 
            'name'
        )
        logger.info(f'branch: {request.user.branch}')
        # Convert products queryset to a dictionary for easy lookup by product ID
        product_prices = {product['name']: product for product in products}

        new_po_items =  []
        for item in purchase_order_items:
            if(item.product):
                product_name = item.product.name  
                product_data = product_prices.get(product_name)
                logger.info(product_name)
                if product_data:
                    item.dealer_price = product_data['dealer_price']
                    item.selling_price = product_data['price']
                else:
                    item.dealer_price = 0  
                    item.selling_price = 0 
                new_po_items.append(item)

        logger.info(f'Purchase order items: {new_po_items}')
        
        return Response(
            {
                'orders':purchase_order_items,
                'purchase_order':purchase_order_serializer.data
            },
            status.HTTP_200_OK
        )

class ProcessReceivedOrder(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        try:
            data = request.data
            edit = data.get('edit')
            order_item_id = data.get('id')
            quantity = data.get('quantity', 0)
            wholesale_price = data.get('wholesale_price', 0)
            selling_price = data.get('selling_price', 0)
            dealer_price = data.get('dealer_price', 0)
            expected_profit = data.get('expected_profit', 0)
            dealer_expected_profit = data.get('dealer_expected_profit', 0)

        except json.JSONDecodeError:
            return Response({'message': 'Invalid JSON payload'}, status.HTTP_400_BAD_REQUEST)

        if edit:
            return edit_purchase_order_item(
                order_item_id, 
                selling_price, 
                dealer_price, 
                expected_profit, 
                dealer_expected_profit, 
                quantity, 
                request
            )

        if quantity == 0:
            return Response({'message': 'Quantity cannot be zero.'}, status.HTTP_400_BAD_REQUEST)

        try:
            order_item = PurchaseOrderItem.objects.get(id=order_item_id)
            order = PurchaseOrder.objects.get(id=order_item.purchase_order.id)
        except PurchaseOrderItem.DoesNotExist:
            return Response({'message': f'Purchase Order Item with ID: {order_item_id} does not exist'}, status.HTTP_400_BAD_REQUEST)

        # Update the order item with received quantity
        order_item.receive_items(quantity)
        order_item.received_quantity = quantity
        order_item.expected_profit = expected_profit
        order_item.dealer_expected_profit = dealer_expected_profit
        order_item.received = True
        order_item.wholesale_price = wholesale_price
        order_item.price = selling_price

        order_item.save()

        # order_item.check_received()
        return Response({'message': 'Inventory updated successfully'}, status.HTTP_200_OK)

class PurchaseOrderDetail(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, order_id):
        try:
            purchase_order = PurchaseOrder.objects.get(id=order_id)
            purchase_order_serializer = PurchaseOrderSerializer(purchase_order)
        except PurchaseOrder.DoesNotExist:
            messages.warning(request, f'Purchase order with ID: {order_id} does not exist')
            return Response({f'Purchase order with ID: {order_id} does not exist'}, status.HTTP_400_BAD_REQUEST)

        items = costAllocationPurchaseOrder.objects.filter(purchase_order=purchase_order)
        expenses = otherExpenses.objects.filter(purchase_order=purchase_order).values()
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order).values()

        total_received_quantity = purchase_order_items.aggregate(Sum('received_quantity'))['received_quantity__sum'] or 0
        total_expected_profit = purchase_order_items.aggregate(Sum('expected_profit'))['expected_profit__sum'] or 0
        total_expected_dealer_profit = purchase_order_items.aggregate(Sum('dealer_expected_profit'))['dealer_expected_profit__sum'] or 0
        total_quantity = items.aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_expense_sum = expenses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0

        products = Inventory.objects.filter(branch=request.user.branch).values(
            'dealer_price', 
            'price', 
            'product__name'
        )

        # Convert products queryset to a dictionary for easy lookup by product ID
        product_prices = {product['product__name']: product for product in products}

        for item in items:
            product_name = item.product  
            product_data = product_prices.get(product_name)

            if product_data:
                item.dealer_price = product_data['dealer_price']
                item.selling_price = product_data['price']
            else:
                item.dealer_price = 0  
                item.selling_price = 0 

        if request.data.get('download') == 'csv':
            return generate_csv_response(items, purchase_order_items)

        return Response({
            'items': items,
            'expenses': expenses,
            'order_items': purchase_order_items,  
            'total_quantity': total_quantity,
            'total_received_quantity': total_received_quantity,
            'total_expected_profit': total_expected_profit,
            'total_expenses': total_expense_sum,
            'purchase_order': purchase_order_serializer.data,
            'total_expected_dealer_profit':total_expected_dealer_profit
        }, status.HTTP_200_OK)
    
class PurchaseOrderStatus(views.APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, order_id):
        try:
            purchase_order = PurchaseOrder.objects.get(id=order_id)
        except PurchaseOrder.DoesNotExist:
            return Response({'error': f'Purchase order with ID: {order_id} does not exist'}, status.HTTP_404_NOT_FOUND)

        try:
            data = request.data
            p_o_status = data['status']
            
            if p_o_status:
                purchase_order.status=p_o_status

                with transaction.atomic():
                    if purchase_order.status == 'received':
                        purchase_order.save()

                        tax_amount = purchase_order.tax_amount
                        payment_method = purchase_order.payment_method

                        if_purchase_order_is_received(
                            request, 
                            purchase_order, 
                            tax_amount,
                            payment_method
                        )
                
                return Response(status.HTTP_202_ACCEPTED)
            else:
                return Response({'message':'Status is required'}, status.HTTP_400_BAD_REQUEST)
        except json.JSONDecodeError:
            return Response({'message': 'Invalid JSON payload'}, status.HTTP_400_BAD_REQUEST)
        
class MarkPurchaseOrderDone(views.APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, po_id):
        purchase_order = PurchaseOrder.objects.get(id=po_id)
        purchase_order.received = True
        purchase_order.save()

        return Response(status.HTTP_202_ACCEPTED)
    
class SalesPriceListPDF(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(request, order_id):
        return Response({f'https://web-production-86a7.up.railway.app/inventory/sales_price_list_pdf/{order_id}/'})
        # try:
        #     purchase_order = PurchaseOrder.objects.get(id=order_id)
        # except PurchaseOrder.DoesNotExist:
        #     messages.warning(request, f'Purchase order with ID: {order_id} does not exist')
        #     return Response({f'Purchase order with ID: {order_id} does not exist'}, status.HTTP_404_NOT_FOUND)

        # items = costAllocationPurchaseOrder.objects.filter(purchase_order=purchase_order)
        # purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)

        # products = Inventory.objects.filter(branch=request.user.branch).values(
        #     'dealer_price', 
        #     'price', 
        #     'product__name',
        #     'product__description',
        #     'quantity'
        # )

        # # Convert products queryset to a dictionary for easy lookup by product ID
        # product_prices = {product['product__name']: product for product in products}
    
        # for item in items:
        #     product_name = item.product
        #     logger.info(product_name)
        #     product_data = product_prices.get(product_name)
        #     description = ''

        #     if product_data:
        #         description = item.description = product_data['product__description'] 

        #     if product_data:
        #         item.dealer_price = product_data['dealer_price']
        #         item.selling_price = product_data['price']
        #         item.description = description
        #     else:
        #         item.dealer_price = 0
        #         item.selling_price = 0
        #         item.description = description
                
        # context = {'items': items}

        # template = get_template('pdf_templates/price_list.html')
        # html = template.render(context)

        # response = HttpResponse(content_type='application/pdf')
        # response['Content-Disposition'] = 'attachment; filename="price_list.pdf"'
        
        # pisa_status = pisa.CreatePDF(html, dest=response)

        # if pisa_status.err:
        #     return Response('Error generating PDF', status.HTTP_400_BAD_REQUEST)
        # return Response(response, status.HTTP_200_OK)

class PurchaseOrderConfirmOrderItem(views.APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, po_id):
        try:
            order_items = PurchaseOrderItem.objects.filter(purchase_order__id=po_id).select_related('product')

            logs = []
            inventory_updates = []

            with transaction.atomic():
                for item in order_items:
                    print(item.received_quantity)
                    print(item.product)
                    logger.info(f'actual product cost {item.actual_unit_cost}')
                    logger.info(f'actual product cost {item.price}')
                    logger.info(f'dummy {item.unit_cost}')
                    inventory_updates.append(
                        Inventory(
                            id=item.product.id,
                            quantity=item.product.quantity + item.received_quantity,
                            price=item.price,
                            dealer_price=item.wholesale_price,
                            cost=item.actual_unit_cost,
                            status=True,
                            disable=False,
                        )
                    )

                    existing_quantity = item.product.quantity 
                    new_quantity = item.received_quantity + item.product.quantity

                    quantity_change = abs(new_quantity -  existing_quantity)

                    print(f' bvbproduct quantity {item.product.quantity}')
                    print(f'update product quantity {item.product.quantity + item.received_quantity}')

                    logs.append(
                        ActivityLog(
                            purchase_order=item.purchase_order,
                            branch=request.user.branch,
                            user=request.user,
                            action='stock in',
                            dealer_price=item.wholesale_price,
                            selling_price=item.price,
                            inventory=item.product,
                            quantity=quantity_change,
                            system_quantity=item.received_quantity,  
                            description=f'Stock in from batch {item.purchase_order.batch}',
                            total_quantity=item.product.quantity + item.received_quantity,
                        )
                    )

                Inventory.objects.bulk_update(
                    inventory_updates, ['quantity', 'price', 'dealer_price', 'cost']
                )

                ActivityLog.objects.bulk_create(logs)

            return Response({'message': 'All purchase order items processed'}, status.HTTP_202_ACCEPTED)
        except Exception as e:
            return Response({'message': str(e)}, status.HTTP_406_NOT_ACCEPTABLE)

class InventoryTransfer(views.APIView):
    def inventory_transfers(request): 
        """
            transfers index. Shows transfers in and out of the branch and the total cost of items transfered in/out
        """
        # q = request.GET.get('q', '') 
        # branch_id = request.GET.get('branch', '')

        # transfer_items = TransferItems.objects.filter(
        #     Q(from_branch=request.user.branch) |
        #     Q(to_branch = request.user.branch),
        #     transfer__delete=False
        # ).annotate(
        #     total_amount = ExpressionWrapper(
        #         Sum(F('quantity') * F('product__cost')),
        #         output_field=FloatField()
        #     )
        # )
        
        transfers = Transfer.objects.filter(
            Q(branch=request.user.branch) |
            Q(transfer_to__in=[request.user.branch]),
            delete=False
        ).annotate(
            total_quantity=Sum('transferitems__quantity'),
            total_amount=ExpressionWrapper(
                Sum(F('transferitems__quantity') * F('transferitems__cost')),
                output_field=FloatField()
            )
        ).order_by('-time').distinct()

        logger.info(f'transfers: {transfers}')
        
        # if q:
        #     transfers = transfers.filter(Q(transfer_ref__icontains=q) | Q(date__icontains=q) )
            
        # if branch_id: 
        #     transfers = transfers.filter(transfer_to__id=branch_id)

        # total_transferred_value = (
        # transfer_items.annotate(total_value=F('quantity') * F('cost'))\
        #     .aggregate(total_sum=Sum('total_value'))['total_sum'] or 0
        # )

        # total_received_value = (
        # transfer_items.annotate(total_value=F('quantity') * F('cost'))\
        #     .aggregate(total_sum=Sum('total_value'))['total_sum'] or 0
        # )

        # logger.info(f'value: {total_transferred_value}, received {total_received_value}')
            
        return render(request, 'transfers.html', {
            'transfers': transfers
            # 'search_query': q, 
            # 'transfer_items':transfer_items,
            # 'transferred_value':total_transferred_value,
            # 'received_value':total_received_value,
            # 'hold_transfers_count':Transfer.objects.filter(
            #         Q(branch=request.user.branch) |
            #         Q(transfer_to__in=[request.user.branch]),
            #         delete=False, 
            #         hold=True
            #     ).count()
            }
        )

class PrintTransfer(views.APIView):
    def get(self, request, transfer_id):
        try:
            transfer = Transfer.objects.get(id=transfer_id)
            transfer_serializer = TransferSerializer(transfer)
            transfer_items = TransferItems.objects.filter(transfer=transfer).values()
        
            return Response({
                'date':datetime.datetime.now(),
                'transfer':transfer_serializer, 
                'transfer_items':transfer_items
            })
        except:
            return Response({'Transfer doesnt exists'}, status.HTTP_406_NOT_ACCEPTABLE)
        
    def post(self, request, transfer_id):
        try:
            transfer = Transfer.objects.get(id=transfer_id)
            transfer_items = TransferItems.objects.filter(transfer=transfer).values(
                'product__name',
                'product__price',
                'quantity',
                'to_branch__name', 
                'product__cost'
            )

            logger.info(transfer_items)
            return Response({'data':transfer_items}, status.HTTP_200_OK)
        except Exception as e:
            return Response({'meesage':f'{e}'}, status.HTTP_400_BAD_REQUEST)
        
class RecieveInventory(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        try:  
            transfer_id = request.data.get('id')
            quantity_received = int(request.data.get('quantity'))
            received = request.data.get('received')

            logger.info(f'transfer item data {received}')

            branch_transfer = get_object_or_404(TransferItems, id=transfer_id, to_branch=request.user.branch)
            transfer_obj = get_object_or_404(Transfer, id=branch_transfer.transfer.id, transfer_to=request.user.branch)

            if quantity_received > branch_transfer.quantity:
                return Response({'message': 'Quantity received cannot be more than quantity transferred'}, status.HTTP_400_BAD_REQUEST)

            if received:
                if quantity_received != branch_transfer.quantity:
                    branch_transfer.over_less_quantity = branch_transfer.quantity - quantity_received
                    branch_transfer.over_less = True
                    branch_transfer.save()

                product, created = Inventory.objects.get_or_create(
                    name=branch_transfer.product.name,
                    branch=request.user.branch,
                    defaults={
                        'cost': branch_transfer.cost,
                        'price': branch_transfer.price,
                        'quantity': quantity_received,
                        'dealer_price': branch_transfer.dealer_price,
                        'description': branch_transfer.product.description or '',
                        'category': branch_transfer.product.category,
                        'tax_type': branch_transfer.product.tax_type,
                        'stock_level_threshold': branch_transfer.product.stock_level_threshold,
                    }
                )

                if not created:
                    product.quantity += quantity_received
                    product.price = branch_transfer.price
                    product.cost = branch_transfer.cost
                    product.dealer_price = branch_transfer.dealer_price
                    product.save()

                logger.info('created')

                ActivityLog.objects.create(
                    branch=request.user.branch,
                    user=request.user,
                    action='stock in',
                    inventory=product,
                    dealer_price=product.dealer_price,
                    selling_price=product.price,
                    system_quantity=product.quantity,
                    quantity=quantity_received,
                    total_quantity=product.quantity,
                    product_transfer=branch_transfer,
                    description=f'received {quantity_received} out of {branch_transfer.quantity}'
                )

                product.batch += f'{branch_transfer.product.batch}, '
                logger.info(f'Product batch numbers: {product.batch}')
                product.save()
                logger.info('product received')

            branch_transfer.quantity_track = branch_transfer.quantity - quantity_received
            branch_transfer.receieved_quantity += quantity_received
            branch_transfer.received_by = request.user
            branch_transfer.received = True
            branch_transfer.description = f'received {quantity_received} out of {branch_transfer.quantity}'
            branch_transfer.save()

            transfer_obj.total_quantity_track -= quantity_received
            transfer_obj.save()

            if not transfer_obj.receive_status:
                transfer_obj.receive_status = True
                transfer_obj.save()

            return Response({'message': 'Product received successfully'}, status.HTTP_202_ACCEPTED)

        except TransferItems.DoesNotExist:
            return Response({'message': 'Invalid transfer ID'}, status.HTTP_400_BAD_REQUEST)
        except Transfer.DoesNotExist:
            return Response({'message': 'Transfer object not found'}, status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'message': str(e)}, status.HTTP_500_INTERNAL_SERVER_ERROR)

class OverListStock(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        search_query = request.GET.get('search_query', '')
        
        transfers =  TransferItems.objects.filter(to_branch=request.user.branch)

        if search_query:
            transfers = transfers.filter(
                Q(transfer__product__name__icontains=search_query)|
                Q(transfer__transfer_ref__icontains=search_query)|
                Q(date__icontains=search_query)
            )
        
        def activity_log(action, inventory, branch_transfer):
            ActivityLog.objects.create(
                branch = request.user.branch,
                user=request.user,
                action= action,
                inventory=inventory,
                quantity=branch_transfer.over_less_quantity,
                total_quantity=inventory.quantity,
                product_transfer=branch_transfer
            )
        
    
        data = request.data
        
        action = data['action']
        transfer_id = data['transfer_id']
        reason = data['reason']
        status_s = data['status']
        branch_loss = data['branch_loss']
        quantity= data['quantity']

        branch_transfer = get_object_or_404(transfers, id=transfer_id)
        transfer = get_object_or_404(Transfer, id=branch_transfer.transfer.id)
        product = Inventory.objects.get(id=branch_transfer.product.id, branch=request.user.branch)
    
        if int(branch_transfer.over_less_quantity) > 0:
            if action == 'write_off':    
                product.quantity += branch_transfer.over_less_quantity 
                product.save()
                
                description='write_off'
                
                activity_log('Stock in', product, branch_transfer)
                
                DefectiveProduct.objects.create(
                    product = product,
                    branch = request.user.branch,
                    quantity = quantity,
                    reason = reason,
                    status = status_s,
                    branch_loss = get_object_or_404(Branch, id=branch_loss),
                )
                
                product.quantity -= branch_transfer.over_less_quantity 
                branch_transfer.over_less_description = description
                branch_transfer.action_by = request.user
                branch_transfer.over_less = False
                
                transfer.defective_status = True
                
                transfer.save()
                branch_transfer.save()
                product.save()
                
                activity_log('write off', product, branch_transfer )     

                return Response({f'{product.name} write-off successfully'}, status.HTTP_200_OK)
            
            if action == 'accept':
                product.quantity += branch_transfer.over_less_quantity 
                product.save()
                description='returned back'
                activity_log('stock in', product, branch_transfer )
                
                branch_transfer.over_less = False
                branch_transfer.over_less_description = description
                branch_transfer.save()
                
                return Response({f'{product.name} accepted back successfully'}, status.HTTP_200_OK)
            
            if action == 'back':
                description=f'transfered to {branch_transfer.to_branch}'
                product.quantity += branch_transfer.over_less_quantity 
                product.save()
                
                activity_log('stock in', product, branch_transfer )
                
                branch_transfer.received = False
                branch_transfer.quantity = branch_transfer.over_less_quantity
                branch_transfer.save()
                
                product.quantity -= branch_transfer.over_less_quantity 
                product.save()
            
                activity_log('transfer', product, branch_transfer )
                
                branch_transfer.over_less = False
                branch_transfer.over_less_description = description
                branch_transfer.save()
                
                return Response(status.HTTP_200_OK)
            
            return Response(status.HTTP_400_BAD_REQUEST)
            
        return Response({'messsage':'Invalid form'}, status.HTTP_400_BAD_REQUEST)
    
class TransferDelete(views.APIView):
    def delete(self, request, transfer_id):
        try:
            logger.info(f'transfer id: {transfer_id}')
            transfer = get_object_or_404(Transfer, id=transfer_id)
            logger.info(f'transfer: {transfer}')

            if transfer.receive_status:
                return Response({'message':f'Cancel failed the transfer is already received.'}, status.HTTP_400_BAD_REQUEST)

            transfer_items = TransferItems.objects.filter(transfer=transfer)
            logger.info(f'transfer: {transfer_items}')

            with transaction.atomic():
                inventory_updates = []
                activity_logs = []
                quantity_updates = {}

                for item in transfer_items:
                    logger.info(f'From branch {item.from_branch}')

                    product = Inventory.objects.get(branch=item.from_branch, id=item.product.id)
                    logger.info(f'Product is: {product}')

                    if product.id not in quantity_updates:
                        quantity_updates[product.id] = {'product': product, 'increment': 0}

                    quantity_updates[product.id]['increment'] += item.quantity

                for update in quantity_updates.values():
                    product = update['product']
                    product.quantity += update['increment']
                    inventory_updates.append(product)

                    activity_logs.append(ActivityLog(
                        invoice=None,
                        product_transfer=None,  
                        branch=request.user.branch,
                        user=request.user,
                        action='transfer cancel',
                        inventory=product,
                        selling_price=None,  
                        dealer_price=None,  
                        quantity=update['increment'],
                        total_quantity=product.quantity,
                        description='Transfer cancelled'
                    ))

                # Perform bulk updates
                Inventory.objects.bulk_update(inventory_updates, ['quantity'])
                ActivityLog.objects.bulk_create(activity_logs)

                transfer.delete = True
                transfer.save()

            return Response(status.HTTP_200_OK)
        except Exception as e:
            return Response({'message':f'{e}'}, status.HTTP_400_BAD_REQUEST)
        
class AddTransferInventory(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        InventorySerializer(data = request.data)
        inventory = Inventory.objects.filter(branch=request.user.branch).values().order_by('-quantity')
        return Response({'inventory':inventory}, status.HTTP_201_CREATED)
    
class TransferDetails(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, transfer_id):
        transfer = TransferItems.objects.filter(id=transfer_id).values(
            'product__name', 'transfer__transfer_ref', 'quantity', 'price', 'from_branch__name', 'to_branch__name'
        )
        return Response(transfer, status.HTTP_200_OK)

class HeldTransferJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, transfer_id):
        transfer_items = Holdtransfer.objects.filter(transfer__id=transfer_id).values(
            'product__name',
            'from_branch__name',
            'to_branch__name',
            'quantity',
            'price',
            'cost',
            'dealer_price'
        )
        return Response(transfer_items, status.HTTP_200_OK)

class HeldTransfers(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        transfers = Transfer.objects.filter(
            Q(branch=request.user.branch),
            delete=False, 
            hold=True
        ).values()
        logger.info(f'held transfers: {transfers}')
        return Response({
            'transfers':transfers
        })
    
class ProcessHeldTransfer(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(request, transfer_id):
        try:
            transfer = Transfer.objects.get(
                id=transfer_id,
                branch=request.user.branch,
                delete=False, 
                hold=True
            )
            transfer_serializer = TransferSerializer(transfer)
            return Response({
                'transfer':transfer_serializer.data
            }, status.HTTP_200_OK)
        except:
            return Response({ f'Transfer with id {transfer_id} not found.'}, status.HTTP_400_BAD_REQUEST)
        
class InventoryPDF(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        category = request.data.get('category', '')

        return Response({'https://web-production-86a7.up.railway.app/inventory/inventory-pdf'})

class InventoryReport(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        view = request.data.get('view', '')
        choice = request.data.get('type', '') 
        time_frame = request.data.get('timeFrame', '')
        branch_id = request.data.get('branch', '')
        product_id = request.data.get('product', '')
        transfer_id = request.data.get('transfer_id', '')
        
        transfers = TransferItems.objects.filter().order_by('-date') 
        
        today = datetime.date.today()
        
        if choice in ['All', '', 'Over/Less']:
            transfers = transfers
        
        if product_id:
            transfers = transfers.filter(product__id=product_id)
        if branch_id:
            transfers = transfers.filter(to_branch_id=branch_id)
        
        def filter_by_date_range(start_date, end_date):
            return transfers.filter(date__range=[start_date, end_date])
        
        date_filters = {
            'All': lambda: transfers, 
            'today': lambda: filter_by_date_range(today, today),
            'yesterday': lambda: filter_by_date_range(today - timedelta(days=1), today - timedelta(days=1)),
            'this week': lambda: filter_by_date_range(today - timedelta(days=today.weekday()), today),
            'this month': lambda: transfers.filter(date__month=today.month, issue_date__year=today.year),
            'this year': lambda: transfers.filter(date__year=today.year),
        }
        
        
        if time_frame in date_filters:
            transfers = date_filters[time_frame]()
            
        if view:
            return Response(transfers.values(
                    'date',
                    'product__name', 
                    'price',
                    'quantity', 
                    'from_branch__name',
                    'from_branch__id',
                    'to_branch__id',
                    'to_branch__name',
                    'received_by__username',
                    'date_received',
                    'description',
                    'received',
                    'declined'
                ), 
                status.HTTP_200_OK
            )
        
        if transfer_id:
            return Response(transfers.filter(id=transfer_id).values(
                    'date',
                    'product__name', 
                    'price',
                    'quantity', 
                    'from_branch__name',
                    'from_branch__id',
                    'to_branch__id',
                    'to_branch__name',
                    'received_by__username',
                    'date_received',
                    'description',
                    'received',
                    'declined'
                ), 
                status.HTTP_200_OK
            )
        
        return Response({'https://web-production-86a7.up.railway.app/inventory/transfers-report'},status.HTTP_200_OK)

class AccessoriesView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, product_id):
        accessories = Accessory.objects.flter(main_product__id=product_id).values('id', 'main_product__name')
        return Response({'data': accessories}, status.HTTP_200_OK)