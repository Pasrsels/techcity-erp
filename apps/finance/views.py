import csv, json
from .models import *
from decimal import Decimal
from io import BytesIO
from apps.users.models import User
from apps.company.models import Branch
from .consumers import CashTransferConsumer 
from xhtml2pdf import pisa 
from django.views import View
from django.db.models import Q
from twilio.rest import Client
from datetime import timedelta
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.contrib import messages
from utils.utils import generate_pdf
from . utils import update_latest_due
from django.http import JsonResponse
from utils.utils import generate_pdf
from asgiref.sync import async_to_sync, sync_to_async
from apps.inventory.models import Inventory, Accessory
from channels.layers import get_channel_layer
import json, datetime, os, boto3, openpyxl 
from utils.account_name_identifier import account_identifier
from .tasks import (
    send_invoice_email_task, 
    send_account_statement_email, 
    send_quotation_email
)
from pytz import timezone as pytz_timezone 
from openpyxl.styles import Alignment, Font
from . utils import calculate_expenses_totals
from django.utils.dateparse import parse_date
from django.templatetags.static import static
from django.db.models import Sum, DecimalField
from apps.inventory.models import ActivityLog, Product
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.core.mail import send_mail, EmailMessage
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from . forms import (
    ExpenseForm, 
    ExpenseCategoryForm, 
    CurrencyForm, 
    InvoiceForm, 
    CustomerForm, 
    TransferForm, 
    CashWithdrawForm, 
    cashWithdrawExpenseForm,
    customerDepositsForm,
    customerDepositsRefundForm,
    cashDepositForm,
    IncomeCategoryForm
)
from django.contrib.auth import authenticate
from loguru import logger
from .tasks import send_expense_creation_notification
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from django.http import FileResponse
import io
from collections import defaultdict
from apps.pos.utils.receipt_signature import generate_receipt_data
from apps.pos.utils.submit_receipt_data import submit_receipt_data
from django.db.models.functions import Coalesce
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Max
from django.utils.dateparse import parse_date
from dotenv import load_dotenv
from apps.settings.models import OfflineReceipt, FiscalDay, FiscalCounter
from utils.zimra import ZIMRA
from utils.zimra_sig_hash import run
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Avg, F, Value, CharField, ExpressionWrapper
import datetime
from itertools import chain
from django.core.paginator import Paginator, EmptyPage
import imghdr, base64
from django.core.files.base import ContentFile
 
# load global zimra instance
zimra = ZIMRA()

load_dotenv()

from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from .models import CashUp, Invoice, Expense
from django.utils import timezone

def get_previous_month():
    first_day_of_current_month = datetime.datetime.now().replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    return last_day_of_previous_month.month

def get_current_month():
    return datetime.datetime.now().month

#image decoding method
def decode_base64_file(data):
    """
    Decodes a base64 file and returns a ContentFile.
    Assumes data is in the format: data:<mime>;base64,<data>
    """
    if not data:
        return None

    try:
        format, imgstr = data.split(';base64,')
        ext = format.split('/')[-1]
        if ext == 'jpeg':
            ext = 'jpg'

        file_name = f"{uuid.uuid4()}.{ext}"
        return ContentFile(base64.b64decode(imgstr), name=file_name)
    except Exception as e:
        logger.error("Failed to decode base64 image:")
        return None

def cashflow_list(request): # to be organised
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        branch = request.GET.get('branch')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        status = request.GET.get('status')

        cashups = CashUp.objects.all()
        
        if branch:
            cashups = cashups.filter(branch__id=int(branch))
        if start_date:
            cashups = cashups.filter(date__gte=start_date)
        if end_date:
            cashups = cashups.filter(date__lte=end_date)
        if status in ['true', 'false']:
            cashups = cashups.filter(status=(status == 'true'))

        html = render_to_string('cashflows/partials/cashup_cards.html', {'cash_ups': cashups})
        
        return JsonResponse({'html': html})

class Finance(View):
    # authentication loginmixin
    template_name = 'finance.html'

    def get(self, request, *args, **kwargs):

        if request.user.role == 'sales':
            return redirect('finance:expenses')
        
        balances = AccountBalance.objects.filter(branch=request.user.branch)
    
        recent_sales = Sale.objects.filter(transaction__branch=request.user.branch).order_by('-date')[:5]

        expenses_by_category = Expense.objects.values('category__name').annotate(
            total_amount=Sum('amount', output_field=DecimalField())
        )
        
        context = {
            'balances': balances,
            'recent_transactions': recent_sales,
            'expenses_by_category': expenses_by_category,
        }
        
        return render(request, self.template_name, context)
    
@login_required
def monthly_installments(request):
    installments = MonthlyInstallment.objects.all().select_related('invoice').order_by('-date')
    return JsonResponse({'installments': list(installments)})

@login_required
def laybys(request):
    laybys = layby.objects.all().select_related('invoice').order_by('-date')
    return JsonResponse({'laybys': list(laybys)})

@login_required
def expenses(request):
    form = ExpenseForm()
    cat_form = ExpenseCategoryForm()

    if request.method == 'GET':
        filter_button = request.GET.get('filter_button')
        filter_option = request.GET.get('filter', 'today')  
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        expenses = Expense.objects.filter(user=request.user).order_by('-issue_date')
        
        # filtered_expenses = filter_expenses(expenses, filter_option, start_date, end_date)

        if request.user.role == 'sale':
            expenses = expenses.filter(user=request.user)
        
        return render(request, 'expenses.html', 
            {
                'form':form,
                'cat_form':cat_form,
                'expenses':filter_expenses,
                'filter_option': expenses,
            }
        )
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
      
            name = data.get('name') 
            amount = data.get('amount')
            category = data.get('category')  
            payment_method = data.get('payment_method', 'cash')
            currency_id = data.get('currency', 'USD')
            branch = request.user.branch.id
            base64_image = data.get('receipt')
            image = decode_base64_file(base64_image)

            is_recurring = data.get('is_recurring') == 'true'
            recurrence_value = data.get('recurrence_value')
            recurrence_unit = data.get('recurrence_unit')
            
            logger.info(branch)

            # Validation
            if not all([name, amount, category, payment_method, currency_id, branch]):
                return JsonResponse({'success': False, 'message': 'Missing required fields.'})

            # Fetch related objects
            try:
                category = ExpenseCategory.objects.get(id=category)
            except ExpenseCategory.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'Category with ID {category} does not exist.'})

            currency = get_object_or_404(Currency, name__icontains='usd')
            branch = get_object_or_404(Branch, id=branch)

            # Get or create account and balance
            account_details = account_identifier(request, currency, payment_method)
            account_name = account_details['account_name']
            account_type = account_details['account_type']

            account, _ = Account.objects.get_or_create(
                name=account_name,
                type=account_type
            )

            account_balance, _ = AccountBalance.objects.get_or_create(
                account=account,
                currency=currency,
                defaults={
                    'currency': currency,
                    'branch': branch,
                    'balance': 0
                }
            )
            
            logger.info(account_balance.balance)

            if account_balance.balance < Decimal(amount):
                return JsonResponse({'success': False, 'message': f'{account_name} has insufficient balance.'})

            # Deduct from balance
            account_balance.balance -= Decimal(amount)
            account_balance.save()

            # Create Expense
            expense = Expense.objects.create(
                description=name,
                amount=amount,
                category=category,
                user=request.user,
                currency=currency,
                payment_method=payment_method ,
                branch=branch,
                is_recurring=is_recurring,
                recurrence_value=int(recurrence_value) if is_recurring and recurrence_value else None,
                recurrence_unit=recurrence_unit if is_recurring else None,
                receipt=image,
            )

            # Create Cashbook entry
            Cashbook.objects.create(
                amount=amount,
                expense=expense,
                currency=currency,
                credit=True,
                description=f'Expense ({expense.description[:20]})',
                branch=branch
            )

            # Send notification (to turn on)
            # send_expense_creation_notification.delay(expense.id)

            return JsonResponse({'success': True, 'message': 'Expense recorded successfully.'})

        except Exception as e:
            logger.exception("Error while recording expense:")
            return JsonResponse({'success': False, 'message': str(e)})
        
@login_required
def get_expenses(request):
    try:
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 10))

        expenses = Expense.objects.select_related('category', 'branch', 'currency') \
                                  .order_by('-issue_date') 
        paginator = Paginator(expenses, limit)

        try:
            paginated_expenses = paginator.page(page)
        except EmptyPage:
            return JsonResponse({
                'data': [],
                'has_next': False
            })

        results = []
        for expense in paginated_expenses:
            results.append({
                'id': expense.id,
                'created_at': expense.issue_date.isoformat(),
                'note': expense.description,
                'amount': float(expense.amount),
                'category': str(expense.category),
                'branch': expense.branch.name,
                'has_receipt': bool(expense.receipt),
                'receipt_url': expense.receipt.url if expense.receipt else None
            })

        return JsonResponse({
            'data': results,
            'has_next': paginated_expenses.has_next()
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def save_expense_split(request):
    try:
        data = json.loads(request.body)
        splits = data.get('splits')
        branch_id = data.get('branch_id')
        expense_id = data.get('expense_id', '')
        
        cash_up = CashUp.objects.filter(date=today, branch__id=int(branch_id)).first() # to put order by
        logger.info(f'Cash up: {cash_up}')
        
        cash_up_expenses = cash_up.expenses.all()
        
        if expense_id:
            record_expense(expense_id,cash_up_expenses, request)
            
        else:
            logger.info(f'Split expenses: {splits}, branch_id = {branch_id}')
            expenses = []
            if not cash_up:
                return JsonResponse({'success': False, 'message': 'No cash up record found for today'}, status=404)

            categories = ExpenseCategory.objects.all()
            
            logger.info(cash_up.expenses)

            for split in splits:
                logger.info(split)
                logger.info(split['amount'])
                exp_obj = cash_up_expenses.get(id=int(split['expense_id']))  # existing expense object
                
                new_expense = Expense(
                    amount=split['amount'],
                    payment_method=exp_obj.payment_method,
                    currency=exp_obj.currency,
                    category=categories.filter(id=split['category_id']).first(),
                    description=exp_obj.description,
                    user=request.user,  
                    branch_id=request.user.branch.id,
                    status=False,
                    # purchase_order=exp_obj.purchase_order,
                    receipt=exp_obj.receipt,
                    is_recurring=exp_obj.is_recurring,
                    recurrence_value= exp_obj.recurrence_value if exp_obj.recurrence_value else None,
                    recurrence_unit= exp_obj.recurrence if exp_obj.recurrence_unit else None
                )
                
                exp_obj.cash_up_status = True
                
                exp_obj.save()
                new_expense.save()
                
                expenses.append(new_expense.id)

        return JsonResponse({
            'success': True,
            'message': 'Expenses split and saved successfully',
            # 'expenses': expenses
        })

    except Exception as e:
        logger.exception("Error in save_expense_split")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
        
def record_expense(expense_id, cash_up_expenses, request):
    logger.info(f'Straight recording of the expense: {expense_id}')
    exp_obj = cash_up_expenses.get(id=int(expense_id))  # existing expense object
            
    new_expense = Expense(
        amount=exp_obj.amount,
        payment_method=exp_obj.payment_method,
        currency=exp_obj.currency,
        category=exp_obj.category,
        description=exp_obj.description,
        user=request.user,  
        branch_id=request.user.branch.id,
        status=False,
        purchase_order=exp_obj.purchase_order,
        receipt=exp_obj.receipt,
        is_recurring=exp_obj.is_recurring,
        recurrence_value= exp_obj.recurrence_value if exp_obj.recurrence_value else None,
        recurrence_unit= exp_obj.recurrence if exp_obj.recurrence_unit else None
    )
    
    exp_obj.cash_up_status = True
    
    exp_obj.save()
    new_expense.save()

@login_required  
def get_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    data = {
        'id': expense.id,
        'amount': expense.amount,
        'description': expense.description,
        'category': expense.category.id
    }
    return JsonResponse({'success': True, 'data': data})

@login_required      
def add_or_edit_expense(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            amount = data.get('amount')
            description = data.get('description')
            category_id = data.get('category')
            expense_id = data.get('id')

            if not amount or not description or not category_id:
                return JsonResponse({'success': False, 'message': 'Missing fields: amount, description, category.'})
            
            category = get_object_or_404(ExpenseCategory, id=category_id)

            if expense_id:  
                expense = get_object_or_404(Expense, id=expense_id)
                before_amount = expense.amount
                
                expense.amount = amount
                expense.description = description
                expense.category = category
                expense.save()
                message = 'Expense successfully updated'
            
                try:
                    cashbook_expense = Cashbook.objects.get(expense=expense)
                    expense_amount = Decimal(expense.amount)
                    if cashbook_expense.amount < expense_amount:
                        cashbook_expense.amount = expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'Expense (update from {before_amount} to {cashbook_expense.amount})'
                    else:
                        cashbook_expense.amount -= cashbook_expense.amount - expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'(update from {before_amount} to {cashbook_expense.amount})'
                    cashbook_expense.save()
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)}, status=400)
            return JsonResponse({'success': True, 'message': message}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
@transaction.atomic()
def add_expense_category(request):
    subcategories = ExpenseCategory.objects.filter(parent__isnull=False).values(
        'id',
        'name'
    )

    logger.info(subcategories)
    

@login_required
def add_expense_category(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

        category_name = data.get('name', '')
        parent_name = data.get('parent', '')
        new_parent_name = data.get('new_parent', '')

        if not category_name:
            return JsonResponse({'success': False, 'message': 'Category name is required.'}, status=400)

        parent_obj = None
        if new_parent_name:
            parent_obj, _ = ExpenseCategory.objects.get_or_create(name=new_parent_name, parent=None)
            logger.info(f'New parent created or found: {parent_obj}')
        elif parent_name:
            parent_obj = ExpenseCategory.objects.filter(name=parent_name, parent=None).first()
            if not parent_obj:
                return JsonResponse({'success': False, 'message': f'Parent category "{parent_name}" not found.'}, status=404)
            logger.info(f'Existing parent found: {parent_obj}')

        if ExpenseCategory.objects.filter(name=category_name, parent=parent_obj).exists():
            return JsonResponse({
                'success': False,
                'message': f'Category "{category_name}" already exists under this parent.'
            }, status=400)

        # Create new child category
        new_category = ExpenseCategory.objects.create(name=category_name, parent=parent_obj)
        logger.info(f'New category created: {new_category} under parent: {parent_obj}')

        return JsonResponse({
            'success': True,
            'id': new_category.id,
            'name': new_category.name
        }, status=201)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

@login_required
def income(request):
    """returns income both from the cash sales and income deposited"""
    pass

def filter_expenses(queryset, filter_option, start_date=None, end_date=None):
    """
    Filter expense queryset based on specified time period
    """
    today = timezone.localtime().date()
    
    if filter_option == 'today':
        return queryset.filter(date=today)
    
    elif filter_option == 'yesterday':
        yesterday = today - timedelta(days=1)
        return queryset.filter(date=yesterday)
    
    elif filter_option == 'this_week':
        start_of_week = today - timedelta(days=today.weekday())  
        end_of_week = start_of_week + timedelta(days=6) 
        return queryset.filter(date__gte=start_of_week, date__lte=end_of_week)
    
    elif filter_option == 'last_week':
        end_of_last_week = today - timedelta(days=today.weekday() + 1)  
        start_of_last_week = end_of_last_week - timedelta(days=6)  
        return queryset.filter(date__gte=start_of_last_week, date__lte=end_of_last_week)
    
    elif filter_option == 'this_month':
        return queryset.filter(date__year=today.year, date__month=today.month)
    
    elif filter_option == 'last_month':
        last_month = today.replace(day=1) - timedelta(days=1)
        return queryset.filter(date__year=last_month.year, date__month=last_month.month)
    
    elif filter_option == 'this_year':
        return queryset.filter(date__year=today.year)
    
    elif filter_option == 'last_year':
        return queryset.filter(date__year=today.year - 1)
    
    elif filter_option == 'custom' and start_date and end_date:
        try:
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                
            return queryset.filter(date__gte=start_date, date__lte=end_date)
        except (ValueError, TypeError):
            return queryset

    return queryset

@login_required
@transaction.atomic
def delete_expense(request, expense_id):
    if request.method == 'DELETE':
        try:
            expense = get_object_or_404(Expense, id=expense_id)
            expense.cancel = True
            expense.save()
            
            Cashbook.objects.create(
                amount=expense.amount,
                debit=True,
                credit=False,
                description=f'Expense ({expense.description}): cancelled'
            )
            return JsonResponse({'success': True, 'message': 'Expense successfully deleted'})
        except Exception as e:
             return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
def update_expense_status(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            expense_id = data.get('id')
            status = data.get('status')

            expense = Expense.objects.get(id=expense_id)
            expense.status = status
            expense.save()

            return JsonResponse({'success': True, 'message': 'Status updated successfully.'})
        except Expense.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Expense not found.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def invoice(request):
    form = InvoiceForm()
    invoices = Invoice.objects.filter(branch=request.user.branch, status=True, cancelled=False).select_related(
        'branch',
        'currency',
        'user'
    ).order_by('-invoice_number')

    query_params = request.GET
    if query_params.get('q'):
        search_query = query_params['q']
        invoices = invoices.filter(
            Q(customer__name__icontains=search_query) |
            Q(invoice_number__icontains=search_query) |
            Q(issue_date__icontains=search_query)
        )

    user_timezone_str = request.user.timezone if hasattr(request.user, 'timezone') else 'UTC'
    user_timezone = pytz_timezone(user_timezone_str)  

    def filter_by_date_range(start_date, end_date):
        start_datetime = user_timezone.localize(
            timezone.datetime.combine(start_date, timezone.datetime.min.time())
        )
        end_datetime = user_timezone.localize(
            timezone.datetime.combine(end_date, timezone.datetime.max.time())
        )
        return invoices.filter(issue_date__range=[start_datetime, end_datetime])

    now = timezone.now().astimezone(user_timezone)
    today = now.date()

    now = timezone.now() 
    today = now.date()  
    
    date_filters = {
        'today': lambda: filter_by_date_range(today, today),
        'yesterday': lambda: filter_by_date_range(today - timedelta(days=1), today - timedelta(days=1)),
        't_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday()), today),
        'l_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday() + 7), today - timedelta(days=today.weekday() + 1)),
        't_month': lambda: invoices.filter(issue_date__month=today.month, issue_date__year=today.year),
        'l_month': lambda: invoices.filter(issue_date__month=today.month - 1 if today.month > 1 else 12, issue_date__year=today.year if today.month > 1 else today.year - 1),
        't_year': lambda: invoices.filter(issue_date__year=today.year),
    }

    if query_params.get('day') in date_filters:
        invoices = date_filters[query_params['day']]()

    total_partial = invoices.filter(payment_status='Partial').aggregate(Sum('amount'))['amount__sum'] or 0
    total_paid = invoices.filter(payment_status='Paid').aggregate(Sum('amount'))['amount__sum'] or 0
    total_amount = invoices.aggregate(Sum('amount'))['amount__sum'] or 0

    grouped_invoices = defaultdict(list)

    for invoice in invoices:

        issue_date = invoice.issue_date.date() 

        if issue_date == today:
            grouped_invoices['Today'].append(invoice)
        elif issue_date == today - timedelta(days=1):
            grouped_invoices['Yesterday'].append(invoice)
        else:
            grouped_invoices[issue_date.strftime('%A, %d %B %Y')].append(invoice)

    return render(request, 'invoices/invoice.html', {
        'form': form,
        'grouped_invoices': dict(grouped_invoices),
        'total_paid': total_paid,
        'total_due': total_partial,
        'total_amount': total_amount,
    })

@login_required
@transaction.atomic 
def update_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    customer_account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(
        CustomerAccountBalances, account=customer_account, currency=invoice.currency
    )

    if request.method == 'POST':
        data = json.loads(request.body)
        amount_paid = Decimal(data['amount_paid'])

        invoice = Invoice.objects.select_for_update().get(pk=invoice.pk)
        customer_account_balance = CustomerAccountBalances.objects.select_for_update().get(pk=customer_account_balance.pk)

        if amount_paid <= 0:
            return JsonResponse({'success': False, 'message': 'Invalid amount paid.'}, status=400)

        if amount_paid >= invoice.amount_due:
            invoice.payment_status = Invoice.PaymentStatus.PAID
            invoice.amount_due = 0
        else:
            invoice.amount_due -= amount_paid

        invoice.amount_paid += amount_paid
        
        # get the latest payment for the invoice
        latest_payment = Payment.objects.filter(invoice=invoice).order_by('-payment_date').first()
        if latest_payment:
            amount_due = latest_payment.amount_due - amount_paid 
        else:
            amount_due = invoice.amount - invoice.amount_paid 

        payment = Payment.objects.create(
            invoice=invoice,
            amount_paid=amount_paid,
            amount_due=amount_due, 
            payment_method=data['payment_method'],
            user=request.user
        )

        account, _ = Account.objects.get_or_create(
            name=f"{request.user.branch} {invoice.currency.name} {payment.payment_method.capitalize()} Account",
            type=Account.AccountType[payment.payment_method.upper()] 
        )
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=invoice.currency,
            branch=request.user.branch,
            defaults={'balance': 0}
        )

        account_balance.balance += amount_paid
        if customer_account_balance.balance < 0:
            customer_account_balance.balance += amount_paid
        else:
            customer_account_balance.balance -= amount_paid

        description = ''
        if invoice.hold_status:
            description = 'Held invoice payment'
            sale = Sale.objects.create(
                date=timezone.now(),
                transaction=invoice,
                total_amount=invoice.amount # invoice delivery amount
            )
            
            VATTransaction.objects.create(
                invoice=invoice,
                vat_type=VATTransaction.VATType.OUTPUT,
                vat_rate=VATRate.objects.get(status=True).rate,
                tax_amount=invoice.vat
            ) 

        else:
            description = 'Invoice payment update'
        
        Cashbook.objects.create(
            issue_date=invoice.issue_date,
            description=f'({description} {invoice.invoice_number})',
            debit=True,
            credit=False,
            amount=invoice.amount_paid,
            currency=invoice.currency,
            branch=invoice.branch
        )

        invoice.hold_status = False
        account_balance.save()
        customer_account_balance.save()
        invoice.save()
        payment.save()
        
        return JsonResponse({'success': True, 'message': 'Invoice successfully updated'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}) 

def update_invoice_amounts(invoice, amount_paid):
    invoice_payments = Payment.objects.filter(invoice=invoice)

    if amount_paid > 0:
        for payment in invoice_payments:
            amount_paid -= payment.amount_due
            payment.save()


@login_required
@transaction.atomic 
def create_invoice(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            invoice_data = data['data'][0]  
            items_data = data['items']
            layby_dates = data.get('layby_dates')
           
            # get currency
            currency = Currency.objects.get(id=invoice_data['currency'])
            
            # create or get accounts
            account_types = {
                'cash': Account.AccountType.CASH,
                'bank': Account.AccountType.BANK,
                'ecocash': Account.AccountType.ECOCASH,
            }

            account_name = f"{request.user.branch} {currency.name} {invoice_data['payment_method'].capitalize()} Account"
            
            account, _ = Account.objects.get_or_create(name=account_name, type=account_types[invoice_data['payment_method']])
            
            account_balance, _ = AccountBalance.objects.get_or_create(
                account=account,
                currency=currency,
                branch=request.user.branch,
                defaults={'balance': 0}  
            )

            logger.info(f"Account Balance: {account_balance}")

            # accounts_receivable
            accounts_receivable, _ = ChartOfAccounts.objects.get_or_create(name="Accounts Receivable")
            
            # VAT rate
            vat_rate = VATRate.objects.get(status=True)

            # customer
            customer = Customer.objects.get(id=int(invoice_data['client_id'])) 
            logger.info(customer)
            
            # customer account
            customer_account = CustomerAccount.objects.get(customer=customer)

            # customer Account + Balances
            customer_account_balance, _ = CustomerAccountBalances.objects.get_or_create(
                account=customer_account,
                currency=currency, 
                defaults={'balance': 0}
            )
            
            # amount_paid = update_latest_due(customer, Decimal(invoice_data['amount_paid']), request, invoice_data['paymentTerms'], customer_account_balance)
            amount_paid = Decimal(invoice_data['amount_paid'])
            
            logger.info(f'Amount paid: {amount_paid}')

            invoice_total_amount = Decimal(invoice_data['payable'])

            # prevent to record greater amount paid than the invoice amount 
            if amount_paid > invoice_total_amount:
                amount_paid = invoice_total_amount
                amount_due = 0
            else:
                amount_paid = amount_paid
                amount_due = invoice_total_amount - amount_paid  
                
            logger.info(f'amount due: {amount_due}')
        

            cogs = COGS.objects.create(amount=Decimal(0))
            
            products_purchased = f"""{', '.join([f'{item['product_name']} x {item['quantity']} ' for item in items_data])}"""
            
            logger.info(products_purchased)
            
            with transaction.atomic():
                invoice = Invoice.objects.create(
                    customer=customer,
                    issue_date=timezone.now(),
                    amount=invoice_total_amount,
                    amount_paid=amount_paid,
                    amount_due=amount_due,
                    vat=Decimal(invoice_data['vat_amount']),
                    payment_status = Invoice.PaymentStatus.PARTIAL if amount_due > 0 else Invoice.PaymentStatus.PAID,
                    branch = request.user.branch,
                    user=request.user,
                    currency=currency,
                    subtotal=invoice_data['subtotal'],
                    reocurring = invoice_data['recourring'],
                    payment_terms = invoice_data['paymentTerms'],
                    hold_status = invoice_data['hold_status'],
                    amount_received = amount_paid,
                    products_purchased = ''
                )
                
           
                category = IncomeCategory.objects.filter(name='sales').first() #to change
                
                income = Income.objects.create(
                    amount=invoice.amount_paid,
                    currency=invoice.currency,
                    category=category,
                    note=invoice.products_purchased,
                    user=invoice.user,
                    branch=invoice.branch,
                    status=False,
                )
                logger.info(f'Income created: {income}')

                # Create Finance Log
                FinanceLog.objects.create(
                    type='income',
                    category='sales',
                    amount=invoice.amount_paid,
                    description=invoice.products_purchased
                )
                logger.info(f'Finance log created for invoice: {invoice}')

                logger.info(f'Invoice created for customer: {invoice}')

                # check if invoice status is hold
                if invoice.hold_status == True:

                    logger.info(f'Processing held invoice: {invoice}')

                    held_invoice(items_data, invoice, request, vat_rate)

                    return JsonResponse({'hold':True, 'message':'Invoice succesfully on hold'})

                # create layby object
                if invoice.payment_terms == 'layby':

                    if amount_due > 0:

                        logger.info(f'Creating layby object for invoice: {invoice}')
                        
                        layby_obj = layby.objects.create(
                            invoice=invoice, 
                            branch=request.user.branch
                        )

                        layby_dates_list = []
                        number_of_dates = len(layby_dates)
                        
                        # calculate amount to be paid for each month
                        amount_per_due_date = (amount_due / number_of_dates) if number_of_dates > 0 else 0

                        logger.info(f'Amount per due date: {amount_per_due_date} : {number_of_dates} : {layby_dates}')

                        for date in layby_dates:

                            obj = laybyDates(
                                layby=layby_obj,
                                due_date=date,
                                amount_due=round(amount_per_due_date, 2),
                            )

                            layby_dates_list.append(obj)
                        
                        laybyDates.objects.bulk_create(layby_dates_list)

                        logger.info(f'Layby object created for invoice: {invoice}')
                
                # create monthly installment object
                if invoice.payment_terms == 'installment':

                    if invoice.reocurring:
                        MonthlyInstallment.objects.create(
                            invoice = invoice,
                            status = False
                        )
                    
                #create a paylater
                if invoice.payment_terms == 'pay later':
                    if amount_due > 0:
                        paylater_obj = Paylater.objects.create(
                            invoice=invoice,
                            amount_due=amount_due,
                            due_date=invoice_data['pay_later_dates'][0] if invoice_data['pay_later_dates'] else timezone.now().date(),
                            payment_method=invoice_data['payment_method']
                        )
                        
                        # Create paylater dates for each interval
                        if invoice_data['pay_later_dates']:
                            logger.info(f'amount_due: {amount_due}')
                            amount_per_interval = round(amount_due / len(invoice_data['pay_later_dates']), 2)
                            logger.info(f'amount_per_interval: {amount_per_interval}')
                            for date in invoice_data['pay_later_dates']:
                                paylaterDates.objects.create(
                                    paylater=paylater_obj,
                                    due_date=date,
                                    amount_due=amount_per_interval,
                                    payment_method=invoice_data['payment_method']
                                )

                # #create transaction
                Transaction.objects.create(
                    date=timezone.now(),
                    description=invoice.products_purchased,
                    account=accounts_receivable,
                    debit=Decimal(invoice_data['payable']),
                    credit=Decimal('0.00'),
                    customer=customer
                )

                logger.info(f'Creating transaction obj for invoice: {invoice}')
            
                # Create InvoiceItem objects
                invoice_items = []
                for item_data in items_data:
                    item = Inventory.objects.get(pk=item_data['inventory_id'])
                    
                    item.quantity -= item_data['quantity']
                    item.save()

                    invoice_items.append(
                        InvoiceItem.objects.create(
                            invoice=invoice,
                            item=item,
                            quantity=item_data['quantity'],
                            unit_price=item_data['price'],
                            vat_rate = vat_rate,
                            total_amount = int(item_data['quantity']) * float(item_data['price']),
                            cash_up_status = False
                        )
                    )

                    print(invoice_items)
                    
                    # cost of sales item
                    COGSItems.objects.get_or_create(
                        invoice=invoice,
                        defaults={'cogs': cogs, 'product': Inventory.objects.get(id=item.id, branch=request.user.branch)}
                    )
                
                    # stock log  
                    ActivityLog.objects.create(
                        branch=request.user.branch,
                        inventory=item,
                        user=request.user,
                        quantity = -item_data['quantity'],
                        total_quantity = item.quantity,
                        action='Sale',
                        invoice=invoice
                    )

                    accessories = Accessory.objects.filter(main_product=item).values('accessory_product', 'accessory_product__quantity')

                    # for acc in accessories:
                    #     COGSItems.objects.get_or_create(
                    #         invoice=invoice,
                    #         defaults={'cogs': cogs, 'product': Inventory.objects.get(id=acc['accessory_product'], branch=request.user.branch)}
                    #     )
                    #     prod_acc = Inventory.objects.get(id = acc['accessory_product'] )
                    #     prod_acc.quantity -= acc.quantity

                    #     logger.info(f'accessory quantity: {acc['accessory_product__quantity']}')

                    #     ActivityLog.objects.create(
                    #         branch=request.user.branch,
                    #         inventory=prod_acc,
                    #         user=request.user,
                    #         quantity=1,
                    #         total_quantity = acc['accessory_product__quantity'],
                    #         action='Sale',
                    #         invoice=invoice
                    #     )
                    #     prod_acc.save()
                        
                # # Create VATTransaction
                VATTransaction.objects.create(
                    invoice=invoice,
                    vat_type=VATTransaction.VATType.OUTPUT,
                    vat_rate=VATRate.objects.get(status=True).rate,
                    tax_amount=invoice_data['vat_amount']
                )                                                          
                # Create Sale object
                sale = Sale.objects.create(
                    date=timezone.now(),
                    transaction=invoice,
                    total_amount=invoice_total_amount
                )
                sale.save()
                
                #payment
                Payment.objects.create(
                    invoice=invoice,
                    amount_paid=amount_paid,
                    payment_method=invoice_data['payment_method'],
                    amount_due=invoice_total_amount - amount_paid,
                    user=request.user
                )

                # calculate total cogs amount
                cogs.amount = COGSItems.objects.filter(cogs=cogs, cogs__date=datetime.datetime.today())\
                                               .aggregate(total=Sum('product__cost'))['total'] or 0
                cogs.save()
                
                # updae account balance
                if invoice.payment_status == 'Partial':
                    customer_account_balance.balance += -amount_due
                    customer_account_balance.save()
                    
                # Update customer balance
                account_balance.balance = Decimal(invoice_data['payable']) + Decimal(account_balance.balance)
                account_balance.save()
                
                # for tax purpose Zimra
                logger.info(invoice_items)

                try:
                    # sig_data, receipt_data = generate_receipt_data(invoice, invoice_items, request)
                    # logger.info(sig_data)
                    # hash_sig_data = run(sig_data)
                    
                    # logger.info(hash_sig_data)
                    # submit_receipt_data(request, receipt_data, hash_sig_data['hash'], hash_sig_data['signature'])
                    
                    invoice_data = invoice_preview_json(request, invoice.id)
                    logger.info(invoice_data)

                except Exception as e:
                    logger.info(e)
                    return JsonResponse({'success': False, 'error': str(e)})

                logger.info(f'inventory creation successfully done: {invoice}')

                return JsonResponse({'success':True, 'invoice_id': invoice.id, 'invoice_data':invoice_data})

        # except (KeyError, json.JSONDecodeError, Customer.DoesNotExist, Inventory.DoesNotExist, Exception) as e:
        #     return JsonResponse({'success': False, 'error': str(e)})
        except Exception as e:
            logger.info(e)

    return render(request, 'invoices/add_invoice.html')

def held_invoice(items_data, invoice, request, vat_rate):
    for item_data in items_data:
        item = Inventory.objects.get(pk=item_data['inventory_id'])
        # product = Product.objects.get(pk=item.product.id)
        item.quantity -= item_data['quantity']
        item.save()
                  
        InvoiceItem.objects.create(
            invoice=invoice,
            item=item,
            quantity=item_data['quantity'],
            unit_price=item_data['price'],
            vat_rate = vat_rate
        )
                    
        # # Create StockTransaction for each sold item
        # stock_transaction = StockTransaction.objects.create(
        #     item=item,
        #     transaction_type=StockTransaction.TransactionType.SALE,
        #     quantity=item_data['quantity'],
        #     unit_price=item.price,
        #     invoice=invoice,
        #     date=timezone.now()
        # )
              
        # stock log  
        ActivityLog.objects.create(
            branch=request.user.branch,
            inventory=item,
            user=request.user,
            quantity=item_data['quantity'],
            total_quantity = item.quantity,
            action='Sale',
            invoice=invoice
        )
        
#credits

@login_required
def paylater(request):
    paylaters = Paylater.objects.all().select_related('invoice', 'invoice__customer').values(
        'id',
        'invoice__invoice_number',
        'invoice__customer__name',
        'due_date',
        'amount_due',
        'amount_paid',
        'paid'
    )
    return JsonResponse({'success':True, 'data':list(paylaters)})

def paylater_details(request, paylater_id):
    logger.info(f'paylater_id: {paylater_id}')
    paylater = Paylater.objects.filter(id=paylater_id).select_related('invoice', 'invoice__customer').values(
        'id',
        'invoice__invoice_number',
        'invoice__customer__name',
        'invoice__amount',
        'due_date',
        'amount_due',
        'invoice__amount_paid',
        'invoice__currency__symbol',
        'paid'
    )
    
    paylater_dates = paylaterDates.objects.filter(paylater=paylater_id).values(
        'id',
        'due_date',
        'amount_due',
        'amount_paid',
        'paid'
    )
    
    return JsonResponse({'success':True, 'data':list(paylater), 'payment_schedule':list(paylater_dates)})

@login_required
def submit_invoice_data_zimra(request):
    try:

        data = json.loads(request.body)
        hash = data.get('hash', '')
        signature = data.get('signature', '') 
        receipt_data = data.get('receipt_data')
        invoice_id = data.get('invoice_id')

        logger.info(receipt_data)

        if not hash:
            return JsonResponse({'success':False,'message':f'Hash data is missing!'}, status=400)

        if not signature:
            return JsonResponse({'success':False,'message':f'Signature data is missing!'}, status=400)
        
        try:
            submit_receipt_data(request, receipt_data, hash, signature)
            logger.info('done')
        except Exception as e:
            logger.info(e)
            return JsonResponse(
                {
                    'success':False,
                    'messsage':f'{e}'
                },
                status=400
            )
        
        invoice_data = invoice_preview_json(request, invoice_id)
        logger.info(invoice_data)

        return JsonResponse({'success':True, 'message':'data received', 'data':invoice_data}, status=200)
    except Exception as e:
        return JsonResponse({'message':f'{e}', 'success':False}, status=200)


@login_required
def get_signature_data(request):
    try:   
        data = json.loads(request.body)
        hash = data.get('hash', '')
        signature = data.get('signature', '') 

        if not hash:
            return JsonResponse({'success':False,'message':f'Hash data is missing!'}, status=400)

        if not signature:
            return JsonResponse({'success':False,'message':f'Signature data is missing!'}, status=400)

        return JsonResponse({'success':True, 'message':'data received'}, status=200)
        
    except Exception as e:
        return JsonResponse({'success':False,'message':f'{e}'}, status=400)
    

@login_required
def held_invoice_view(request):
    form = InvoiceForm()
    invoices = Invoice.objects.filter(branch=request.user.branch, status=True, hold_status =True).order_by('-invoice_number')
    logger.info(f'Held invoices: {invoices}')
    return render(request, 'invoices/held_invoices.html', {'invoices':invoices, 'form':form})


def create_invoice_pdf(invoice):
    # Buffer to hold the PDF
    buffer = io.BytesIO()
    
    # Create the PDF object, using the buffer as its "file."
    pdf = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    heading_style = styles['Heading1']
    
    # Company logo and header
    elements.append(Paragraph('<b>Tech City</b>', heading_style))
    elements.append(Paragraph('See • Touch • Own Quality', normal_style))
    elements.append(Paragraph(f'Invoice Number: {invoice.invoice_number}', normal_style))
    
    elements.append(Spacer(1, 12))
    
    # Table Data (Items)
    data = [
        ['Q.', 'Description', 'Amount'],
        [1, 'Hp (hp 250)', 'USD 250.00'],
        ['Sub Total', '', 'USD 250.00'],
        ['Discount', '', 'USD 0.00'],
        ['VAT @15%', '', 'USD 37.50'],
        ['Delivery Charge', '', 'USD 0.00'],
        ['Previous Due', '', 'USD 75.00'],
        ['Current Due', '', 'USD 287.50'],
        ['Total Balance', '', 'USD 362.50'],
        ['Amount Paid', '', 'USD 362.50'],
        ['Due Amount', '', 'USD 0.00']
    ]
    
    # Create table
    table = Table(data, colWidths=[0.5*inch, 3*inch, 1.5*inch])
    
    # Add style to the table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
    ]))
    
    elements.append(table)
    
    elements.append(Spacer(1, 24))
    
    # Terms and conditions
    terms = '''
    All laptop in-built batteries attract 1 month warranty.
    Non in-built batteries attract 48hrs warranty.
    Warranty for all preowned laptops is 5 months. Tech City does not warranty laptops if damaged by water, liquids, or short circuits.
    Any withdrawn deposits for any purchase will attract 10percent administration fee.
    Tech City only accepts exchanges on faulty laptops.
    '''
    elements.append(Paragraph('Terms and Conditions', heading_style))
    elements.append(Paragraph(terms, normal_style))
    
    elements.append(Spacer(1, 12))
    elements.append(Paragraph('Thanks for your purchase!', normal_style))
    
    # Build PDF
    pdf.build(elements)
    
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'Invoice_{invoice.invoice_number}.pdf')


@login_required
@transaction.atomic
def invoice_returns(request, invoice_id): # dont forget the payments
    invoice = get_object_or_404(Invoice, id=invoice_id)
    account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

    sale = get_object_or_404(Sale, transaction=invoice)
    invoice_payment = get_object_or_404(Payment, invoice=invoice)
    stock_transactions = invoice.stocktransaction_set.all()  
    vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)
    activity = ActivityLog.objects.filter(invoice=invoice)

    if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
        customer_account_balance.balance -= invoice.amount_due

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account = get_object_or_404(
        Account, 
        name=f"{request.user.branch} {invoice.currency.name} {invoice_payment.payment_method.capitalize()} Account", 
        type=account_types.get(invoice_payment.payment_method, None) 
    )
    account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
    account_balance.balance -= invoice.amount_paid

    for stock_transaction in activity:
        product = Inventory.objects.get(product=stock_transaction.inventory, branch=request.user.branch)
        product.quantity += stock_transaction.quantity
        product.save()

        logger.info(f'product quantity {product.quantity}')
        logger.info(f'stock quantity {stock_transaction.quantity}')

        ActivityLog.objects.create(
            invoice=invoice,
            product_transfer=None,
            branch=request.user.branch,
            user=request.user,
            action='returns',
            inventory=product,
            quantity=stock_transaction.quantity,
            total_quantity=product.quantity
        )

    InvoiceItem.objects.filter(invoice=invoice).delete() 
    StockTransaction.objects.filter(invoice=invoice).delete()
    Payment.objects.filter(invoice=invoice).delete()

    account_balance.save()
    customer_account_balance.save()
    sale.delete()
    vat_transaction.delete()
    invoice.invoice_return=True
    invoice.save()

    return JsonResponse({'message': f'Invoice {invoice.invoice_number} successfully deleted'})
    

@login_required
@transaction.atomic
def delete_invoice(request, invoice_id):
    try:
        invoice = get_object_or_404(Invoice, id=invoice_id)
        account = get_object_or_404(CustomerAccount, customer=invoice.customer)
        customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

        sale = get_object_or_404(Sale, transaction=invoice)
        payments = Payment.objects.filter(invoice=invoice)  
        vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)
        activity = ActivityLog.objects.filter(invoice=invoice)
        
        with transaction.atomic():
            if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
                customer_account_balance.balance -= invoice.amount_due

            account_types = {
                'cash': Account.AccountType.CASH,
                'bank': Account.AccountType.BANK,
                'ecocash': Account.AccountType.ECOCASH,
            }

            for payment in payments:
                account = get_object_or_404(
                    Account, 
                    name=f"{request.user.branch} {invoice.currency.name} {payment.payment_method.capitalize()} Account", 
                    type=account_types.get(payment.payment_method, None)
                )
                account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
                account_balance.balance -= payment.amount_due
                account_balance.save()

            for stock_transaction in activity:
                product = Inventory.objects.get(id=stock_transaction.inventory.id, branch=request.user.branch)
                product.quantity += abs(stock_transaction.quantity)
                product.save()

                logger.info(f'product quantity {stock_transaction.quantity}')

                ActivityLog.objects.create(
                    invoice=invoice,
                    product_transfer=None,
                    branch=request.user.branch,
                    user=request.user,
                    action='sale return',
                    inventory=product,
                    quantity=stock_transaction.quantity,
                    total_quantity=product.quantity
                )

            InvoiceItem.objects.filter(invoice=invoice).delete() 
            StockTransaction.objects.filter(invoice=invoice).delete()
            payments.delete()
            customer_account_balance.save()
            sale.delete()
            vat_transaction.delete()
            invoice.cancelled = True
            invoice.save()

            logger.info(f'Invoice {invoice.invoice_number} successfully deleted')

        return JsonResponse({'success': True, 'message': f'Invoice {invoice.invoice_number} successfully deleted'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f"{e}"})
      
@login_required       
def invoice_details(request, invoice_id):
    invoice = Invoice.objects.filter(id=invoice_id, branch=request.user.branch).values(
        'invoice_number',
        'customer__id', 
        'customer__name', 
        'products_purchased', 
        'payment_status', 
        'amount'
    )
    return JsonResponse(list(invoice), safe=False)


@login_required
def layby_data(request):
    if request.method == 'GET':
        layby_data = layby.objects.all().select_related(
            'invoice',
            'branch'
        ).values()
        return JsonResponse(list(layby_data), safe=False)
    
    if request.method == 'POST':
        logger.info(f'layby data')
        data = json.loads(request.body)

        invoice_id = data.get('invoice_id')

        if not invoice_id:
            return JsonResponse({'success': False, 'message': 'Invoice ID is required.'})

        laby_dates = laybyDates.objects.filter(layby__invoice__id=invoice_id).values()
        
        logger.info(laby_dates)
        return JsonResponse({'success': True, 'data': list(laby_dates)})

@login_required
@transaction.atomic
def layby_payment(request, layby_date_id):
    try:
        data = json.loads(request.body)
        amount_paid = data.get('amount_paid')
        payment_method = data.get('payment_method')

        layby_date = laybyDates.objects.get(id=layby_date_id)
        layby_obj = layby.objects.get(id=layby_date.layby.id)
        invoice = layby_obj.invoice

        account = CustomerAccount.objects.get(customer=invoice.customer)
        customer_account_balance = CustomerAccountBalances.objects.get(account=account, currency=invoice.currency)
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        customer_account_balance.balance -= amount_paid

        account_name = f"{request.user.branch} {invoice.currency.name} {'cash'.capitalize()} Account"
        account = Account.objects.get(name=account_name, type=account_types['cash'])
        account_balance = AccountBalance.objects.get(account=account, currency=invoice.currency, branch=request.user.branch)

        account_balance.balance -= amount_paid

        amount_paid = layby_date.amount_paid
        amount_due = layby_date.amount_due

        with transaction.atomic():

            account_balance.save()
            customer_account_balance.save()

            # create a payment object
            Payment.objects.create(
                invoice=invoice,
                amount_paid=amount_paid,
                amount_due=amount_due, 
                payment_method=payment_method,
                user=request.user
            )

            # create a cash book object
            Cashbook.objects.create(
                issue_date=timezone.now(),
                description=f'Layby payment ({layby_date.layby.invoice.invoice_number})',
                debit=False,
                credit=True,
                amount=amount_paid,
                currency=layby_date.layby.invoice.currency,
                branch=request.user.branch
            )

            if amount_paid >= amount_due:
                layby_date.paid = True
                layby_date.save()
                layby_obj.fully_paid = True
                layby_obj.save()
                invoice.payment_status = Invoice.PaymentStatus.PAID
                invoice.save()

                layby.check_payment_status()

                return JsonResponse({'success': True, 'message': 'Layby payment successfully completed.'})
            else:
                return JsonResponse({'success': False, 'message': 'Invalid amount paid.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'{e}'})
    

@login_required
def customer(request):
    if request.method == 'GET':
        customers = Customer.objects.all().values()
        return JsonResponse(list(customers), safe=False)
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        
        # validation_errors = validate_customer_data(data)
        # if validation_errors:
        #     return JsonResponse({'success': False, 'message': 'Validation errors occurred:', 'errors': validation_errors})
    
        if Customer.objects.filter(phone_number=data['phonenumber']).exists():
            return JsonResponse({'success': False, 'message': 'Customer exists'})
        else:
            customer = Customer.objects.create(
                name=data['name'],
                email=data['email'],
                address=data['address'],
                phone_number=data['phonenumber'],
                branch=request.user.branch
            )
            account = CustomerAccount.objects.create(customer=customer)
            
            logger.info(account)

        balances_to_create = [
            CustomerAccountBalances(account=account, currency=currency, balance=0) 
            for currency in Currency.objects.all()
        ]
        CustomerAccountBalances.objects.bulk_create(balances_to_create)
    

        return JsonResponse({'success': True, 'message': 'Customer successfully created'})

    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

def validate_customer_data(data):
    errors = {}
    if 'name' not in data or len(data['name']) < 2:
        errors['name'] = 'Name is required and must be at least 2 characters long.'

    if 'email' not in data or not validate_email(data['email']):
        errors['email'] = 'A valid email address is required.'

    if 'address' not in data:
        errors['address'] = 'Address is required.'

    if 'phonenumber' not in data:
        errors['phonenumber'] = 'Phone number is required.'

    return errors

def validate_email(email):
    import re
    email_regex = r"^[a-zA-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[a-zA-Z0-9-]+(?:\.[a-zA-Z0-9-]+)*$"
    return bool(re.match(email_regex, email))

@login_required
def customer_list(request):
    search_query = request.GET.get('q', '')
    
    customers = Customer.objects.filter(branch=request.user.branch)
    accounts = CustomerAccountBalances.objects.all()
    
    total_balances_per_currency = CustomerAccountBalances.objects.filter(account__customer__branch=request.user.branch).values('currency__name').annotate(
        total_balance=Sum('balance')
    )
    
    if search_query:
        customers = CustomerAccount.objects.filter(Q(customer__name__icontains=search_query))
        
    if 'receivable' in request.GET:
        negative_balances_per_currency = CustomerAccountBalances.objects.filter(account__customer__branch=request.user.branch, balance__lt=0) \
            .values('currency') \
            .annotate(total_balance=Sum('balance'))

        customers = Customer.objects.filter(
            id__in=negative_balances_per_currency.values('account__customer_id'),
        ).distinct()
        
        total_balances_per_currency = negative_balances_per_currency.values('currency__name').annotate(
            total_balance=Sum('balance')
        )
        
        logger.info(f'Customers:{total_balances_per_currency.values}')

    if 'download' in request.GET: 
        customers = Customer.objects.all() 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=customers.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header_title in enumerate(['Customer Name', 'Phone Number', 'Email', 'Account Balance'], start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.alignment = header_alignment
            
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header_title), 20)

        customer_accounts = CustomerAccountBalances.objects.all()
        for customer in customer_accounts:
            worksheet.append(
                [
                    customer.account.customer.name, 
                    customer.account.customer.phone_number, 
                    customer.account.customer.email, 
                    customer.balance if customer.balance else 0,
                ]
            )  
            
        workbook.save(response)
        return response
        
    return render(request, 'customers/customers.html', {
        'customers':customers, 
        'accounts':accounts,
        'total_balances_per_currency':total_balances_per_currency,
    })

@login_required
def update_customer(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)

    if request.method == 'POST':  
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f'{customer.name} details updated successfully')  #
            return redirect('finance:customer_list')  
    else:
        form = CustomerForm(instance=customer)  

    return render(request, 'customers/update_customer.html', {'form': form, 'customer': customer}) 

def delete_customer(request, customer_id):
    if request.method == 'DELETE':
        customer = get_object_or_404(Customer, pk=customer_id)

        customer_name = customer.name  
        customer.delete()
        messages.success(request, f'{customer_name} deleted successfully.')
        return JsonResponse({'status': 'success', 'message': f'Customer {customer_name} deleted successfully.'})  
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})  
    

@login_required
def customer_account(request, customer_id):
    form = customerDepositsForm()
    refund_form = customerDepositsRefundForm()
    customer = get_object_or_404(Customer, id=customer_id)

    account = CustomerAccountBalances.objects.filter(account__customer=customer)

    invoices = Invoice.objects.filter(
        customer=customer, 
        branch=request.user.branch, 
        status=True
    )
    
    invoice_payments = Payment.objects.filter(
        invoice__branch=request.user.branch, 
        invoice__customer=customer
    ).order_by('-payment_date')

    filters = Q()
    if request.GET.get('q'):
        filters &= Q(payment_status=request.GET['q'])
    if request.GET.get('search_query'):
        search_query = request.GET['search_query']
        filters &= (Q(invoice_number__icontains=search_query) | Q(issue_date__icontains=search_query))

    invoices = invoices.filter(filters)

    if request.GET.get('email_bool'):
        send_account_statement_email(customer.id, request.user.branch.id, request.user.id)
        return JsonResponse({'message': 'Email sent'})

    return render(request, 'customer.html', {
        'form':form,
        'account': account,
        'invoices': invoices,
        'customer': customer,
        'refund_form':refund_form,
        'invoice_count': invoices.count(),
        'invoice_payments': invoice_payments,
        'paid': invoices.filter(payment_status='Paid').count(),  
        'due': invoices.filter(payment_status='Partial').count(), 
    })


@login_required
@transaction.atomic
def add_customer_deposit(request, customer_id):
    # payload
    """
        customer_id
        amount
        currency
        payment_method
        reason
        payment_reference
    """
    
    try: 
        # get payload
        data = json.loads(request.body)
        customer_id = data.get('customer_id')
        amount = data.get('amount')
        currency = data.get('currency')
        payment_method = data.get('payment_method')
        reason = data.get('reason')
        payment_reference = data.get('payment_reference')        
        
        # payment_reference validation
        if CustomerDeposits.objects.filter(payment_reference=payment_reference).exists():
            return JsonResponse(
                {
                    'success':False,
                    'message': f'Payment reference: {payment_reference} exists'
                }
            )   
                                                   
        # get currency
        currency = Currency.objects.get(id=currency)
        
        # get account types
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        
        account_name = f"{request.user.branch} {currency.name} {payment_method.capitalize()} Account"
        
        
        account, _ = Account.objects.get_or_create(name=account_name, type=account_types[payment_method])
        
        # get or create the account balances
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=currency,
            branch=request.user.branch,
            defaults={'balance': 0}  
        )
        
        account_balance.balance += Decimal(amount)
        account_balance.save()
        logger.info(f"[FINANCE]: deposit -> System {account}")
        
        # check if customer exits
        customer = get_object_or_404(Customer, id=customer_id)  
        logger.info(f"[FINANCE]: deposit -> customer {customer}")
        customer_account = CustomerAccount.objects.get(customer=customer)
        
        customer_account_bal_object, _ = CustomerAccountBalances.objects.get_or_create(
                account=customer_account,
                currency=currency, 
                defaults={'balance': 0}
        )  
        logger.info(f"[FINANCE]: deposit -> customer account object {customer_account_bal_object}")
        
        # effect customer deposit
        customer_deposit = CustomerDeposits.objects.create(
            customer_account=customer_account_bal_object,
            amount=amount,
            currency=currency,
            payment_method=payment_method,
            reason=reason,
            payment_reference=payment_reference,
            cashier=request.user,
            branch=request.user.branch
        )
        
        # effect customer account balances
        customer_account_bal_object.balance += amount
        
        customer_account_bal_object.save()
        
        Cashbook.objects.create(
            issue_date=customer_deposit.date_created,
            description=f'{customer_deposit.payment_method.upper()} deposit ({customer_deposit.customer_account.account.customer.name})',
            debit=True,
            credit=False,
            amount=customer_deposit.amount,
            currency=customer_deposit.currency,
            branch=customer_deposit.branch
        )

        return JsonResponse(
            {
                "success":True,
                "message": f"Customer Deposit of {currency} {amount:2f} has been successfull",
            },
            status=200
        )
    except Exception as e:
        return JsonResponse(
            {
                "message": f"{e}",
                'success':False
            },status=500)


@login_required    
def deposits_list(request):
    deposits = CustomerDeposits.objects.filter(branch=request.user.branch).order_by('-date_created')
    return render(request, 'deposits.html', {
        'deposits':deposits,
        'total_deposits': deposits.aggregate(Sum('amount'))['amount__sum'] or 0,
    })

@login_required
@transaction.atomic
def refund_customer_deposit(request, deposit_id):
    try:
        deposit = CustomerDeposits.objects.get(id=deposit_id)
    except CustomerDeposits.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Deposit not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        amount = Decimal(data.get('amount', 0))
        if amount <= 0:
            return JsonResponse({'success': False, 'message': 'Invalid amount'}, status=400)
    except (json.JSONDecodeError, TypeError, ValueError):
        return JsonResponse({'success': False, 'message': 'Invalid input data'}, status=400)

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"

    try:
        account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
        account_balance = AccountBalance.objects.get(
            account=account,
            currency=deposit.currency,
            branch=request.user.branch,
        )
    except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    if amount > deposit.amount:
        return JsonResponse({'success': False, 'message': 'Refund amount exceeds deposit amount'}, status=400)
    
    account_balance.balance -= amount
    diff_amount = deposit.amount - amount

    if diff_amount == 0:
        deposit.delete()
    else:
        deposit.amount = diff_amount
        deposit.save()

    Cashbook.objects.create(
        issue_date=datetime.date.today(),
        description=f'{deposit.payment_method.upper()} deposit refund ({deposit.customer_account.account.customer.name})',
        debit=False,
        credit=True,
        amount=amount,
        currency=deposit.currency,
        branch=deposit.branch
    )

    account_balance.save()

    return JsonResponse({'success': True}, status=200)

        
@login_required
@transaction.atomic
def edit_customer_deposit(request, deposit_id):
    try:
        deposit = CustomerDeposits.objects.get(id=deposit_id)
    except CustomerDeposits.DoesNotExist:
        messages.warning(request, 'Deposit not found')
        return redirect('finance:customer_account', deposit.customer_account.account.customer.id)
    
    if request.method == 'POST':
        form = customerDepositsForm(request.POST)
        if not form.is_valid():
            messages.warning(request, 'Invalid form submission')
            return redirect('finance:edit_customer_deposit', deposit_id)

        amount = Decimal(request.POST.get('amount'))
        if amount <= 0:
            messages.warning(request, 'Amount cannot be zero or negative')
            return redirect('finance:edit_customer_deposit', deposit_id)

        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"
        
        try:
            account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
            account_balance = AccountBalance.objects.get(
                account=account,
                currency=deposit.currency,
                branch=request.user.branch,
            )
        except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
            messages.warning(request, str(e))
            return redirect('finance:edit_customer_deposit', deposit_id)
        
        adj_amount = amount - deposit.amount

        if adj_amount != 0:
            if adj_amount > 0:
                account_balance.balance += adj_amount
                debit, credit = True, False
            else:
                account_balance.balance += adj_amount 
                debit, credit = False, True

            Cashbook.objects.create(
                issue_date=datetime.date.today(),
                description=f'{deposit.payment_method.upper()} deposit adjustment ({deposit.customer_account.account.customer.name})',
                debit=debit,
                credit=credit,
                amount=abs(adj_amount),
                currency=deposit.currency,
                branch=deposit.branch
            )

            account_balance.save()
            deposit.amount = amount
            deposit.save()
            messages.success(request, 'Customer deposit successfully updated')
            return redirect('finance:customer', deposit.customer_account.account.customer.id)
    else:
        form = customerDepositsForm(instance=deposit)

    return render(request, 'customers/edit_deposit.html', {'form': form})
    

@login_required
def customer_deposits(request): 
    customer_id = request.GET.get('customer_id')
    
    if customer_id: 
        deposits = CustomerDeposits.objects.filter(branch=request.user.branch).values(
            'customer_account__account__customer_id',
            'date_created',
            'amount', 
            'reason',
            'currency__name', 
            'currency__symbol', 
            'payment_method',
            'payment_reference',
            'cashier__username', 
            'id'
        ).order_by('-date_created')
        return JsonResponse(list(deposits), safe=False)
    else:
        return JsonResponse({
            'success':False,
            'message':f'{customer_id} was not provided'
        })

@login_required
def customer_account_transactions_json(request):
    customer_id = request.GET.get('customer_id')
    transaction_type = request.GET.get('type')

    customer = get_object_or_404(Customer, id=customer_id)  

    if transaction_type == 'invoices':
        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        ).order_by('-issue_date').values(
            'issue_date',
            'invoice_number',
            'products_purchased', 
            'amount_paid', 
            'amount_due', 
            'amount', 
            'user__username',
            'payment_status'
        )
        return JsonResponse(list(invoices), safe=False)
    else:
        return JsonResponse({'message': 'Invalid transaction type.'}, status=400)  

@login_required
def customer_account_payments_json(request):
    customer_id = request.GET.get('customer_id')
    transaction_type = request.GET.get('type')

    customer = get_object_or_404(Customer, id=customer_id)

    if transaction_type == 'invoice_payments':
        invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
            invoice__branch=request.user.branch, 
            invoice__customer=customer
        ).order_by('-payment_date').values(
            'invoice__products_purchased',
            'payment_date',
            'invoice__invoice_number',
            'invoice__currency__symbol', 
            'invoice__payment_status',
            'invoice__amount_due',
            'invoice__amount', 
            'user__username', 
            'amount_paid', 
            'amount_due'
        )
        return JsonResponse(list(invoice_payments), safe=False)
    else:
        return JsonResponse({'message': 'Invalid transaction type.'}, status=400)  


@login_required
def customer_account_json(request, customer_id):
    account = CustomerAccountBalances.objects.filter(account__customer__id=customer_id).values(
        'currency__symbol', 'balance'
    )   
    return JsonResponse(list(account), safe=False)

@login_required
def print_account_statement(request, customer_id):
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        
        account = CustomerAccountBalances.objects.filter(account__customer=customer)
        
        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        )
    except:
        messages.warning(request, 'Error in processing the request')
        return redirect('finance:customer')

    invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
        invoice__branch=request.user.branch, 
        invoice__customer=customer
    ).order_by('-payment_date')
    
    return render(request, 'customers/print_customer_statement.html', {
        'customer':customer,
        'account':account,
        'invoices':invoices, 
        'invoice_payments':invoice_payments
    })

# currency views  
@login_required  
def currency(request):
    return render(request, 'currency/currency.html')

@login_required
def currency_json(request):
    currency_id = request.GET.get('id', '')
    currency = Currency.objects.filter(id=currency_id).values()
    return JsonResponse(list(currency), safe=False)


@login_required
def add_currency(request):
    if request.method == 'POST':
        form = CurrencyForm(request.POST)
        if form.is_valid():
            default = request.POST['default']
            try:
                form.save()
                messages.success(request, 'Currency added successfully!')  
            except Exception as e: 
                messages.error(request, f'Error adding currency: {e}')
            return redirect('finance:currency') 
    else:
        form = CurrencyForm()

    return render(request, 'currency/currency_add.html', {'form': form})


@login_required
def update_currency(request, currency_id):
    currency = get_object_or_404(Currency, id=currency_id)  

    if request.method == 'POST': 
        form = CurrencyForm(request.POST, instance=currency)  
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Currency updated successfully') 
            except Exception as e: 
                messages.error(request, f'Error updating currency: {e}')
            return redirect('finance:currency')  
    else:
        form = CurrencyForm(instance=currency) 

    return render(request, 'currency/currency_add.html', {'form': form})

@login_required
def delete_currency(request, currency_id):
    if request.method == 'POST': 
        currency = get_object_or_404(Currency, id=currency_id)
        
        try:
            if currency.invoice_set.exists() or currency.accountbalance_set.exists() or currency.expense_set.exists():  
                raise Exception("Currency is in use and cannot be deleted.")

            currency.delete()
            return JsonResponse({'message': 'Currency deleted successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'message':'Deletion Failed'})


@login_required
def finance_settings(request):
    return render(request, 'settings/settings.html')
    
# Reports
@login_required
def expenses_report(request):
    
    template_name = 'reports/expenses.html'
    
    search = request.GET.get('search', '')
    start_date_str = request.GET.get('startDate', '')
    end_date_str = request.GET.get('endDate', '')
    category_id = request.GET.get('category', '')
   
    if start_date_str and end_date_str:
        try:
            end_date = datetime.date.fromisoformat(end_date_str)
            start_date = datetime.date.fromisoformat(start_date_str)
        except ValueError:
            return JsonResponse({'messgae':'Invalid date format. Please use YYYY-MM-DD.'})
    else:
        start_date = ''
        end_date= ''
        
    try:
        category_id = int(category_id) if category_id else None
    except ValueError:
        return JsonResponse({'messgae':'Invalid category or search ID.'})

    expenses = Expense.objects.all()  
    
    if search:
        expenses = expenses.filter(Q('amount=search'))
    if start_date:
        start_date = parse_date(start_date_str)
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        end_date = parse_date(end_date_str)
        expenses = expenses.filter(date__lte=end_date)
    if category_id:
        expenses = expenses.filter(category__id=category_id)
    
    return generate_pdf(
        template_name,
        {
            'title': 'Expenses', 
            'date_range': f"{start_date} to {end_date}", 
            'report_date': datetime.date.today(),
            'total_expenses':calculate_expenses_totals(expenses),
            'expenses':expenses
        }
    )


@login_required 
def invoice_preview(request, invoice_id):
    invoice = Invoice.objects.get(id=invoice_id)
    invoice_items = InvoiceItem.objects.filter(invoice=invoice)
    return render(request, 'Pos/printable_receipt.html', {'invoice_id':invoice_id, 'invoice':invoice, 'invoice_items':invoice_items})

@login_required
def remove_item(request, item_id):
    if request.method == 'DELETE':
        try:
            item = InvoiceItem.objects.get(id=item_id)
            item.delete()
            return JsonResponse({'success': True})
        except InvoiceItem.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Item not found'}, status=404)

@login_required
def replace_item(request, item_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        new_item_id = data.get('newItemId')

        try:
            item = InvoiceItem.objects.get(id=item_id)
            new_item = InvoiceItem.objects.get(id=new_item_id)

            # Replace the old item with the new one
            item.name = new_item.name  # Update other relevant fields as needed
            item.price = new_item.price
            item.quantity = new_item.quantity
            item.save()

            return JsonResponse({'success': True})
        except (InvoiceItem.DoesNotExist, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid item'}, status=404)
        
@login_required
def invoice_preview_data(request, invoice_id):
    try:
        invoice = Invoice.objects.get(id=invoice_id)
    except Invoice.DoesNotExist:
        return JsonResponse({"error": "Invoice not found"}, status=404)

    dates = {}
    if invoice.payment_terms == 'layby':
        dates = laybyDates.objects.filter(layby__invoice=invoice).values('due_date')

    invoice_items = InvoiceItem.objects.filter(invoice=invoice).values(
        'item__name',
        'quantity',
        'item__description',
        'total_amount',
        'unit_price'
    )

    invoice_dict = {}
    invoice_dict['customer_name'] = invoice.customer.name
    invoice_dict['customer_email'] = invoice.customer.email
    invoice_dict['customer_cell'] = invoice.customer.phone_number
    invoice_dict['customer_address'] = invoice.customer.address
    invoice_dict['currency_symbol'] = invoice.currency.symbol
    invoice_dict['amount_paid'] = invoice.amount_paid
    invoice_dict['payment_terms'] = invoice.payment_terms
    invoice_dict['amount'] = invoice.amount
    invoice_dict['invoice_number'] = invoice.invoice_number
    invoice_dict['receipt_hash'] = invoice.receipt_hash
    invoice_dict['subtotal'] = invoice.subtotal
    invoice_dict['vat'] = round(invoice.vat, 2)
    invoice_dict['device_id'] = os.getenv("DEVICE_ID")
    invoice_dict['device_serial_number'] = os.getenv("DEVICE_SERIAL_NUMBER")
    invoice_dict['code'] =  invoice.code
    invoice_dict['fiscal_day'] = invoice.fiscal_day

    if invoice.branch:
        invoice_dict['branch_name'] = invoice.branch.name
        invoice_dict['branch_phone'] = invoice.branch.phonenumber
        invoice_dict['branch_email'] = invoice.branch.email

    invoice_dict['user_username'] = invoice.user.username
    invoice_dict['receipt_signature'] = invoice.receiptServerSignature if invoice.receiptServerSignature else None

    # Safely serialize qr_code
    if invoice.qr_code and hasattr(invoice.qr_code, 'url'):
        try:
            invoice_dict['qr_code'] = request.build_absolute_uri(invoice.qr_code.url)
            logger.info(invoice_dict['qr_code'])
        except Exception as e:
            invoice_dict['qr_code'] = None
            logger.info(f"Error generating QR code URL: {e}")
    else:
        invoice_dict['qr_code'] = None

    invoice_data = {
        'invoice': invoice_dict,
        'invoice_items': list(invoice_items),
        'dates': list(dates)
    }
    return JsonResponse(invoice_data, safe=False)
        
def invoice_preview_json(request, invoice_id):
    try:
        invoice = Invoice.objects.get(id=invoice_id)
    except Invoice.DoesNotExist:
        return JsonResponse({"error": "Invoice not found"}, status=404)

    dates = {}
    if invoice.payment_terms == 'layby':
        dates = laybyDates.objects.filter(layby__invoice=invoice).values('due_date')

    invoice_items = InvoiceItem.objects.filter(invoice=invoice).values(
        'item__name',
        'quantity',
        'item__description',
        'total_amount',
        'unit_price'
    )

    invoice_dict = {}
    invoice_dict['customer_name'] = invoice.customer.name
    invoice_dict['customer_email'] = invoice.customer.email
    invoice_dict['customer_cell'] = invoice.customer.phone_number
    invoice_dict['customer_address'] = invoice.customer.address
    invoice_dict['currency_symbol'] = invoice.currency.symbol
    invoice_dict['amount_paid'] = invoice.amount_paid
    invoice_dict['payment_terms'] = invoice.payment_terms
    invoice_dict['amount'] = invoice.amount
    invoice_dict['invoice_number'] = invoice.invoice_number
    invoice_dict['receipt_hash'] = invoice.receipt_hash
    invoice_dict['subtotal'] = invoice.subtotal
    invoice_dict['vat'] = round(invoice.vat, 2)
    invoice_dict['device_id'] = os.getenv("DEVICE_ID")
    invoice_dict['device_serial_number'] = os.getenv("DEVICE_SERIAL_NUMBER")
    invoice_dict['code'] =  invoice.code
    invoice_dict['fiscal_day'] = invoice.fiscal_day

    if invoice.branch:
        invoice_dict['branch_name'] = invoice.branch.name
        invoice_dict['branch_phone'] = invoice.branch.phonenumber
        invoice_dict['branch_email'] = invoice.branch.email

    invoice_dict['user_username'] = invoice.user.username
    invoice_dict['receipt_signature'] = invoice.receiptServerSignature if invoice.receiptServerSignature else None

    # Safely serialize qr_code
    if invoice.qr_code and hasattr(invoice.qr_code, 'url'):
        try:
            invoice_dict['qr_code'] = request.build_absolute_uri(invoice.qr_code.url)
            logger.info(invoice_dict['qr_code'])
        except Exception as e:
            invoice_dict['qr_code'] = None
            logger.info(f"Error generating QR code URL: {e}")
    else:
        invoice_dict['qr_code'] = None

    invoice_data = {
        'invoice': invoice_dict,
        'invoice_items': list(invoice_items),
        'dates': list(dates)
    }
    return invoice_data

@login_required
def invoice_pdf(request):
    template_name = 'reports/invoice.html'
    invoice_id = request.GET.get('id', '')
    if invoice_id:
        try:
            invoice = get_object_or_404(Invoice, pk=invoice_id)

            invoice_items = InvoiceItem.objects.filter(invoice=invoice)
            
        except Invoice.DoesNotExist:
            return HttpResponse("Invoice not found")
    else:
        return HttpResponse("Invoice ID is required")
    
    return generate_pdf(
        template_name,
        {
            'title': 'Invoice', 
            'report_date': datetime.date.today(),
            'invoice':invoice,
            'invoice_items':invoice_items
        }
    )
   
# emails
@login_required
def send_invoice_email(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        invoice_id = data['invoice_id']
        invoice = Invoice.objects.get(id=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        account = CustomerAccount.objects.get(customer__id = invoice.customer.id)
        
        html_string = render_to_string('Pos/receipt.html', {'invoice': invoice, 'invoice_items':invoice_items, 'account':account})
        buffer = BytesIO()

        pisa.CreatePDF(html_string, dest=buffer) 

        email = EmailMessage(
            'Your Invoice',
            'Please find your invoice attached.',
            'your_email@example.com',
            ['recipient_email@example.com'],
        )
        
        buffer.seek(0)
        email.attach(f'invoice_{invoice.invoice_number}.pdf', buffer.getvalue(), 'application/pdf')

        # Send the email
        email.send()

        task = send_invoice_email_task.delay(data['invoice_id']) 
        task_id = task.id 
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=invoice_{invoice.invoice_number}.pdf'
        
        return response
    return JsonResponse({'success': False, 'error':'error'})


#whatsapp
@login_required
def send_invoice_whatsapp(request, invoice_id):
    try:
        
        invoice = Invoice.objects.get(pk=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        img = settings.STATIC_URL + "/assets/logo.png"
    
        html_string = render_to_string('Pos/invoice_template.html', {'invoice': invoice, 'request':request, 'invoice_items':invoice_items, 'img':img})
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if not pisa_status.err:
          
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            invoice_filename = f"invoice_{invoice.invoice_number}.pdf"
            s3.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f"invoices/{invoice_filename}",
                Body=pdf_buffer.getvalue(),
                ContentType="application/pdf",
                ACL="public-read",
            )
            s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

            account_sid = 'AC6890aa7c095ce1315c4a3a86f13bb403'
            auth_token = '897e02139a624574c5bd175aa7aaf628'
            client = Client(account_sid, auth_token)
            from_whatsapp_number = 'whatsapp:' + '+14155238886'
            to_whatsapp_number = 'whatsapp:' + '+263778587612'

            message = client.messages.create(
                from_=from_whatsapp_number,
                body="Your invoice is attached.",
                to=to_whatsapp_number,
                media_url=s3_url
            )
            logger.info(f"WhatsApp message SID: {message.sid}")
            return JsonResponse({"success": True, "message_sid": message.sid})
        else:
            logger.error(f"PDF generation error for Invoice ID: {invoice_id}")
            return JsonResponse({"error": "PDF generation failed"})
    except Invoice.DoesNotExist:
        logger.error(f"Invoice not found with ID: {invoice_id}")
        return JsonResponse({"error": "Invoice not found"})
    except Exception as e:
        logger.exception(f"Error sending invoice via WhatsApp: {e}")
        return JsonResponse({"error": "Error sending invoice via WhatsApp"})
    
@login_required
@transaction.atomic()
def end_of_day(request):
    today = timezone.now().date()
    
    user_timezone_str = request.user.timezone if hasattr(request.user, 'timezone') else 'UTC'
    user_timezone = pytz_timezone(user_timezone_str)  

    def filter_by_date_range(start_date, end_date):
        start_datetime = user_timezone.localize(
            timezone.datetime.combine(start_date, timezone.datetime.min.time())
        )
        end_datetime = user_timezone.localize(
            timezone.datetime.combine(end_date, timezone.datetime.max.time())
        )
        return Invoice.objects.filter(branch=request.user.branch, issue_date__range=[start_datetime, end_datetime])

    now = timezone.now().astimezone(user_timezone)
    today = now.date()
    
    invoices = filter_by_date_range(today, today)
    withdrawals = CashWithdraw.objects.filter(user__branch=request.user.branch, date=today, status=False)
    
    total_cash_amounts = [
        {
            'total_invoices_amount': invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
            'total_withdrawals_amount': withdrawals.aggregate(Sum('amount'))['amount__sum'] or 0
        }
    ]

    sold_inventory = (
        ActivityLog.objects
        .filter(invoice__branch=request.user.branch, timestamp__date=today, action='Sale')
        .values('inventory__id', 'inventory__name')
        .annotate(quantity_sold=Sum('quantity'))
    )

    if request.method == 'GET':
        all_inventory = Inventory.objects.filter(branch=request.user.branch, status=True).values(
            'id', 'name', 'quantity'
        )

        inventory_data = []
        for item in sold_inventory:
            sold_info = next((inv for inv in all_inventory if item['inventory__id'] == inv['id']), None)
            
            if sold_info:
                inventory_data.append({
                    'id': item['inventory__id'],
                    'name': item['inventory__name'],
                    'initial_quantity': item['quantity_sold'] + sold_info['quantity'] if sold_info else 0,
                    'quantity_sold': item['quantity_sold'],
                    'remaining_quantity': sold_info['quantity'] if sold_info else 0,
                    'physical_count': None
                })
           
        return JsonResponse({'inventory': inventory_data, 'total_cash_amounts': total_cash_amounts})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            logger.info(data)
            inventory_data = []
            cashed_amount = data['cash_input']
            physical_counts = data['physical_counts']
            
            if not cashed_amount:
                return JsonResponse({'success': False, 'error': 'Cash input is required.'})
            
            if not physical_counts:
                return JsonResponse({'success': False, 'error': 'Physical counts are required.'})

            with transaction.atomic():
                for item in physical_counts:
                    try:
                        inventory = Inventory.objects.get(id=int(item['item_id']), branch=request.user.branch, status=True)
                        inventory.physical_count = item['physical_count']
                        inventory.save()

                        sold_info = next((i for i in sold_inventory if i['inventory__id'] == inventory.id), None)
                        inventory_data.append({
                            'id': inventory.id,
                            'name': inventory.name,
                            'initial_quantity': inventory.quantity,
                            'quantity_sold': sold_info['quantity_sold'] if sold_info else 0,
                            'remaining_quantity': inventory.quantity - (sold_info['quantity_sold'] if sold_info else 0),
                            'physical_count': inventory.physical_count,
                            'difference': inventory.physical_count - (inventory.quantity - (sold_info['quantity_sold'] if sold_info else 0))
                        })
                    except Inventory.DoesNotExist:
                        return JsonResponse({'success': False, 'error': f'Inventory item with id {item["inventory_id"]} does not exist.'})

                today_min = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
                today_max = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                
                # Invoice data
                invoices = Invoice.objects.filter(branch=request.user.branch, issue_date__range=(today_min, today_max))
                partial_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PARTIAL)
                paid_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PAID)
            
                # Expenses
                expenses = Expense.objects.filter(branch=request.user.branch, issue_date__range=(today_min, today_max))
                
                confirmed_expenses = expenses.filter(status=True)
                unconfirmed_expenses = expenses.filter(status=False)
                
                # Calculate totals for CashUp
                total_sales = paid_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0 
                total_partial = partial_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
                total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
                expected_cash = total_sales - total_expenses
                
                short_fall = expected_cash - Decimal(cashed_amount) # to take into account the total_partial sales 
               
                sales_person_cash_up = CashUp.objects.filter(
                    created_by=request.user, 
                    branch=request.user.branch, 
                    created_at__date=today, 
                    status=False
                ).first()
                
                logger.info(sales_person_cash_up)
                
                if sales_person_cash_up:
                    logger.info(f'Deleting cash up for salesperson: {request.user.username}')
                    sales_person_cash_up.delete()
                    
                # Create CashUp entry
                cashup = CashUp.objects.create(
                    date=today,
                    branch=request.user.branch,
                    expected_cash=expected_cash,
                    cashed_amount=cashed_amount,
                    received_amount=0,  
                    short_fall=short_fall,
                    balance=0,  
                    created_by=request.user,
                    status=False 
                )

                logger.info(cashup.expected_cash)

                items = Invoice.objects.filter(
                    payment_status=Invoice.PaymentStatus.PAID, 
                    issue_date__range=(today_min, today_max),
                    branch=request.user.branch
                )
                
                cashup.sales.set(items)

                cashup.expenses.set(expenses) #to change to confirmed expenses (later)

                # Create UserAccount object
                user_account, _ = UserAccount.objects.get_or_create(
                    user=request.user,
                    defaults={
                        'balance': Decimal('0.00'),
                        'total_credits': Decimal('0.00'),
                        'total_debits': Decimal('0.00'),
                        'last_transaction_date': timezone.now()
                    }
                )

                # Create UserTransaction object
                user_transaction = UserTransaction.objects.create(
                    account=user_account,
                    transaction_type='Cash',
                    amount=total_cash_amounts[0]['total_invoices_amount'] - total_cash_amounts[0]['total_withdrawals_amount'],
                    description='End of day transaction',
                    debit = expected_cash,
                    credit = 0,
                )

                # Update UserAccount balance
                user_account.balance += user_transaction.amount
                user_account.total_debits += user_transaction.amount
                user_account.last_transaction_date = timezone.now()
                user_account.save()


                # # Generate PDF report
                # html_string = render_to_string('day_report.html', {
                #     'request': request,
                #     'invoices': invoices,
                #     'expenses': expenses,
                #     'date': today,
                #     'inventory_data': inventory_data,
                #     'total_sales': total_sales,
                #     'partial_payments': total_partial,
                #     'total_paid_invoices': paid_invoices.count(),
                #     'total_partial_invoices': partial_invoices.count(),
                #     'total_expenses': total_expenses,
                #     'confirmed_expenses': confirmed_expenses,
                #     'unconfirmed_expenses': unconfirmed_expenses,
                #     'account_balances': AccountBalance.objects.filter(branch=request.user.branch),
                #     'cashup': cashup
                # })
                
                # pdf_buffer = BytesIO()
                # pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
                
                # if not pisa_status.err:
                #     filename = f"{request.user.branch.name}_today_report_{today}.pdf"
                #     return JsonResponse({
                #         "success": True,
                #         "cashup_id": cashup.id,
                #         "expected_cash": float(expected_cash)
                #     })
                # else:
                #     return JsonResponse({"success": False, "error": "Error generating PDF."}

                return JsonResponse({
                        "success": True,
                        "cashup_id": cashup.id,
                        # "expected_cash": float(expected_cash)
                })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data.'})
        except Exception as e:
            logger.exception(f"Error processing request: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def invoice_payment_track(request):
    invoice_id = request.GET.get('invoice_id', '')
    
    if invoice_id:
        payments = Payment.objects.filter(invoice__id=invoice_id).order_by('-payment_date').values(
            'payment_date', 'amount_paid', 'payment_method', 'user__username'
        )
    return JsonResponse(list(payments), safe=False)

@login_required
def day_report(request, inventory_data):
    today_min = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_max = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # invoice data
    invoices = Invoice.objects.filter(branch=request.user.branch, issue_date__range=(today_min, today_max))
    
    partial_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PARTIAL)
    paid_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PAID)
    
    # expenses
    expenses = Expense.objects.filter(branch=request.user.branch, date=datetime.date.today())
    
    confirmed_expenses = expenses.filter(status=True)
    unconfirmed_expenses = expenses.filter(staus=False)
    
    # accounts
    account_balances = AccountBalance.objects.filter(branch=request.user.branch)
    
    try:
        html_string = render_to_string('day_report.html',{
                'request':request,
                'invoices':invoices,
                'date': datetime.date.today(),
                'inventory_data': inventory_data,
                'total_sales': paid_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'partial_payments': partial_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'total_paid_invoices': paid_invoices.count(),
                'expenses':expenses,
                'total_partial_invoices': partial_invoices.count(),
                'total_expenses': confirmed_expenses.aggregate(Sum('amount'))['amount__sum'] or 0,
                'confirmed_expenses': confirmed_expenses,
                'unconfirmed_expenses': unconfirmed_expenses,
                'account_balances': account_balances,
            })
        
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if not pisa_status.err:
            # Save PDF to S3 and get URL
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            invoice_filename = f"{request.user.branch} today's ({datetime.date.today}) report.pdf"
            s3.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f"daily_reports/{invoice_filename}",
                Body=pdf_buffer.getvalue(),
                ContentType="application/pdf",
                ACL="public-read",
            )
            s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

            # Send WhatsApp Message with Twilio
            account_sid = settings.TWILIO_ACCOUNT_SID
            auth_token = settings.TWILIO_AUTH_TOKEN
            client = Client(account_sid, auth_token)
            from_whatsapp_number = 'whatsapp:' + '+14155238886'
            to_whatsapp_number = 'whatsapp:' + '+263778587612'

            message = client.messages.create(
                from_=from_whatsapp_number,
                body="Today's report.",
                to=to_whatsapp_number,
                media_url=s3_url
            )

            logger.info(f"WhatsApp message SID: {message.sid}")
            return JsonResponse({"success": True, "message_sid": message.sid})
    except Exception as e:
        logger.exception(f"Error sending invoice via WhatsApp: {e}")
        return JsonResponse({"error": "Error sending invoice via WhatsApp"})

   
@login_required 
@transaction.atomic 
def cash_transfer(request):
    form = TransferForm()
    transfers = CashTransfers.objects.filter(branch=request.user.branch).select_related(
        'user',
        'currency',
        'to',
        'branch',
        'from_branch'
    )
    
    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    if request.method == 'POST':
        form = TransferForm(request.POST)
        
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.user = request.user
            transfer.notification_type = 'Expense'
            transfer.from_branch = request.user.branch
            transfer.branch = request.user.branch
            transfer.received_status = False
            
            account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"
            logger

            with transaction.atomic():
                try:
                    account = Account.objects.get(name=account_name, type=account_types[transfer.transfer_method.lower()])
                    account_balance = AccountBalance.objects.select_for_update().get(
                        account=account,
                        currency=transfer.currency,
                        branch=request.user.branch
                    )
                
                    if account_balance.balance < transfer.amount:
                        messages.error(request, "Insufficient funds in the account.")
                        return redirect('finance:cash_transfer')  

                    account_balance.balance -= transfer.amount
                    account_balance.save()
                    transfer.save()  

                    Cashbook.objects.create(
                        issue_date=transfer.date,
                        description=f'Cash Transfer to {transfer.to.name}',
                        debit=False,
                        credit=True,
                        amount=transfer.amount,
                        currency=transfer.currency,
                        branch=transfer.branch
                    )
                    
                    messages.success(request, 'Money successfully transferred.')
                    return redirect('finance:cash_transfer')  
          
                except Exception as e:
                    messages.error(request, f"{e}")
                return redirect('finance:cash_transfer')  
        else:
            messages.error(request, "Invalid form data. Please correct the errors.")
    return render(request, 'transfers/cash_transfers.html', {'form': form, 'transfers':transfers})

@login_required
def finance_notifications_json(request):
    notifications = FinanceNotifications.objects.filter(status=True).values(
        'transfer__id', 
        'transfer__to',
        'expense__id',
        'expense__branch',
        'invoice__id',
        'invoice__branch',
        'notification',
        'notification_type',
        'id'
    )
    return JsonResponse(list(notifications), safe=False)


@login_required
@transaction.atomic
def cash_transfer_list(request):
    search_query = request.GET.get('q', '')
    transfers = CashTransfers.objects.filter(to=request.user.branch.id)
    
    if search_query:
        transfers = transfers.filter(Q(date__icontains=search_query))
        
    return render(request, 'transfers/cash_transfers_list.html', {'transfers':transfers, 'search_query':search_query})

@login_required
@transaction.atomic
def receive_money_transfer(request, transfer_id):
    if transfer_id:
        transfer = get_object_or_404(CashTransfers, id=transfer_id)
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        
        account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"

        with transaction.atomic():
            try:
                account, _ = Account.objects.get_or_create(name=account_name, type=account_types[transfer.transfer_method.lower()])
            
                account_balance, _ = AccountBalance.objects.get_or_create(
                    account=account,
                    currency=transfer.currency,
                    branch=request.user.branch
                )

                Cashbook.objects.create(
                    issue_date=transfer.date,
                    description=f'Cash Transfer from {transfer.from_branch.name}',
                    debit=True,
                    credit=False,
                    amount=transfer.amount,
                    currency=transfer.currency,
                    branch=transfer.to
                )

                account_balance.balance += transfer.amount
                account_balance.save()
                
                transfer.received_status = True
                transfer.save() 

                return JsonResponse({'message':True})  
            
            except Exception as e:
                return JsonResponse({'success':False, 'message':f"{e}"}) 
    return JsonResponse({'message':"Transfer ID is needed"})  


@login_required
@transaction.atomic
def create_quotation(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        qoute_data = data['data'][0]  
        items_data = data['items']
        
        customer = Customer.objects.get(id=int(qoute_data['client_id']))
        currency = Currency.objects.get(id=qoute_data['currency'])
        
        qoute = Qoutation.objects.create(
            customer = customer,
            amount =  Decimal(qoute_data['subtotal']),
            branch = request.user.branch,
            currency = currency,
            qoute_reference = Qoutation.generate_qoute_number(request.user.branch.name),
            products = ', '.join([f'{item['product_name']} x {item['quantity']}' for item in items_data])
        )
        
        for item_data in items_data:
            item = Inventory.objects.get(pk=item_data['inventory_id'])
            
            QoutationItems.objects.create(
                qoute=qoute,
                product=item,
                unit_price=item.price,
                quantity=item_data['quantity'],
                total_amount= item.price * item_data['quantity'],
            )
        return JsonResponse({'success': True, 'qoute_id': qoute.id})
    return JsonResponse({'success': False})

@login_required        
def qoutation_list(request):
    search_query = request.GET.get('q', '')
    qoutations = Qoutation.objects.filter(branch=request.user.branch).order_by('-date')
 
    if search_query:

        qoutations = qoutations.filter(
            Q(customer__name__icontains=search_query)|
            Q(products__icontains=search_query)|
            Q(date__icontains=search_query)|
            Q(qoute_reference__icontains=search_query)
        )
        
    return render(request, 'qoutations.html', {'qoutations':qoutations, 'search_query':search_query})
        
@login_required 
def qoute_preview(request, qoutation_id):
    qoute = Qoutation.objects.get(id=qoutation_id)
    qoute_items = QoutationItems.objects.filter(qoute=qoute)
    return render(request, 'qoute.html', {'qoute':qoute, 'qoute_items':qoute_items})

@login_required
def qoute_preview_modal(request, qoutation_id):
    try:
        qoute = Qoutation.objects.get(id=qoutation_id)
        logger.info(qoute)
        qoute_items = QoutationItems.objects.filter(qoute=qoute)
        logger.info(f'qoute items: {qoute_items.values()}')
        html = render_to_string('qoutations/partial_preview.html', {
            'qoute': qoute,
            'qoute_items': qoute_items
        }, request=request)  
        
        logger.info(html)

        return JsonResponse({'success': True, 'html': html}, status=200)
        
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success': False, 'message':str(e)}, status=400)

@login_required
def delete_qoute(request, qoutation_id):
    qoute = get_object_or_404(Qoutation, id=qoutation_id)
    qoute.delete()
    return JsonResponse({'success':True, 'message':'Qoutation successfully deleted'}, status=200)


# @login_required
# def cashbook_data(request):
#     """AJAX endpoint for cashbook data with filters and pagination"""
#     logger.info('Processing cashbook data request')
    
#     page = int(request.GET.get('page', 1))
#     per_page = int(request.GET.get('per_page', 20))
#     filter_option = request.GET.get('filter', 'this_week')
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     search_query = request.GET.get('search', '')
    
#     logger.info(f'filter: {filter_option}')

#     now = timezone.now()
#     end_date = now

#     if filter_option == 'today':
#         start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
#     elif filter_option == 'this_week':
#         start_date = now - timedelta(days=now.weekday())
#     elif filter_option == 'yesterday':
#         start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
#     elif filter_option == 'this_month':
#         start_date = now.replace(day=1)
#     elif filter_option == 'last_month':
#         start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
#     elif filter_option == 'this_year':
#         start_date = now.replace(month=1, day=1)
#     elif filter_option == 'custom':
#         if start_date and end_date:
#             start_date = datetime.strptime(start_date, '%Y-%m-%d')
#             end_date = datetime.strptime(end_date, '%Y-%m-%d')
#         else:
#             start_date = now - timedelta(days=now.weekday())
#             end_date = now

#     entries = Cashbook.objects.filter(
#         issue_date__gte=start_date,
#         issue_date__lte=end_date,
#         branch=request.user.branch
#     )
    
#     logger.info(f'Found {entries.count()} entries')

#     if search_query:
#         entries = entries.filter(
#             Q(description__icontains=search_query) |
#             Q(accountant__icontains=search_query) |
#             Q(manager__icontains=search_query) |
#             Q(director__icontains=search_query)
#         )

#     entries = entries.order_by('-issue_date')

#     total_entries = entries.count()
#     total_pages = (total_entries + per_page - 1) // per_page
#     start_index = (page - 1) * per_page
#     end_index = start_index + per_page

#     paginated_entries = entries[start_index:end_index]

#     balance = 0
#     entries_data = []
#     for entry in paginated_entries:
#         if entry.debit:
#             balance += entry.amount
#         elif entry.credit:
#             balance -= entry.amount

#         entries_data.append({
#             'id': entry.id,
#             'date': entry.issue_date.strftime('%Y-%m-%d %H:%M'),
#             'description': entry.description,
#             'debit': float(entry.amount) if entry.debit else None,
#             'credit': float(entry.amount) if entry.credit else None,
#             'balance': float(balance),
#             'accountant': entry.accountant,
#             'manager': entry.manager,
#             'director': entry.director,
#             'status': entry.status
#         })

#     return JsonResponse({
#         'entries': entries_data,
#         'pagination': {
#             'current_page': page,
#             'total_pages': total_pages,
#             'total_entries': total_entries,
#             'has_next': page < total_pages,
#             'has_previous': page > 1
#         }
#     })

@login_required
def cashbook_data(request):
    """AJAX endpoint for cashbook data with filters and pagination"""
    logger.info('Processing cashbook data request')
    
    # Get parameters from either GET or POST request
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            page = int(data.get('page', 1))
            per_page = int(data.get('per_page', 20))
            filter_option = data.get('filter', 'this_week')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            search_query = data.get('search', '')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    else:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        filter_option = request.GET.get('filter', 'this_week')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        search_query = request.GET.get('search', '')
    
    logger.info(f'filter: {filter_option}')

    now = timezone.now()
    end_date = now

    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now

    entries = Cashbook.objects.filter(
        issue_date__gte=start_date,
        issue_date__lte=end_date,
        branch=request.user.branch
    ).select_related('created_by', 'branch', 'updated_by', 'currency', 'invoice', 'expense').order_by('-issue_date')
    
    logger.info(f'Found {entries.count()} entries')

    if search_query:
        entries = entries.filter(
            Q(description__icontains=search_query) |
            Q(accountant__icontains=search_query) |
            Q(manager__icontains=search_query) |
            Q(director__icontains=search_query)
        )

    entries = entries.order_by('-issue_date')
    
    # Calculate totals
    total_cash_in = entries.filter(debit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
    total_cash_out = entries.filter(credit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
    total_balance = total_cash_in - total_cash_out

    total_entries = entries.count()
    total_pages = (total_entries + per_page - 1) // per_page
    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    paginated_entries = entries[start_index:end_index]

    balance = 0
    entries_data = []
    for entry in paginated_entries:
        if entry.debit:
            balance += entry.amount
        elif entry.credit:
            balance -= entry.amount
        logger.info(f'Entry: {entry.issue_date.strftime('%Y-%m-%d %H:%M')}')
        entries_data.append({
            'id': entry.id,
            'date': entry.issue_date.strftime('%Y-%m-%d %H:%M'),
            'description': entry.description,
            'debit': float(entry.amount) if entry.debit else None,
            'credit': float(entry.amount) if entry.credit else None,
            'balance': float(balance),
            'accountant': entry.accountant,
            'manager': entry.manager,
            'director': entry.director,
            'status': entry.status,
            'created_by': entry.created_by.first_name
        })

    return JsonResponse({
        'entries': entries_data,
        'totals': {
            'cash_in': float(total_cash_in),
            'cash_out': float(total_cash_out),
            'balance': float(total_balance)
        },
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'total_entries': total_entries,
            'has_next': page < total_pages,
            'has_previous': page > 1
        }
    }) 
    
@login_required
def download_cashbook_report(request):
    filter_option = request.GET.get('filter', 'this_week')
    now = datetime.datetime.now()
    end_date = now
    
    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    else:
        start_date = now - timedelta(days=now.weekday())
        end_date = now

    entries = Cashbook.objects.filter(date__gte=start_date, date__lte=end_date, branch=request.user.branch).order_by('date')

    # Create a CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="cashbook_report_{filter_option}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Description', 'Expenses', 'Income', 'Balance'])

    balance = 0  
    for entry in entries:
        if entry.debit:
            balance += entry.amount
        elif entry.credit:
            balance -= entry.amount

        writer.writerow([
            entry.issue_date,
            entry.description,
            entry.amount if entry.debit else '',
            entry.amount if entry.credit else '',
            balance,
            entry.accountant,
            entry.manager,
            entry.director
        ])

    return response


@login_required
def cashbook_note(request):
    #payload
    """
        entry_id:id,
        note:str
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            entry_id = data.get('entry_id')
            note = data.get('note')
            
            entry = Cashbook.objects.get(id=entry_id)
            entry.note = note
            
            entry.save()
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}.'}, status=400)
        return JsonResponse({'success':False, 'message':'Note successfully saved.'}, status=201)
    return JsonResponse({'success':False, 'message':'Invalid request.'}, status=405)


@login_required
def cashbook_note_view(request, entry_id):
    entry = get_object_or_404(Cashbook, id=entry_id)
    
    if request.method == 'GET':
        notes = entry.notes.all().order_by('timestamp')
        notes_data = [
            {'user': note.user.username, 'note': note.note, 'timestamp': note.timestamp.strftime("%Y-%m-%d %H:%M:%S")}
            for note in notes
        ]
        return JsonResponse({'success': True, 'notes': notes_data})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            note_text = data.get('note')
            CashBookNote.objects.create(entry=entry, user=request.user, note=note_text)
            return JsonResponse({'success': True, 'message': 'Note successfully added.'}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request.'}, status=405)
    
@login_required
def cancel_transaction(request):
    #payload
    """
        entry_id:id,
    """
    try:
        data = json.loads(request.body)
        entry_id = int(data.get('entry_id'))
        
        logger.info(entry_id)
        
        entry = Cashbook.objects.get(id=entry_id)
        
        entry.cancelled = True
        
        if entry.director:
            entry.director = False
        elif entry.manager:
            entry.manager = False
        elif entry.accountant:
            entry.accountant = False
            
        entry.save()
        logger.info(entry)
        return JsonResponse({'success': True}, status=201)
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
def update_transaction_status(request, pk):
    if request.method == 'POST':
        entry = get_object_or_404(Cashbook, pk=pk)
        
        data = json.loads(request.body)
        
        status = data.get('status')
        field = data.get('field')  

        if field in ['manager', 'accountant', 'director']:
            setattr(entry, field, status)

            if entry.cancelled:
                entry.cancelled = False
            entry.save()
            return JsonResponse({'success': True, 'status': getattr(entry, field)})
        
    return JsonResponse({'success': False}, status=400)   
    
@login_required
def cashWithdrawals(request):
    search_query = request.GET.get('q', '')
    selected_query = request.GET.get('sq', '')
    
    withdrawals = CashWithdraw.objects.all().order_by('-date')
    
    if search_query:
        withdrawals = withdrawals.filter(
            Q(user__branch__name__icontains=search_query)|
            Q(amount__icontains=search_query)|
            Q(date__icontains=search_query)|
            Q(reason__icontains=search_query)
        )
    if selected_query:
        withdrawals = CashWithdraw.objects.filter(deleted=True).order_by('-date')
        
    if 'download' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=withdrawals.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header_title in enumerate(['Date', 'User', 'Amount', 'Reason', 'Status'], start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.alignment = header_alignment
            
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header_title), 20)

        withdrawals = CashWithdraw.objects.all().order_by('-date')
        for withdrawal in withdrawals:
            worksheet.append(
                [
                    withdrawal.date,
                    withdrawal.user.username,
                    withdrawal.amount,
                    withdrawal.reason,
                    'Canceled' if withdrawal.deleted else 'Expensed' if withdrawal.status else 'pending'
                ])  
            
        workbook.save(response)
        return response
    
    form = CashWithdrawForm()
    expense_form = cashWithdrawExpenseForm()
    
    if request.method == 'POST':
        form = CashWithdrawForm(request.POST)
        
        if form.is_valid():
             
            password = form.cleaned_data['password']
            currency = form.cleaned_data['currency']
            amount = form.cleaned_data['amount']
            
            user = authenticate(username=request.user.username, password=password)
            
            if user is None:
                messages.warning(request, 'Incorrect password')
                return redirect('finance:withdrawals')
            
            cw_obj = form.save(commit=False)
            cw_obj.user = user
            cw_obj.save()
            
            account_name = f"{request.user.branch} {currency.name} {'Cash'} Account"
            
            try:
                account = Account.objects.get(name=account_name, type=Account.AccountType.CASH)
            except Account.DoesNotExist:
                messages.error(request, f'{account_name} doesnt exists')
                return redirect('finance:withdrawals')

            try:
                account_balance = AccountBalance.objects.get(account=account,  branch=request.user.branch)
            except AccountBalance.DoesNotExist:
                messages.error(request, f'Account Balances for account {account_name} doesnt exists')
                return redirect('finance:withdrawals')
            
            account_balance.balance -= Decimal(amount)
            account_balance.save()
            messages.success(request, 'Cash Withdrawal Successfully saved')
        else:
            messages.error(request, 'Invalid form data')
    return render(request, 'cashWithdaraws/withdrawals.html', 
        {
            'withdrawals':withdrawals,
            'count': withdrawals.filter(status=False, deleted=False).count(),
            'expense_form':expense_form,
            'form':form,
        }
    )

@login_required
@transaction.atomic
def cash_withdrawal_to_expense(request):
    if request.method == 'GET':
        cwte_id = request.GET.get('id', '')
        withdrawals = CashWithdraw.objects.filter(id=cwte_id).values(
            'user__branch__name', 'amount', 'reason', 'currency__id', 'user__id'
        )
        return JsonResponse(list(withdrawals), safe=False)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        
        withdrawal_data = data['withdrawal'][0]
        
        reason = data['reason']
        category_id = data['category_id']
        withdrawal_id = data['withdrawal_id']
        currency_id = withdrawal_data['currency__id']
        branch_name = withdrawal_data['user__branch__name']
        amount = withdrawal_data['amount']
        
        try:
            currency = Currency.objects.get(id=currency_id)
            branch = Branch.objects.get(name=branch_name)
            withdrawal = CashWithdraw.objects.get(id=withdrawal_id)
            category = ExpenseCategory.objects.get(id=category_id)
        except:
            return JsonResponse({'success':False,'message':'Invalid form data here'})
        
        Expense.objects.create(
            category=category,
            amount=amount,
            branch=branch,
            user = request.user,
            currency = currency,
            description=f'Cash withdrawal: {reason}',
            status=True,
            issue_date=withdrawal.date,
            payment_method='cash'
        )
        
        logger.info(withdrawal_id)
        withdrawal.status=True
        withdrawal.save()
        
        return JsonResponse({'success':True, 'message':'Successfully added to expenses'}, status=201)
    return JsonResponse({'success':False, 'message':'Invalid form data'}, status=400)
       
@login_required
def delete_withdrawal(request, withdrawal_id):
    try:
        withdrawal = CashWithdraw.objects.get(id=withdrawal_id)
    except User.DoesNotExist:
        messages.warning(request, 'Withdrawal doesnt exist')
        return redirect('finance:withdrawals')
    
    account_name = f"{request.user.branch} {withdrawal.currency.name} {'Cash'} Account"
    
    try:
        account = Account.objects.get(name=account_name, type=Account.AccountType.CASH)
    except Account.DoesNotExist:
        messages.error(request, f'{account_name} doesnt exists')
        return redirect('finance:withdrawals')
    
    try:
        account_balance = AccountBalance.objects.get(account=account,  branch=request.user.branch)
    except AccountBalance.DoesNotExist:
        messages.error(request, f'Account Balances for account {account_name} doesnt exists')
        return redirect('finance:withdrawals')
    
    account_balance.balance += Decimal(withdrawal.amount)
    account_balance.save()
    withdrawal.deleted=True
    withdrawal.save()
    
    messages.success(request, 'Withdrawal successfully deleted')
    return redirect('finance:withdrawals')
    
    
@login_required
def days_data(request):
    current_month = get_current_month()

    sales = Sale.objects.filter(date__month=current_month)
    cogs = COGSItems.objects.filter(date__month=current_month)

    first_day = min(sales.first().date, cogs.first().date)
    
    def get_week_data(queryset, start_date, end_date, amount_field):
        week_data = queryset.filter(date__gte=start_date, date__lt=end_date).values(amount_field, 'date')
        logger.info(week_data)
        total = week_data.aggregate(total=Sum(amount_field))['total'] or 0
        return week_data, total

    data = {}
    for week in range(1, 5):
        week_start = first_day + timedelta(days=(week-1)*7)
        week_end = week_start + timedelta(days=7)

        logger.info(week_start)
        logger.info(week_end)

        sales_data, sales_total = get_week_data(sales, week_start, week_end, 'total_amount')
        cogs_data, cogs_total = get_week_data(cogs, week_start, week_end, 'product__cost')
        
        data[f'week {week}'] = {
            'sales': list(sales_data),
            'cogs': list(cogs_data),
            'total_sales': sales_total,
            'total_cogs': cogs_total
        }

    return JsonResponse(data)

@login_required
def income_json(request):
    current_month = get_current_month()
    today = datetime.date.today()
    
    month = request.GET.get('month', current_month)
    day = request.GET.get('day', today.day)

    sales = Sale.objects.filter(transaction__branch=request.user.branch)
    
    if request.GET.get('filter') == 'today':
        sales_total = sales.filter(date=today).aggregate(Sum('total_amount'))
    else:
        sales_total = sales.filter(date__month=month).aggregate(Sum('total_amount'))

    return JsonResponse({'sales_total': sales_total['total_amount__sum'] or 0})


@login_required
def expense_json(request):
    current_month = get_current_month()
    today = datetime.date.today()
    
    month = request.GET.get('month', current_month)
    day = request.GET.get('day', today.day)

    expenses = Expense.objects.filter(branch=request.user.branch)
    
    if request.GET.get('filter') == 'today':
        expense_total = expenses.filter(issue_date=today, status=False).aggregate(Sum('amount'))
    else:
        expense_total = expenses.filter(issue_date__month=month, status=False).aggregate(Sum('amount'))
    
    return JsonResponse({'expense_total': expense_total['amount__sum'] or 0})


@login_required
def pl_overview(request):
    filter_option = request.GET.get('filter')
    today = datetime.date.today()
    previous_month = get_previous_month()
    current_year = today.year
    current_month = today.month

    sales = Sale.objects.filter(transaction__branch=request.user.branch)
    expenses = Expense.objects.filter(branch=request.user.branch)
    cogs = COGSItems.objects.filter(invoice__branch=request.user.branch)

    if filter_option == 'today':
        date_filter = today
    elif filter_option == 'last_week':
        last_week_start = today - datetime.timedelta(days=today.weekday() + 7)
        last_week_end = last_week_start + datetime.timedelta(days=6)
        date_filter = (last_week_start, last_week_end)
    elif filter_option == 'this_month':
        date_filter = (datetime.date(current_year, current_month, 1), today)
    elif filter_option == 'year':
        year = int(request.GET.get('year', current_year))
        date_filter = (datetime.date(year, 1, 1), datetime.date(year, 12, 31))
    else:
        date_filter = (datetime.date(current_year, current_month, 1), today)

    if filter_option == 'today':
        current_month_sales = sales.filter(date=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(issue_date=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.objects.filter(date=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    elif filter_option == 'last_week':
        current_month_sales = sales.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(issue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    else:
        current_month_sales = sales.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(dissue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0

    previous_month_sales = sales.filter(date__year=current_year, date__month=previous_month).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
    previous_month_expenses = expenses.filter(issue_date__year=current_year, issue_date__month=previous_month).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
    previous_cogs =  cogs.filter(date__year=current_year, date__month=previous_month).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    
    current_net_income = current_month_sales
    previous_net_income = previous_month_sales 
    current_expenses = current_month_expenses 
    
    current_gross_profit = current_month_sales - cogs_total
    previous_gross_profit = previous_month_sales - previous_cogs
    
    current_net_profit = current_gross_profit - current_month_expenses
    previous_net_profit = previous_gross_profit - previous_month_expenses

    current_gross_profit_margin = (current_gross_profit / current_month_sales * 100) if current_month_sales != 0 else 0
    previous_gross_profit_margin = (previous_gross_profit / previous_month_sales * 100) if previous_month_sales != 0 else 0
    
    # net_income_change = calculate_percentage_change(current_net_income, previous_net_income)
    # gross_profit_change = calculate_percentage_change(current_gross_profit, previous_gross_profit)
    # gross_profit_margin_change = calculate_percentage_change(current_gross_profit_margin, previous_gross_profit_margin)


    data = {
        'net_profit':current_net_profit,
        'cogs_total':cogs_total,
        'current_expenses':current_expenses,
        'current_net_profit': current_net_profit,
        'previous_net_profit':previous_net_profit,
        'current_net_income': current_net_income,
        'previous_net_income': previous_net_income,
        'current_gross_profit': current_gross_profit,
        'previous_gross_profit': previous_gross_profit,
        'current_gross_profit_margin': f'{current_gross_profit_margin:.2f}',
        'previous_gross_profit_margin': previous_gross_profit_margin,
    }
    
    return JsonResponse(data)

@login_required
def cash_deposit(request):
    if request.method == 'GET':
        deposits = CashDeposit.objects.all()
        return render(request, 'cash_deposit.html', 
            {
                'form':cashDepositForm(),
                'deposits':deposits
            }
        )
    
@login_required
def vat(request):
    if request.method == 'GET':
        
        filter_option = request.GET.get('filter', 'today')
        download = request.GET.get('download')
        
        now = datetime.datetime.now()
        end_date = now
        
        if filter_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_week':
            start_date = now - timedelta(days=now.weekday())
        elif filter_option == 'yesterday':
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_month':
            start_date = now.replace(day=1)
        elif filter_option == 'last_month':
            start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        elif filter_option == 'this_year':
            start_date = now.replace(month=1, day=1)
        elif filter_option == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now
            
        vat_transactions = VATTransaction.objects.filter(date__gte=start_date, date__lte=end_date).order_by('-date')
        
        if download:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="vat_report_{filter_option}.csv"'

            writer = csv.writer(response)
            writer.writerow(['Date', 'Description', 'Status', 'Input', 'Output'])

            balance = 0
            for transaction in vat_transactions:

                if transaction.vat_type == 'Input':
                    balance += transaction.tax_amount
                else:
                    balance -= transaction.tax_amount

                writer.writerow([
                    transaction.date,
                    transaction.invoice.invoice_number if transaction.invoice else transaction.purchase_order.order_number,
                    transaction.tax_amount if transaction.vat_type == 'Input' else  '',
                    transaction.tax_amount if transaction.vat_type == 'Output' else  ''
                ])

            writer.writerow(['Total', '', '', balance])
            
            return response
        return render(request, 'vat.html', 
            {
                'filter_option':filter_option,
                'vat_transactions':vat_transactions
            }
        )
    
    if request.method == 'POST':
        # payload 
        {
            'date_from':'date',
            'date_to':'date'
        }
        try:
            data = json.loads(request.body)
            
            date_to = data.get('date_to')
            date_from = data.get('date_from')

            vat_transactions = VATTransaction.objects.filter(
                date__gte=date_from, 
                date__lte=date_to
            )
            
            vat_transactions.update(paid=True)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status = 400)
        return JsonResponse({'success':False, 'message':'VAT successfully paid'}, status = 200)


@login_required
def cash_flow(request):
    """
    View to display a comprehensive financial overview including sales, income, and expenses.
    """
    today = datetime.datetime.today()
    income_form = IncomeCategoryForm()
    

    filter_type = request.GET.get('filter_type', 'today')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if filter_type == 'today':
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'weekly':
        start_date = (today - datetime.timedelta(days=today.weekday())).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'monthly':
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'yearly':
        start_date = today.replace(month=1, day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'custom':
        if not start_date:
            start_date = today.strftime('%Y-%m-%d')
        if not end_date:
            end_date = today.strftime('%Y-%m-%d')
    
    start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d')
    end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d')

    end_date_query = end_date_obj + datetime.timedelta(days=1)
        
    # Query for invoice items in the date range
    invoice_items = InvoiceItem.objects.filter(
        invoice__issue_date__date__gte=start_date_obj,
        invoice__issue_date__date__lt=end_date_query
    )
    
    # Query for other income and expenses
    income = Income.objects.filter(
        created_at__date__gte=start_date_obj,
        created_at__date__lt=end_date_query
    )
    
    expenses = Expense.objects.filter(
        issue_date__date__gte=start_date_obj,
        issue_date__date__lt=end_date_query
    )
    
    logs = FinanceLog.objects.filter(
        date__gte=start_date_obj.date(),
        date__lt=end_date_query.date()
    )
    
    # Normalize invoice items for timeline
    normalized_sales = invoice_items.annotate(
        type_label=Value('sale', output_field=CharField()),
        category_name=F('item__description'), 
        parent_category=Value('Sales', output_field=CharField()),
        datetime=F('invoice__issue_date'),
        source=Value('Invoice', output_field=CharField()),
        amount=F('total_amount')  
    ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')
    
    # Normalize income entries
    normalized_incomes = income.annotate(
        type_label=Value('income', output_field=CharField()),
        category_name=F('category__name'),
        parent_category=F('category__parent__name'),
        datetime=F('created_at'),
        source=Value('Income', output_field=CharField())
    ).values('datetime',  'sale__invoice_items__item__name', 'amount', 'type_label', 'category_name', 'parent_category', 'source', 'note')
    
    # Normalize expense entries
    normalized_expenses = expenses.annotate(
        type_label=Value('expense', output_field=CharField()),
        category_name=F('category__name'),
        parent_category=F('category__parent__name'),
        datetime=F('issue_date'),
        source=Value('Expense', output_field=CharField())
    ).values('datetime', 'amount', 'description', 'type_label', 'category_name', 'parent_category', 'source')
    
    # Combine and sort by datetime (chronological timeline of all financial activity)
    combined_cashflow = sorted(
        chain(normalized_incomes, normalized_expenses),
        key=lambda x: x['datetime']
    )
    
    product_sales = invoice_items.values(
        'item__id', 
        'item__name',
        'item__description'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('total_amount'),
        average_price=Avg('unit_price'),
        total_vat=Sum('vat_amount')
    ).order_by('-total_revenue')
    
    # Calculate totals
    sales_total = invoice_items.aggregate(total=Sum('total_amount'))['total'] or 0
    income_total = income.aggregate(total=Sum('amount'))['total'] or 0
    expenses_total = expenses.aggregate(total=Sum('amount'))['total'] or 0
    total_income = sales_total 
    balance = total_income - expenses_total
    
    # Group expenses by category for summary
    expenses_by_category = expenses.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Group income by category for summary
    income_by_category = income.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')


     # Add percentage to each expense category
    for category in expenses_by_category:
        if expenses_total > 0:
            category['percentage'] = (category['total'] / expenses_total) * 100
        else:
            category['percentage'] = 0
    
    # Group income by category for summary
    income_by_category = income.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Add percentage to each income category
    for category in income_by_category:
        if income_total > 0:
            category['percentage'] = (category['total'] / income_total) * 100
        else:
            category['percentage'] = 0

    # categories
    expenses_categories = ExpenseCategory.objects.all()
    income_categories = IncomeCategory.objects.all()
    
    cash_ups = CashUp.objects.all().select_related(
        'branch', 'created_by'
    ).prefetch_related(
        'sales', 'expenses'
    ).values(
        'expected_cash',
        'branch__id',
        'branch__name',
        'received_amount',
        'sales',
        'expenses',
        'created_by__username',
        'sales_status',
        'expenses_status',
        'status',
        'date'
    ).order_by('-created_at')
    
    context = {
        # Time filter data
        'start_date': start_date,
        'end_date': end_date,
        'filter_type': filter_type,
        
        # Raw data
        'sales': invoice_items,
        'income': income,
        'expenses': expenses,
        'logs': logs,
        
        # Summaries
        'product_sales': product_sales,
        'expenses_by_category': expenses_by_category,
        'income_by_category': income_by_category,
        'combined_cashflow': combined_cashflow,
        
        # Totals
        'sales_total': sales_total,
        'income_total': income_total,
        'expenses_total': expenses_total,
        'total_income': total_income,
        'balance': balance,

        # categories 
        'expenses_categories':expenses_categories,
        'income_categories':income_categories,
        
        # branches
        'cash_ups':cash_ups,
    }
    
    return render(request, 'cashflow.html', context)
    
# @login_required
# def cash_flow(request):
#     from itertools import chain
#     today = datetime.datetime.today()

#     # Income & Sales
#     sales = InvoiceItems.objects.filter(sale__issue_date__date=today)
#     income = Income.objects.filter(created_at__date=today)
#     logs = FinanceLog.objects.filter(date=today)

#     # Expenses
#     expenses = Expense.objects.filter(issue_date__date=today)

#     # Normalize income entries
#     normalized_incomes = income.annotate(
#         type_label=models.Value('income', output_field=models.CharField()),
#         category_name=models.F('category__name'),
#         parent_category=models.F('category__parent__name'),
#         datetime=models.F('created_at'),
#         source=models.Value('Income', output_field=models.CharField())
#     ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')

#     # Normalize expense entries
#     normalized_expenses = expenses.annotate(
#         type_label=models.Value('expense', output_field=models.CharField()),
#         category_name=models.F('category__name'),
#         parent_category=models.F('category__parent__name'),
#         datetime=models.F('issue_date'),
#         source=models.Value('Expense', output_field=models.CharField())
#     ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')

#     # Combine and sort by datetime
#     combined_cashflow = sorted(
#         chain(normalized_incomes, normalized_expenses),
#         key=lambda x: x['datetime']
#     )

#     # Totals
#     sales_total = sales.aggregate(total=Sum('amount'))['total'] or 0
#     income_total = income.aggregate(total=Sum('amount'))['total'] or 0
#     expenses_total = expenses.aggregate(total=Sum('amount'))['total'] or 0
#     total_income = sales_total + income_total
#     balance = total_income - expenses_total

#     context = {
#         'sales': sales,
#         'sales_total': sales_total,
#         'income': income,
#         'income_total': income_total,
#         'expenses_total': expenses_total,
#         'total_income': total_income,
#         'combined_cashflow': combined_cashflow,  
#         'balance':balance,
#         'logs':logs
#         # 'grouped_expenses': grouped_expenses,
#         # 'grouped_income': grouped_income,
#     }

#     return render(request, 'cashflow.html', context)

@login_required
def confirm_cash_up(request):
    try:
        data = json.loads(request.body)
        cash_up_id = data.get('cash_up_id')
        amount_received = Decimal(data.get('amount_received', 0))
        account_type = data.get('account_type')
        assigned_to = data.get('assigned_to')
        notes = data.get('notes', '')
 
        cash_up = CashUp.objects.get(id=cash_up_id)
        
        shortfall = cash_up.expected_cash - amount_received
        
        with transaction.atomic():
            cash_up.status = True
            cash_up.received_amount = amount_received
            cash_up.short_fall = shortfall
            cash_up.notes = notes
            cash_up.save()

            if shortfall > 0 and assigned_to:
                if account_type == 'employee':
                    try:
                        salesperson = User.objects.get(id=assigned_to)
                        user_account, _ = UserAccount.objects.get_or_create(
                            user=salesperson,
                            defaults={
                                'balance': Decimal('0.00'),
                                'total_credits': Decimal('0.00'),
                                'total_debits': Decimal('0.00'),
                                'last_transaction_date': timezone.now()
                            }
                        )

                        UserTransaction.objects.create(
                            account=user_account,
                            transaction_type='Shortfall',
                            amount=shortfall,
                            description=f'Cash up shortfall from {cash_up.branch.name} - {notes}',
                            debit=shortfall,
                            credit=0
                        )
                        
                        user_account.balance += shortfall
                        user_account.total_debits += shortfall
                        user_account.last_transaction_date = timezone.now()
                        user_account.save()

                    except User.DoesNotExist:
                        return JsonResponse({
                            'success': False,
                            'message': 'Selected salesperson not found'
                        }, status=400)
                elif account_type == 'company':
                    try:
                        # Create a company account transaction
                        last_lost_account = LossAccount.objects.filter().last()
                        
                        if last_lost_account:
                            balance = last_lost_account.balance + Decimal(shortfall)
                        
                        LossAccount.objects.create(
                            name=f'Cash up shortfall from {cash_up.branch.name} - {notes}',
                            balance = balance,
                            amount=shortfall
                        )

                    except UserAccount.DoesNotExist:
                        return JsonResponse({
                            'success': False,
                            'message': 'Selected company account not found'
                        }, status=400)

        return JsonResponse({
            'success': True,
            'message': 'Cash up confirmed successfully'
        }, status=200)

    except CashUp.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Cash up not found'
        }, status=404)
    except Exception as e:
        logger.error(f'Error in confirm_cash_up: {str(e)}')
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def get_recorded_cash_ups(request):
    try:
        data = json.loads(request.body)
        branch_id = data.get('branch_id')
        date_range = data.get('date_range')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        queryset = CashUp.objects.filter(status='recorded')
        
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)

        now = datetime.now()
        if date_range == 'today':
            queryset = queryset.filter(created_at__date=now.date())
        elif date_range == 'yesterday':
            yesterday = now - timedelta(days=1)
            queryset = queryset.filter(created_at__date=yesterday.date())
        elif date_range == 'this_week':
            start_of_week = now - timedelta(days=now.weekday())
            queryset = queryset.filter(created_at__date__gte=start_of_week.date())
        elif date_range == 'last_week':
            start_of_week = now - timedelta(days=now.weekday() + 7)
            end_of_week = start_of_week + timedelta(days=6)
            queryset = queryset.filter(created_at__date__gte=start_of_week.date(), created_at__date__lte=end_of_week.date())
        elif date_range == 'this_month':
            start_of_month = now.replace(day=1)
            queryset = queryset.filter(created_at__date__gte=start_of_month.date())
        elif date_range == 'last_month':
            start_of_month = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
            end_of_month = now.replace(day=1) - timedelta(days=1)
            queryset = queryset.filter(created_at__date__gte=start_of_month.date(), created_at__date__lte=end_of_month.date())
        elif date_range == 'custom' and start_date and end_date:
            queryset = queryset.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

        queryset = queryset.order_by('-created_at')

        cash_ups = list(queryset.values(
            'id',
            'branch__name',
            'created_at',
            'expected_cash',
            'received_amount',
            'short_fall',
            'recorded_by__username'
        ))

        return JsonResponse({
            'success': True,
            'cash_ups': cash_ups
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@login_required
@transaction.atomic
def cashflow_create(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            amount = data.get('amount')
            cash_up_id = data.get('cash_up_id')

            with transaction.atomic():
                cash_up = CashUp.objects.get(id=cash_up_id)
                cash_up.received_amount = Decimal(amount)
                cash_up.status = True
                cash_up.save()

                total_income = sum(sale.unit_price * sale.quantity for sale in cash_up.sales.all())
                total_expenses = sum(expense.amount for expense in cash_up.expenses.all())

                for sale in cash_up.sales.all():
                    logger.info(f'cash sale: {sale.unit_price * sale.quantity}', {sale})

                logger.info(f'total income: {total_income}')

                for expenses in cash_up.expenses.all():
                    logger.info(f'cash expense: {expenses}')
                
                logger.info(f'total expenses: {total_expenses}')

                expense_category, _ = MainExpenseCategory.objects.get_or_create(name='Expense')
                income_category, _ = MainIncomeCategory.objects.get_or_create(name='Income')
                
                # for income
                Cashflow.objects.create(
                    branch=cash_up.branch,
                    total=amount,
                    date=datetime.datetime.now(),
                    status=False,
                    cash_up=cash_up,
                    income = total_income,
                    category=income_category,
                    created_by=request.user
                )

                # for expense
                Cashflow.objects.create(
                    branch=cash_up.branch,
                    total=amount,
                    date=datetime.datetime.now(),
                    status=False,
                    cash_up=cash_up,
                    expense = total_expenses,
                    category = expense_category,
                    created_by=request.user
                )

                # if amount > total_income - total_expenses:
                #     user_account, _ = UserAccount.objects.get_or_create(
                #         user=request.user, 
                #         defaults={
                #             'balance': Decimal('0.00'),
                #             'total_credits': Decimal('0.00'),
                #             'total_debits': Decimal('0.00'),
                #             'last_transaction_date':datetime.datetime.now()
                #         }
                #     )

                #     transaction = UserTransaction.objects.create(
                #         account=user_account,
                #         branch=cash_up.branch,
                #         amount=amount - total_income + total_expenses,
                #         transaction_type=UserTransaction.TransactionType.CASH,
                #         description='Cashup deficit',
                #         created_by=request.user
                #     )

                #     user_account.balance += transaction.amount
                #     user_account.total_credits = 0
                #     user_account.total_debits = transaction.amount


                return JsonResponse({'success':True, 'message':'Cashflow successfully created'}, status=201)

        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        

@login_required
@transaction.atomic
def undo_record_cashflow(request):
    try:
        data = json.loads(request.body)
        sale_type = data.get('type', '')
        sale_id = data.get('id', '')
        branch_id = int(data.get('branch', ''))

        branch = get_object_or_404(Branch, id=branch_id)

        if sale_type != 'sale':
            return JsonResponse({'success': False, 'message': 'Unsupported type'}, status=400)

        sale = Invoice.objects.filter(id=sale_id).first()
        if not sale:
            return JsonResponse({'success': False, 'message': 'Sale not found'}, status=404)

        income = Income.objects.filter(sale=sale, branch=branch).first()
        if income:
            income.delete()

        sale.cash_up_status = False
        sale.save()

        return JsonResponse({
            'success': True,
            'message': 'Cashflow undo successful',
            'id': sale_id,
            'cash_up_status': False
        }, status=200)

    except Exception as e:
        logger.exception("Error undoing cashflow")
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

        
@login_required
@transaction.atomic
def record_cashflow(request):
    try:
        data = json.loads(request.body)
        type = data.get('type', '')
        id = data.get('id', '')
        branch = int(data.get('branch', ''))
        category = (data.get('category'))
        
        # get branch
        branch = get_object_or_404(Branch, id=branch)

        category = None
        if type == 'sale':
            sale = Invoice.objects.filter(id=id).first()
      
            category = IncomeCategory.objects.filter(name__iexact=category).first()

            if not category:
                new_main_category = IncomeCategory.objects.create(
                    name='sales',
                    parent=None
                )

                new_sub_category = IncomeCategory.objects.create(
                    name="sales",
                    parent=new_main_category
                )
                category = new_sub_category


        with transaction.atomic():

            Income.objects.create(
                amount = sale.amount_paid,
                currency = sale.currency,
                note = sale.products_purchased,
                branch = request.user.branch,
                status = False,
                sale = sale,
                user=request.user,
                category = category
            )

            sale.cash_up_status = True
            sale.save()
            
            return JsonResponse(
                {
                    'success':True, 
                    'message':'Sale recorded succesfully', 
                    'id':id,
                    'cash_up_status':True
                }, status=200)
            
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
    
@login_required
def check_cashup_status(request, cash_up_id):
    logger.info(cash_up_id)
    cash_up = CashUp.objects.filter(id=cash_up_id).first()
    
    if cash_up:
        if True:
            cash_up.status = True
            cash_up.save()
            return JsonResponse({
                'success':True,
                'status':True
            }, status=200)
        return JsonResponse({
                'success':True,
                'status':False
            }, status=200)
    return JsonResponse({
            'success':False,
            'message':'Error occured'
        }, status=400)
    

@login_required
def banking(request):
    return render(request, 'cashbook/banking.html')

def create_bank_account(request): # to account for permissions
    """ creates a bank account """
    try:
        data = json.loads(request.body)
        name = data.get('name', '')
        branch = data.get('branch', '')
        
        if not name:
            return JsonResponse({'successs':False, 'message':'Bank name is missing, status = 400'})
        
        if not branch:
            return JsonResponse({'successs':False, 'message':'Branch name is missing, status = 400'})
        
        branch = Branch.objects.get(id=branch)
        
        account = BankAccount()
        account.name = name
        account.branch = branch
        account.user = request.user
        account.save()
        
        logger.info(f'User: {account.user.username} created bank account: {name}')
        
    except Branch.DoesNotExist:
        return JsonResponse({'success':False, 'message':'Branch does not exists'})
    except Exception as e:
        return JsonResponse({'success':False, 'message':str(e)}, status=500)

@login_required
def banking_data(request):
    """ Shows the banking transactions in different banks """
    
    # notes
    # add a bank account 
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            page = int(data.get('page', 1))
            per_page = int(data.get('per_page', 20))
            filter_option = data.get('filter', 'this_week')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            search_query = data.get('search', '')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    else:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        filter_option = request.GET.get('filter', 'this_week')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        search_query = request.GET.get('search', '')
    
    logger.info(f'filter: {filter_option}')

    now = timezone.now()
    end_date = now

    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now

    entries = Cashbook.objects.filter(
        issue_date__gte=start_date,
        issue_date__lte=end_date,
        branch=request.user.branch
    ).select_related('created_by', 'branch', 'updated_by', 'currency', 'invoice', 'expense').order_by('-issue_date')
    
    logger.info(f'Found {entries.count()} entries')

    if search_query:
        entries = entries.filter(
            Q(description__icontains=search_query) |
            Q(accountant__icontains=search_query) |
            Q(manager__icontains=search_query) |
            Q(director__icontains=search_query)
        )

    entries = entries.order_by('-issue_date')
    
    # Calculate totals
    total_cash_in = entries.filter(debit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
    total_cash_out = entries.filter(credit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
    total_balance = total_cash_in - total_cash_out

    total_entries = entries.count()
    total_pages = (total_entries + per_page - 1) // per_page
    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    paginated_entries = entries[start_index:end_index]

    balance = 0
    entries_data = []
    for entry in paginated_entries:
        if entry.debit:
            balance += entry.amount
        elif entry.credit:
            balance -= entry.amount
        logger.info(f'Entry: {entry.issue_date.strftime('%Y-%m-%d %H:%M')}')
        entries_data.append({
            'id': entry.id,
            'date': entry.issue_date.strftime('%Y-%m-%d %H:%M'),
            'description': entry.description,
            'debit': float(entry.amount) if entry.debit else None,
            'credit': float(entry.amount) if entry.credit else None,
            'balance': float(balance),
            'accountant': entry.accountant,
            'manager': entry.manager,
            'director': entry.director,
            'status': entry.status,
            'created_by': entry.created_by.first_name
        })

    return JsonResponse({
        'entries': entries_data,
        'totals': {
            'cash_in': float(total_cash_in),
            'cash_out': float(total_cash_out),
            'balance': float(total_balance)
        },
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'total_entries': total_entries,
            'has_next': page < total_pages,
            'has_previous': page > 1
        }
    }) 
    
@login_required
def get_incomes(request):
    page_number = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('limit', 10))  

    incomes = Income.objects.filter(branch=request.user.branch).order_by('-created_at')
    paginator = Paginator(incomes, per_page)
    page_obj = paginator.get_page(page_number)

    income_data = [
        {
            'id': income.id,
            'created_at': income.created_at.strftime('%Y-%m-%d %H:%M'),
            'amount': str(income.amount),
            'category': str(income.category.name),
            'branch':income.branch.name,
            'note': income.note,
            'user': income.user.get_full_name() or income.user.username,
            'is_recurring': income.is_recurring,
            'recurrence': f"{income.recurrence_value} {income.recurrence_unit}" if income.is_recurring else '',
        }
        for income in page_obj.object_list
    ]

    return JsonResponse({
        'data': income_data,
        'has_next': page_obj.has_next()
    })
    
@transaction.atomic
@login_required
def record_income(request):
    try:
        data = json.loads(request.body)
        logger.info(data)

        name = data.get('name')
        amount = data.get('amount')
        category_name = data.get('category')
        branch_id = data.get('branch')
        r_value = data.get('r_value')
        r_unit = data.get('r_unit')

        if not all([name, amount, category_name, branch_id]):
            return JsonResponse({'success': False, 'message': 'Missing required fields'}, status=400)

        parent_category, _ = IncomeCategory.objects.get_or_create(name="Manual", parent=None)
        category, _ = IncomeCategory.objects.get_or_create(name=category_name, parent=parent_category)

        branch = Branch.objects.get(id=branch_id)
        currency = Currency.objects.filter(name__icontains="usd").first() #to be dynamic
        
        logger.info(currency)

        Income.objects.create(
            amount=amount,
            currency_id=currency.id,
            category=category,
            note=name,
            user=request.user,
            branch=branch,
            is_recurring=bool(r_value),
            recurrence_value=r_value if r_value else None,
            recurrence_unit=r_unit if r_unit else None
        )
        
        Cashbook.objects.create(
            amount=amount,
            description=f"Income: {name} -> {category.name}",
            debit=True,
            credit=False,
            branch=branch,
            created_by=request.user,
            updated_by=request.user,
            issue_date=timezone.now()
        )
        
        logger.info(f'Income recorded successfully: {Income.objects.last()}')

        return JsonResponse({'success': True, 'message': 'Income recorded successfully'}, status=200)
    except Exception as e:
        logger.error("Income record error: %s", e)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


import csv
from .models import *
from decimal import Decimal
from io import BytesIO
from apps.users.models import User
from apps.company.models import Branch
from .consumers import CashTransferConsumer 
from xhtml2pdf import pisa 
from django.views import View
from django.db.models import Q
from twilio.rest import Client
from datetime import timedelta
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.contrib import messages
from utils.utils import generate_pdf
from . utils import update_latest_due
from django.http import JsonResponse
from utils.utils import generate_pdf
from asgiref.sync import async_to_sync, sync_to_async
from apps.inventory.models import Inventory, Accessory
from channels.layers import get_channel_layer
import json, datetime, os, boto3, openpyxl 
from utils.account_name_identifier import account_identifier
from .tasks import (
    send_invoice_email_task, 
    send_account_statement_email, 
    send_quotation_email
)
from pytz import timezone as pytz_timezone 
from openpyxl.styles import Alignment, Font
from . utils import calculate_expenses_totals
from django.utils.dateparse import parse_date
from django.templatetags.static import static
from django.db.models import Sum, DecimalField
from apps.inventory.models import ActivityLog, Product
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.core.mail import send_mail, EmailMessage
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from . forms import (
    ExpenseForm, 
    ExpenseCategoryForm, 
    CurrencyForm, 
    InvoiceForm, 
    CustomerForm, 
    TransferForm, 
    CashWithdrawForm, 
    cashWithdrawExpenseForm,
    customerDepositsForm,
    customerDepositsRefundForm,
    cashDepositForm,
    IncomeCategoryForm
)
from django.contrib.auth import authenticate
from loguru import logger
from .tasks import send_expense_creation_notification
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from django.http import FileResponse
import io
from collections import defaultdict
from apps.pos.utils.receipt_signature import generate_receipt_data
from apps.pos.utils.submit_receipt_data import submit_receipt_data
from django.db.models.functions import Coalesce
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Max
from django.utils.dateparse import parse_date
from dotenv import load_dotenv
from apps.settings.models import OfflineReceipt, FiscalDay, FiscalCounter
from utils.zimra import ZIMRA
from utils.zimra_sig_hash import run
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Avg, F, Value, CharField, ExpressionWrapper
import datetime
from itertools import chain
from django.core.paginator import Paginator, EmptyPage
import imghdr, base64
from django.core.files.base import ContentFile
 
# load global zimra instance
zimra = ZIMRA()

load_dotenv()

from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from .models import CashUp, Invoice, Expense
from django.utils import timezone

def get_previous_month():
    first_day_of_current_month = datetime.datetime.now().replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    return last_day_of_previous_month.month

def get_current_month():
    return datetime.datetime.now().month

#image decoding method
def decode_base64_file(data):
    """
    Decodes a base64 file and returns a ContentFile.
    Assumes data is in the format: data:<mime>;base64,<data>
    """
    if not data:
        return None

    try:
        format, imgstr = data.split(';base64,')
        ext = format.split('/')[-1]
        if ext == 'jpeg':
            ext = 'jpg'

        file_name = f"{uuid.uuid4()}.{ext}"
        return ContentFile(base64.b64decode(imgstr), name=file_name)
    except Exception as e:
        logger.error("Failed to decode base64 image:")
        return None

def cashflow_list(request): # to be organised
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        branch = request.GET.get('branch')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        status = request.GET.get('status')

        cashups = CashUp.objects.all()
        
        if branch:
            cashups = cashups.filter(branch__id=int(branch))
        if start_date:
            cashups = cashups.filter(date__gte=start_date)
        if end_date:
            cashups = cashups.filter(date__lte=end_date)
        if status in ['true', 'false']:
            cashups = cashups.filter(status=(status == 'true'))

        html = render_to_string('cashflows/partials/cashup_cards.html', {'cash_ups': cashups})
        
        return JsonResponse({'html': html})

class Finance(View):
    # authentication loginmixin
    template_name = 'finance.html'

    def get(self, request, *args, **kwargs):

        if request.user.role == 'sales':
            return redirect('finance:expenses')
        
        balances = AccountBalance.objects.filter(branch=request.user.branch)
    
        recent_sales = Sale.objects.filter(transaction__branch=request.user.branch).order_by('-date')[:5]

        expenses_by_category = Expense.objects.values('category__name').annotate(
            total_amount=Sum('amount', output_field=DecimalField())
        )
        
        context = {
            'balances': balances,
            'recent_transactions': recent_sales,
            'expenses_by_category': expenses_by_category,
        }
        
        return render(request, self.template_name, context)
    
@login_required
def monthly_installments(request):
    installments = MonthlyInstallment.objects.all().select_related('invoice').order_by('-date')
    return JsonResponse({'installments': list(installments)})

@login_required
def laybys(request):
    laybys = layby.objects.all().select_related('invoice').order_by('-date')
    return JsonResponse({'laybys': list(laybys)})

@login_required
def expenses(request):
    form = ExpenseForm()
    cat_form = ExpenseCategoryForm()

    if request.method == 'GET':
        filter_button = request.GET.get('filter_button')
        filter_option = request.GET.get('filter', 'today')  
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        expenses = Expense.objects.filter(user=request.user).order_by('-issue_date')
        
        # filtered_expenses = filter_expenses(expenses, filter_option, start_date, end_date)

        if request.user.role == 'sale':
            expenses = expenses.filter(user=request.user)
        
        return render(request, 'expenses.html', 
            {
                'form':form,
                'cat_form':cat_form,
                'expenses':filter_expenses,
                'filter_option': expenses,
            }
        )
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
      
            name = data.get('name') 
            amount = data.get('amount')
            category = data.get('category')  
            payment_method = data.get('payment_method', 'cash')
            currency_id = data.get('currency', 'USD')
            branch = request.user.branch.id
            base64_image = data.get('receipt')
            image = decode_base64_file(base64_image)

            is_recurring = data.get('is_recurring') == 'true'
            recurrence_value = data.get('recurrence_value')
            recurrence_unit = data.get('recurrence_unit')
            
            logger.info(branch)

            # Validation
            if not all([name, amount, category, payment_method, currency_id, branch]):
                return JsonResponse({'success': False, 'message': 'Missing required fields.'})

            # Fetch related objects
            try:
                category = ExpenseCategory.objects.get(id=category)
            except ExpenseCategory.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'Category with ID {category} does not exist.'})

            currency = get_object_or_404(Currency, name__icontains='usd')
            branch = get_object_or_404(Branch, id=branch)

            # Get or create account and balance
            account_details = account_identifier(request, currency, payment_method)
            account_name = account_details['account_name']
            account_type = account_details['account_type']

            account, _ = Account.objects.get_or_create(
                name=account_name,
                type=account_type
            )

            account_balance, _ = AccountBalance.objects.get_or_create(
                account=account,
                currency=currency,
                defaults={
                    'currency': currency,
                    'branch': branch,
                    'balance': 0
                }
            )
            
            logger.info(account_balance.balance)

            if account_balance.balance < Decimal(amount):
                return JsonResponse({'success': False, 'message': f'{account_name} has insufficient balance.'})

            # Deduct from balance
            account_balance.balance -= Decimal(amount)
            account_balance.save()

            # Create Expense
            expense = Expense.objects.create(
                description=name,
                amount=amount,
                category=category,
                user=request.user,
                currency=currency,
                payment_method=payment_method ,
                branch=branch,
                is_recurring=is_recurring,
                recurrence_value=int(recurrence_value) if is_recurring and recurrence_value else None,
                recurrence_unit=recurrence_unit if is_recurring else None,
                receipt=image,
            )

            # Create Cashbook entry
            Cashbook.objects.create(
                amount=amount,
                expense=expense,
                currency=currency,
                credit=True,
                description=f'Expense ({expense.description[:20]})',
                branch=branch
            )

            # Send notification (to turn on)
            # send_expense_creation_notification.delay(expense.id)

            return JsonResponse({'success': True, 'message': 'Expense recorded successfully.'})

        except Exception as e:
            logger.exception("Error while recording expense:")
            return JsonResponse({'success': False, 'message': str(e)})
        
@login_required
def get_expenses(request):
    try:
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 10))

        expenses = Expense.objects.select_related('category', 'branch', 'currency') \
                                  .order_by('-issue_date') 
        paginator = Paginator(expenses, limit)

        try:
            paginated_expenses = paginator.page(page)
        except EmptyPage:
            return JsonResponse({
                'data': [],
                'has_next': False
            })

        results = []
        for expense in paginated_expenses:
            results.append({
                'id': expense.id,
                'created_at': expense.issue_date.isoformat(),
                'note': expense.description,
                'amount': float(expense.amount),
                'category': str(expense.category),
                'branch': expense.branch.name,
                'has_receipt': bool(expense.receipt),
                'receipt_url': expense.receipt.url if expense.receipt else None
            })

        return JsonResponse({
            'data': results,
            'has_next': paginated_expenses.has_next()
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def save_expense_split(request):
    try:
        data = json.loads(request.body)
        splits = data.get('splits')
        branch_id = data.get('branch_id')
        expense_id = data.get('expense_id', '')
        
        cash_up = CashUp.objects.filter(date=today, branch__id=int(branch_id)).first() # to put order by
        logger.info(f'Cash up: {cash_up}')
        
        cash_up_expenses = cash_up.expenses.all()
        
        if expense_id:
            record_expense(expense_id,cash_up_expenses, request)
            
        else:
            logger.info(f'Split expenses: {splits}, branch_id = {branch_id}')
            expenses = []
            if not cash_up:
                return JsonResponse({'success': False, 'message': 'No cash up record found for today'}, status=404)

            categories = ExpenseCategory.objects.all()
            
            logger.info(cash_up.expenses)

            for split in splits:
                logger.info(split)
                logger.info(split['amount'])
                exp_obj = cash_up_expenses.get(id=int(split['expense_id']))  # existing expense object
                
                new_expense = Expense(
                    amount=split['amount'],
                    payment_method=exp_obj.payment_method,
                    currency=exp_obj.currency,
                    category=categories.filter(id=split['category_id']).first(),
                    description=exp_obj.description,
                    user=request.user,  
                    branch_id=request.user.branch.id,
                    status=False,
                    # purchase_order=exp_obj.purchase_order,
                    receipt=exp_obj.receipt,
                    is_recurring=exp_obj.is_recurring,
                    recurrence_value= exp_obj.recurrence_value if exp_obj.recurrence_value else None,
                    recurrence_unit= exp_obj.recurrence if exp_obj.recurrence_unit else None
                )
                
                exp_obj.cash_up_status = True
                
                exp_obj.save()
                new_expense.save()
                
                expenses.append(new_expense.id)

        return JsonResponse({
            'success': True,
            'message': 'Expenses split and saved successfully',
            # 'expenses': expenses
        })

    except Exception as e:
        logger.exception("Error in save_expense_split")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
        
def record_expense(expense_id, cash_up_expenses, request):
    logger.info(f'Straight recording of the expense: {expense_id}')
    exp_obj = cash_up_expenses.get(id=int(expense_id))  # existing expense object
            
    new_expense = Expense(
        amount=exp_obj.amount,
        payment_method=exp_obj.payment_method,
        currency=exp_obj.currency,
        category=exp_obj.category,
        description=exp_obj.description,
        user=request.user,  
        branch_id=request.user.branch.id,
        status=False,
        purchase_order=exp_obj.purchase_order,
        receipt=exp_obj.receipt,
        is_recurring=exp_obj.is_recurring,
        recurrence_value= exp_obj.recurrence_value if exp_obj.recurrence_value else None,
        recurrence_unit= exp_obj.recurrence if exp_obj.recurrence_unit else None
    )
    
    exp_obj.cash_up_status = True
    
    exp_obj.save()
    new_expense.save()

@login_required  
def get_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    data = {
        'id': expense.id,
        'amount': expense.amount,
        'description': expense.description,
        'category': expense.category.id
    }
    return JsonResponse({'success': True, 'data': data})

@login_required      
def add_or_edit_expense(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            amount = data.get('amount')
            description = data.get('description')
            category_id = data.get('category')
            expense_id = data.get('id')

            if not amount or not description or not category_id:
                return JsonResponse({'success': False, 'message': 'Missing fields: amount, description, category.'})
            
            category = get_object_or_404(ExpenseCategory, id=category_id)

            if expense_id:  
                expense = get_object_or_404(Expense, id=expense_id)
                before_amount = expense.amount
                
                expense.amount = amount
                expense.description = description
                expense.category = category
                expense.save()
                message = 'Expense successfully updated'
            
                try:
                    cashbook_expense = Cashbook.objects.get(expense=expense)
                    expense_amount = Decimal(expense.amount)
                    if cashbook_expense.amount < expense_amount:
                        cashbook_expense.amount = expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'Expense (update from {before_amount} to {cashbook_expense.amount})'
                    else:
                        cashbook_expense.amount -= cashbook_expense.amount - expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'(update from {before_amount} to {cashbook_expense.amount})'
                    cashbook_expense.save()
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)}, status=400)
            return JsonResponse({'success': True, 'message': message}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
@transaction.atomic()
def add_expense_category(request):
    subcategories = ExpenseCategory.objects.filter(parent__isnull=False).values(
        'id',
        'name'
    )

    logger.info(subcategories)
    

@login_required
def add_expense_category(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

        category_name = data.get('name', '')
        parent_name = data.get('parent', '')
        new_parent_name = data.get('new_parent', '')

        if not category_name:
            return JsonResponse({'success': False, 'message': 'Category name is required.'}, status=400)

        parent_obj = None
        if new_parent_name:
            parent_obj, _ = ExpenseCategory.objects.get_or_create(name=new_parent_name, parent=None)
            logger.info(f'New parent created or found: {parent_obj}')
        elif parent_name:
            parent_obj = ExpenseCategory.objects.filter(name=parent_name, parent=None).first()
            if not parent_obj:
                return JsonResponse({'success': False, 'message': f'Parent category "{parent_name}" not found.'}, status=404)
            logger.info(f'Existing parent found: {parent_obj}')

        if ExpenseCategory.objects.filter(name=category_name, parent=parent_obj).exists():
            return JsonResponse({
                'success': False,
                'message': f'Category "{category_name}" already exists under this parent.'
            }, status=400)

        # Create new child category
        new_category = ExpenseCategory.objects.create(name=category_name, parent=parent_obj)
        logger.info(f'New category created: {new_category} under parent: {parent_obj}')

        return JsonResponse({
            'success': True,
            'id': new_category.id,
            'name': new_category.name
        }, status=201)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

@login_required
def income(request):
    """returns income both from the cash sales and income deposited"""
    pass

def filter_expenses(queryset, filter_option, start_date=None, end_date=None):
    """
    Filter expense queryset based on specified time period
    """
    today = timezone.localtime().date()
    
    if filter_option == 'today':
        return queryset.filter(date=today)
    
    elif filter_option == 'yesterday':
        yesterday = today - timedelta(days=1)
        return queryset.filter(date=yesterday)
    
    elif filter_option == 'this_week':
        start_of_week = today - timedelta(days=today.weekday())  
        end_of_week = start_of_week + timedelta(days=6) 
        return queryset.filter(date__gte=start_of_week, date__lte=end_of_week)
    
    elif filter_option == 'last_week':
        end_of_last_week = today - timedelta(days=today.weekday() + 1)  
        start_of_last_week = end_of_last_week - timedelta(days=6)  
        return queryset.filter(date__gte=start_of_last_week, date__lte=end_of_last_week)
    
    elif filter_option == 'this_month':
        return queryset.filter(date__year=today.year, date__month=today.month)
    
    elif filter_option == 'last_month':
        last_month = today.replace(day=1) - timedelta(days=1)
        return queryset.filter(date__year=last_month.year, date__month=last_month.month)
    
    elif filter_option == 'this_year':
        return queryset.filter(date__year=today.year)
    
    elif filter_option == 'last_year':
        return queryset.filter(date__year=today.year - 1)
    
    elif filter_option == 'custom' and start_date and end_date:
        try:
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                
            return queryset.filter(date__gte=start_date, date__lte=end_date)
        except (ValueError, TypeError):
            return queryset

    return queryset

@login_required
@transaction.atomic
def delete_expense(request, expense_id):
    if request.method == 'DELETE':
        try:
            expense = get_object_or_404(Expense, id=expense_id)
            expense.cancel = True
            expense.save()
            
            Cashbook.objects.create(
                amount=expense.amount,
                debit=True,
                credit=False,
                description=f'Expense ({expense.description}): cancelled'
            )
            return JsonResponse({'success': True, 'message': 'Expense successfully deleted'})
        except Exception as e:
             return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
def update_expense_status(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            expense_id = data.get('id')
            status = data.get('status')

            expense = Expense.objects.get(id=expense_id)
            expense.status = status
            expense.save()

            return JsonResponse({'success': True, 'message': 'Status updated successfully.'})
        except Expense.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Expense not found.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def invoice(request):
    form = InvoiceForm()
    invoices = Invoice.objects.filter(branch=request.user.branch, status=True, cancelled=False).select_related(
        'branch',
        'currency',
        'user'
    ).order_by('-invoice_number')

    query_params = request.GET
    if query_params.get('q'):
        search_query = query_params['q']
        invoices = invoices.filter(
            Q(customer__name__icontains=search_query) |
            Q(invoice_number__icontains=search_query) |
            Q(issue_date__icontains=search_query)
        )

    user_timezone_str = request.user.timezone if hasattr(request.user, 'timezone') else 'UTC'
    user_timezone = pytz_timezone(user_timezone_str)  

    def filter_by_date_range(start_date, end_date):
        start_datetime = user_timezone.localize(
            timezone.datetime.combine(start_date, timezone.datetime.min.time())
        )
        end_datetime = user_timezone.localize(
            timezone.datetime.combine(end_date, timezone.datetime.max.time())
        )
        return invoices.filter(issue_date__range=[start_datetime, end_datetime])

    now = timezone.now().astimezone(user_timezone)
    today = now.date()

    now = timezone.now() 
    today = now.date()  
    
    date_filters = {
        'today': lambda: filter_by_date_range(today, today),
        'yesterday': lambda: filter_by_date_range(today - timedelta(days=1), today - timedelta(days=1)),
        't_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday()), today),
        'l_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday() + 7), today - timedelta(days=today.weekday() + 1)),
        't_month': lambda: invoices.filter(issue_date__month=today.month, issue_date__year=today.year),
        'l_month': lambda: invoices.filter(issue_date__month=today.month - 1 if today.month > 1 else 12, issue_date__year=today.year if today.month > 1 else today.year - 1),
        't_year': lambda: invoices.filter(issue_date__year=today.year),
    }

    if query_params.get('day') in date_filters:
        invoices = date_filters[query_params['day']]()

    total_partial = invoices.filter(payment_status='Partial').aggregate(Sum('amount'))['amount__sum'] or 0
    total_paid = invoices.filter(payment_status='Paid').aggregate(Sum('amount'))['amount__sum'] or 0
    total_amount = invoices.aggregate(Sum('amount'))['amount__sum'] or 0

    grouped_invoices = defaultdict(list)

    for invoice in invoices:

        issue_date = invoice.issue_date.date() 

        if issue_date == today:
            grouped_invoices['Today'].append(invoice)
        elif issue_date == today - timedelta(days=1):
            grouped_invoices['Yesterday'].append(invoice)
        else:
            grouped_invoices[issue_date.strftime('%A, %d %B %Y')].append(invoice)

    return render(request, 'invoices/invoice.html', {
        'form': form,
        'grouped_invoices': dict(grouped_invoices),
        'total_paid': total_paid,
        'total_due': total_partial,
        'total_amount': total_amount,
    })

@login_required
@transaction.atomic 
def update_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    customer_account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(
        CustomerAccountBalances, account=customer_account, currency=invoice.currency
    )

    if request.method == 'POST':
        data = json.loads(request.body)
        amount_paid = Decimal(data['amount_paid'])

        invoice = Invoice.objects.select_for_update().get(pk=invoice.pk)
        customer_account_balance = CustomerAccountBalances.objects.select_for_update().get(pk=customer_account_balance.pk)

        if amount_paid <= 0:
            return JsonResponse({'success': False, 'message': 'Invalid amount paid.'}, status=400)

        if amount_paid >= invoice.amount_due:
            invoice.payment_status = Invoice.PaymentStatus.PAID
            invoice.amount_due = 0
        else:
            invoice.amount_due -= amount_paid

        invoice.amount_paid += amount_paid
        
        # get the latest payment for the invoice
        latest_payment = Payment.objects.filter(invoice=invoice).order_by('-payment_date').first()
        if latest_payment:
            amount_due = latest_payment.amount_due - amount_paid 
        else:
            amount_due = invoice.amount - invoice.amount_paid 

        payment = Payment.objects.create(
            invoice=invoice,
            amount_paid=amount_paid,
            amount_due=amount_due, 
            payment_method=data['payment_method'],
            user=request.user
        )

        account, _ = Account.objects.get_or_create(
            name=f"{request.user.branch} {invoice.currency.name} {payment.payment_method.capitalize()} Account",
            type=Account.AccountType[payment.payment_method.upper()] 
        )
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=invoice.currency,
            branch=request.user.branch,
            defaults={'balance': 0}
        )

        account_balance.balance += amount_paid
        if customer_account_balance.balance < 0:
            customer_account_balance.balance += amount_paid
        else:
            customer_account_balance.balance -= amount_paid

        description = ''
        if invoice.hold_status:
            description = 'Held invoice payment'
            sale = Sale.objects.create(
                date=timezone.now(),
                transaction=invoice,
                total_amount=invoice.amount # invoice delivery amount
            )
            
            VATTransaction.objects.create(
                invoice=invoice,
                vat_type=VATTransaction.VATType.OUTPUT,
                vat_rate=VATRate.objects.get(status=True).rate,
                tax_amount=invoice.vat
            ) 

        else:
            description = 'Invoice payment update'
        
        Cashbook.objects.create(
            issue_date=invoice.issue_date,
            description=f'({description} {invoice.invoice_number})',
            debit=True,
            credit=False,
            amount=invoice.amount_paid,
            currency=invoice.currency,
            branch=invoice.branch
        )

        invoice.hold_status = False
        account_balance.save()
        customer_account_balance.save()
        invoice.save()
        payment.save()
        
        return JsonResponse({'success': True, 'message': 'Invoice successfully updated'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}) 

def update_invoice_amounts(invoice, amount_paid):
    invoice_payments = Payment.objects.filter(invoice=invoice)

    if amount_paid > 0:
        for payment in invoice_payments:
            amount_paid -= payment.amount_due
            payment.save()

def held_invoice(items_data, invoice, request, vat_rate):
    for item_data in items_data:
        item = Inventory.objects.get(pk=item_data['inventory_id'])
        # product = Product.objects.get(pk=item.product.id)
        item.quantity -= item_data['quantity']
        item.save()
                  
        InvoiceItem.objects.create(
            invoice=invoice,
            item=item,
            quantity=item_data['quantity'],
            unit_price=item_data['price'],
            vat_rate = vat_rate
        )
                    
        # # Create StockTransaction for each sold item
        # stock_transaction = StockTransaction.objects.create(
        #     item=item,
        #     transaction_type=StockTransaction.TransactionType.SALE,
        #     quantity=item_data['quantity'],
        #     unit_price=item.price,
        #     invoice=invoice,
        #     date=timezone.now()
        # )
              
        # stock log  
        ActivityLog.objects.create(
            branch=request.user.branch,
            inventory=item,
            user=request.user,
            quantity=item_data['quantity'],
            total_quantity = item.quantity,
            action='Sale',
            invoice=invoice
        )

@login_required
def submit_invoice_data_zimra(request):
    try:

        data = json.loads(request.body)
        hash = data.get('hash', '')
        signature = data.get('signature', '') 
        receipt_data = data.get('receipt_data')
        invoice_id = data.get('invoice_id')

        logger.info(receipt_data)

        if not hash:
            return JsonResponse({'success':False,'message':f'Hash data is missing!'}, status=400)

        if not signature:
            return JsonResponse({'success':False,'message':f'Signature data is missing!'}, status=400)
        
        try:
            submit_receipt_data(request, receipt_data, hash, signature)
            logger.info('done')
        except Exception as e:
            logger.info(e)
            return JsonResponse(
                {
                    'success':False,
                    'messsage':f'{e}'
                },
                status=400
            )
        
        invoice_data = invoice_preview_json(request, invoice_id)
        logger.info(invoice_data)

        return JsonResponse({'success':True, 'message':'data received', 'data':invoice_data}, status=200)
    except Exception as e:
        return JsonResponse({'message':f'{e}', 'success':False}, status=200)


@login_required
def get_signature_data(request):
    try:   
        data = json.loads(request.body)
        hash = data.get('hash', '')
        signature = data.get('signature', '') 

        if not hash:
            return JsonResponse({'success':False,'message':f'Hash data is missing!'}, status=400)

        if not signature:
            return JsonResponse({'success':False,'message':f'Signature data is missing!'}, status=400)

        return JsonResponse({'success':True, 'message':'data received'}, status=200)
        
    except Exception as e:
        return JsonResponse({'success':False,'message':f'{e}'}, status=400)
    

@login_required
def held_invoice_view(request):
    form = InvoiceForm()
    invoices = Invoice.objects.filter(branch=request.user.branch, status=True, hold_status =True).order_by('-invoice_number')
    logger.info(f'Held invoices: {invoices}')
    return render(request, 'invoices/held_invoices.html', {'invoices':invoices, 'form':form})


def create_invoice_pdf(invoice):
    # Buffer to hold the PDF
    buffer = io.BytesIO()
    
    # Create the PDF object, using the buffer as its "file."
    pdf = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    heading_style = styles['Heading1']
    
    # Company logo and header
    elements.append(Paragraph('<b>Tech City</b>', heading_style))
    elements.append(Paragraph('See • Touch • Own Quality', normal_style))
    elements.append(Paragraph(f'Invoice Number: {invoice.invoice_number}', normal_style))
    
    elements.append(Spacer(1, 12))
    
    # Table Data (Items)
    data = [
        ['Q.', 'Description', 'Amount'],
        [1, 'Hp (hp 250)', 'USD 250.00'],
        ['Sub Total', '', 'USD 250.00'],
        ['Discount', '', 'USD 0.00'],
        ['VAT @15%', '', 'USD 37.50'],
        ['Delivery Charge', '', 'USD 0.00'],
        ['Previous Due', '', 'USD 75.00'],
        ['Current Due', '', 'USD 287.50'],
        ['Total Balance', '', 'USD 362.50'],
        ['Amount Paid', '', 'USD 362.50'],
        ['Due Amount', '', 'USD 0.00']
    ]
    
    # Create table
    table = Table(data, colWidths=[0.5*inch, 3*inch, 1.5*inch])
    
    # Add style to the table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
    ]))
    
    elements.append(table)
    
    elements.append(Spacer(1, 24))
    
    # Terms and conditions
    terms = '''
    All laptop in-built batteries attract 1 month warranty.
    Non in-built batteries attract 48hrs warranty.
    Warranty for all preowned laptops is 5 months. Tech City does not warranty laptops if damaged by water, liquids, or short circuits.
    Any withdrawn deposits for any purchase will attract 10percent administration fee.
    Tech City only accepts exchanges on faulty laptops.
    '''
    elements.append(Paragraph('Terms and Conditions', heading_style))
    elements.append(Paragraph(terms, normal_style))
    
    elements.append(Spacer(1, 12))
    elements.append(Paragraph('Thanks for your purchase!', normal_style))
    
    # Build PDF
    pdf.build(elements)
    
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'Invoice_{invoice.invoice_number}.pdf')


@login_required
@transaction.atomic
def invoice_returns(request, invoice_id): # dont forget the payments
    invoice = get_object_or_404(Invoice, id=invoice_id)
    account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

    sale = get_object_or_404(Sale, transaction=invoice)
    invoice_payment = get_object_or_404(Payment, invoice=invoice)
    stock_transactions = invoice.stocktransaction_set.all()  
    vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)
    activity = ActivityLog.objects.filter(invoice=invoice)

    if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
        customer_account_balance.balance -= invoice.amount_due

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account = get_object_or_404(
        Account, 
        name=f"{request.user.branch} {invoice.currency.name} {invoice_payment.payment_method.capitalize()} Account", 
        type=account_types.get(invoice_payment.payment_method, None) 
    )
    account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
    account_balance.balance -= invoice.amount_paid

    for stock_transaction in activity:
        product = Inventory.objects.get(product=stock_transaction.inventory, branch=request.user.branch)
        product.quantity += stock_transaction.quantity
        product.save()

        logger.info(f'product quantity {product.quantity}')
        logger.info(f'stock quantity {stock_transaction.quantity}')

        ActivityLog.objects.create(
            invoice=invoice,
            product_transfer=None,
            branch=request.user.branch,
            user=request.user,
            action='returns',
            inventory=product,
            quantity=stock_transaction.quantity,
            total_quantity=product.quantity
        )

    InvoiceItem.objects.filter(invoice=invoice).delete() 
    StockTransaction.objects.filter(invoice=invoice).delete()
    Payment.objects.filter(invoice=invoice).delete()

    account_balance.save()
    customer_account_balance.save()
    sale.delete()
    vat_transaction.delete()
    invoice.invoice_return=True
    invoice.save()

    return JsonResponse({'message': f'Invoice {invoice.invoice_number} successfully deleted'})
    

@login_required
@transaction.atomic
def delete_invoice(request, invoice_id):
    try:
        invoice = get_object_or_404(Invoice, id=invoice_id)
        account = get_object_or_404(CustomerAccount, customer=invoice.customer)
        customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

        sale = get_object_or_404(Sale, transaction=invoice)
        payments = Payment.objects.filter(invoice=invoice)  
        vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)
        activity = ActivityLog.objects.filter(invoice=invoice)
        
        with transaction.atomic():
            if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
                customer_account_balance.balance -= invoice.amount_due

            account_types = {
                'cash': Account.AccountType.CASH,
                'bank': Account.AccountType.BANK,
                'ecocash': Account.AccountType.ECOCASH,
            }

            for payment in payments:
                account = get_object_or_404(
                    Account, 
                    name=f"{request.user.branch} {invoice.currency.name} {payment.payment_method.capitalize()} Account", 
                    type=account_types.get(payment.payment_method, None)
                )
                account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
                account_balance.balance -= payment.amount_due
                account_balance.save()

            for stock_transaction in activity:
                product = Inventory.objects.get(id=stock_transaction.inventory.id, branch=request.user.branch)
                product.quantity += abs(stock_transaction.quantity)
                product.save()

                logger.info(f'product quantity {stock_transaction.quantity}')

                ActivityLog.objects.create(
                    invoice=invoice,
                    product_transfer=None,
                    branch=request.user.branch,
                    user=request.user,
                    action='sale return',
                    inventory=product,
                    quantity=stock_transaction.quantity,
                    total_quantity=product.quantity
                )

            InvoiceItem.objects.filter(invoice=invoice).delete() 
            StockTransaction.objects.filter(invoice=invoice).delete()
            payments.delete()
            customer_account_balance.save()
            sale.delete()
            vat_transaction.delete()
            invoice.cancelled = True
            invoice.save()

            logger.info(f'Invoice {invoice.invoice_number} successfully deleted')

        return JsonResponse({'success': True, 'message': f'Invoice {invoice.invoice_number} successfully deleted'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f"{e}"})
      
@login_required       
def invoice_details(request, invoice_id):
    invoice = Invoice.objects.filter(id=invoice_id, branch=request.user.branch).values(
        'invoice_number',
        'customer__id', 
        'customer__name', 
        'products_purchased', 
        'payment_status', 
        'amount'
    )
    return JsonResponse(list(invoice), safe=False)


@login_required
def layby_data(request):
    if request.method == 'GET':
        layby_data = layby.objects.all().select_related(
            'invoice',
            'branch'
        ).values()
        return JsonResponse(list(layby_data), safe=False)
    
    if request.method == 'POST':
        logger.info(f'layby data')
        data = json.loads(request.body)

        invoice_id = data.get('invoice_id')

        if not invoice_id:
            return JsonResponse({'success': False, 'message': 'Invoice ID is required.'})

        laby_dates = laybyDates.objects.filter(layby__invoice__id=invoice_id).values()
        
        logger.info(laby_dates)
        return JsonResponse({'success': True, 'data': list(laby_dates)})

@login_required
@transaction.atomic
def layby_payment(request, layby_date_id):
    try:
        data = json.loads(request.body)
        amount_paid = data.get('amount_paid')
        payment_method = data.get('payment_method')

        layby_date = laybyDates.objects.get(id=layby_date_id)
        layby_obj = layby.objects.get(id=layby_date.layby.id)
        invoice = layby_obj.invoice

        account = CustomerAccount.objects.get(customer=invoice.customer)
        customer_account_balance = CustomerAccountBalances.objects.get(account=account, currency=invoice.currency)
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        customer_account_balance.balance -= amount_paid

        account_name = f"{request.user.branch} {invoice.currency.name} {'cash'.capitalize()} Account"
        account = Account.objects.get(name=account_name, type=account_types['cash'])
        account_balance = AccountBalance.objects.get(account=account, currency=invoice.currency, branch=request.user.branch)

        account_balance.balance -= amount_paid

        amount_paid = layby_date.amount_paid
        amount_due = layby_date.amount_due

        with transaction.atomic():

            account_balance.save()
            customer_account_balance.save()

            # create a payment object
            Payment.objects.create(
                invoice=invoice,
                amount_paid=amount_paid,
                amount_due=amount_due, 
                payment_method=payment_method,
                user=request.user
            )

            # create a cash book object
            Cashbook.objects.create(
                issue_date=timezone.now(),
                description=f'Layby payment ({layby_date.layby.invoice.invoice_number})',
                debit=False,
                credit=True,
                amount=amount_paid,
                currency=layby_date.layby.invoice.currency,
                branch=request.user.branch
            )

            if amount_paid >= amount_due:
                layby_date.paid = True
                layby_date.save()
                layby_obj.fully_paid = True
                layby_obj.save()
                invoice.payment_status = Invoice.PaymentStatus.PAID
                invoice.save()

                layby.check_payment_status()

                return JsonResponse({'success': True, 'message': 'Layby payment successfully completed.'})
            else:
                return JsonResponse({'success': False, 'message': 'Invalid amount paid.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'{e}'})

@login_required
def customer(request):
    if request.method == 'GET':
        customers = Customer.objects.all().values()
        return JsonResponse(list(customers), safe=False)
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        
        # validation_errors = validate_customer_data(data)
        # if validation_errors:
        #     return JsonResponse({'success': False, 'message': 'Validation errors occurred:', 'errors': validation_errors})
    
        if Customer.objects.filter(phone_number=data['phonenumber']).exists():
            return JsonResponse({'success': False, 'message': 'Customer exists'})
        else:
            customer = Customer.objects.create(
                name=data['name'],
                email=data['email'],
                address=data['address'],
                phone_number=data['phonenumber'],
                branch=request.user.branch
            )
            account = CustomerAccount.objects.create(customer=customer)
            
            logger.info(account)

        balances_to_create = [
            CustomerAccountBalances(account=account, currency=currency, balance=0) 
            for currency in Currency.objects.all()
        ]
        CustomerAccountBalances.objects.bulk_create(balances_to_create)
    

        return JsonResponse({'success': True, 'message': 'Customer successfully created'})

    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

def validate_customer_data(data):
    errors = {}
    if 'name' not in data or len(data['name']) < 2:
        errors['name'] = 'Name is required and must be at least 2 characters long.'

    if 'email' not in data or not validate_email(data['email']):
        errors['email'] = 'A valid email address is required.'

    if 'address' not in data:
        errors['address'] = 'Address is required.'

    if 'phonenumber' not in data:
        errors['phonenumber'] = 'Phone number is required.'

    return errors

def validate_email(email):
    import re
    email_regex = r"^[a-zA-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[a-zA-Z0-9-]+(?:\.[a-zA-Z0-9-]+)*$"
    return bool(re.match(email_regex, email))

@login_required
def customer_list(request):
    search_query = request.GET.get('q', '')
    
    customers = Customer.objects.filter(branch=request.user.branch)
    accounts = CustomerAccountBalances.objects.all()
    
    total_balances_per_currency = CustomerAccountBalances.objects.filter(account__customer__branch=request.user.branch).values('currency__name').annotate(
        total_balance=Sum('balance')
    )
    
    if search_query:
        customers = CustomerAccount.objects.filter(Q(customer__name__icontains=search_query))
        
    if 'receivable' in request.GET:
        negative_balances_per_currency = CustomerAccountBalances.objects.filter(account__customer__branch=request.user.branch, balance__lt=0) \
            .values('currency') \
            .annotate(total_balance=Sum('balance'))

        customers = Customer.objects.filter(
            id__in=negative_balances_per_currency.values('account__customer_id'),
        ).distinct()
        
        total_balances_per_currency = negative_balances_per_currency.values('currency__name').annotate(
            total_balance=Sum('balance')
        )
        
        logger.info(f'Customers:{total_balances_per_currency.values}')

    if 'download' in request.GET: 
        customers = Customer.objects.all() 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=customers.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header_title in enumerate(['Customer Name', 'Phone Number', 'Email', 'Account Balance'], start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.alignment = header_alignment
            
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header_title), 20)

        customer_accounts = CustomerAccountBalances.objects.all()
        for customer in customer_accounts:
            worksheet.append(
                [
                    customer.account.customer.name, 
                    customer.account.customer.phone_number, 
                    customer.account.customer.email, 
                    customer.balance if customer.balance else 0,
                ]
            )  
            
        workbook.save(response)
        return response
        
    return render(request, 'customers/customers.html', {
        'customers':customers, 
        'accounts':accounts,
        'total_balances_per_currency':total_balances_per_currency,
    })

@login_required
def update_customer(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)

    if request.method == 'POST':  
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f'{customer.name} details updated successfully')  #
            return redirect('finance:customer_list')  
    else:
        form = CustomerForm(instance=customer)  

    return render(request, 'customers/update_customer.html', {'form': form, 'customer': customer}) 

def delete_customer(request, customer_id):
    if request.method == 'DELETE':
        customer = get_object_or_404(Customer, pk=customer_id)

        customer_name = customer.name  
        customer.delete()
        messages.success(request, f'{customer_name} deleted successfully.')
        return JsonResponse({'status': 'success', 'message': f'Customer {customer_name} deleted successfully.'})  
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})  
    

@login_required
def customer_account(request, customer_id):
    form = customerDepositsForm()
    refund_form = customerDepositsRefundForm()
    customer = get_object_or_404(Customer, id=customer_id)

    account = CustomerAccountBalances.objects.filter(account__customer=customer)

    invoices = Invoice.objects.filter(
        customer=customer, 
        branch=request.user.branch, 
        status=True
    )
    
    invoice_payments = Payment.objects.filter(
        invoice__branch=request.user.branch, 
        invoice__customer=customer
    ).order_by('-payment_date')

    filters = Q()
    if request.GET.get('q'):
        filters &= Q(payment_status=request.GET['q'])
    if request.GET.get('search_query'):
        search_query = request.GET['search_query']
        filters &= (Q(invoice_number__icontains=search_query) | Q(issue_date__icontains=search_query))

    invoices = invoices.filter(filters)

    if request.GET.get('email_bool'):
        send_account_statement_email(customer.id, request.user.branch.id, request.user.id)
        return JsonResponse({'message': 'Email sent'})

    return render(request, 'customer.html', {
        'form':form,
        'account': account,
        'invoices': invoices,
        'customer': customer,
        'refund_form':refund_form,
        'invoice_count': invoices.count(),
        'invoice_payments': invoice_payments,
        'paid': invoices.filter(payment_status='Paid').count(),  
        'due': invoices.filter(payment_status='Partial').count(), 
    })


@login_required
@transaction.atomic
def add_customer_deposit(request, customer_id):
    # payload
    """
        customer_id
        amount
        currency
        payment_method
        reason
        payment_reference
    """
    
    try: 
        # get payload
        data = json.loads(request.body)
        customer_id = data.get('customer_id')
        amount = data.get('amount')
        currency = data.get('currency')
        payment_method = data.get('payment_method')
        reason = data.get('reason')
        payment_reference = data.get('payment_reference')        
        
        # payment_reference validation
        if CustomerDeposits.objects.filter(payment_reference=payment_reference).exists():
            return JsonResponse(
                {
                    'success':False,
                    'message': f'Payment reference: {payment_reference} exists'
                }
            )   
                                                   
        # get currency
        currency = Currency.objects.get(id=currency)
        
        # get account types
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        
        account_name = f"{request.user.branch} {currency.name} {payment_method.capitalize()} Account"
        
        
        account, _ = Account.objects.get_or_create(name=account_name, type=account_types[payment_method])
        
        # get or create the account balances
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=currency,
            branch=request.user.branch,
            defaults={'balance': 0}  
        )
        
        account_balance.balance += Decimal(amount)
        account_balance.save()
        logger.info(f"[FINANCE]: deposit -> System {account}")
        
        # check if customer exits
        customer = get_object_or_404(Customer, id=customer_id)  
        logger.info(f"[FINANCE]: deposit -> customer {customer}")
        customer_account = CustomerAccount.objects.get(customer=customer)
        
        customer_account_bal_object, _ = CustomerAccountBalances.objects.get_or_create(
                account=customer_account,
                currency=currency, 
                defaults={'balance': 0}
        )  
        logger.info(f"[FINANCE]: deposit -> customer account object {customer_account_bal_object}")
        
        # effect customer deposit
        customer_deposit = CustomerDeposits.objects.create(
            customer_account=customer_account_bal_object,
            amount=amount,
            currency=currency,
            payment_method=payment_method,
            reason=reason,
            payment_reference=payment_reference,
            cashier=request.user,
            branch=request.user.branch
        )
        
        # effect customer account balances
        customer_account_bal_object.balance += amount
        
        customer_account_bal_object.save()
        
        Cashbook.objects.create(
            issue_date=customer_deposit.date_created,
            description=f'{customer_deposit.payment_method.upper()} deposit ({customer_deposit.customer_account.account.customer.name})',
            debit=True,
            credit=False,
            amount=customer_deposit.amount,
            currency=customer_deposit.currency,
            branch=customer_deposit.branch
        )

        return JsonResponse(
            {
                "success":True,
                "message": f"Customer Deposit of {currency} {amount:2f} has been successfull",
            },
            status=200
        )
    except Exception as e:
        return JsonResponse(
            {
                "message": f"{e}",
                'success':False
            },status=500)


@login_required    
def deposits_list(request):
    deposits = CustomerDeposits.objects.filter(branch=request.user.branch).order_by('-date_created')
    return render(request, 'deposits.html', {
        'deposits':deposits,
        'total_deposits': deposits.aggregate(Sum('amount'))['amount__sum'] or 0,
    })

@login_required
@transaction.atomic
def refund_customer_deposit(request, deposit_id):
    try:
        deposit = CustomerDeposits.objects.get(id=deposit_id)
    except CustomerDeposits.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Deposit not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        amount = Decimal(data.get('amount', 0))
        if amount <= 0:
            return JsonResponse({'success': False, 'message': 'Invalid amount'}, status=400)
    except (json.JSONDecodeError, TypeError, ValueError):
        return JsonResponse({'success': False, 'message': 'Invalid input data'}, status=400)

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"

    try:
        account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
        account_balance = AccountBalance.objects.get(
            account=account,
            currency=deposit.currency,
            branch=request.user.branch,
        )
    except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    if amount > deposit.amount:
        return JsonResponse({'success': False, 'message': 'Refund amount exceeds deposit amount'}, status=400)
    
    account_balance.balance -= amount
    diff_amount = deposit.amount - amount

    if diff_amount == 0:
        deposit.delete()
    else:
        deposit.amount = diff_amount
        deposit.save()

    Cashbook.objects.create(
        issue_date=datetime.date.today(),
        description=f'{deposit.payment_method.upper()} deposit refund ({deposit.customer_account.account.customer.name})',
        debit=False,
        credit=True,
        amount=amount,
        currency=deposit.currency,
        branch=deposit.branch
    )

    account_balance.save()

    return JsonResponse({'success': True}, status=200)

        
@login_required
@transaction.atomic
def edit_customer_deposit(request, deposit_id):
    try:
        deposit = CustomerDeposits.objects.get(id=deposit_id)
    except CustomerDeposits.DoesNotExist:
        messages.warning(request, 'Deposit not found')
        return redirect('finance:customer_account', deposit.customer_account.account.customer.id)
    
    if request.method == 'POST':
        form = customerDepositsForm(request.POST)
        if not form.is_valid():
            messages.warning(request, 'Invalid form submission')
            return redirect('finance:edit_customer_deposit', deposit_id)

        amount = Decimal(request.POST.get('amount'))
        if amount <= 0:
            messages.warning(request, 'Amount cannot be zero or negative')
            return redirect('finance:edit_customer_deposit', deposit_id)

        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"
        
        try:
            account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
            account_balance = AccountBalance.objects.get(
                account=account,
                currency=deposit.currency,
                branch=request.user.branch,
            )
        except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
            messages.warning(request, str(e))
            return redirect('finance:edit_customer_deposit', deposit_id)
        
        adj_amount = amount - deposit.amount

        if adj_amount != 0:
            if adj_amount > 0:
                account_balance.balance += adj_amount
                debit, credit = True, False
            else:
                account_balance.balance += adj_amount 
                debit, credit = False, True

            Cashbook.objects.create(
                issue_date=datetime.date.today(),
                description=f'{deposit.payment_method.upper()} deposit adjustment ({deposit.customer_account.account.customer.name})',
                debit=debit,
                credit=credit,
                amount=abs(adj_amount),
                currency=deposit.currency,
                branch=deposit.branch
            )

            account_balance.save()
            deposit.amount = amount
            deposit.save()
            messages.success(request, 'Customer deposit successfully updated')
            return redirect('finance:customer', deposit.customer_account.account.customer.id)
    else:
        form = customerDepositsForm(instance=deposit)

    return render(request, 'customers/edit_deposit.html', {'form': form})
    

@login_required
def customer_deposits(request): 
    customer_id = request.GET.get('customer_id')
    
    if customer_id: 
        deposits = CustomerDeposits.objects.filter(branch=request.user.branch).values(
            'customer_account__account__customer_id',
            'date_created',
            'amount', 
            'reason',
            'currency__name', 
            'currency__symbol', 
            'payment_method',
            'payment_reference',
            'cashier__username', 
            'id'
        ).order_by('-date_created')
        return JsonResponse(list(deposits), safe=False)
    else:
        return JsonResponse({
            'success':False,
            'message':f'{customer_id} was not provided'
        })

@login_required
def customer_account_transactions_json(request):
    customer_id = request.GET.get('customer_id')
    transaction_type = request.GET.get('type')

    customer = get_object_or_404(Customer, id=customer_id)  

    if transaction_type == 'invoices':
        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        ).order_by('-issue_date').values(
            'issue_date',
            'invoice_number',
            'products_purchased', 
            'amount_paid', 
            'amount_due', 
            'amount', 
            'user__username',
            'payment_status'
        )
        return JsonResponse(list(invoices), safe=False)
    else:
        return JsonResponse({'message': 'Invalid transaction type.'}, status=400)  

@login_required
def customer_account_payments_json(request):
    customer_id = request.GET.get('customer_id')
    transaction_type = request.GET.get('type')

    customer = get_object_or_404(Customer, id=customer_id)

    if transaction_type == 'invoice_payments':
        invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
            invoice__branch=request.user.branch, 
            invoice__customer=customer
        ).order_by('-payment_date').values(
            'invoice__products_purchased',
            'payment_date',
            'invoice__invoice_number',
            'invoice__currency__symbol', 
            'invoice__payment_status',
            'invoice__amount_due',
            'invoice__amount', 
            'user__username', 
            'amount_paid', 
            'amount_due'
        )
        return JsonResponse(list(invoice_payments), safe=False)
    else:
        return JsonResponse({'message': 'Invalid transaction type.'}, status=400)  


@login_required
def customer_account_json(request, customer_id):
    account = CustomerAccountBalances.objects.filter(account__customer__id=customer_id).values(
        'currency__symbol', 'balance'
    )   
    return JsonResponse(list(account), safe=False)

@login_required
def print_account_statement(request, customer_id):
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        
        account = CustomerAccountBalances.objects.filter(account__customer=customer)
        
        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        )
    except:
        messages.warning(request, 'Error in processing the request')
        return redirect('finance:customer')

    invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
        invoice__branch=request.user.branch, 
        invoice__customer=customer
    ).order_by('-payment_date')
    
    return render(request, 'customers/print_customer_statement.html', {
        'customer':customer,
        'account':account,
        'invoices':invoices, 
        'invoice_payments':invoice_payments
    })

# currency views  
@login_required  
def currency(request):
    return render(request, 'currency/currency.html')

@login_required
def currency_json(request):
    currency_id = request.GET.get('id', '')
    currency = Currency.objects.filter(id=currency_id).values()
    return JsonResponse(list(currency), safe=False)


@login_required
def add_currency(request):
    if request.method == 'POST':
        form = CurrencyForm(request.POST)
        if form.is_valid():
            default = request.POST['default']
            try:
                form.save()
                messages.success(request, 'Currency added successfully!')  
            except Exception as e: 
                messages.error(request, f'Error adding currency: {e}')
            return redirect('finance:currency') 
    else:
        form = CurrencyForm()

    return render(request, 'currency/currency_add.html', {'form': form})


@login_required
def update_currency(request, currency_id):
    currency = get_object_or_404(Currency, id=currency_id)  

    if request.method == 'POST': 
        form = CurrencyForm(request.POST, instance=currency)  
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Currency updated successfully') 
            except Exception as e: 
                messages.error(request, f'Error updating currency: {e}')
            return redirect('finance:currency')  
    else:
        form = CurrencyForm(instance=currency) 

    return render(request, 'currency/currency_add.html', {'form': form})

@login_required
def delete_currency(request, currency_id):
    if request.method == 'POST': 
        currency = get_object_or_404(Currency, id=currency_id)
        
        try:
            if currency.invoice_set.exists() or currency.accountbalance_set.exists() or currency.expense_set.exists():  
                raise Exception("Currency is in use and cannot be deleted.")

            currency.delete()
            return JsonResponse({'message': 'Currency deleted successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'message':'Deletion Failed'})


@login_required
def finance_settings(request):
    return render(request, 'settings/settings.html')
    
# Reports
@login_required
def expenses_report(request):
    
    template_name = 'reports/expenses.html'
    
    search = request.GET.get('search', '')
    start_date_str = request.GET.get('startDate', '')
    end_date_str = request.GET.get('endDate', '')
    category_id = request.GET.get('category', '')
   
    if start_date_str and end_date_str:
        try:
            end_date = datetime.date.fromisoformat(end_date_str)
            start_date = datetime.date.fromisoformat(start_date_str)
        except ValueError:
            return JsonResponse({'messgae':'Invalid date format. Please use YYYY-MM-DD.'})
    else:
        start_date = ''
        end_date= ''
        
    try:
        category_id = int(category_id) if category_id else None
    except ValueError:
        return JsonResponse({'messgae':'Invalid category or search ID.'})

    expenses = Expense.objects.all()  
    
    if search:
        expenses = expenses.filter(Q('amount=search'))
    if start_date:
        start_date = parse_date(start_date_str)
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        end_date = parse_date(end_date_str)
        expenses = expenses.filter(date__lte=end_date)
    if category_id:
        expenses = expenses.filter(category__id=category_id)
    
    return generate_pdf(
        template_name,
        {
            'title': 'Expenses', 
            'date_range': f"{start_date} to {end_date}", 
            'report_date': datetime.date.today(),
            'total_expenses':calculate_expenses_totals(expenses),
            'expenses':expenses
        }
    )


@login_required 
def invoice_preview(request, invoice_id):
    invoice = Invoice.objects.get(id=invoice_id)
    invoice_items = InvoiceItem.objects.filter(invoice=invoice)
    return render(request, 'Pos/printable_receipt.html', {'invoice_id':invoice_id, 'invoice':invoice, 'invoice_items':invoice_items})

@login_required
def remove_item(request, item_id):
    if request.method == 'DELETE':
        try:
            item = InvoiceItem.objects.get(id=item_id)
            item.delete()
            return JsonResponse({'success': True})
        except InvoiceItem.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Item not found'}, status=404)

@login_required
def replace_item(request, item_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        new_item_id = data.get('newItemId')

        try:
            item = InvoiceItem.objects.get(id=item_id)
            new_item = InvoiceItem.objects.get(id=new_item_id)

            # Replace the old item with the new one
            item.name = new_item.name  # Update other relevant fields as needed
            item.price = new_item.price
            item.quantity = new_item.quantity
            item.save()

            return JsonResponse({'success': True})
        except (InvoiceItem.DoesNotExist, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid item'}, status=404)
        
def invoice_preview_json(request, invoice_id):
    from django.core.serializers.json import DjangoJSONEncoder
    try:
        invoice = Invoice.objects.get(id=invoice_id)
    except Invoice.DoesNotExist:
        return JsonResponse({"error": "Invoice not found"}, status=404)

    dates = {}
    if invoice.payment_terms == 'layby':
        dates = laybyDates.objects.filter(layby__invoice=invoice).values('due_date')

    invoice_items = InvoiceItem.objects.filter(invoice=invoice).values(
        'item__name',
        'quantity',
        'item__description',
        'total_amount',
        'unit_price'
    )

    invoice_dict = {}
    invoice_dict['customer_name'] = invoice.customer.name
    invoice_dict['customer_email'] = invoice.customer.email
    invoice_dict['customer_cell'] = invoice.customer.phone_number
    invoice_dict['customer_address'] = invoice.customer.address
    invoice_dict['currency_symbol'] = invoice.currency.symbol
    invoice_dict['amount_paid'] = invoice.amount_paid
    invoice_dict['payment_terms'] = invoice.payment_terms
    invoice_dict['amount'] = invoice.amount
    invoice_dict['invoice_number'] = invoice.invoice_number
    invoice_dict['receipt_hash'] = invoice.receipt_hash
    invoice_dict['subtotal'] = invoice.subtotal
    invoice_dict['vat'] = round(invoice.vat, 2)
    invoice_dict['device_id'] = os.getenv("DEVICE_ID")
    invoice_dict['device_serial_number'] = os.getenv("DEVICE_SERIAL_NUMBER")
    invoice_dict['code'] =  invoice.code
    invoice_dict['fiscal_day'] = invoice.fiscal_day

    if invoice.branch:
        invoice_dict['branch_name'] = invoice.branch.name
        invoice_dict['branch_phone'] = invoice.branch.phonenumber
        invoice_dict['branch_email'] = invoice.branch.email

    invoice_dict['user_username'] = invoice.user.username
    invoice_dict['receipt_signature'] = invoice.receiptServerSignature if invoice.receiptServerSignature else None

    # Safely serialize qr_code
    if invoice.qr_code and hasattr(invoice.qr_code, 'url'):
        try:
            invoice_dict['qr_code'] = request.build_absolute_uri(invoice.qr_code.url)
            logger.info(invoice_dict['qr_code'])
        except Exception as e:
            invoice_dict['qr_code'] = None
            logger.info(f"Error generating QR code URL: {e}")
    else:
        invoice_dict['qr_code'] = None

    invoice_data = {
        'invoice': invoice_dict,
        'invoice_items': list(invoice_items),
        'dates': list(dates)
    }
    return invoice_data

@login_required
def invoice_pdf(request):
    template_name = 'reports/invoice.html'
    invoice_id = request.GET.get('id', '')
    if invoice_id:
        try:
            invoice = get_object_or_404(Invoice, pk=invoice_id)

            invoice_items = InvoiceItem.objects.filter(invoice=invoice)
            
        except Invoice.DoesNotExist:
            return HttpResponse("Invoice not found")
    else:
        return HttpResponse("Invoice ID is required")
    
    return generate_pdf(
        template_name,
        {
            'title': 'Invoice', 
            'report_date': datetime.date.today(),
            'invoice':invoice,
            'invoice_items':invoice_items
        }
    )
   
# emails
@login_required
def send_invoice_email(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        invoice_id = data['invoice_id']
        invoice = Invoice.objects.get(id=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        account = CustomerAccount.objects.get(customer__id = invoice.customer.id)
        
        html_string = render_to_string('Pos/receipt.html', {'invoice': invoice, 'invoice_items':invoice_items, 'account':account})
        buffer = BytesIO()

        pisa.CreatePDF(html_string, dest=buffer) 

        email = EmailMessage(
            'Your Invoice',
            'Please find your invoice attached.',
            'your_email@example.com',
            ['recipient_email@example.com'],
        )
        
        buffer.seek(0)
        email.attach(f'invoice_{invoice.invoice_number}.pdf', buffer.getvalue(), 'application/pdf')

        # Send the email
        email.send()

        task = send_invoice_email_task.delay(data['invoice_id']) 
        task_id = task.id 
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=invoice_{invoice.invoice_number}.pdf'
        
        return response
    return JsonResponse({'success': False, 'error':'error'})


#whatsapp
@login_required
def send_invoice_whatsapp(request, invoice_id):
    try:
        
        invoice = Invoice.objects.get(pk=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        img = settings.STATIC_URL + "/assets/logo.png"
    
        html_string = render_to_string('Pos/invoice_template.html', {'invoice': invoice, 'request':request, 'invoice_items':invoice_items, 'img':img})
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if not pisa_status.err:
          
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            invoice_filename = f"invoice_{invoice.invoice_number}.pdf"
            s3.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f"invoices/{invoice_filename}",
                Body=pdf_buffer.getvalue(),
                ContentType="application/pdf",
                ACL="public-read",
            )
            s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

            account_sid = 'AC6890aa7c095ce1315c4a3a86f13bb403'
            auth_token = '897e02139a624574c5bd175aa7aaf628'
            client = Client(account_sid, auth_token)
            from_whatsapp_number = 'whatsapp:' + '+14155238886'
            to_whatsapp_number = 'whatsapp:' + '+263778587612'

            message = client.messages.create(
                from_=from_whatsapp_number,
                body="Your invoice is attached.",
                to=to_whatsapp_number,
                media_url=s3_url
            )
            logger.info(f"WhatsApp message SID: {message.sid}")
            return JsonResponse({"success": True, "message_sid": message.sid})
        else:
            logger.error(f"PDF generation error for Invoice ID: {invoice_id}")
            return JsonResponse({"error": "PDF generation failed"})
    except Invoice.DoesNotExist:
        logger.error(f"Invoice not found with ID: {invoice_id}")
        return JsonResponse({"error": "Invoice not found"})
    except Exception as e:
        logger.exception(f"Error sending invoice via WhatsApp: {e}")
        return JsonResponse({"error": "Error sending invoice via WhatsApp"})
    
@login_required
def invoice_payment_track(request):
    invoice_id = request.GET.get('invoice_id', '')
    
    if invoice_id:
        payments = Payment.objects.filter(invoice__id=invoice_id).order_by('-payment_date').values(
            'payment_date', 'amount_paid', 'payment_method', 'user__username'
        )
    return JsonResponse(list(payments), safe=False)

@login_required
def day_report(request, inventory_data):
    today_min = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_max = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # invoice data
    invoices = Invoice.objects.filter(branch=request.user.branch, issue_date__range=(today_min, today_max))
    
    partial_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PARTIAL)
    paid_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PAID)
    
    # expenses
    expenses = Expense.objects.filter(branch=request.user.branch, date=datetime.date.today())
    
    confirmed_expenses = expenses.filter(status=True)
    unconfirmed_expenses = expenses.filter(staus=False)
    
    # accounts
    account_balances = AccountBalance.objects.filter(branch=request.user.branch)
    
    try:
        html_string = render_to_string('day_report.html',{
                'request':request,
                'invoices':invoices,
                'date': datetime.date.today(),
                'inventory_data': inventory_data,
                'total_sales': paid_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'partial_payments': partial_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'total_paid_invoices': paid_invoices.count(),
                'expenses':expenses,
                'total_partial_invoices': partial_invoices.count(),
                'total_expenses': confirmed_expenses.aggregate(Sum('amount'))['amount__sum'] or 0,
                'confirmed_expenses': confirmed_expenses,
                'unconfirmed_expenses': unconfirmed_expenses,
                'account_balances': account_balances,
            })
        
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if not pisa_status.err:
            # Save PDF to S3 and get URL
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            invoice_filename = f"{request.user.branch} today's ({datetime.date.today}) report.pdf"
            s3.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f"daily_reports/{invoice_filename}",
                Body=pdf_buffer.getvalue(),
                ContentType="application/pdf",
                ACL="public-read",
            )
            s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

            # Send WhatsApp Message with Twilio
            account_sid = settings.TWILIO_ACCOUNT_SID
            auth_token = settings.TWILIO_AUTH_TOKEN
            client = Client(account_sid, auth_token)
            from_whatsapp_number = 'whatsapp:' + '+14155238886'
            to_whatsapp_number = 'whatsapp:' + '+263778587612'

            message = client.messages.create(
                from_=from_whatsapp_number,
                body="Today's report.",
                to=to_whatsapp_number,
                media_url=s3_url
            )

            logger.info(f"WhatsApp message SID: {message.sid}")
            return JsonResponse({"success": True, "message_sid": message.sid})
    except Exception as e:
        logger.exception(f"Error sending invoice via WhatsApp: {e}")
        return JsonResponse({"error": "Error sending invoice via WhatsApp"})

   
@login_required 
@transaction.atomic 
def cash_transfer(request):
    form = TransferForm()
    transfers = CashTransfers.objects.filter(branch=request.user.branch).select_related(
        'user',
        'currency',
        'to',
        'branch',
        'from_branch'
    )
    
    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    if request.method == 'POST':
        form = TransferForm(request.POST)
        
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.user = request.user
            transfer.notification_type = 'Expense'
            transfer.from_branch = request.user.branch
            transfer.branch = request.user.branch
            transfer.received_status = False
            
            account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"
            logger

            with transaction.atomic():
                try:
                    account = Account.objects.get(name=account_name, type=account_types[transfer.transfer_method.lower()])
                    account_balance = AccountBalance.objects.select_for_update().get(
                        account=account,
                        currency=transfer.currency,
                        branch=request.user.branch
                    )
                
                    if account_balance.balance < transfer.amount:
                        messages.error(request, "Insufficient funds in the account.")
                        return redirect('finance:cash_transfer')  

                    account_balance.balance -= transfer.amount
                    account_balance.save()
                    transfer.save()  

                    Cashbook.objects.create(
                        issue_date=transfer.date,
                        description=f'Cash Transfer to {transfer.to.name}',
                        debit=False,
                        credit=True,
                        amount=transfer.amount,
                        currency=transfer.currency,
                        branch=transfer.branch
                    )
                    
                    messages.success(request, 'Money successfully transferred.')
                    return redirect('finance:cash_transfer')  
          
                except Exception as e:
                    messages.error(request, f"{e}")
                return redirect('finance:cash_transfer')  
        else:
            messages.error(request, "Invalid form data. Please correct the errors.")
    return render(request, 'transfers/cash_transfers.html', {'form': form, 'transfers':transfers})

@login_required
def finance_notifications_json(request):
    notifications = FinanceNotifications.objects.filter(status=True).values(
        'transfer__id', 
        'transfer__to',
        'expense__id',
        'expense__branch',
        'invoice__id',
        'invoice__branch',
        'notification',
        'notification_type',
        'id'
    )
    return JsonResponse(list(notifications), safe=False)


@login_required
@transaction.atomic
def cash_transfer_list(request):
    search_query = request.GET.get('q', '')
    transfers = CashTransfers.objects.filter(to=request.user.branch.id)
    
    if search_query:
        transfers = transfers.filter(Q(date__icontains=search_query))
        
    return render(request, 'transfers/cash_transfers_list.html', {'transfers':transfers, 'search_query':search_query})

@login_required
@transaction.atomic
def receive_money_transfer(request, transfer_id):
    if transfer_id:
        transfer = get_object_or_404(CashTransfers, id=transfer_id)
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        
        account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"

        with transaction.atomic():
            try:
                account, _ = Account.objects.get_or_create(name=account_name, type=account_types[transfer.transfer_method.lower()])
            
                account_balance, _ = AccountBalance.objects.get_or_create(
                    account=account,
                    currency=transfer.currency,
                    branch=request.user.branch
                )

                Cashbook.objects.create(
                    issue_date=transfer.date,
                    description=f'Cash Transfer from {transfer.from_branch.name}',
                    debit=True,
                    credit=False,
                    amount=transfer.amount,
                    currency=transfer.currency,
                    branch=transfer.to
                )

                account_balance.balance += transfer.amount
                account_balance.save()
                
                transfer.received_status = True
                transfer.save() 

                return JsonResponse({'message':True})  
            
            except Exception as e:
                return JsonResponse({'success':False, 'message':f"{e}"}) 
    return JsonResponse({'message':"Transfer ID is needed"})  


@login_required
@transaction.atomic
def create_quotation(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        qoute_data = data['data'][0]  
        items_data = data['items']
        
        customer = Customer.objects.get(id=int(qoute_data['client_id']))
        currency = Currency.objects.get(id=qoute_data['currency'])
        
        qoute = Qoutation.objects.create(
            customer = customer,
            amount =  Decimal(qoute_data['subtotal']),
            branch = request.user.branch,
            currency = currency,
            qoute_reference = Qoutation.generate_qoute_number(request.user.branch.name),
            products = ', '.join([f'{item['product_name']} x {item['quantity']}' for item in items_data])
        )
        
        for item_data in items_data:
            item = Inventory.objects.get(pk=item_data['inventory_id'])
            
            QoutationItems.objects.create(
                qoute=qoute,
                product=item,
                unit_price=item.price,
                quantity=item_data['quantity'],
                total_amount= item.price * item_data['quantity'],
            )
        return JsonResponse({'success': True, 'qoute_id': qoute.id})
    return JsonResponse({'success': False})

@login_required        
def qoutation_list(request):
    search_query = request.GET.get('q', '')
    qoutations = Qoutation.objects.filter(branch=request.user.branch).order_by('-date')
 
    if search_query:

        qoutations = qoutations.filter(
            Q(customer__name__icontains=search_query)|
            Q(products__icontains=search_query)|
            Q(date__icontains=search_query)|
            Q(qoute_reference__icontains=search_query)
        )
        
    return render(request, 'qoutations.html', {'qoutations':qoutations, 'search_query':search_query})
        
@login_required 
def qoute_preview(request, qoutation_id):
    qoute = Qoutation.objects.get(id=qoutation_id)
    qoute_items = QoutationItems.objects.filter(qoute=qoute)
    return render(request, 'qoute.html', {'qoute':qoute, 'qoute_items':qoute_items})

@login_required
def qoute_preview_modal(request, qoutation_id):
    try:
        qoute = Qoutation.objects.get(id=qoutation_id)
        logger.info(qoute)
        qoute_items = QoutationItems.objects.filter(qoute=qoute)
        logger.info(f'qoute items: {qoute_items.values()}')
        html = render_to_string('qoutations/partial_preview.html', {
            'qoute': qoute,
            'qoute_items': qoute_items
        }, request=request)  
        
        logger.info(html)

        return JsonResponse({'success': True, 'html': html}, status=200)
        
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success': False, 'message':str(e)}, status=400)

@login_required
def delete_qoute(request, qoutation_id):
    qoute = get_object_or_404(Qoutation, id=qoutation_id)
    qoute.delete()
    return JsonResponse({'success':True, 'message':'Qoutation successfully deleted'}, status=200)

@login_required
def cashbook_data(request):
    """AJAX endpoint for cashbook data with filters and pagination"""
    logger.info('Processing cashbook data request')

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            page = int(data.get('page', 1))
            per_page = int(data.get('per_page', 20))
            filter_option = data.get('filter', 'this_week')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            search_query = data.get('search', '')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    else:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        filter_option = request.GET.get('filter', 'this_week')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        search_query = request.GET.get('search', '')
    
    logger.info(f'filter: {filter_option}')

    now = timezone.now()
    end_date = now

    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now

    entries = Cashbook.objects.filter(
        issue_date__gte=start_date,
        issue_date__lte=end_date,
        branch=request.user.branch
    ).select_related('created_by', 'branch', 'updated_by', 'currency', 'invoice', 'expense').order_by('-issue_date')
    
    logger.info(f'Found {entries.count()} entries')

    if search_query:
        entries = entries.filter(
            Q(description__icontains=search_query) |
            Q(accountant__icontains=search_query) |
            Q(manager__icontains=search_query) |
            Q(director__icontains=search_query)
        )

    entries = entries.order_by('-issue_date')
    
    # Calculate totals
    total_cash_in = entries.filter(debit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
    total_cash_out = entries.filter(credit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
    total_balance = total_cash_in - total_cash_out

    total_entries = entries.count()
    total_pages = (total_entries + per_page - 1) // per_page
    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    paginated_entries = entries[start_index:end_index]

    balance = 0
    entries_data = []
    for entry in paginated_entries:
        if entry.debit:
            balance += entry.amount
        elif entry.credit:
            balance -= entry.amount
        logger.info(f'Entry: {entry.issue_date.strftime('%Y-%m-%d %H:%M')}')
        entries_data.append({
            'id': entry.id,
            'date': entry.issue_date.strftime('%Y-%m-%d %H:%M'),
            'description': entry.description,
            'debit': float(entry.amount) if entry.debit else None,
            'credit': float(entry.amount) if entry.credit else None,
            'balance': float(balance),
            'accountant': entry.accountant,
            'manager': entry.manager,
            'director': entry.director,
            'status': entry.status,
            'created_by': entry.created_by.first_name
        })

    return JsonResponse({
        'entries': entries_data,
        'totals': {
            'cash_in': float(total_cash_in),
            'cash_out': float(total_cash_out),
            'balance': float(total_balance)
        },
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'total_entries': total_entries,
            'has_next': page < total_pages,
            'has_previous': page > 1
        }
    }) 
    
@login_required
def download_cashbook_report(request):
    filter_option = request.GET.get('filter', 'this_week')
    now = datetime.datetime.now()
    end_date = now
    
    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    else:
        start_date = now - timedelta(days=now.weekday())
        end_date = now

    entries = Cashbook.objects.filter(date__gte=start_date, date__lte=end_date, branch=request.user.branch).order_by('date')

    # Create a CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="cashbook_report_{filter_option}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Description', 'Expenses', 'Income', 'Balance'])

    balance = 0  
    for entry in entries:
        if entry.debit:
            balance += entry.amount
        elif entry.credit:
            balance -= entry.amount

        writer.writerow([
            entry.issue_date,
            entry.description,
            entry.amount if entry.debit else '',
            entry.amount if entry.credit else '',
            balance,
            entry.accountant,
            entry.manager,
            entry.director
        ])

    return response


@login_required
def cashbook_note(request):
    #payload
    """
        entry_id:id,
        note:str
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            entry_id = data.get('entry_id')
            note = data.get('note')
            
            entry = Cashbook.objects.get(id=entry_id)
            entry.note = note
            
            entry.save()
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}.'}, status=400)
        return JsonResponse({'success':False, 'message':'Note successfully saved.'}, status=201)
    return JsonResponse({'success':False, 'message':'Invalid request.'}, status=405)


@login_required
def cashbook_note_view(request, entry_id):
    entry = get_object_or_404(Cashbook, id=entry_id)
    
    if request.method == 'GET':
        notes = entry.notes.all().order_by('timestamp')
        notes_data = [
            {'user': note.user.username, 'note': note.note, 'timestamp': note.timestamp.strftime("%Y-%m-%d %H:%M:%S")}
            for note in notes
        ]
        return JsonResponse({'success': True, 'notes': notes_data})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            note_text = data.get('note')
            CashBookNote.objects.create(entry=entry, user=request.user, note=note_text)
            return JsonResponse({'success': True, 'message': 'Note successfully added.'}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request.'}, status=405)
    
@login_required
def cancel_transaction(request):
    #payload
    """
        entry_id:id,
    """
    try:
        data = json.loads(request.body)
        entry_id = int(data.get('entry_id'))
        
        logger.info(entry_id)
        
        entry = Cashbook.objects.get(id=entry_id)
        
        entry.cancelled = True
        
        if entry.director:
            entry.director = False
        elif entry.manager:
            entry.manager = False
        elif entry.accountant:
            entry.accountant = False
            
        entry.save()
        logger.info(entry)
        return JsonResponse({'success': True}, status=201)
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
def update_transaction_status(request, pk):
    if request.method == 'POST':
        entry = get_object_or_404(Cashbook, pk=pk)
        
        data = json.loads(request.body)
        
        status = data.get('status')
        field = data.get('field')  

        if field in ['manager', 'accountant', 'director']:
            setattr(entry, field, status)

            if entry.cancelled:
                entry.cancelled = False
            entry.save()
            return JsonResponse({'success': True, 'status': getattr(entry, field)})
        
    return JsonResponse({'success': False}, status=400)   
    
@login_required
def cashWithdrawals(request):
    search_query = request.GET.get('q', '')
    selected_query = request.GET.get('sq', '')
    
    withdrawals = CashWithdraw.objects.all().order_by('-date')
    
    if search_query:
        withdrawals = withdrawals.filter(
            Q(user__branch__name__icontains=search_query)|
            Q(amount__icontains=search_query)|
            Q(date__icontains=search_query)|
            Q(reason__icontains=search_query)
        )
    if selected_query:
        withdrawals = CashWithdraw.objects.filter(deleted=True).order_by('-date')
        
    if 'download' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=withdrawals.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header_title in enumerate(['Date', 'User', 'Amount', 'Reason', 'Status'], start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.alignment = header_alignment
            
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header_title), 20)

        withdrawals = CashWithdraw.objects.all().order_by('-date')
        for withdrawal in withdrawals:
            worksheet.append(
                [
                    withdrawal.date,
                    withdrawal.user.username,
                    withdrawal.amount,
                    withdrawal.reason,
                    'Canceled' if withdrawal.deleted else 'Expensed' if withdrawal.status else 'pending'
                ])  
            
        workbook.save(response)
        return response
    
    form = CashWithdrawForm()
    expense_form = cashWithdrawExpenseForm()
    
    if request.method == 'POST':
        form = CashWithdrawForm(request.POST)
        
        if form.is_valid():
             
            password = form.cleaned_data['password']
            currency = form.cleaned_data['currency']
            amount = form.cleaned_data['amount']
            
            user = authenticate(username=request.user.username, password=password)
            
            if user is None:
                messages.warning(request, 'Incorrect password')
                return redirect('finance:withdrawals')
            
            cw_obj = form.save(commit=False)
            cw_obj.user = user
            cw_obj.save()
            
            account_name = f"{request.user.branch} {currency.name} {'Cash'} Account"
            
            try:
                account = Account.objects.get(name=account_name, type=Account.AccountType.CASH)
            except Account.DoesNotExist:
                messages.error(request, f'{account_name} doesnt exists')
                return redirect('finance:withdrawals')

            try:
                account_balance = AccountBalance.objects.get(account=account,  branch=request.user.branch)
            except AccountBalance.DoesNotExist:
                messages.error(request, f'Account Balances for account {account_name} doesnt exists')
                return redirect('finance:withdrawals')
            
            account_balance.balance -= Decimal(amount)
            account_balance.save()
            messages.success(request, 'Cash Withdrawal Successfully saved')
        else:
            messages.error(request, 'Invalid form data')
    return render(request, 'cashWithdaraws/withdrawals.html', 
        {
            'withdrawals':withdrawals,
            'count': withdrawals.filter(status=False, deleted=False).count(),
            'expense_form':expense_form,
            'form':form,
        }
    )

@login_required
@transaction.atomic
def cash_withdrawal_to_expense(request):
    if request.method == 'GET':
        cwte_id = request.GET.get('id', '')
        withdrawals = CashWithdraw.objects.filter(id=cwte_id).values(
            'user__branch__name', 'amount', 'reason', 'currency__id', 'user__id'
        )
        return JsonResponse(list(withdrawals), safe=False)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        
        withdrawal_data = data['withdrawal'][0]
        
        reason = data['reason']
        category_id = data['category_id']
        withdrawal_id = data['withdrawal_id']
        currency_id = withdrawal_data['currency__id']
        branch_name = withdrawal_data['user__branch__name']
        amount = withdrawal_data['amount']
        
        try:
            currency = Currency.objects.get(id=currency_id)
            branch = Branch.objects.get(name=branch_name)
            withdrawal = CashWithdraw.objects.get(id=withdrawal_id)
            category = ExpenseCategory.objects.get(id=category_id)
        except:
            return JsonResponse({'success':False,'message':'Invalid form data here'})
        
        Expense.objects.create(
            category=category,
            amount=amount,
            branch=branch,
            user = request.user,
            currency = currency,
            description=f'Cash withdrawal: {reason}',
            status=True,
            issue_date=withdrawal.date,
            payment_method='cash'
        )
        
        logger.info(withdrawal_id)
        withdrawal.status=True
        withdrawal.save()
        
        return JsonResponse({'success':True, 'message':'Successfully added to expenses'}, status=201)
    return JsonResponse({'success':False, 'message':'Invalid form data'}, status=400)
       
@login_required
def delete_withdrawal(request, withdrawal_id):
    try:
        withdrawal = CashWithdraw.objects.get(id=withdrawal_id)
    except User.DoesNotExist:
        messages.warning(request, 'Withdrawal doesnt exist')
        return redirect('finance:withdrawals')
    
    account_name = f"{request.user.branch} {withdrawal.currency.name} {'Cash'} Account"
    
    try:
        account = Account.objects.get(name=account_name, type=Account.AccountType.CASH)
    except Account.DoesNotExist:
        messages.error(request, f'{account_name} doesnt exists')
        return redirect('finance:withdrawals')
    
    try:
        account_balance = AccountBalance.objects.get(account=account,  branch=request.user.branch)
    except AccountBalance.DoesNotExist:
        messages.error(request, f'Account Balances for account {account_name} doesnt exists')
        return redirect('finance:withdrawals')
    
    account_balance.balance += Decimal(withdrawal.amount)
    account_balance.save()
    withdrawal.deleted=True
    withdrawal.save()
    
    messages.success(request, 'Withdrawal successfully deleted')
    return redirect('finance:withdrawals')
    
    
@login_required
def days_data(request):
    current_month = get_current_month()

    sales = Sale.objects.filter(date__month=current_month)
    cogs = COGSItems.objects.filter(date__month=current_month)

    first_day = min(sales.first().date, cogs.first().date)
    
    def get_week_data(queryset, start_date, end_date, amount_field):
        week_data = queryset.filter(date__gte=start_date, date__lt=end_date).values(amount_field, 'date')
        logger.info(week_data)
        total = week_data.aggregate(total=Sum(amount_field))['total'] or 0
        return week_data, total

    data = {}
    for week in range(1, 5):
        week_start = first_day + timedelta(days=(week-1)*7)
        week_end = week_start + timedelta(days=7)

        logger.info(week_start)
        logger.info(week_end)

        sales_data, sales_total = get_week_data(sales, week_start, week_end, 'total_amount')
        cogs_data, cogs_total = get_week_data(cogs, week_start, week_end, 'product__cost')
        
        data[f'week {week}'] = {
            'sales': list(sales_data),
            'cogs': list(cogs_data),
            'total_sales': sales_total,
            'total_cogs': cogs_total
        }

    return JsonResponse(data)

@login_required
def income_json(request):
    current_month = get_current_month()
    today = datetime.date.today()
    
    month = request.GET.get('month', current_month)
    day = request.GET.get('day', today.day)

    sales = Sale.objects.filter(transaction__branch=request.user.branch)
    
    if request.GET.get('filter') == 'today':
        sales_total = sales.filter(date=today).aggregate(Sum('total_amount'))
    else:
        sales_total = sales.filter(date__month=month).aggregate(Sum('total_amount'))

    return JsonResponse({'sales_total': sales_total['total_amount__sum'] or 0})


@login_required
def expense_json(request):
    current_month = get_current_month()
    today = datetime.date.today()
    
    month = request.GET.get('month', current_month)
    day = request.GET.get('day', today.day)

    expenses = Expense.objects.filter(branch=request.user.branch)
    
    if request.GET.get('filter') == 'today':
        expense_total = expenses.filter(issue_date=today, status=False).aggregate(Sum('amount'))
    else:
        expense_total = expenses.filter(issue_date__month=month, status=False).aggregate(Sum('amount'))
    
    return JsonResponse({'expense_total': expense_total['amount__sum'] or 0})


@login_required
def pl_overview(request):
    filter_option = request.GET.get('filter')
    today = datetime.date.today()
    previous_month = get_previous_month()
    current_year = today.year
    current_month = today.month

    sales = Sale.objects.filter(transaction__branch=request.user.branch)
    expenses = Expense.objects.filter(branch=request.user.branch)
    cogs = COGSItems.objects.filter(invoice__branch=request.user.branch)

    if filter_option == 'today':
        date_filter = today
    elif filter_option == 'last_week':
        last_week_start = today - datetime.timedelta(days=today.weekday() + 7)
        last_week_end = last_week_start + datetime.timedelta(days=6)
        date_filter = (last_week_start, last_week_end)
    elif filter_option == 'this_month':
        date_filter = (datetime.date(current_year, current_month, 1), today)
    elif filter_option == 'year':
        year = int(request.GET.get('year', current_year))
        date_filter = (datetime.date(year, 1, 1), datetime.date(year, 12, 31))
    else:
        date_filter = (datetime.date(current_year, current_month, 1), today)

    if filter_option == 'today':
        current_month_sales = sales.filter(date=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(issue_date=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.objects.filter(date=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    elif filter_option == 'last_week':
        current_month_sales = sales.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(issue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    else:
        current_month_sales = sales.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(dissue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0

    previous_month_sales = sales.filter(date__year=current_year, date__month=previous_month).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
    previous_month_expenses = expenses.filter(issue_date__year=current_year, issue_date__month=previous_month).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
    previous_cogs =  cogs.filter(date__year=current_year, date__month=previous_month).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    
    current_net_income = current_month_sales
    previous_net_income = previous_month_sales 
    current_expenses = current_month_expenses 
    
    current_gross_profit = current_month_sales - cogs_total
    previous_gross_profit = previous_month_sales - previous_cogs
    
    current_net_profit = current_gross_profit - current_month_expenses
    previous_net_profit = previous_gross_profit - previous_month_expenses

    current_gross_profit_margin = (current_gross_profit / current_month_sales * 100) if current_month_sales != 0 else 0
    previous_gross_profit_margin = (previous_gross_profit / previous_month_sales * 100) if previous_month_sales != 0 else 0
    
    # net_income_change = calculate_percentage_change(current_net_income, previous_net_income)
    # gross_profit_change = calculate_percentage_change(current_gross_profit, previous_gross_profit)
    # gross_profit_margin_change = calculate_percentage_change(current_gross_profit_margin, previous_gross_profit_margin)


    data = {
        'net_profit':current_net_profit,
        'cogs_total':cogs_total,
        'current_expenses':current_expenses,
        'current_net_profit': current_net_profit,
        'previous_net_profit':previous_net_profit,
        'current_net_income': current_net_income,
        'previous_net_income': previous_net_income,
        'current_gross_profit': current_gross_profit,
        'previous_gross_profit': previous_gross_profit,
        'current_gross_profit_margin': f'{current_gross_profit_margin:.2f}',
        'previous_gross_profit_margin': previous_gross_profit_margin,
    }
    
    return JsonResponse(data)

@login_required
def cash_deposit(request):
    if request.method == 'GET':
        deposits = CashDeposit.objects.all()
        return render(request, 'cash_deposit.html', 
            {
                'form':cashDepositForm(),
                'deposits':deposits
            }
        )
    
@login_required
def vat(request):
    if request.method == 'GET':
        
        filter_option = request.GET.get('filter', 'today')
        download = request.GET.get('download')
        
        now = datetime.datetime.now()
        end_date = now
        
        if filter_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_week':
            start_date = now - timedelta(days=now.weekday())
        elif filter_option == 'yesterday':
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_month':
            start_date = now.replace(day=1)
        elif filter_option == 'last_month':
            start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        elif filter_option == 'this_year':
            start_date = now.replace(month=1, day=1)
        elif filter_option == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now
            
        vat_transactions = VATTransaction.objects.filter(date__gte=start_date, date__lte=end_date).order_by('-date')
        
        if download:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="vat_report_{filter_option}.csv"'

            writer = csv.writer(response)
            writer.writerow(['Date', 'Description', 'Status', 'Input', 'Output'])

            balance = 0
            for transaction in vat_transactions:

                if transaction.vat_type == 'Input':
                    balance += transaction.tax_amount
                else:
                    balance -= transaction.tax_amount

                writer.writerow([
                    transaction.date,
                    transaction.invoice.invoice_number if transaction.invoice else transaction.purchase_order.order_number,
                    transaction.tax_amount if transaction.vat_type == 'Input' else  '',
                    transaction.tax_amount if transaction.vat_type == 'Output' else  ''
                ])

            writer.writerow(['Total', '', '', balance])
            
            return response
        return render(request, 'vat.html', 
            {
                'filter_option':filter_option,
                'vat_transactions':vat_transactions
            }
        )
    
    if request.method == 'POST':
        # payload 
        {
            'date_from':'date',
            'date_to':'date'
        }
        try:
            data = json.loads(request.body)
            
            date_to = data.get('date_to')
            date_from = data.get('date_from')

            vat_transactions = VATTransaction.objects.filter(
                date__gte=date_from, 
                date__lte=date_to
            )
            
            vat_transactions.update(paid=True)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status = 400)
        return JsonResponse({'success':False, 'message':'VAT successfully paid'}, status = 200)


@login_required
def cash_flow(request):
    """
    View to display a comprehensive financial overview including sales, income, and expenses.
    """
    today = datetime.datetime.today()
    income_form = IncomeCategoryForm()
    

    filter_type = request.GET.get('filter_type', 'today')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if filter_type == 'today':
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'weekly':
        start_date = (today - datetime.timedelta(days=today.weekday())).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'monthly':
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'yearly':
        start_date = today.replace(month=1, day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'custom':
        if not start_date:
            start_date = today.strftime('%Y-%m-%d')
        if not end_date:
            end_date = today.strftime('%Y-%m-%d')
    
    start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d')
    end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d')

    end_date_query = end_date_obj + datetime.timedelta(days=1)
        
    # Query for invoice items in the date range
    invoice_items = InvoiceItem.objects.filter(
        invoice__issue_date__date__gte=start_date_obj,
        invoice__issue_date__date__lt=end_date_query
    )
    
    # Query for other income and expenses
    income = Income.objects.filter(
        created_at__date__gte=start_date_obj,
        created_at__date__lt=end_date_query
    )
    
    expenses = Expense.objects.filter(
        issue_date__date__gte=start_date_obj,
        issue_date__date__lt=end_date_query
    )
    
    logs = FinanceLog.objects.filter(
        date__gte=start_date_obj.date(),
        date__lt=end_date_query.date()
    )
    
    # Normalize invoice items for timeline
    normalized_sales = invoice_items.annotate(
        type_label=Value('sale', output_field=CharField()),
        category_name=F('item__description'), 
        parent_category=Value('Sales', output_field=CharField()),
        datetime=F('invoice__issue_date'),
        source=Value('Invoice', output_field=CharField()),
        amount=F('total_amount')  
    ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')
    
    # Normalize income entries
    normalized_incomes = income.annotate(
        type_label=Value('income', output_field=CharField()),
        category_name=F('category__name'),
        parent_category=F('category__parent__name'),
        datetime=F('created_at'),
        source=Value('Income', output_field=CharField())
    ).values('datetime',  'sale__invoice_items__item__name', 'amount', 'type_label', 'category_name', 'parent_category', 'source', 'note')
    
    # Normalize expense entries
    normalized_expenses = expenses.annotate(
        type_label=Value('expense', output_field=CharField()),
        category_name=F('category__name'),
        parent_category=F('category__parent__name'),
        datetime=F('issue_date'),
        source=Value('Expense', output_field=CharField())
    ).values('datetime', 'amount', 'description', 'type_label', 'category_name', 'parent_category', 'source')
    
    # Combine and sort by datetime (chronological timeline of all financial activity)
    combined_cashflow = sorted(
        chain(normalized_incomes, normalized_expenses),
        key=lambda x: x['datetime']
    )
    
    product_sales = invoice_items.values(
        'item__id', 
        'item__name',
        'item__description'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('total_amount'),
        average_price=Avg('unit_price'),
        total_vat=Sum('vat_amount')
    ).order_by('-total_revenue')
    
    # Calculate totals
    sales_total = invoice_items.aggregate(total=Sum('total_amount'))['total'] or 0
    income_total = income.aggregate(total=Sum('amount'))['total'] or 0
    expenses_total = expenses.aggregate(total=Sum('amount'))['total'] or 0
    total_income = sales_total 
    balance = total_income - expenses_total
    
    # Group expenses by category for summary
    expenses_by_category = expenses.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Group income by category for summary
    income_by_category = income.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')


     # Add percentage to each expense category
    for category in expenses_by_category:
        if expenses_total > 0:
            category['percentage'] = (category['total'] / expenses_total) * 100
        else:
            category['percentage'] = 0
    
    # Group income by category for summary
    income_by_category = income.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Add percentage to each income category
    for category in income_by_category:
        if income_total > 0:
            category['percentage'] = (category['total'] / income_total) * 100
        else:
            category['percentage'] = 0

    # categories
    expenses_categories = ExpenseCategory.objects.all()
    income_categories = IncomeCategory.objects.all()
    
    cash_ups = CashUp.objects.all().select_related(
        'branch', 'created_by'
    ).prefetch_related(
        'sales', 'expenses'
    ).values(
        'expected_cash',
        'branch__id',
        'branch__name',
        'received_amount',
        'sales',
        'expenses',
        'created_by__username',
        'sales_status',
        'expenses_status',
        'status',
        'date'
    ).order_by('-created_at')
    
    context = {
        # Time filter data
        'start_date': start_date,
        'end_date': end_date,
        'filter_type': filter_type,
        
        # Raw data
        'sales': invoice_items,
        'income': income,
        'expenses': expenses,
        'logs': logs,
        
        # Summaries
        'product_sales': product_sales,
        'expenses_by_category': expenses_by_category,
        'income_by_category': income_by_category,
        'combined_cashflow': combined_cashflow,
        
        # Totals
        'sales_total': sales_total,
        'income_total': income_total,
        'expenses_total': expenses_total,
        'total_income': total_income,
        'balance': balance,

        # categories 
        'expenses_categories':expenses_categories,
        'income_categories':income_categories,
        
        # branches
        'cash_ups':cash_ups,
    }
    
    return render(request, 'cashflow.html', context)
    
# @login_required
# def cash_flow(request):
#     from itertools import chain
#     today = datetime.datetime.today()

#     # Income & Sales
#     sales = InvoiceItems.objects.filter(sale__issue_date__date=today)
#     income = Income.objects.filter(created_at__date=today)
#     logs = FinanceLog.objects.filter(date=today)

#     # Expenses
#     expenses = Expense.objects.filter(issue_date__date=today)

#     # Normalize income entries
#     normalized_incomes = income.annotate(
#         type_label=models.Value('income', output_field=models.CharField()),
#         category_name=models.F('category__name'),
#         parent_category=models.F('category__parent__name'),
#         datetime=models.F('created_at'),
#         source=models.Value('Income', output_field=models.CharField())
#     ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')

#     # Normalize expense entries
#     normalized_expenses = expenses.annotate(
#         type_label=models.Value('expense', output_field=models.CharField()),
#         category_name=models.F('category__name'),
#         parent_category=models.F('category__parent__name'),
#         datetime=models.F('issue_date'),
#         source=models.Value('Expense', output_field=models.CharField())
#     ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')

#     # Combine and sort by datetime
#     combined_cashflow = sorted(
#         chain(normalized_incomes, normalized_expenses),
#         key=lambda x: x['datetime']
#     )

#     # Totals
#     sales_total = sales.aggregate(total=Sum('amount'))['total'] or 0
#     income_total = income.aggregate(total=Sum('amount'))['total'] or 0
#     expenses_total = expenses.aggregate(total=Sum('amount'))['total'] or 0
#     total_income = sales_total + income_total
#     balance = total_income - expenses_total

#     context = {
#         'sales': sales,
#         'sales_total': sales_total,
#         'income': income,
#         'income_total': income_total,
#         'expenses_total': expenses_total,
#         'total_income': total_income,
#         'combined_cashflow': combined_cashflow,  
#         'balance':balance,
#         'logs':logs
#         # 'grouped_expenses': grouped_expenses,
#         # 'grouped_income': grouped_income,
#     }

#     return render(request, 'cashflow.html', context)

@login_required
def get_recorded_cash_ups(request):
    try:
        data = json.loads(request.body)
        branch_id = data.get('branch_id')
        date_range = data.get('date_range')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        queryset = CashUp.objects.filter(status=True)

        logger.info(f'cash ups: {queryset}')
        logger.info(f'date ranger: {date_range}')
        logger.info(f'start date: {start_date}')
        logger.info(f'end date: {end_date}')
        
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)

        now = datetime.datetime.now()
        if date_range == 'today':
            queryset = queryset.filter(created_at__date=now.date())
        elif date_range == 'yesterday':
            yesterday = now - timedelta(days=1)
            queryset = queryset.filter(created_at__date=yesterday.date())
        elif date_range == 'this_week':
            start_of_week = now - timedelta(days=now.weekday())
            queryset = queryset.filter(created_at__date__gte=start_of_week.date())
        elif date_range == 'last_week':
            start_of_week = now - timedelta(days=now.weekday() + 7)
            end_of_week = start_of_week + timedelta(days=6)
            queryset = queryset.filter(created_at__date__gte=start_of_week.date(), created_at__date__lte=end_of_week.date())
        elif date_range == 'this_month':
            start_of_month = now.replace(day=1)
            queryset = queryset.filter(created_at__date__gte=start_of_month.date())
        elif date_range == 'last_month':
            start_of_month = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
            end_of_month = now.replace(day=1) - timedelta(days=1)
            queryset = queryset.filter(created_at__date__gte=start_of_month.date(), created_at__date__lte=end_of_month.date())
        elif date_range == 'custom' and start_date and end_date:
            queryset = queryset.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

        queryset = queryset.order_by('-created_at')

        cash_ups = list(queryset.values(
            'id',
            'branch__name',
            'created_at',
            'expected_cash',
            'received_amount',
            'short_fall',
            'created_by__username'
        ))

        return JsonResponse({
            'success': True,
            'cash_ups': cash_ups
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@login_required
@transaction.atomic
def cashflow_create(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            amount = data.get('amount')
            cash_up_id = data.get('cash_up_id')

            with transaction.atomic():
                cash_up = CashUp.objects.get(id=cash_up_id)
                cash_up.received_amount = Decimal(amount)
                cash_up.status = True
                cash_up.save()

                total_income = sum(sale.unit_price * sale.quantity for sale in cash_up.sales.all())
                total_expenses = sum(expense.amount for expense in cash_up.expenses.all())

                for sale in cash_up.sales.all():
                    logger.info(f'cash sale: {sale.unit_price * sale.quantity}', {sale})

                logger.info(f'total income: {total_income}')

                for expenses in cash_up.expenses.all():
                    logger.info(f'cash expense: {expenses}')
                
                logger.info(f'total expenses: {total_expenses}')

                expense_category, _ = MainExpenseCategory.objects.get_or_create(name='Expense')
                income_category, _ = MainIncomeCategory.objects.get_or_create(name='Income')
                
                # for income
                Cashflow.objects.create(
                    branch=cash_up.branch,
                    total=amount,
                    date=datetime.datetime.now(),
                    status=False,
                    cash_up=cash_up,
                    income = total_income,
                    category=income_category,
                    created_by=request.user
                )

                # for expense
                Cashflow.objects.create(
                    branch=cash_up.branch,
                    total=amount,
                    date=datetime.datetime.now(),
                    status=False,
                    cash_up=cash_up,
                    expense = total_expenses,
                    category = expense_category,
                    created_by=request.user
                )

                # if amount > total_income - total_expenses:
                #     user_account, _ = UserAccount.objects.get_or_create(
                #         user=request.user, 
                #         defaults={
                #             'balance': Decimal('0.00'),
                #             'total_credits': Decimal('0.00'),
                #             'total_debits': Decimal('0.00'),
                #             'last_transaction_date':datetime.datetime.now()
                #         }
                #     )

                #     transaction = UserTransaction.objects.create(
                #         account=user_account,
                #         branch=cash_up.branch,
                #         amount=amount - total_income + total_expenses,
                #         transaction_type=UserTransaction.TransactionType.CASH,
                #         description='Cashup deficit',
                #         created_by=request.user
                #     )

                #     user_account.balance += transaction.amount
                #     user_account.total_credits = 0
                #     user_account.total_debits = transaction.amount


                return JsonResponse({'success':True, 'message':'Cashflow successfully created'}, status=201)

        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        


@login_required
@transaction.atomic
def record_cashflow(request):
    try:
        data = json.loads(request.body)
        logger.info(data)
        type = data.get('type', '')
        id = data.get('id', '')
        branch = int(data.get('branch', ''))
        category = (data.get('category'))
        
        # get branch
        branch = get_object_or_404(Branch, id=branch)

        category = None
        if type == 'sale':
            sale = Invoice.objects.filter(id=id).first()
      
            category = IncomeCategory.objects.filter(name__iexact=category).first()

            if not category:
                new_main_category = IncomeCategory.objects.create(
                    name='sales',
                    parent=None
                )

                new_sub_category = IncomeCategory.objects.create(
                    name="sales",
                    parent=new_main_category
                )
                category = new_sub_category
        else:
            # expense = Expense.objects.filter(branch=branch, id=id)
            pass

        with transaction.atomic():

            Income.objects.create(
                amount = sale.amount_paid,
                currency = sale.currency,
                note = sale.products_purchased,
                branch = request.user.branch,
                status = False,
                sale = sale,
                user=request.user,
                category = category
            )

            sale.cash_up_status = True
            sale.save()
            
            return JsonResponse(
                {
                    'success':True, 
                    'message':'Sale recorded succesfully', 
                    'id':id,
                    'cash_up_status':True
                }, status=200)
            
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
    


@login_required
def check_cashup_status(request, cash_up_id):
    logger.info(cash_up_id)
    cash_up = CashUp.objects.filter(id=cash_up_id).first()
    
    if cash_up:
        if True:
            cash_up.status = True
            cash_up.save()
            return JsonResponse({
                'success':True,
                'status':True
            }, status=200)
        return JsonResponse({
                'success':True,
                'status':False
            }, status=200)
    return JsonResponse({
            'success':False,
            'message':'Error occured'
        }, status=400)
    
@login_required
def get_incomes(request):
    page_number = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('limit', 10))  

    incomes = Income.objects.filter(branch=request.user.branch).order_by('-created_at')
    paginator = Paginator(incomes, per_page)
    page_obj = paginator.get_page(page_number)

    income_data = [
        {
            'id': income.id,
            'created_at': income.created_at.strftime('%Y-%m-%d %H:%M'),
            'amount': str(income.amount),
            'category': str(income.category.name),
            'branch':income.branch.name,
            'note': income.note,
            'user': income.user.get_full_name() or income.user.username,
            'is_recurring': income.is_recurring,
            'recurrence': f"{income.recurrence_value} {income.recurrence_unit}" if income.is_recurring else '',
        }
        for income in page_obj.object_list
    ]

    return JsonResponse({
        'data': income_data,
        'has_next': page_obj.has_next()
    })
    
@transaction.atomic
@login_required
def record_income(request):
    try:
        data = json.loads(request.body)
        logger.info(data)

        name = data.get('name')
        amount = data.get('amount')
        category_name = data.get('category')
        branch_id = data.get('branch')
        r_value = data.get('r_value')
        r_unit = data.get('r_unit')

        if not all([name, amount, category_name, branch_id]):
            return JsonResponse({'success': False, 'message': 'Missing required fields'}, status=400)

        parent_category, _ = IncomeCategory.objects.get_or_create(name="Manual", parent=None)
        category, _ = IncomeCategory.objects.get_or_create(name=category_name, parent=parent_category)

        branch = Branch.objects.get(id=branch_id)
        currency = Currency.objects.filter(name__icontains="usd").first() #to be dynamic
        
        logger.info(currency)

        Income.objects.create(
            amount=amount,
            currency_id=currency.id,
            category=category,
            note=name,
            user=request.user,
            branch=branch,
            is_recurring=bool(r_value),
            recurrence_value=r_value if r_value else None,
            recurrence_unit=r_unit if r_unit else None
        )
        
        Cashbook.objects.create(
            amount=amount,
            description=f"Income: {name} -> {category.name}",
            debit=True,
            credit=False,
            branch=branch,
            created_by=request.user,
            updated_by=request.user,
            issue_date=timezone.now()
        )
        
        logger.info(f'Income recorded successfully: {Income.objects.last()}')

        return JsonResponse({'success': True, 'message': 'Income recorded successfully'}, status=200)
    except Exception as e:
        logger.error("Income record error: %s", e)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@transaction.atomic
def record_cashflow_transaction(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            income = float(data.get('IncomeAmount', 0))
            expense_amount = float(data.get('ExpenseAmount', 0))
            transaction_type = data.get('type', '')
            categories = data.get('categories', {})

            category = categories.get('category', {})
            subcategory = categories.get('subcategory', {})
            name = categories.get('name', {})

            if not all([category, subcategory, name]):
                return JsonResponse({
                    'success': False, 
                    'message': 'Missing required category information.'
                }, status=400)

            if not transaction_type or transaction_type not in ['income', 'expense']:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid transaction type.'
                }, status=400)
            
            with transaction.atomic():
                if transaction_type == 'income':
                    if not income or income <= 0:
                        return JsonResponse({
                            'success': False,
                            'message': 'Invalid amount.'
                        }, status=400)
                    
                    logger.info(f'Creating Income amount: {income}')
                    logger.info(f'Categories data: {categories}')

                    cash_flow_name, _ = CashFlowName.objects.get_or_create(
                        name=name.get('value')
                    )
                    sub_category, _ = IncomeSubCategory.objects.get_or_create(
                        name=subcategory.get('value')
                    )
                    main_category, _ = MainIncomeCategory.objects.get_or_create(
                        name=category.get('value'),
                        defaults={'sub_income_category': sub_category}
                    )

                    main_category.save()

                    logger.info(f'cash_flow name: {cash_flow_name}')
                    logger.info(f'sub category: {sub_category}')
                    logger.info(f'main category: {main_category.id} type: {type(main_category)}, main category sub: {main_category.sub_income_category}')

                    object = Cashflow.objects.create(
                        name=cash_flow_name,
                        branch=request.user.branch,
                        total=income,
                        date=datetime.datetime.now(),
                        status=False,
                        income=income,
                        expense=0,
                        income_category=main_category,
                        created_by=request.user
                    )

                    logger.info(f'Income created: {object}.')

                    return JsonResponse({
                        'success': True,
                        'message': 'Income cashflow successfully created'
                    }, status=201)
                
                else:  # expense
                    if not expense_amount or expense_amount <= 0:
                        return JsonResponse({
                            'success': False,
                            'message': 'Invalid amount.'
                        }, status=400)
                    
                    logger.info(f'Creating Expense amount: {expense_amount}')
                    
                    # Create or get the required objects
                    cash_flow_name, _ = CashFlowName.objects.get_or_create(
                        name=name.get('value')
                    )
                    sub_category, _ = ExpenseSubCategory.objects.get_or_create(
                        name=subcategory.get('value')
                    )
                    main_category, _ = MainExpenseCategory.objects.get_or_create(
                        name=category.get('value'),
                        defaults={'sub_expense': sub_category}
                    )

                    main_category.save()                    

                    logger.info(f'cash_flow name: {cash_flow_name}')
                    logger.info(f'sub category: {sub_category}')
                    logger.info(f'main category: {main_category.id} type: {type(main_category)}, main category sub: {main_category.sub_expense}')

                    object = Cashflow.objects.create(
                        name=cash_flow_name,
                        branch=request.user.branch,
                        total=expense_amount,
                        date=datetime.datetime.now(),
                        status=False,
                        expense=expense_amount,
                        income=0,
                        expense_category=main_category,
                        created_by=request.user
                    )

                    logger.info(f'Expense created: {object}.')

                    return JsonResponse({
                        'success': True,
                        'message': 'Expense cashflow successfully created'
                    }, status=201)

        except Exception as e:
            logger.error(f"Error recording transaction: {e}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
        
@login_required
def get_cashflow_categories(request):
    if request.method == 'GET':
        main_income_category = MainIncomeCategory.objects.all()
        main_expense_category = MainExpenseCategory.objects.all()
        return JsonResponse({'income': main_income_category, "expense":main_expense_category}, status == 200)
    return JsonResponse({'success': True, 'message': 'Invalid request'}, status == 500)

@login_required
def add_income_category(request):
    try:
        data = json.loads(request.body)
        name = data.get("name")
        parent_id = data.get("parent_id")
        new_parent_name = data.get("new_parent_name")

        if not name:
            return JsonResponse({"success": False, "error": "Name is required."})

        if new_parent_name:
            parent = IncomeCategory.objects.create(name=new_parent_name)
        elif parent_id:
            parent = IncomeCategory.objects.get(id=parent_id)
        else:
            parent = None

        category = IncomeCategory.objects.create(name=name, parent=parent)
        return JsonResponse({"success": True, "id": category.id, "name": str(category)})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

@login_required
def get_branch_data(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    
    cashup = CashUp.objects.filter(branch=branch).order_by('-date').first()
    
    if cashup:
        categories = {}
        logger.info(cashup.sales.all())
        for sale in cashup.sales.all():
            cat_name = sale.item.name
            logger.info(sale)
            if sale.invoice.amount_paid > 0:
                if cat_name not in categories:
                    categories[cat_name] = {
                        'product': cat_name,
                        'expense': 0,
                        'income': float(sale.quantity * sale.unit_price),
                        'total': float(sale.quantity * sale.unit_price)  
                    }
                else:
                    categories[cat_name]['income'] += float(sale.quantity * sale.unit_price)
                    categories[cat_name]['total'] += float(sale.quantity * sale.unit_price)
import csv
from .models import *
from decimal import Decimal
from io import BytesIO
from apps.users.models import User
from apps.company.models import Branch
from .consumers import CashTransferConsumer 
from xhtml2pdf import pisa 
from django.views import View
from django.db.models import Q
from twilio.rest import Client
from datetime import timedelta
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.contrib import messages
from utils.utils import generate_pdf
from . utils import update_latest_due
from django.http import JsonResponse
from utils.utils import generate_pdf
from asgiref.sync import async_to_sync, sync_to_async
from apps.inventory.models import Inventory, Accessory
from channels.layers import get_channel_layer
import json, datetime, os, boto3, openpyxl 
from utils.account_name_identifier import account_identifier
from .tasks import (
    send_invoice_email_task, 
    send_account_statement_email, 
    send_quotation_email
)
from pytz import timezone as pytz_timezone 
from openpyxl.styles import Alignment, Font
from . utils import calculate_expenses_totals
from django.utils.dateparse import parse_date
from django.templatetags.static import static
from django.db.models import Sum, DecimalField
from apps.inventory.models import ActivityLog, Product
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.core.mail import send_mail, EmailMessage
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from . forms import (
    ExpenseForm, 
    ExpenseCategoryForm, 
    CurrencyForm, 
    InvoiceForm, 
    CustomerForm, 
    TransferForm, 
    CashWithdrawForm, 
    cashWithdrawExpenseForm,
    customerDepositsForm,
    customerDepositsRefundForm,
    cashDepositForm,
    IncomeCategoryForm
)
from django.contrib.auth import authenticate
from loguru import logger
from .tasks import send_expense_creation_notification
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from django.http import FileResponse
import io
from collections import defaultdict
from apps.pos.utils.receipt_signature import generate_receipt_data
from apps.pos.utils.submit_receipt_data import submit_receipt_data
from django.db.models.functions import Coalesce
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Max
from django.utils.dateparse import parse_date
from dotenv import load_dotenv
from apps.settings.models import OfflineReceipt, FiscalDay, FiscalCounter
from utils.zimra import ZIMRA
from utils.zimra_sig_hash import run
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Avg, F, Value, CharField, ExpressionWrapper
import datetime
from itertools import chain
from django.core.paginator import Paginator, EmptyPage
import imghdr, base64
from django.core.files.base import ContentFile
 
# load global zimra instance
zimra = ZIMRA()

load_dotenv()

from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from .models import CashUp, Invoice, Expense
from django.utils import timezone

def get_previous_month():
    first_day_of_current_month = datetime.datetime.now().replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    return last_day_of_previous_month.month

def get_current_month():
    return datetime.datetime.now().month

#image decoding method
def decode_base64_file(data):
    """
    Decodes a base64 file and returns a ContentFile.
    Assumes data is in the format: data:<mime>;base64,<data>
    """
    if not data:
        return None

    try:
        format, imgstr = data.split(';base64,')
        ext = format.split('/')[-1]
        if ext == 'jpeg':
            ext = 'jpg'

        file_name = f"{uuid.uuid4()}.{ext}"
        return ContentFile(base64.b64decode(imgstr), name=file_name)
    except Exception as e:
        logger.error("Failed to decode base64 image:")
        return None

def cashflow_list(request): # to be organised
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        branch = request.GET.get('branch')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        status = request.GET.get('status')

        cashups = CashUp.objects.all()
        
        if branch:
            cashups = cashups.filter(branch__id=int(branch))
        if start_date:
            cashups = cashups.filter(date__gte=start_date)
        if end_date:
            cashups = cashups.filter(date__lte=end_date)
        if status in ['true', 'false']:
            cashups = cashups.filter(status=(status == 'true'))

        html = render_to_string('cashflows/partials/cashup_cards.html', {'cash_ups': cashups})
        
        return JsonResponse({'html': html})

class Finance(View):
    # authentication loginmixin
    template_name = 'finance.html'

    def get(self, request, *args, **kwargs):

        if request.user.role == 'sales':
            return redirect('finance:expenses')
        
        balances = AccountBalance.objects.filter(branch=request.user.branch)
    
        recent_sales = Sale.objects.filter(transaction__branch=request.user.branch).order_by('-date')[:5]

        expenses_by_category = Expense.objects.values('category__name').annotate(
            total_amount=Sum('amount', output_field=DecimalField())
        )
        
        context = {
            'balances': balances,
            'recent_transactions': recent_sales,
            'expenses_by_category': expenses_by_category,
        }
        
        return render(request, self.template_name, context)
    
@login_required
def monthly_installments(request):
    installments = MonthlyInstallment.objects.all().select_related('invoice').order_by('-date')
    return JsonResponse({'installments': list(installments)})

@login_required
def laybys(request):
    laybys = layby.objects.all().select_related('invoice').order_by('-date')
    return JsonResponse({'laybys': list(laybys)})

@login_required
def expenses(request):
    form = ExpenseForm()
    cat_form = ExpenseCategoryForm()

    if request.method == 'GET':
        filter_button = request.GET.get('filter_button')
        filter_option = request.GET.get('filter', 'today')  
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        expenses = Expense.objects.filter(user=request.user).order_by('-issue_date')
        
        # filtered_expenses = filter_expenses(expenses, filter_option, start_date, end_date)

        if request.user.role == 'sale':
            expenses = expenses.filter(user=request.user)
        
        return render(request, 'expenses.html', 
            {
                'form':form,
                'cat_form':cat_form,
                'expenses':filter_expenses,
                'filter_option': expenses,
            }
        )
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
      
            name = data.get('name') 
            amount = data.get('amount')
            category = data.get('category')  
            payment_method = data.get('payment_method', 'cash')
            currency_id = data.get('currency', 'USD')
            branch = request.user.branch.id
            base64_image = data.get('receipt')
            image = decode_base64_file(base64_image)

            is_recurring = data.get('is_recurring') == 'true'
            recurrence_value = data.get('recurrence_value')
            recurrence_unit = data.get('recurrence_unit')
            
            logger.info(branch)

            # Validation
            if not all([name, amount, category, payment_method, currency_id, branch]):
                return JsonResponse({'success': False, 'message': 'Missing required fields.'})

            # Fetch related objects
            try:
                category = ExpenseCategory.objects.get(id=category)
            except ExpenseCategory.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'Category with ID {category} does not exist.'})

            currency = get_object_or_404(Currency, name__icontains='usd')
            branch = get_object_or_404(Branch, id=branch)

            # Get or create account and balance
            account_details = account_identifier(request, currency, payment_method)
            account_name = account_details['account_name']
            account_type = account_details['account_type']

            account, _ = Account.objects.get_or_create(
                name=account_name,
                type=account_type
            )

            account_balance, _ = AccountBalance.objects.get_or_create(
                account=account,
                currency=currency,
                defaults={
                    'currency': currency,
                    'branch': branch,
                    'balance': 0
                }
            )
            
            logger.info(account_balance.balance)

            if account_balance.balance < Decimal(amount):
                return JsonResponse({'success': False, 'message': f'{account_name} has insufficient balance.'})

            # Deduct from balance
            account_balance.balance -= Decimal(amount)
            account_balance.save()

            # Create Expense
            expense = Expense.objects.create(
                description=name,
                amount=amount,
                category=category,
                user=request.user,
                currency=currency,
                payment_method=payment_method ,
                branch=branch,
                is_recurring=is_recurring,
                recurrence_value=int(recurrence_value) if is_recurring and recurrence_value else None,
                recurrence_unit=recurrence_unit if is_recurring else None,
                receipt=image,
            )

            # Create Cashbook entry
            Cashbook.objects.create(
                amount=amount,
                expense=expense,
                currency=currency,
                credit=True,
                description=f'Expense ({expense.description[:20]})',
                branch=branch
            )

            # Send notification (to turn on)
            # send_expense_creation_notification.delay(expense.id)

            return JsonResponse({'success': True, 'message': 'Expense recorded successfully.'})

        except Exception as e:
            logger.exception("Error while recording expense:")
            return JsonResponse({'success': False, 'message': str(e)})
        
@login_required
def get_expenses(request):
    try:
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 10))

        expenses = Expense.objects.select_related('category', 'branch', 'currency') \
                                  .order_by('-issue_date') 
        paginator = Paginator(expenses, limit)

        try:
            paginated_expenses = paginator.page(page)
        except EmptyPage:
            return JsonResponse({
                'data': [],
                'has_next': False
            })

        results = []
        for expense in paginated_expenses:
            results.append({
                'id': expense.id,
                'created_at': expense.issue_date.isoformat(),
                'note': expense.description,
                'amount': float(expense.amount),
                'category': str(expense.category),
                'branch': expense.branch.name,
                'has_receipt': bool(expense.receipt),
                'receipt_url': expense.receipt.url if expense.receipt else None
            })

        return JsonResponse({
            'data': results,
            'has_next': paginated_expenses.has_next()
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def save_expense_split(request):
    try:
        data = json.loads(request.body)
        splits = data.get('splits')
        branch_id = data.get('branch_id')
        expense_id = data.get('expense_id', '')
        
        cash_up = CashUp.objects.filter(date=today, branch__id=int(branch_id)).first() # to put order by
        logger.info(f'Cash up: {cash_up}')
        
        cash_up_expenses = cash_up.expenses.all()
        
        if expense_id:
            record_expense(expense_id,cash_up_expenses, request)
            
        else:
            logger.info(f'Split expenses: {splits}, branch_id = {branch_id}')
            expenses = []
            if not cash_up:
                return JsonResponse({'success': False, 'message': 'No cash up record found for today'}, status=404)

            categories = ExpenseCategory.objects.all()
            
            logger.info(cash_up.expenses)

            for split in splits:
                logger.info(split)
                logger.info(split['amount'])
                exp_obj = cash_up_expenses.get(id=int(split['expense_id']))  # existing expense object
                
                new_expense = Expense(
                    amount=split['amount'],
                    payment_method=exp_obj.payment_method,
                    currency=exp_obj.currency,
                    category=categories.filter(id=split['category_id']).first(),
                    description=exp_obj.description,
                    user=request.user,  
                    branch_id=request.user.branch.id,
                    status=False,
                    # purchase_order=exp_obj.purchase_order,
                    receipt=exp_obj.receipt,
                    is_recurring=exp_obj.is_recurring,
                    recurrence_value= exp_obj.recurrence_value if exp_obj.recurrence_value else None,
                    recurrence_unit= exp_obj.recurrence if exp_obj.recurrence_unit else None
                )
                
                exp_obj.cash_up_status = True
                
                exp_obj.save()
                new_expense.save()
                
                expenses.append(new_expense.id)

        return JsonResponse({
            'success': True,
            'message': 'Expenses split and saved successfully',
            # 'expenses': expenses
        })

    except Exception as e:
        logger.exception("Error in save_expense_split")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
        
def record_expense(expense_id, cash_up_expenses, request):
    logger.info(f'Straight recording of the expense: {expense_id}')
    exp_obj = cash_up_expenses.get(id=int(expense_id))  # existing expense object
            
    new_expense = Expense(
        amount=exp_obj.amount,
        payment_method=exp_obj.payment_method,
        currency=exp_obj.currency,
        category=exp_obj.category,
        description=exp_obj.description,
        user=request.user,  
        branch_id=request.user.branch.id,
        status=False,
        purchase_order=exp_obj.purchase_order,
        receipt=exp_obj.receipt,
        is_recurring=exp_obj.is_recurring,
        recurrence_value= exp_obj.recurrence_value if exp_obj.recurrence_value else None,
        recurrence_unit= exp_obj.recurrence if exp_obj.recurrence_unit else None
    )
    
    exp_obj.cash_up_status = True
    
    exp_obj.save()
    new_expense.save()

@login_required  
def get_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    data = {
        'id': expense.id,
        'amount': expense.amount,
        'description': expense.description,
        'category': expense.category.id
    }
    return JsonResponse({'success': True, 'data': data})

@login_required      
def add_or_edit_expense(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            amount = data.get('amount')
            description = data.get('description')
            category_id = data.get('category')
            expense_id = data.get('id')

            if not amount or not description or not category_id:
                return JsonResponse({'success': False, 'message': 'Missing fields: amount, description, category.'})
            
            category = get_object_or_404(ExpenseCategory, id=category_id)

            if expense_id:  
                expense = get_object_or_404(Expense, id=expense_id)
                before_amount = expense.amount
                
                expense.amount = amount
                expense.description = description
                expense.category = category
                expense.save()
                message = 'Expense successfully updated'
            
                try:
                    cashbook_expense = Cashbook.objects.get(expense=expense)
                    expense_amount = Decimal(expense.amount)
                    if cashbook_expense.amount < expense_amount:
                        cashbook_expense.amount = expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'Expense (update from {before_amount} to {cashbook_expense.amount})'
                    else:
                        cashbook_expense.amount -= cashbook_expense.amount - expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'(update from {before_amount} to {cashbook_expense.amount})'
                    cashbook_expense.save()
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)}, status=400)
            return JsonResponse({'success': True, 'message': message}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
@transaction.atomic()
def add_expense_category(request):
    subcategories = ExpenseCategory.objects.filter(parent__isnull=False).values(
        'id',
        'name'
    )

    logger.info(subcategories)
    

@login_required
def add_expense_category(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

        category_name = data.get('name', '')
        parent_name = data.get('parent', '')
        new_parent_name = data.get('new_parent', '')

        if not category_name:
            return JsonResponse({'success': False, 'message': 'Category name is required.'}, status=400)

        parent_obj = None
        if new_parent_name:
            parent_obj, _ = ExpenseCategory.objects.get_or_create(name=new_parent_name, parent=None)
            logger.info(f'New parent created or found: {parent_obj}')
        elif parent_name:
            parent_obj = ExpenseCategory.objects.filter(name=parent_name, parent=None).first()
            if not parent_obj:
                return JsonResponse({'success': False, 'message': f'Parent category "{parent_name}" not found.'}, status=404)
            logger.info(f'Existing parent found: {parent_obj}')

        if ExpenseCategory.objects.filter(name=category_name, parent=parent_obj).exists():
            return JsonResponse({
                'success': False,
                'message': f'Category "{category_name}" already exists under this parent.'
            }, status=400)

        # Create new child category
        new_category = ExpenseCategory.objects.create(name=category_name, parent=parent_obj)
        logger.info(f'New category created: {new_category} under parent: {parent_obj}')

        return JsonResponse({
            'success': True,
            'id': new_category.id,
            'name': new_category.name
        }, status=201)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

@login_required
def income(request):
    """returns income both from the cash sales and income deposited"""
    pass

def filter_expenses(queryset, filter_option, start_date=None, end_date=None):
    """
    Filter expense queryset based on specified time period
    """
    today = timezone.localtime().date()
    
    if filter_option == 'today':
        return queryset.filter(date=today)
    
    elif filter_option == 'yesterday':
        yesterday = today - timedelta(days=1)
        return queryset.filter(date=yesterday)
    
    elif filter_option == 'this_week':
        start_of_week = today - timedelta(days=today.weekday())  
        end_of_week = start_of_week + timedelta(days=6) 
        return queryset.filter(date__gte=start_of_week, date__lte=end_of_week)
    
    elif filter_option == 'last_week':
        end_of_last_week = today - timedelta(days=today.weekday() + 1)  
        start_of_last_week = end_of_last_week - timedelta(days=6)  
        return queryset.filter(date__gte=start_of_last_week, date__lte=end_of_last_week)
    
    elif filter_option == 'this_month':
        return queryset.filter(date__year=today.year, date__month=today.month)
    
    elif filter_option == 'last_month':
        last_month = today.replace(day=1) - timedelta(days=1)
        return queryset.filter(date__year=last_month.year, date__month=last_month.month)
    
    elif filter_option == 'this_year':
        return queryset.filter(date__year=today.year)
    
    elif filter_option == 'last_year':
        return queryset.filter(date__year=today.year - 1)
    
    elif filter_option == 'custom' and start_date and end_date:
        try:
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                
            return queryset.filter(date__gte=start_date, date__lte=end_date)
        except (ValueError, TypeError):
            return queryset

    return queryset

@login_required
@transaction.atomic
def delete_expense(request, expense_id):
    if request.method == 'DELETE':
        try:
            expense = get_object_or_404(Expense, id=expense_id)
            expense.cancel = True
            expense.save()
            
            Cashbook.objects.create(
                amount=expense.amount,
                debit=True,
                credit=False,
                description=f'Expense ({expense.description}): cancelled'
            )
            return JsonResponse({'success': True, 'message': 'Expense successfully deleted'})
        except Exception as e:
             return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
def update_expense_status(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            expense_id = data.get('id')
            status = data.get('status')

            expense = Expense.objects.get(id=expense_id)
            expense.status = status
            expense.save()

            return JsonResponse({'success': True, 'message': 'Status updated successfully.'})
        except Expense.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Expense not found.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

@login_required
def invoice(request):
    form = InvoiceForm()
    invoices = Invoice.objects.filter(branch=request.user.branch, status=True, cancelled=False).select_related(
        'branch',
        'currency',
        'user'
    ).order_by('-invoice_number')

    invoice_items = InvoiceItem.objects.filter(invoice__branch=request.user.branch).select_related(
        'invoice',
    )

    query_params = request.GET
    if query_params.get('q'):
        search_query = query_params['q']
        invoices = invoices.filter(
            Q(customer__name__icontains=search_query) |
            Q(invoice_number__icontains=search_query) |
            Q(issue_date__icontains=search_query)
        )

    user_timezone_str = request.user.timezone if hasattr(request.user, 'timezone') else 'UTC'
    user_timezone = pytz_timezone(user_timezone_str)  

    def filter_by_date_range(start_date, end_date):
        start_datetime = user_timezone.localize(
            timezone.datetime.combine(start_date, timezone.datetime.min.time())
        )
        end_datetime = user_timezone.localize(
            timezone.datetime.combine(end_date, timezone.datetime.max.time())
        )
        return invoices.filter(issue_date__range=[start_datetime, end_datetime])

    now = timezone.now().astimezone(user_timezone)
    today = now.date()

    now = timezone.now() 
    today = now.date()  
    
    date_filters = {
        'today': lambda: filter_by_date_range(today, today),
        'yesterday': lambda: filter_by_date_range(today - timedelta(days=1), today - timedelta(days=1)),
        't_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday()), today),
        'l_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday() + 7), today - timedelta(days=today.weekday() + 1)),
        't_month': lambda: invoices.filter(issue_date__month=today.month, issue_date__year=today.year),
        'l_month': lambda: invoices.filter(issue_date__month=today.month - 1 if today.month > 1 else 12, issue_date__year=today.year if today.month > 1 else today.year - 1),
        't_year': lambda: invoices.filter(issue_date__year=today.year),
    }

    if query_params.get('day') in date_filters:
        invoices = date_filters[query_params['day']]()

    total_partial = invoices.filter(payment_status='Partial').aggregate(Sum('amount'))['amount__sum'] or 0
    total_paid = invoices.filter(payment_status='Paid').aggregate(Sum('amount'))['amount__sum'] or 0
    total_amount = invoices.aggregate(Sum('amount'))['amount__sum'] or 0

    grouped_invoices = defaultdict(lambda: {'invoices': [], 'total_amount': 0, 'amount_due': 0})

    for invoice in invoices:
        issue_date = invoice.issue_date.date() 

        if issue_date == today:
            date_key = 'Today'
        elif issue_date == today - timedelta(days=1):
            date_key = 'Yesterday'
        else:
            date_key = issue_date.strftime('%A, %d %B %Y')
        
        grouped_invoices[date_key]['invoices'].append(invoice)

        grouped_invoices[date_key]['total_amount'] += invoice.amount_paid
        
        if invoice.payment_status == 'Paid':
            amount_due = 0
        elif invoice.payment_status == 'Partial':
            paid_amount = getattr(invoice, 'amount_paid', 0)
            amount_due = invoice.amount - paid_amount
        else:  
            amount_due = invoice.amount
            
        grouped_invoices[date_key]['amount_due'] += amount_due

    ordered_grouped_invoices = {}
    
    if 'Today' in grouped_invoices:
        ordered_grouped_invoices['Today'] = grouped_invoices['Today']

    if 'Yesterday' in grouped_invoices:
        ordered_grouped_invoices['Yesterday'] = grouped_invoices['Yesterday']
    
    remaining_dates = [(k, v) for k, v in grouped_invoices.items() if k not in ['Today', 'Yesterday']]
   
    for date_key, data in remaining_dates:
        ordered_grouped_invoices[date_key] = data

    return render(request, 'invoices/invoice.html', {
        'form': form,
        'grouped_invoices': ordered_grouped_invoices,
        'total_paid': total_paid,
        'total_due': total_partial,
        'total_amount': total_amount,
        'invoice_items':invoice_items
    })
    
@login_required
@transaction.atomic 
def update_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    customer_account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(
        CustomerAccountBalances, account=customer_account, currency=invoice.currency
    )

    if request.method == 'POST':
        data = json.loads(request.body)
        amount_paid = Decimal(data['amount_paid'])

        invoice = Invoice.objects.select_for_update().get(pk=invoice.pk)
        customer_account_balance = CustomerAccountBalances.objects.select_for_update().get(pk=customer_account_balance.pk)

        if amount_paid <= 0:
            return JsonResponse({'success': False, 'message': 'Invalid amount paid.'}, status=400)

        if amount_paid >= invoice.amount_due:
            invoice.payment_status = Invoice.PaymentStatus.PAID
            invoice.amount_due = 0
        else:
            invoice.amount_due -= amount_paid

        invoice.amount_paid += amount_paid
        
        # get the latest payment for the invoice
        latest_payment = Payment.objects.filter(invoice=invoice).order_by('-payment_date').first()
        if latest_payment:
            amount_due = latest_payment.amount_due - amount_paid 
        else:
            amount_due = invoice.amount - invoice.amount_paid 

        payment = Payment.objects.create(
            invoice=invoice,
            amount_paid=amount_paid,
            amount_due=amount_due, 
            payment_method=data['payment_method'],
            user=request.user
        )

        account, _ = Account.objects.get_or_create(
            name=f"{request.user.branch} {invoice.currency.name} {payment.payment_method.capitalize()} Account",
            type=Account.AccountType[payment.payment_method.upper()] 
        )
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=invoice.currency,
            branch=request.user.branch,
            defaults={'balance': 0}
        )

        account_balance.balance += amount_paid
        if customer_account_balance.balance < 0:
            customer_account_balance.balance += amount_paid
        else:
            customer_account_balance.balance -= amount_paid

        description = ''
        if invoice.hold_status:
            description = 'Held invoice payment'
            sale = Sale.objects.create(
                date=timezone.now(),
                transaction=invoice,
                total_amount=invoice.amount # invoice delivery amount
            )
            
            VATTransaction.objects.create(
                invoice=invoice,
                vat_type=VATTransaction.VATType.OUTPUT,
                vat_rate=VATRate.objects.get(status=True).rate,
                tax_amount=invoice.vat
            ) 

        else:
            description = 'Invoice payment update'
        
        Cashbook.objects.create(
            issue_date=invoice.issue_date,
            description=f'({description} {invoice.invoice_number})',
            debit=True,
            credit=False,
            amount=invoice.amount_paid,
            currency=invoice.currency,
            branch=invoice.branch
        )

        invoice.hold_status = False
        account_balance.save()
        customer_account_balance.save()
        invoice.save()
        payment.save()
        
        return JsonResponse({'success': True, 'message': 'Invoice successfully updated'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}) 

def update_invoice_amounts(invoice, amount_paid):
    invoice_payments = Payment.objects.filter(invoice=invoice)

    if amount_paid > 0:
        for payment in invoice_payments:
            amount_paid -= payment.amount_due
            payment.save()


# @login_required
# @transaction.atomic 
# def create_invoice(request):
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             invoice_data = data['data'][0]  
#             items_data = data['items']
#             layby_dates = data.get('layby_dates')

#             # logger.info(f'Invoice data: {data}')
           
#             # get currency
#             currency = Currency.objects.get(id=invoice_data['currency'])
            
#             # create or get accounts
#             account_types = {
#                 'cash': Account.AccountType.CASH,
#                 'bank': Account.AccountType.BANK,
#                 'ecocash': Account.AccountType.ECOCASH,
#             }

#             account_name = f"{request.user.branch} {currency.name} {invoice_data['payment_method'].capitalize()} Account"
            
#             account, _ = Account.objects.get_or_create(name=account_name, type=account_types[invoice_data['payment_method']])
            
#             account_balance, _ = AccountBalance.objects.get_or_create(
#                 account=account,
#                 currency=currency,
#                 branch=request.user.branch,
#                 defaults={'balance': 0}  
#             )

#             logger.info(f"Account Balance: {account_balance}")

#             # accounts_receivable
#             accounts_receivable, _ = ChartOfAccounts.objects.get_or_create(name="Accounts Receivable")
            
#             # VAT rate
#             vat_rate = VATRate.objects.get(status=True)

#             # customer
#             customer = Customer.objects.get(id=int(invoice_data['client_id'])) 
#             logger.info(customer)
            
#             # customer account
#             customer_account = CustomerAccount.objects.get(customer=customer)

#             # customer Account + Balances
#             customer_account_balance, _ = CustomerAccountBalances.objects.get_or_create(
#                 account=customer_account,
#                 currency=currency, 
#                 defaults={'balance': 0}
#             )
            
#             amount_paid = invoice_data['amount_paid']
            
#             logger.info(f'Amount paid: {amount_paid}')
            
#             # amount_paid = update_latest_due(customer, Decimal(invoice_data['amount_paid']), request, invoice_data['paymentTerms'], customer_account_balance)
#             # to be revised a lot
#             invoice_total_amount = Decimal(invoice_data['payable'])

#             # prevent to record greater amount paid than the invoice amount 
#             if amount_paid > invoice_total_amount:
#                 amount_paid = invoice_total_amount
#                 amount_due = 0
#             else:
#                 amount_paid = amount_paid
#                 amount_due = invoice_total_amount - amount_paid  
                
#             logger.info(f'amount due: {amount_due}')
        

#             cogs = COGS.objects.create(amount=Decimal(0))
            
#             products_purchased = f"""{', '.join([f'{item['product_name']} x {item['quantity']} ' for item in items_data])}"""
            
#             logger.info(products_purchased)
            
#             with transaction.atomic():
#                 invoice = Invoice.objects.create(
#                     invoice_number=Invoice.generate_invoice_number(request.user.branch.name),
#                     customer=customer,
#                     issue_date=timezone.now(),
#                     amount=invoice_total_amount,
#                     amount_paid=amount_paid,
#                     amount_due=amount_due,
#                     vat=Decimal(invoice_data['vat_amount']),
#                     payment_status = Invoice.PaymentStatus.PARTIAL if amount_due > 0 else Invoice.PaymentStatus.PAID,
#                     branch = request.user.branch,
#                     user=request.user,
#                     currency=currency,
#                     subtotal=invoice_data['subtotal'],
#                     reocurring = invoice_data['recourring'],
#                     products_purchased = products_purchased,
#                     payment_terms = invoice_data['paymentTerms'],
#                     hold_status = invoice_data['hold_status'],
#                     amount_received = amount_paid
#                 )

#                 logger.info(f'Invoice created for customer: {invoice}')

#                 # check if invoice status is hold
#                 if invoice.hold_status == True:

#                     logger.info(f'Processing held invoice: {invoice}')

#                     held_invoice(items_data, invoice, request, vat_rate)

#                     return JsonResponse({'hold':True, 'message':'Invoice succesfully on hold'})

#                 # create layby object
#                 if invoice.payment_terms == 'layby':

#                     if amount_due > 0:

#                         logger.info(f'Creating layby object for invoice: {invoice}')
                        
#                         layby_obj = layby.objects.create(
#                             invoice=invoice, 
#                             branch=request.user.branch
#                         )

#                         layby_dates_list = []
#                         number_of_dates = len(layby_dates)
                        
#                         # calculate amount to be paid for each month
#                         amount_per_due_date = (amount_due / number_of_dates) if number_of_dates > 0 else 0

#                         logger.info(f'Amount per due date: {amount_per_due_date} : {number_of_dates} : {layby_dates}')

#                         for date in layby_dates:

#                             obj = laybyDates(
#                                 layby=layby_obj,
#                                 due_date=date,
#                                 amount_due=round(amount_per_due_date, 2),
#                             )

#                             layby_dates_list.append(obj)
                        
#                         laybyDates.objects.bulk_create(layby_dates_list)

#                         logger.info(f'Layby object created for invoice: {invoice}')
                
#                 # create monthly installment object
#                 if invoice.payment_terms == 'installment':

#                     if invoice.reocurring:
#                         MonthlyInstallment.objects.create(
#                             invoice = invoice,
#                             status = False
#                         )
                    
#                 #create a paylater
#                 if invoice.payment_terms == 'pay later':
#                     if amount_due > 0:
#                         paylater_obj = Paylater.objects.create(
#                             invoice=invoice,
#                             amount_due=amount_due,
#                             due_date=invoice_data['pay_later_dates'][0] if invoice_data['pay_later_dates'] else timezone.now().date(),
#                             payment_method=invoice_data['payment_method']
#                         )
                        
#                         # Create paylater dates for each interval
#                         if invoice_data['pay_later_dates']:
#                             amount_per_interval = amount_due / len(invoice_data['pay_later_dates'])
#                             for date in invoice_data['pay_later_dates']:
#                                 logger.info(date)
#                                 paylaterDates.objects.create(
#                                     paylater=paylater_obj,
#                                     due_date=date,
#                                     amount_due=amount_per_interval,
#                                     payment_method=invoice_data['payment_method']
#                                 )

#                 # #create transaction
#                 Transaction.objects.create(
#                     date=timezone.now(),
#                     description=invoice.products_purchased,
#                     account=accounts_receivable,
#                     debit=Decimal(invoice_data['payable']),
#                     credit=Decimal('0.00'),
#                     customer=customer
#                 )

#                 logger.info(f'Creating transaction obj for invoice: {invoice}')
            
#                 # Create InvoiceItem objects
#                 invoice_items = []
#                 for item_data in items_data:
#                     item = Inventory.objects.get(pk=item_data['inventory_id'])
                    
#                     item.quantity -= item_data['quantity']
#                     item.save()

#                     invoice_items.append(
#                         InvoiceItem.objects.create(
#                             invoice=invoice,
#                             item=item,
#                             quantity=item_data['quantity'],
#                             unit_price=item_data['price'],
#                             vat_rate = vat_rate,
#                             total_amount = int(item_data['quantity']) * float(item_data['price']),
#                             cash_up_status = False
#                         )
#                     )
                    
#                     # cost of sales item
#                     COGSItems.objects.get_or_create(
#                         invoice=invoice,
#                         defaults={'cogs': cogs, 'product': Inventory.objects.get(id=item.id, branch=request.user.branch)}
#                     )
                
#                     # stock log  
#                     ActivityLog.objects.create(
#                         branch=request.user.branch,
#                         inventory=item,
#                         user=request.user,
#                         quantity = -item_data['quantity'],
#                         total_quantity = item.quantity,
#                         action='Sale',
#                         invoice=invoice
#                     )

#                     accessories = Accessory.objects.filter(main_product=item).values('accessory_product', 'accessory_product__quantity')

#                     # for acc in accessories:
#                     #     COGSItems.objects.get_or_create(
#                     #         invoice=invoice,
#                     #         defaults={'cogs': cogs, 'product': Inventory.objects.get(id=acc['accessory_product'], branch=request.user.branch)}
#                     #     )
#                     #     prod_acc = Inventory.objects.get(id = acc['accessory_product'] )
#                     #     prod_acc.quantity -= acc.quantity

#                     #     logger.info(f'accessory quantity: {acc['accessory_product__quantity']}')

#                     #     ActivityLog.objects.create(
#                     #         branch=request.user.branch,
#                     #         inventory=prod_acc,
#                     #         user=request.user,
#                     #         quantity=1,
#                     #         total_quantity = acc['accessory_product__quantity'],
#                     #         action='Sale',
#                     #         invoice=invoice
#                     #     )
#                     #     prod_acc.save()
                        
#                 # # Create VATTransaction
#                 VATTransaction.objects.create(
#                     invoice=invoice,
#                     vat_type=VATTransaction.VATType.OUTPUT,
#                     vat_rate=VATRate.objects.get(status=True).rate,
#                     tax_amount=invoice_data['vat_amount']
#                 )                                                          
#                 # Create Sale object
#                 sale = Sale.objects.create(
#                     date=timezone.now(),
#                     transaction=invoice,
#                     total_amount=invoice_total_amount
#                 )
#                 sale.save()
                
#                 #payment
#                 Payment.objects.create(
#                     invoice=invoice,
#                     amount_paid=amount_paid,
#                     payment_method=invoice_data['payment_method'],
#                     amount_due=invoice_total_amount - amount_paid,
#                     user=request.user
#                 )

#                 # calculate total cogs amount
#                 cogs.amount = COGSItems.objects.filter(cogs=cogs, cogs__date=datetime.datetime.today())\
#                                                .aggregate(total=Sum('product__cost'))['total'] or 0
#                 cogs.save()
                
#                 # updae account balance
#                 if invoice.payment_status == 'Partial':
#                     customer_account_balance.balance += -amount_due
#                     customer_account_balance.save()
                    
#                 # Update customer balance
#                 account_balance.balance = Decimal(invoice_data['payable']) + Decimal(account_balance.balance)
#                 account_balance.save()
                
#                 # for tax purpose Zimra
#                 logger.info(invoice_items)

#                 try:
#                     # sig_data, receipt_data = generate_receipt_data(invoice, invoice_items, request)
#                     # logger.info(sig_data)
#                     # hash_sig_data = run(sig_data)
                    
#                     # logger.info(hash_sig_data)
#                     # submit_receipt_data(request, receipt_data, hash_sig_data['hash'], hash_sig_data['signature'])
                    
#                     invoice_data = invoice_preview_json(request, invoice.id)
#                     logger.info(invoice_data)

#                 except Exception as e:
#                     logger.info(e)
#                     return JsonResponse({'success': False, 'error': str(e)})

#                 logger.info(f'inventory creation successfully done: {invoice}')

#                 return JsonResponse({'success':True, 'invoice_id': invoice.id, 'invoice_data':invoice_data})

#         # except (KeyError, json.JSONDecodeError, Customer.DoesNotExist, Inventory.DoesNotExist, Exception) as e:
#         #     return JsonResponse({'success': False, 'error': str(e)})
#         except Exception as e:
#             logger.info(e)

    return render(request, 'invoices/add_invoice.html')

def held_invoice(items_data, invoice, request, vat_rate):
    for item_data in items_data:
        item = Inventory.objects.get(pk=item_data['inventory_id'])
        # product = Product.objects.get(pk=item.product.id)
        item.quantity -= item_data['quantity']
        item.save()
                  
        InvoiceItem.objects.create(
            invoice=invoice,
            item=item,
            quantity=item_data['quantity'],
            unit_price=item_data['price'],
            vat_rate = vat_rate
        )
                    
        # # Create StockTransaction for each sold item
        # stock_transaction = StockTransaction.objects.create(
        #     item=item,
        #     transaction_type=StockTransaction.TransactionType.SALE,
        #     quantity=item_data['quantity'],
        #     unit_price=item.price,
        #     invoice=invoice,
        #     date=timezone.now()
        # )
              
        # stock log  
        ActivityLog.objects.create(
            branch=request.user.branch,
            inventory=item,
            user=request.user,
            quantity=item_data['quantity'],
            total_quantity = item.quantity,
            action='Sale',
            invoice=invoice
        )

@login_required
def submit_invoice_data_zimra(request):
    try:

        data = json.loads(request.body)
        hash = data.get('hash', '')
        signature = data.get('signature', '') 
        receipt_data = data.get('receipt_data')
        invoice_id = data.get('invoice_id')

        logger.info(receipt_data)

        if not hash:
            return JsonResponse({'success':False,'message':f'Hash data is missing!'}, status=400)

        if not signature:
            return JsonResponse({'success':False,'message':f'Signature data is missing!'}, status=400)
        
        try:
            submit_receipt_data(request, receipt_data, hash, signature)
            logger.info('done')
        except Exception as e:
            logger.info(e)
            return JsonResponse(
                {
                    'success':False,
                    'messsage':f'{e}'
                },
                status=400
            )
        
        invoice_data = invoice_preview_json(request, invoice_id)
        logger.info(invoice_data)

        return JsonResponse({'success':True, 'message':'data received', 'data':invoice_data}, status=200)
    except Exception as e:
        return JsonResponse({'message':f'{e}', 'success':False}, status=200)


@login_required
def get_signature_data(request):
    try:   
        data = json.loads(request.body)
        hash = data.get('hash', '')
        signature = data.get('signature', '') 

        if not hash:
            return JsonResponse({'success':False,'message':f'Hash data is missing!'}, status=400)

        if not signature:
            return JsonResponse({'success':False,'message':f'Signature data is missing!'}, status=400)

        return JsonResponse({'success':True, 'message':'data received'}, status=200)
        
    except Exception as e:
        return JsonResponse({'success':False,'message':f'{e}'}, status=400)
    

@login_required
def held_invoice_view(request):
    form = InvoiceForm()
    invoices = Invoice.objects.filter(branch=request.user.branch, status=True, hold_status =True).order_by('-invoice_number')
    logger.info(f'Held invoices: {invoices}')
    return render(request, 'invoices/held_invoices.html', {'invoices':invoices, 'form':form})


def create_invoice_pdf(invoice):
    # Buffer to hold the PDF
    buffer = io.BytesIO()
    
    # Create the PDF object, using the buffer as its "file."
    pdf = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    heading_style = styles['Heading1']
    
    # Company logo and header
    elements.append(Paragraph('<b>Tech City</b>', heading_style))
    elements.append(Paragraph('See • Touch • Own Quality', normal_style))
    elements.append(Paragraph(f'Invoice Number: {invoice.invoice_number}', normal_style))
    
    elements.append(Spacer(1, 12))
    
    # Table Data (Items)
    data = [
        ['Q.', 'Description', 'Amount'],
        [1, 'Hp (hp 250)', 'USD 250.00'],
        ['Sub Total', '', 'USD 250.00'],
        ['Discount', '', 'USD 0.00'],
        ['VAT @15%', '', 'USD 37.50'],
        ['Delivery Charge', '', 'USD 0.00'],
        ['Previous Due', '', 'USD 75.00'],
        ['Current Due', '', 'USD 287.50'],
        ['Total Balance', '', 'USD 362.50'],
        ['Amount Paid', '', 'USD 362.50'],
        ['Due Amount', '', 'USD 0.00']
    ]
    
    # Create table
    table = Table(data, colWidths=[0.5*inch, 3*inch, 1.5*inch])
    
    # Add style to the table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
    ]))
    
    elements.append(table)
    
    elements.append(Spacer(1, 24))
    
    # Terms and conditions
    terms = '''
    All laptop in-built batteries attract 1 month warranty.
    Non in-built batteries attract 48hrs warranty.
    Warranty for all preowned laptops is 5 months. Tech City does not warranty laptops if damaged by water, liquids, or short circuits.
    Any withdrawn deposits for any purchase will attract 10percent administration fee.
    Tech City only accepts exchanges on faulty laptops.
    '''
    elements.append(Paragraph('Terms and Conditions', heading_style))
    elements.append(Paragraph(terms, normal_style))
    
    elements.append(Spacer(1, 12))
    elements.append(Paragraph('Thanks for your purchase!', normal_style))
    
    # Build PDF
    pdf.build(elements)
    
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'Invoice_{invoice.invoice_number}.pdf')


@login_required
@transaction.atomic
def invoice_returns(request, invoice_id): # dont forget the payments
    invoice = get_object_or_404(Invoice, id=invoice_id)
    account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

    sale = get_object_or_404(Sale, transaction=invoice)
    invoice_payment = get_object_or_404(Payment, invoice=invoice)
    stock_transactions = invoice.stocktransaction_set.all()  
    vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)
    activity = ActivityLog.objects.filter(invoice=invoice)

    if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
        customer_account_balance.balance -= invoice.amount_due

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account = get_object_or_404(
        Account, 
        name=f"{request.user.branch} {invoice.currency.name} {invoice_payment.payment_method.capitalize()} Account", 
        type=account_types.get(invoice_payment.payment_method, None) 
    )
    account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
    account_balance.balance -= invoice.amount_paid

    for stock_transaction in activity:
        product = Inventory.objects.get(product=stock_transaction.inventory, branch=request.user.branch)
        product.quantity += stock_transaction.quantity
        product.save()

        logger.info(f'product quantity {product.quantity}')
        logger.info(f'stock quantity {stock_transaction.quantity}')

        ActivityLog.objects.create(
            invoice=invoice,
            product_transfer=None,
            branch=request.user.branch,
            user=request.user,
            action='returns',
            inventory=product,
            quantity=stock_transaction.quantity,
            total_quantity=product.quantity
        )

    InvoiceItem.objects.filter(invoice=invoice).delete() 
    StockTransaction.objects.filter(invoice=invoice).delete()
    Payment.objects.filter(invoice=invoice).delete()

    account_balance.save()
    customer_account_balance.save()
    sale.delete()
    vat_transaction.delete()
    invoice.invoice_return=True
    invoice.save()

    return JsonResponse({'message': f'Invoice {invoice.invoice_number} successfully deleted'})
    

@login_required
@transaction.atomic
def delete_invoice(request, invoice_id):
    try:
        invoice = get_object_or_404(Invoice, id=invoice_id)
        account = get_object_or_404(CustomerAccount, customer=invoice.customer)
        customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

        sale = get_object_or_404(Sale, transaction=invoice)
        payments = Payment.objects.filter(invoice=invoice)  
        vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)
        activity = ActivityLog.objects.filter(invoice=invoice)
        
        with transaction.atomic():
            if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
                customer_account_balance.balance -= invoice.amount_due

            account_types = {
                'cash': Account.AccountType.CASH,
                'bank': Account.AccountType.BANK,
                'ecocash': Account.AccountType.ECOCASH,
            }

            for payment in payments:
                account = get_object_or_404(
                    Account, 
                    name=f"{request.user.branch} {invoice.currency.name} {payment.payment_method.capitalize()} Account", 
                    type=account_types.get(payment.payment_method, None)
                )
                account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
                account_balance.balance -= payment.amount_due
                account_balance.save()

            for stock_transaction in activity:
                product = Inventory.objects.get(id=stock_transaction.inventory.id, branch=request.user.branch)
                product.quantity += abs(stock_transaction.quantity)
                product.save()

                logger.info(f'product quantity {stock_transaction.quantity}')

                ActivityLog.objects.create(
                    invoice=invoice,
                    product_transfer=None,
                    branch=request.user.branch,
                    user=request.user,
                    action='sale return',
                    inventory=product,
                    quantity=stock_transaction.quantity,
                    total_quantity=product.quantity
                )

            InvoiceItem.objects.filter(invoice=invoice).delete() 
            StockTransaction.objects.filter(invoice=invoice).delete()
            payments.delete()
            customer_account_balance.save()
            sale.delete()
            vat_transaction.delete()
            invoice.cancelled = True
            invoice.save()

            logger.info(f'Invoice {invoice.invoice_number} successfully deleted')

        return JsonResponse({'success': True, 'message': f'Invoice {invoice.invoice_number} successfully deleted'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f"{e}"})
      
@login_required       
def invoice_details(request, invoice_id):
    invoice = Invoice.objects.filter(id=invoice_id, branch=request.user.branch).values(
        'invoice_number',
        'customer__id', 
        'customer__name', 
        'products_purchased', 
        'payment_status', 
        'amount'
    )
    return JsonResponse(list(invoice), safe=False)


@login_required
def layby_data(request):
    if request.method == 'GET':
        layby_data = layby.objects.all().select_related(
            'invoice',
            'branch'
        ).values()
        return JsonResponse(list(layby_data), safe=False)
    
    if request.method == 'POST':
        logger.info(f'layby data')
        data = json.loads(request.body)

        invoice_id = data.get('invoice_id')

        if not invoice_id:
            return JsonResponse({'success': False, 'message': 'Invoice ID is required.'})

        laby_dates = laybyDates.objects.filter(layby__invoice__id=invoice_id).values()
        
        logger.info(laby_dates)
        return JsonResponse({'success': True, 'data': list(laby_dates)})

@login_required
@transaction.atomic
def layby_payment(request, layby_date_id):
    try:
        data = json.loads(request.body)
        amount_paid = data.get('amount_paid')
        payment_method = data.get('payment_method')

        layby_date = laybyDates.objects.get(id=layby_date_id)
        layby_obj = layby.objects.get(id=layby_date.layby.id)
        invoice = layby_obj.invoice

        account = CustomerAccount.objects.get(customer=invoice.customer)
        customer_account_balance = CustomerAccountBalances.objects.get(account=account, currency=invoice.currency)
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        customer_account_balance.balance -= amount_paid

        account_name = f"{request.user.branch} {invoice.currency.name} {'cash'.capitalize()} Account"
        account = Account.objects.get(name=account_name, type=account_types['cash'])
        account_balance = AccountBalance.objects.get(account=account, currency=invoice.currency, branch=request.user.branch)

        account_balance.balance -= amount_paid

        amount_paid = layby_date.amount_paid
        amount_due = layby_date.amount_due

        with transaction.atomic():

            account_balance.save()
            customer_account_balance.save()

            # create a payment object
            Payment.objects.create(
                invoice=invoice,
                amount_paid=amount_paid,
                amount_due=amount_due, 
                payment_method=payment_method,
                user=request.user
            )

            # create a cash book object
            Cashbook.objects.create(
                issue_date=timezone.now(),
                description=f'Layby payment ({layby_date.layby.invoice.invoice_number})',
                debit=False,
                credit=True,
                amount=amount_paid,
                currency=layby_date.layby.invoice.currency,
                branch=request.user.branch
            )

            if amount_paid >= amount_due:
                layby_date.paid = True
                layby_date.save()
                layby_obj.fully_paid = True
                layby_obj.save()
                invoice.payment_status = Invoice.PaymentStatus.PAID
                invoice.save()

                layby.check_payment_status()

                return JsonResponse({'success': True, 'message': 'Layby payment successfully completed.'})
            else:
                return JsonResponse({'success': False, 'message': 'Invalid amount paid.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'{e}'})
    

# @login_required
# def paylater(request):
#     if request.method == 'GET':
#         paylater_data = Paylater.objects.all().select_related(
#             'invoice',
#             'branch'
#         ).values()
#         return JsonResponse(list(paylater_data), safe=False)
    
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             amount_paid = data.get('amount_paid')
#             payment_method = data.get('payment_method')
#             invoice_id = data.get('invoice_id')
#             paylater_id = data.get('paylater_id')

#             paylater = Paylater.objects.get(id=paylater_id)
            
#             invoice = Invoice.objects.get(id=invoice_id)
#             account = CustomerAccount.objects.get(customer=invoice.customer)
#             customer_account_balance = CustomerAccountBalances.objects.get(account=account, currency=invoice.currency)
            
#             account_types = {
#                 'cash': Account.AccountType.CASH,
#                 'bank': Account.AccountType.BANK,
#                 'ecocash': Account.AccountType.ECOCASH,
#             }
            
#             customer_account_balance.balance -= amount_paid
            
#             account_name = f"{request.user.branch} {invoice.currency.name} {'cash'.capitalize()} Account"
#             account = Account.objects.get(name=account_name, type=account_types['cash'])
#             account_balance = AccountBalance.objects.get(account=account, currency=invoice.currency, branch=request.user.branch)
#             account_balance.balance -= amount_paid
            
#             amount_due = invoice.total_amount - invoice.amount_paid
            
#             with transaction.atomic():
#                 account_balance.save()
#                 customer_account_balance.save()
                
#                 # Create a payment object
#                 Payment.objects.create(
#                     invoice=invoice,
#                     amount_paid=amount_paid,
#                     amount_due=amount_due,
#                     payment_method=payment_method,
#                     user=request.user
#                 )
                
#                 # Create a cash book object
#                 Cashbook.objects.create(
#                     issue_date=timezone.now(),
#                     description=f'Payment ({invoice.invoice_number})',
#                     debit=False,
#                     credit=True,
#                     amount=amount_paid,
#                     currency=invoice.currency,
#                     branch=request.user.branch
#                 )
                
#                 # Update invoice payment status if fully paid
#                 if amount_paid >= amount_due:
#                     invoice.payment_status = Invoice.PaymentStatus.PAID
#                     invoice.save()

#                     # paylater.amount_paid = amount_paid

#                     # if paylater.amount_paid < paylater.amount_due:
#                     #     paylater.amount_due = abs(amount_due - amount_paid)

#                     # paylater.paid = True
#                     # paylater.save()
                    
#                     # create paylater dates
#                     create_paylater(request, paylater_dates, invoice_id)
#                     return JsonResponse({'status': 'success', 'message': 'Paylater payment successfully completed.'})
#         except Exception as e:
#             return JsonResponse({'status': 'error', 'message': str(e)})
    
#     return JsonResponse({'status': 'success'})

# how many days before
# remimder every how many days 5 days or 20
# reminder days before 

def create_paylater(request, paylater_dates, invoice_id):
    invoice = Invoice.objects.get(id=invoice_id)
    
    paid = check_paylater_payments(invoice_id)
    
    if paid:
        return JsonResponse({'status': 'success', 'message': 'Paylater already paid.'})
    
    paylater = Paylater.objects.create(
        invoice=invoice,
        amount_due=invoice.total_amount,
        due_date=invoice.issue_date + timedelta(days=5),
        payment_method=invoice.payment_method,
        amount_paid = invoice.amount_paid,
        paid = paid
    )
    
    for paylater_date in paylater_dates:
        paylater_date = paylaterDates.objects.create(
            paylater=paylater,
            amount_due=paylater_date.amount_due,
            due_date=paylater_date.due_date,
            payment_method=paylater_date.payment_method,
        )
    
    return JsonResponse({'status': 'success', 'message': 'Paylater created successfully.'})

def check_paylater_payments(paylater_id):
    
    paylater_dates = paylaterDates.objects.filter(paylater=paylater_id)
    paylater_total = paylater_dates.aggregate(total_amount=Sum('amount_paid'))['total_amount']
    paid = paylater_total == paylater.invoice.total_amount
    
    return paid
        
@login_required
def customer(request):
    if request.method == 'GET':
        customers = Customer.objects.all().values()
        return JsonResponse(list(customers), safe=False)
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        
        # validation_errors = validate_customer_data(data)
        # if validation_errors:
        #     return JsonResponse({'success': False, 'message': 'Validation errors occurred:', 'errors': validation_errors})
    
        if Customer.objects.filter(phone_number=data['phonenumber']).exists():
            return JsonResponse({'success': False, 'message': 'Customer exists'})
        else:
            customer = Customer.objects.create(
                name=data['name'],
                email=data['email'],
                address=data['address'],
                phone_number=data['phonenumber'],
                branch=request.user.branch
            )
            account = CustomerAccount.objects.create(customer=customer)
            
            logger.info(account)

        balances_to_create = [
            CustomerAccountBalances(account=account, currency=currency, balance=0) 
            for currency in Currency.objects.all()
        ]
        CustomerAccountBalances.objects.bulk_create(balances_to_create)
    

        return JsonResponse({'success': True, 'message': 'Customer successfully created'})

    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

def validate_customer_data(data):
    errors = {}
    if 'name' not in data or len(data['name']) < 2:
        errors['name'] = 'Name is required and must be at least 2 characters long.'

    if 'email' not in data or not validate_email(data['email']):
        errors['email'] = 'A valid email address is required.'

    if 'address' not in data:
        errors['address'] = 'Address is required.'

    if 'phonenumber' not in data:
        errors['phonenumber'] = 'Phone number is required.'

    return errors

def validate_email(email):
    import re
    email_regex = r"^[a-zA-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[a-zA-Z0-9-]+(?:\.[a-zA-Z0-9-]+)*$"
    return bool(re.match(email_regex, email))

@login_required
def customer_list(request):
    search_query = request.GET.get('q', '')
    
    customers = Customer.objects.filter(branch=request.user.branch)
    accounts = CustomerAccountBalances.objects.all()
    
    total_balances_per_currency = CustomerAccountBalances.objects.filter(account__customer__branch=request.user.branch).values('currency__name').annotate(
        total_balance=Sum('balance')
    )
    
    if search_query:
        customers = CustomerAccount.objects.filter(Q(customer__name__icontains=search_query))
        
    if 'receivable' in request.GET:
        negative_balances_per_currency = CustomerAccountBalances.objects.filter(account__customer__branch=request.user.branch, balance__lt=0) \
            .values('currency') \
            .annotate(total_balance=Sum('balance'))

        customers = Customer.objects.filter(
            id__in=negative_balances_per_currency.values('account__customer_id'),
        ).distinct()
        
        total_balances_per_currency = negative_balances_per_currency.values('currency__name').annotate(
            total_balance=Sum('balance')
        )
        
        logger.info(f'Customers:{total_balances_per_currency.values}')

    if 'download' in request.GET: 
        customers = Customer.objects.all() 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=customers.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header_title in enumerate(['Customer Name', 'Phone Number', 'Email', 'Account Balance'], start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.alignment = header_alignment
            
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header_title), 20)

        customer_accounts = CustomerAccountBalances.objects.all()
        for customer in customer_accounts:
            worksheet.append(
                [
                    customer.account.customer.name, 
                    customer.account.customer.phone_number, 
                    customer.account.customer.email, 
                    customer.balance if customer.balance else 0,
                ]
            )  
            
        workbook.save(response)
        return response
        
    return render(request, 'customers/customers.html', {
        'customers':customers, 
        'accounts':accounts,
        'total_balances_per_currency':total_balances_per_currency,
    })

@login_required
def update_customer(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)

    if request.method == 'POST':  
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f'{customer.name} details updated successfully')  #
            return redirect('finance:customer_list')  
    else:
        form = CustomerForm(instance=customer)  

    return render(request, 'customers/update_customer.html', {'form': form, 'customer': customer}) 

def delete_customer(request, customer_id):
    if request.method == 'DELETE':
        customer = get_object_or_404(Customer, pk=customer_id)

        customer_name = customer.name  
        customer.delete()
        messages.success(request, f'{customer_name} deleted successfully.')
        return JsonResponse({'status': 'success', 'message': f'Customer {customer_name} deleted successfully.'})  
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})  
    

@login_required
def customer_account(request, customer_id):
    form = customerDepositsForm()
    refund_form = customerDepositsRefundForm()
    customer = get_object_or_404(Customer, id=customer_id)

    account = CustomerAccountBalances.objects.filter(account__customer=customer)

    invoices = Invoice.objects.filter(
        customer=customer, 
        branch=request.user.branch, 
        status=True
    )
    
    invoice_payments = Payment.objects.filter(
        invoice__branch=request.user.branch, 
        invoice__customer=customer
    ).order_by('-payment_date')

    filters = Q()
    if request.GET.get('q'):
        filters &= Q(payment_status=request.GET['q'])
    if request.GET.get('search_query'):
        search_query = request.GET['search_query']
        filters &= (Q(invoice_number__icontains=search_query) | Q(issue_date__icontains=search_query))

    invoices = invoices.filter(filters)

    if request.GET.get('email_bool'):
        send_account_statement_email(customer.id, request.user.branch.id, request.user.id)
        return JsonResponse({'message': 'Email sent'})

    return render(request, 'customer.html', {
        'form':form,
        'account': account,
        'invoices': invoices,
        'customer': customer,
        'refund_form':refund_form,
        'invoice_count': invoices.count(),
        'invoice_payments': invoice_payments,
        'paid': invoices.filter(payment_status='Paid').count(),  
        'due': invoices.filter(payment_status='Partial').count(), 
    })


@login_required
@transaction.atomic
def add_customer_deposit(request, customer_id):
    # payload
    """
        customer_id
        amount
        currency
        payment_method
        reason
        payment_reference
    """
    
    try: 
        # get payload
        data = json.loads(request.body)
        customer_id = data.get('customer_id')
        amount = data.get('amount')
        currency = data.get('currency')
        payment_method = data.get('payment_method')
        reason = data.get('reason')
        payment_reference = data.get('payment_reference')        
        
        # payment_reference validation
        if CustomerDeposits.objects.filter(payment_reference=payment_reference).exists():
            return JsonResponse(
                {
                    'success':False,
                    'message': f'Payment reference: {payment_reference} exists'
                }
            )   
                                                   
        # get currency
        currency = Currency.objects.get(id=currency)
        
        # get account types
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        
        account_name = f"{request.user.branch} {currency.name} {payment_method.capitalize()} Account"
        
        
        account, _ = Account.objects.get_or_create(name=account_name, type=account_types[payment_method])
        
        # get or create the account balances
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=currency,
            branch=request.user.branch,
            defaults={'balance': 0}  
        )
        
        account_balance.balance += Decimal(amount)
        account_balance.save()
        logger.info(f"[FINANCE]: deposit -> System {account}")
        
        # check if customer exits
        customer = get_object_or_404(Customer, id=customer_id)  
        logger.info(f"[FINANCE]: deposit -> customer {customer}")
        customer_account = CustomerAccount.objects.get(customer=customer)
        
        customer_account_bal_object, _ = CustomerAccountBalances.objects.get_or_create(
                account=customer_account,
                currency=currency, 
                defaults={'balance': 0}
        )  
        logger.info(f"[FINANCE]: deposit -> customer account object {customer_account_bal_object}")
        
        # effect customer deposit
        customer_deposit = CustomerDeposits.objects.create(
            customer_account=customer_account_bal_object,
            amount=amount,
            currency=currency,
            payment_method=payment_method,
            reason=reason,
            payment_reference=payment_reference,
            cashier=request.user,
            branch=request.user.branch
        )
        
        # effect customer account balances
        customer_account_bal_object.balance += amount
        
        customer_account_bal_object.save()
        
        Cashbook.objects.create(
            issue_date=customer_deposit.date_created,
            description=f'{customer_deposit.payment_method.upper()} deposit ({customer_deposit.customer_account.account.customer.name})',
            debit=True,
            credit=False,
            amount=customer_deposit.amount,
            currency=customer_deposit.currency,
            branch=customer_deposit.branch
        )

        return JsonResponse(
            {
                "success":True,
                "message": f"Customer Deposit of {currency} {amount:2f} has been successfull",
            },
            status=200
        )
    except Exception as e:
        return JsonResponse(
            {
                "message": f"{e}",
                'success':False
            },status=500)


@login_required    
def deposits_list(request):
    deposits = CustomerDeposits.objects.filter(branch=request.user.branch).order_by('-date_created')
    return render(request, 'deposits.html', {
        'deposits':deposits,
        'total_deposits': deposits.aggregate(Sum('amount'))['amount__sum'] or 0,
    })

@login_required
@transaction.atomic
def refund_customer_deposit(request, deposit_id):
    try:
        deposit = CustomerDeposits.objects.get(id=deposit_id)
    except CustomerDeposits.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Deposit not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        amount = Decimal(data.get('amount', 0))
        if amount <= 0:
            return JsonResponse({'success': False, 'message': 'Invalid amount'}, status=400)
    except (json.JSONDecodeError, TypeError, ValueError):
        return JsonResponse({'success': False, 'message': 'Invalid input data'}, status=400)

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"

    try:
        account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
        account_balance = AccountBalance.objects.get(
            account=account,
            currency=deposit.currency,
            branch=request.user.branch,
        )
    except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    if amount > deposit.amount:
        return JsonResponse({'success': False, 'message': 'Refund amount exceeds deposit amount'}, status=400)
    
    account_balance.balance -= amount
    diff_amount = deposit.amount - amount

    if diff_amount == 0:
        deposit.delete()
    else:
        deposit.amount = diff_amount
        deposit.save()

    Cashbook.objects.create(
        issue_date=datetime.date.today(),
        description=f'{deposit.payment_method.upper()} deposit refund ({deposit.customer_account.account.customer.name})',
        debit=False,
        credit=True,
        amount=amount,
        currency=deposit.currency,
        branch=deposit.branch
    )

    account_balance.save()

    return JsonResponse({'success': True}, status=200)

        
@login_required
@transaction.atomic
def edit_customer_deposit(request, deposit_id):
    try:
        deposit = CustomerDeposits.objects.get(id=deposit_id)
    except CustomerDeposits.DoesNotExist:
        messages.warning(request, 'Deposit not found')
        return redirect('finance:customer_account', deposit.customer_account.account.customer.id)
    
    if request.method == 'POST':
        form = customerDepositsForm(request.POST)
        if not form.is_valid():
            messages.warning(request, 'Invalid form submission')
            return redirect('finance:edit_customer_deposit', deposit_id)

        amount = Decimal(request.POST.get('amount'))
        if amount <= 0:
            messages.warning(request, 'Amount cannot be zero or negative')
            return redirect('finance:edit_customer_deposit', deposit_id)

        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"
        
        try:
            account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
            account_balance = AccountBalance.objects.get(
                account=account,
                currency=deposit.currency,
                branch=request.user.branch,
            )
        except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
            messages.warning(request, str(e))
            return redirect('finance:edit_customer_deposit', deposit_id)
        
        adj_amount = amount - deposit.amount

        if adj_amount != 0:
            if adj_amount > 0:
                account_balance.balance += adj_amount
                debit, credit = True, False
            else:
                account_balance.balance += adj_amount 
                debit, credit = False, True

            Cashbook.objects.create(
                issue_date=datetime.date.today(),
                description=f'{deposit.payment_method.upper()} deposit adjustment ({deposit.customer_account.account.customer.name})',
                debit=debit,
                credit=credit,
                amount=abs(adj_amount),
                currency=deposit.currency,
                branch=deposit.branch
            )

            account_balance.save()
            deposit.amount = amount
            deposit.save()
            messages.success(request, 'Customer deposit successfully updated')
            return redirect('finance:customer', deposit.customer_account.account.customer.id)
    else:
        form = customerDepositsForm(instance=deposit)

    return render(request, 'customers/edit_deposit.html', {'form': form})
    

@login_required
def customer_deposits(request): 
    customer_id = request.GET.get('customer_id')
    
    if customer_id: 
        deposits = CustomerDeposits.objects.filter(branch=request.user.branch).values(
            'customer_account__account__customer_id',
            'date_created',
            'amount', 
            'reason',
            'currency__name', 
            'currency__symbol', 
            'payment_method',
            'payment_reference',
            'cashier__username', 
            'id'
        ).order_by('-date_created')
        return JsonResponse(list(deposits), safe=False)
    else:
        return JsonResponse({
            'success':False,
            'message':f'{customer_id} was not provided'
        })

@login_required
def customer_account_transactions_json(request):
    customer_id = request.GET.get('customer_id')
    transaction_type = request.GET.get('type')

    customer = get_object_or_404(Customer, id=customer_id)  

    if transaction_type == 'invoices':
        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        ).order_by('-issue_date').values(
            'issue_date',
            'invoice_number',
            'products_purchased', 
            'amount_paid', 
            'amount_due', 
            'amount', 
            'user__username',
            'payment_status'
        )
        return JsonResponse(list(invoices), safe=False)
    else:
        return JsonResponse({'message': 'Invalid transaction type.'}, status=400)  

@login_required
def customer_account_payments_json(request):
    customer_id = request.GET.get('customer_id')
    transaction_type = request.GET.get('type')

    customer = get_object_or_404(Customer, id=customer_id)

    if transaction_type == 'invoice_payments':
        invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
            invoice__branch=request.user.branch, 
            invoice__customer=customer
        ).order_by('-payment_date').values(
            'invoice__products_purchased',
            'payment_date',
            'invoice__invoice_number',
            'invoice__currency__symbol', 
            'invoice__payment_status',
            'invoice__amount_due',
            'invoice__amount', 
            'user__username', 
            'amount_paid', 
            'amount_due'
        )
        return JsonResponse(list(invoice_payments), safe=False)
    else:
        return JsonResponse({'message': 'Invalid transaction type.'}, status=400)  


@login_required
def customer_account_json(request, customer_id):
    account = CustomerAccountBalances.objects.filter(account__customer__id=customer_id).values(
        'currency__symbol', 'balance'
    )   
    return JsonResponse(list(account), safe=False)

@login_required
def print_account_statement(request, customer_id):
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        
        account = CustomerAccountBalances.objects.filter(account__customer=customer)
        
        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        )
    except:
        messages.warning(request, 'Error in processing the request')
        return redirect('finance:customer')

    invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
        invoice__branch=request.user.branch, 
        invoice__customer=customer
    ).order_by('-payment_date')
    
    return render(request, 'customers/print_customer_statement.html', {
        'customer':customer,
        'account':account,
        'invoices':invoices, 
        'invoice_payments':invoice_payments
    })

# currency views  
@login_required  
def currency(request):
    return render(request, 'currency/currency.html')

@login_required
def currency_json(request):
    currency_id = request.GET.get('id', '')
    currency = Currency.objects.filter(id=currency_id).values()
    return JsonResponse(list(currency), safe=False)


@login_required
def add_currency(request):
    if request.method == 'POST':
        form = CurrencyForm(request.POST)
        if form.is_valid():
            default = request.POST['default']
            try:
                form.save()
                messages.success(request, 'Currency added successfully!')  
            except Exception as e: 
                messages.error(request, f'Error adding currency: {e}')
            return redirect('finance:currency') 
    else:
        form = CurrencyForm()

    return render(request, 'currency/currency_add.html', {'form': form})


@login_required
def update_currency(request, currency_id):
    currency = get_object_or_404(Currency, id=currency_id)  

    if request.method == 'POST': 
        form = CurrencyForm(request.POST, instance=currency)  
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Currency updated successfully') 
            except Exception as e: 
                messages.error(request, f'Error updating currency: {e}')
            return redirect('finance:currency')  
    else:
        form = CurrencyForm(instance=currency) 

    return render(request, 'currency/currency_add.html', {'form': form})

@login_required
def delete_currency(request, currency_id):
    if request.method == 'POST': 
        currency = get_object_or_404(Currency, id=currency_id)
        
        try:
            if currency.invoice_set.exists() or currency.accountbalance_set.exists() or currency.expense_set.exists():  
                raise Exception("Currency is in use and cannot be deleted.")

            currency.delete()
            return JsonResponse({'message': 'Currency deleted successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'message':'Deletion Failed'})


@login_required
def finance_settings(request):
    return render(request, 'settings/settings.html')
    
# Reports
@login_required
def expenses_report(request):
    
    template_name = 'reports/expenses.html'
    
    search = request.GET.get('search', '')
    start_date_str = request.GET.get('startDate', '')
    end_date_str = request.GET.get('endDate', '')
    category_id = request.GET.get('category', '')
   
    if start_date_str and end_date_str:
        try:
            end_date = datetime.date.fromisoformat(end_date_str)
            start_date = datetime.date.fromisoformat(start_date_str)
        except ValueError:
            return JsonResponse({'messgae':'Invalid date format. Please use YYYY-MM-DD.'})
    else:
        start_date = ''
        end_date= ''
        
    try:
        category_id = int(category_id) if category_id else None
    except ValueError:
        return JsonResponse({'messgae':'Invalid category or search ID.'})

    expenses = Expense.objects.all()  
    
    if search:
        expenses = expenses.filter(Q('amount=search'))
    if start_date:
        start_date = parse_date(start_date_str)
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        end_date = parse_date(end_date_str)
        expenses = expenses.filter(date__lte=end_date)
    if category_id:
        expenses = expenses.filter(category__id=category_id)
    
    return generate_pdf(
        template_name,
        {
            'title': 'Expenses', 
            'date_range': f"{start_date} to {end_date}", 
            'report_date': datetime.date.today(),
            'total_expenses':calculate_expenses_totals(expenses),
            'expenses':expenses
        }
    )


@login_required 
def invoice_preview(request, invoice_id):
    invoice = Invoice.objects.get(id=invoice_id)
    invoice_items = InvoiceItem.objects.filter(invoice=invoice)
    return render(request, 'Pos/printable_receipt.html', {'invoice_id':invoice_id, 'invoice':invoice, 'invoice_items':invoice_items})

@login_required
def remove_item(request, item_id):
    if request.method == 'DELETE':
        try:
            item = InvoiceItem.objects.get(id=item_id)
            item.delete()
            return JsonResponse({'success': True})
        except InvoiceItem.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Item not found'}, status=404)

@login_required
def replace_item(request, item_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        new_item_id = data.get('newItemId')

        try:
            item = InvoiceItem.objects.get(id=item_id)
            new_item = InvoiceItem.objects.get(id=new_item_id)

            # Replace the old item with the new one
            item.name = new_item.name  # Update other relevant fields as needed
            item.price = new_item.price
            item.quantity = new_item.quantity
            item.save()

            return JsonResponse({'success': True})
        except (InvoiceItem.DoesNotExist, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid item'}, status=404)
        
def invoice_preview_json(request, invoice_id):
    from django.core.serializers.json import DjangoJSONEncoder
    try:
        invoice = Invoice.objects.get(id=invoice_id)
    except Invoice.DoesNotExist:
        return JsonResponse({"error": "Invoice not found"}, status=404)

    dates = {}
    if invoice.payment_terms == 'layby':
        dates = laybyDates.objects.filter(layby__invoice=invoice).values('due_date')

    invoice_items = InvoiceItem.objects.filter(invoice=invoice).values(
        'item__name',
        'quantity',
        'item__description',
        'total_amount',
        'unit_price'
    )

    invoice_dict = {}
    invoice_dict['customer_name'] = invoice.customer.name
    invoice_dict['customer_email'] = invoice.customer.email
    invoice_dict['customer_cell'] = invoice.customer.phone_number
    invoice_dict['customer_address'] = invoice.customer.address
    invoice_dict['currency_symbol'] = invoice.currency.symbol
    invoice_dict['amount_paid'] = invoice.amount_paid
    invoice_dict['payment_terms'] = invoice.payment_terms
    invoice_dict['amount'] = invoice.amount
    invoice_dict['invoice_number'] = invoice.invoice_number
    invoice_dict['receipt_hash'] = invoice.receipt_hash
    invoice_dict['subtotal'] = invoice.subtotal
    invoice_dict['vat'] = round(invoice.vat, 2)
    invoice_dict['device_id'] = os.getenv("DEVICE_ID")
    invoice_dict['device_serial_number'] = os.getenv("DEVICE_SERIAL_NUMBER")
    invoice_dict['code'] =  invoice.code
    invoice_dict['fiscal_day'] = invoice.fiscal_day

    if invoice.branch:
        invoice_dict['branch_name'] = invoice.branch.name
        invoice_dict['branch_phone'] = invoice.branch.phonenumber
        invoice_dict['branch_email'] = invoice.branch.email

    invoice_dict['user_username'] = invoice.user.username
    invoice_dict['receipt_signature'] = invoice.receiptServerSignature if invoice.receiptServerSignature else None

    # Safely serialize qr_code
    if invoice.qr_code and hasattr(invoice.qr_code, 'url'):
        try:
            invoice_dict['qr_code'] = request.build_absolute_uri(invoice.qr_code.url)
            logger.info(invoice_dict['qr_code'])
        except Exception as e:
            invoice_dict['qr_code'] = None
            logger.info(f"Error generating QR code URL: {e}")
    else:
        invoice_dict['qr_code'] = None

    invoice_data = {
        'invoice': invoice_dict,
        'invoice_items': list(invoice_items),
        'dates': list(dates)
    }
    return invoice_data

@login_required
def invoice_pdf(request):
    template_name = 'reports/invoice.html'
    invoice_id = request.GET.get('id', '')
    if invoice_id:
        try:
            invoice = get_object_or_404(Invoice, pk=invoice_id)

            invoice_items = InvoiceItem.objects.filter(invoice=invoice)
            
        except Invoice.DoesNotExist:
            return HttpResponse("Invoice not found")
    else:
        return HttpResponse("Invoice ID is required")
    
    return generate_pdf(
        template_name,
        {
            'title': 'Invoice', 
            'report_date': datetime.date.today(),
            'invoice':invoice,
            'invoice_items':invoice_items
        }
    )
   
# emails
@login_required
def send_invoice_email(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        invoice_id = data['invoice_id']
        invoice = Invoice.objects.get(id=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        account = CustomerAccount.objects.get(customer__id = invoice.customer.id)
        
        html_string = render_to_string('Pos/receipt.html', {'invoice': invoice, 'invoice_items':invoice_items, 'account':account})
        buffer = BytesIO()

        pisa.CreatePDF(html_string, dest=buffer) 

        email = EmailMessage(
            'Your Invoice',
            'Please find your invoice attached.',
            'your_email@example.com',
            ['recipient_email@example.com'],
        )
        
        buffer.seek(0)
        email.attach(f'invoice_{invoice.invoice_number}.pdf', buffer.getvalue(), 'application/pdf')

        # Send the email
        email.send()

        task = send_invoice_email_task.delay(data['invoice_id']) 
        task_id = task.id 
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=invoice_{invoice.invoice_number}.pdf'
        
        return response
    return JsonResponse({'success': False, 'error':'error'})


#whatsapp
@login_required
def send_invoice_whatsapp(request, invoice_id):
    try:
        
        invoice = Invoice.objects.get(pk=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        img = settings.STATIC_URL + "/assets/logo.png"
    
        html_string = render_to_string('Pos/invoice_template.html', {'invoice': invoice, 'request':request, 'invoice_items':invoice_items, 'img':img})
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if not pisa_status.err:
          
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            invoice_filename = f"invoice_{invoice.invoice_number}.pdf"
            s3.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f"invoices/{invoice_filename}",
                Body=pdf_buffer.getvalue(),
                ContentType="application/pdf",
                ACL="public-read",
            )
            s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

            account_sid = 'AC6890aa7c095ce1315c4a3a86f13bb403'
            auth_token = '897e02139a624574c5bd175aa7aaf628'
            client = Client(account_sid, auth_token)
            from_whatsapp_number = 'whatsapp:' + '+14155238886'
            to_whatsapp_number = 'whatsapp:' + '+263778587612'

            message = client.messages.create(
                from_=from_whatsapp_number,
                body="Your invoice is attached.",
                to=to_whatsapp_number,
                media_url=s3_url
            )
            logger.info(f"WhatsApp message SID: {message.sid}")
            return JsonResponse({"success": True, "message_sid": message.sid})
        else:
            logger.error(f"PDF generation error for Invoice ID: {invoice_id}")
            return JsonResponse({"error": "PDF generation failed"})
    except Invoice.DoesNotExist:
        logger.error(f"Invoice not found with ID: {invoice_id}")
        return JsonResponse({"error": "Invoice not found"})
    except Exception as e:
        logger.exception(f"Error sending invoice via WhatsApp: {e}")
        return JsonResponse({"error": "Error sending invoice via WhatsApp"})
    
@login_required
def invoice_payment_track(request):
    invoice_id = request.GET.get('invoice_id', '')
    
    if invoice_id:
        payments = Payment.objects.filter(invoice__id=invoice_id).order_by('-payment_date').values(
            'payment_date', 'amount_paid', 'payment_method', 'user__username'
        )
    return JsonResponse(list(payments), safe=False)

@login_required
def day_report(request, inventory_data):
    today_min = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_max = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # invoice data
    invoices = Invoice.objects.filter(branch=request.user.branch, issue_date__range=(today_min, today_max))
    
    partial_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PARTIAL)
    paid_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PAID)
    
    # expenses
    expenses = Expense.objects.filter(branch=request.user.branch, date=datetime.date.today())
    
    confirmed_expenses = expenses.filter(status=True)
    unconfirmed_expenses = expenses.filter(staus=False)
    
    # accounts
    account_balances = AccountBalance.objects.filter(branch=request.user.branch)
    
    try:
        html_string = render_to_string('day_report.html',{
                'request':request,
                'invoices':invoices,
                'date': datetime.date.today(),
                'inventory_data': inventory_data,
                'total_sales': paid_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'partial_payments': partial_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'total_paid_invoices': paid_invoices.count(),
                'expenses':expenses,
                'total_partial_invoices': partial_invoices.count(),
                'total_expenses': confirmed_expenses.aggregate(Sum('amount'))['amount__sum'] or 0,
                'confirmed_expenses': confirmed_expenses,
                'unconfirmed_expenses': unconfirmed_expenses,
                'account_balances': account_balances,
            })
        
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if not pisa_status.err:
            # Save PDF to S3 and get URL
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            invoice_filename = f"{request.user.branch} today's ({datetime.date.today}) report.pdf"
            s3.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f"daily_reports/{invoice_filename}",
                Body=pdf_buffer.getvalue(),
                ContentType="application/pdf",
                ACL="public-read",
            )
            s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

            # Send WhatsApp Message with Twilio
            account_sid = settings.TWILIO_ACCOUNT_SID
            auth_token = settings.TWILIO_AUTH_TOKEN
            client = Client(account_sid, auth_token)
            from_whatsapp_number = 'whatsapp:' + '+14155238886'
            to_whatsapp_number = 'whatsapp:' + '+263778587612'

            message = client.messages.create(
                from_=from_whatsapp_number,
                body="Today's report.",
                to=to_whatsapp_number,
                media_url=s3_url
            )

            logger.info(f"WhatsApp message SID: {message.sid}")
            return JsonResponse({"success": True, "message_sid": message.sid})
    except Exception as e:
        logger.exception(f"Error sending invoice via WhatsApp: {e}")
        return JsonResponse({"error": "Error sending invoice via WhatsApp"})

   
@login_required 
@transaction.atomic 
def cash_transfer(request):
    form = TransferForm()
    transfers = CashTransfers.objects.filter(branch=request.user.branch).select_related(
        'user',
        'currency',
        'to',
        'branch',
        'from_branch'
    )
    
    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    if request.method == 'POST':
        form = TransferForm(request.POST)
        
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.user = request.user
            transfer.notification_type = 'Expense'
            transfer.from_branch = request.user.branch
            transfer.branch = request.user.branch
            transfer.received_status = False
            
            account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"
            logger

            with transaction.atomic():
                try:
                    account = Account.objects.get(name=account_name, type=account_types[transfer.transfer_method.lower()])
                    account_balance = AccountBalance.objects.select_for_update().get(
                        account=account,
                        currency=transfer.currency,
                        branch=request.user.branch
                    )
                
                    if account_balance.balance < transfer.amount:
                        messages.error(request, "Insufficient funds in the account.")
                        return redirect('finance:cash_transfer')  

                    account_balance.balance -= transfer.amount
                    account_balance.save()
                    transfer.save()  

                    Cashbook.objects.create(
                        issue_date=transfer.date,
                        description=f'Cash Transfer to {transfer.to.name}',
                        debit=False,
                        credit=True,
                        amount=transfer.amount,
                        currency=transfer.currency,
                        branch=transfer.branch
                    )
                    
                    messages.success(request, 'Money successfully transferred.')
                    return redirect('finance:cash_transfer')  
          
                except Exception as e:
                    messages.error(request, f"{e}")
                return redirect('finance:cash_transfer')  
        else:
            messages.error(request, "Invalid form data. Please correct the errors.")
    return render(request, 'transfers/cash_transfers.html', {'form': form, 'transfers':transfers})

@login_required
def finance_notifications_json(request):
    notifications = FinanceNotifications.objects.filter(status=True).values(
        'transfer__id', 
        'transfer__to',
        'expense__id',
        'expense__branch',
        'invoice__id',
        'invoice__branch',
        'notification',
        'notification_type',
        'id'
    )
    return JsonResponse(list(notifications), safe=False)


@login_required
@transaction.atomic
def cash_transfer_list(request):
    search_query = request.GET.get('q', '')
    transfers = CashTransfers.objects.filter(to=request.user.branch.id)
    
    if search_query:
        transfers = transfers.filter(Q(date__icontains=search_query))
        
    return render(request, 'transfers/cash_transfers_list.html', {'transfers':transfers, 'search_query':search_query})

@login_required
@transaction.atomic
def receive_money_transfer(request, transfer_id):
    if transfer_id:
        transfer = get_object_or_404(CashTransfers, id=transfer_id)
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        
        account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"

        with transaction.atomic():
            try:
                account, _ = Account.objects.get_or_create(name=account_name, type=account_types[transfer.transfer_method.lower()])
            
                account_balance, _ = AccountBalance.objects.get_or_create(
                    account=account,
                    currency=transfer.currency,
                    branch=request.user.branch
                )

                Cashbook.objects.create(
                    issue_date=transfer.date,
                    description=f'Cash Transfer from {transfer.from_branch.name}',
                    debit=True,
                    credit=False,
                    amount=transfer.amount,
                    currency=transfer.currency,
                    branch=transfer.to
                )

                account_balance.balance += transfer.amount
                account_balance.save()
                
                transfer.received_status = True
                transfer.save() 

                return JsonResponse({'message':True})  
            
            except Exception as e:
                return JsonResponse({'success':False, 'message':f"{e}"}) 
    return JsonResponse({'message':"Transfer ID is needed"})  


@login_required
@transaction.atomic
def create_quotation(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        qoute_data = data['data'][0]  
        items_data = data['items']
        
        customer = Customer.objects.get(id=int(qoute_data['client_id']))
        currency = Currency.objects.get(id=qoute_data['currency'])
        
        qoute = Qoutation.objects.create(
            customer = customer,
            amount =  Decimal(qoute_data['subtotal']),
            branch = request.user.branch,
            currency = currency,
            qoute_reference = Qoutation.generate_qoute_number(request.user.branch.name),
            products = ', '.join([f'{item['product_name']} x {item['quantity']}' for item in items_data])
        )
        
        for item_data in items_data:
            item = Inventory.objects.get(pk=item_data['inventory_id'])
            
            QoutationItems.objects.create(
                qoute=qoute,
                product=item,
                unit_price=item.price,
                quantity=item_data['quantity'],
                total_amount= item.price * item_data['quantity'],
            )
        return JsonResponse({'success': True, 'qoute_id': qoute.id})
    return JsonResponse({'success': False})

@login_required        
def qoutation_list(request):
    search_query = request.GET.get('q', '')
    qoutations = Qoutation.objects.filter(branch=request.user.branch).order_by('-date')
 
    if search_query:

        qoutations = qoutations.filter(
            Q(customer__name__icontains=search_query)|
            Q(products__icontains=search_query)|
            Q(date__icontains=search_query)|
            Q(qoute_reference__icontains=search_query)
        )
        
    return render(request, 'qoutations.html', {'qoutations':qoutations, 'search_query':search_query})
        
@login_required 
def qoute_preview(request, qoutation_id):
    qoute = Qoutation.objects.get(id=qoutation_id)
    qoute_items = QoutationItems.objects.filter(qoute=qoute)
    return render(request, 'qoute.html', {'qoute':qoute, 'qoute_items':qoute_items})

@login_required
def qoute_preview_modal(request, qoutation_id):
    try:
        qoute = Qoutation.objects.get(id=qoutation_id)
        logger.info(qoute)
        qoute_items = QoutationItems.objects.filter(qoute=qoute)
        logger.info(f'qoute items: {qoute_items.values()}')
        html = render_to_string('qoutations/partial_preview.html', {
            'qoute': qoute,
            'qoute_items': qoute_items
        }, request=request)  
        
        logger.info(html)

        return JsonResponse({'success': True, 'html': html}, status=200)
        
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success': False, 'message':str(e)}, status=400)

@login_required
def delete_qoute(request, qoutation_id):
    qoute = get_object_or_404(Qoutation, id=qoutation_id)
    qoute.delete()
    return JsonResponse({'success':True, 'message':'Qoutation successfully deleted'}, status=200)

@login_required
def cashbook_view(request):
    """Main view to render the cashbook page"""
    currency = Currency.objects.filter(default=True).first()
    cash_up = CashUp.objects.filter(status=False)

    branches_pending_totals = defaultdict(float)
    total = 0
    branches_data = {}

    for cash in cash_up:
        total += cash.expected_cash
        branches_pending_totals[cash.branch.name] += float(cash.expected_cash)
        branches_data[cash.branch.id] = {
            'id': cash.branch.id,
            'name': cash.branch.name,
            'total': branches_pending_totals[cash.branch.name]
        }

    return render(request, 'cashbook.html', {
        'currency': currency,
        'all_totals': total,
        'cash_up_count': cash_up.count(),
        'to_date': cash_up.last().created_at.date if cash_up else '',
        'from_date':cash_up.first().created_at.date if cash_up else '',
        'branches_data': list(branches_data.values())
    })

@login_required
def cashbook_data(request):
    """AJAX endpoint for cashbook data with filters and pagination"""
    logger.info('Processing cashbook data request')
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            page = int(data.get('page', 1))
            per_page = int(data.get('per_page', 20))
            filter_option = data.get('filter', 'this_week')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            search_query = data.get('search', '')
            currency = data.get('currency')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    else:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        filter_option = request.GET.get('filter', 'this_week')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        search_query = request.GET.get('search', '')
        currency = data.get('currency', '')
    
    logger.info(f'filter Option: {filter_option}, currency: {currency}')

    now = timezone.now()
    end_date = now

    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now

    entries = Cashbook.objects.filter(
        issue_date__gte=start_date,
        issue_date__lte=end_date,
        currency__id=currency,
        branch=request.user.branch
    ).select_related('created_by', 'branch', 'updated_by', 'currency', 'invoice', 'expense').order_by('-issue_date')
    
    logger.info(f'Found {entries.count()} entries')

    if search_query:
        entries = entries.filter(
            Q(description__icontains=search_query) |
            Q(accountant__icontains=search_query) |
            Q(manager__icontains=search_query) |
            Q(director__icontains=search_query)
        )

    entries = entries.order_by('-issue_date')
    
    # Calculate totals
    total_cash_in = entries.filter(debit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
    total_cash_out = entries.filter(credit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
    total_balance = total_cash_in - total_cash_out

    total_entries = entries.count()
    total_pages = (total_entries + per_page - 1) // per_page
    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    paginated_entries = entries[start_index:end_index]
    
    
    # usd and other currencies balances
    currency_cash_in_balances = []
    currency_cash_out_balances = []
    currency_net_balances = []
    
    for currency in Currency.objects.all():
        cash_in = entries.filter(currency=currency, debit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
        cash_out = entries.filter(currency=currency, credit=True, cancelled=False).aggregate(total=Sum('amount'))['total'] or 0
        
        currency_cash_in_balances.append({'name': currency.name, 'amount':cash_in})
        currency_cash_out_balances.append({'name': currency.name, 'amount':cash_out})
        currency_net_balances.append({'name': currency.name, 'amount':cash_in - cash_out})
    
    logger.info(f'balances: {currency_net_balances} /n')
    logger.info(f'cash in{currency_cash_in_balances}')
    logger.info(f'cash out {currency_cash_out_balances}')
     
    balance = 0
    entries_data = []
    for entry in paginated_entries:
        if entry.debit:
            balance += entry.amount
        elif entry.credit:
            balance -= entry.amount
        
        entries_data.append({
            'id': entry.id,
            'date': entry.issue_date.strftime('%Y-%m-%d %H:%M'),
            'description': entry.description,
            'debit': float(entry.amount) if entry.debit else None,
            'credit': float(entry.amount) if entry.credit else None,
            'balance': float(balance),
            'accountant': entry.accountant,
            'manager': entry.manager,
            'director': entry.director,
            'status': entry.status,
            'created_by': entry.created_by.first_name
        })

    return JsonResponse({
        'entries': entries_data,
        'totals': {
            'cash_in': currency_cash_in_balances,
            'cash_out': currency_cash_out_balances,
            'balance': currency_net_balances
        },
        'pagination': {
            'current_page': page,
            'total_pages': total_pages,
            'total_entries': total_entries,
            'has_next': page < total_pages,
            'has_previous': page > 1
        }
    }) 
    
@login_required
def download_cashbook_report(request):
    filter_option = request.GET.get('filter', 'this_week')
    now = datetime.datetime.now()
    end_date = now
    
    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    else:
        start_date = now - timedelta(days=now.weekday())
        end_date = now

    entries = Cashbook.objects.filter(date__gte=start_date, date__lte=end_date, branch=request.user.branch).order_by('date')

    # Create a CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="cashbook_report_{filter_option}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Description', 'Expenses', 'Income', 'Balance'])

    balance = 0  
    for entry in entries:
        if entry.debit:
            balance += entry.amount
        elif entry.credit:
            balance -= entry.amount

        writer.writerow([
            entry.issue_date,
            entry.description,
            entry.amount if entry.debit else '',
            entry.amount if entry.credit else '',
            balance,
            entry.accountant,
            entry.manager,
            entry.director
        ])

    return response


@login_required
def cashbook_note(request):
    #payload
    """
        entry_id:id,
        note:str
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            entry_id = data.get('entry_id')
            note = data.get('note')
            
            entry = Cashbook.objects.get(id=entry_id)
            entry.note = note
            
            entry.save()
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}.'}, status=400)
        return JsonResponse({'success':False, 'message':'Note successfully saved.'}, status=201)
    return JsonResponse({'success':False, 'message':'Invalid request.'}, status=405)


@login_required
def cashbook_note_view(request, entry_id):
    entry = get_object_or_404(Cashbook, id=entry_id)
    
    if request.method == 'GET':
        notes = entry.notes.all().order_by('timestamp')
        notes_data = [
            {'user': note.user.username, 'note': note.note, 'timestamp': note.timestamp.strftime("%Y-%m-%d %H:%M:%S")}
            for note in notes
        ]
        return JsonResponse({'success': True, 'notes': notes_data})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            note_text = data.get('note')
            CashBookNote.objects.create(entry=entry, user=request.user, note=note_text)
            return JsonResponse({'success': True, 'message': 'Note successfully added.'}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request.'}, status=405)
    
@login_required
def cancel_transaction(request):
    #payload
    """
        entry_id:id,
    """
    try:
        data = json.loads(request.body)
        entry_id = int(data.get('entry_id'))
        
        logger.info(entry_id)
        
        entry = Cashbook.objects.get(id=entry_id)
        
        entry.cancelled = True
        
        if entry.director:
            entry.director = False
        elif entry.manager:
            entry.manager = False
        elif entry.accountant:
            entry.accountant = False
            
        entry.save()
        logger.info(entry)
        return JsonResponse({'success': True}, status=201)
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
def update_transaction_status(request, pk):
    if request.method == 'POST':
        entry = get_object_or_404(Cashbook, pk=pk)
        
        data = json.loads(request.body)
        
        status = data.get('status')
        field = data.get('field')  

        if field in ['manager', 'accountant', 'director']:
            setattr(entry, field, status)

            if entry.cancelled:
                entry.cancelled = False
            entry.save()
            return JsonResponse({'success': True, 'status': getattr(entry, field)})
        
    return JsonResponse({'success': False}, status=400)   
    
@login_required
def cashWithdrawals(request):
    search_query = request.GET.get('q', '')
    selected_query = request.GET.get('sq', '')
    
    withdrawals = CashWithdraw.objects.all().order_by('-date')
    
    if search_query:
        withdrawals = withdrawals.filter(
            Q(user__branch__name__icontains=search_query)|
            Q(amount__icontains=search_query)|
            Q(date__icontains=search_query)|
            Q(reason__icontains=search_query)
        )
    if selected_query:
        withdrawals = CashWithdraw.objects.filter(deleted=True).order_by('-date')
        
    if 'download' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=withdrawals.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header_title in enumerate(['Date', 'User', 'Amount', 'Reason', 'Status'], start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.alignment = header_alignment
            
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header_title), 20)

        withdrawals = CashWithdraw.objects.all().order_by('-date')
        for withdrawal in withdrawals:
            worksheet.append(
                [
                    withdrawal.date,
                    withdrawal.user.username,
                    withdrawal.amount,
                    withdrawal.reason,
                    'Canceled' if withdrawal.deleted else 'Expensed' if withdrawal.status else 'pending'
                ])  
            
        workbook.save(response)
        return response
    
    form = CashWithdrawForm()
    expense_form = cashWithdrawExpenseForm()
    
    if request.method == 'POST':
        form = CashWithdrawForm(request.POST)
        
        if form.is_valid():
             
            password = form.cleaned_data['password']
            currency = form.cleaned_data['currency']
            amount = form.cleaned_data['amount']
            
            user = authenticate(username=request.user.username, password=password)
            
            if user is None:
                messages.warning(request, 'Incorrect password')
                return redirect('finance:withdrawals')
            
            cw_obj = form.save(commit=False)
            cw_obj.user = user
            cw_obj.save()
            
            account_name = f"{request.user.branch} {currency.name} {'Cash'} Account"
            
            try:
                account = Account.objects.get(name=account_name, type=Account.AccountType.CASH)
            except Account.DoesNotExist:
                messages.error(request, f'{account_name} doesnt exists')
                return redirect('finance:withdrawals')

            try:
                account_balance = AccountBalance.objects.get(account=account,  branch=request.user.branch)
            except AccountBalance.DoesNotExist:
                messages.error(request, f'Account Balances for account {account_name} doesnt exists')
                return redirect('finance:withdrawals')
            
            account_balance.balance -= Decimal(amount)
            account_balance.save()
            messages.success(request, 'Cash Withdrawal Successfully saved')
        else:
            messages.error(request, 'Invalid form data')
    return render(request, 'cashWithdaraws/withdrawals.html', 
        {
            'withdrawals':withdrawals,
            'count': withdrawals.filter(status=False, deleted=False).count(),
            'expense_form':expense_form,
            'form':form,
        }
    )

@login_required
@transaction.atomic
def cash_withdrawal_to_expense(request):
    if request.method == 'GET':
        cwte_id = request.GET.get('id', '')
        withdrawals = CashWithdraw.objects.filter(id=cwte_id).values(
            'user__branch__name', 'amount', 'reason', 'currency__id', 'user__id'
        )
        return JsonResponse(list(withdrawals), safe=False)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        
        withdrawal_data = data['withdrawal'][0]
        
        reason = data['reason']
        category_id = data['category_id']
        withdrawal_id = data['withdrawal_id']
        currency_id = withdrawal_data['currency__id']
        branch_name = withdrawal_data['user__branch__name']
        amount = withdrawal_data['amount']
        
        try:
            currency = Currency.objects.get(id=currency_id)
            branch = Branch.objects.get(name=branch_name)
            withdrawal = CashWithdraw.objects.get(id=withdrawal_id)
            category = ExpenseCategory.objects.get(id=category_id)
        except:
            return JsonResponse({'success':False,'message':'Invalid form data here'})
        
        Expense.objects.create(
            category=category,
            amount=amount,
            branch=branch,
            user = request.user,
            currency = currency,
            description=f'Cash withdrawal: {reason}',
            status=True,
            issue_date=withdrawal.date,
            payment_method='cash'
        )
        
        logger.info(withdrawal_id)
        withdrawal.status=True
        withdrawal.save()
        
        return JsonResponse({'success':True, 'message':'Successfully added to expenses'}, status=201)
    return JsonResponse({'success':False, 'message':'Invalid form data'}, status=400)
       
@login_required
def delete_withdrawal(request, withdrawal_id):
    try:
        withdrawal = CashWithdraw.objects.get(id=withdrawal_id)
    except User.DoesNotExist:
        messages.warning(request, 'Withdrawal doesnt exist')
        return redirect('finance:withdrawals')
    
    account_name = f"{request.user.branch} {withdrawal.currency.name} {'Cash'} Account"
    
    try:
        account = Account.objects.get(name=account_name, type=Account.AccountType.CASH)
    except Account.DoesNotExist:
        messages.error(request, f'{account_name} doesnt exists')
        return redirect('finance:withdrawals')
    
    try:
        account_balance = AccountBalance.objects.get(account=account,  branch=request.user.branch)
    except AccountBalance.DoesNotExist:
        messages.error(request, f'Account Balances for account {account_name} doesnt exists')
        return redirect('finance:withdrawals')
    
    account_balance.balance += Decimal(withdrawal.amount)
    account_balance.save()
    withdrawal.deleted=True
    withdrawal.save()
    
    messages.success(request, 'Withdrawal successfully deleted')
    return redirect('finance:withdrawals')
    
    
@login_required
def days_data(request):
    current_month = get_current_month()

    sales = Sale.objects.filter(date__month=current_month)
    cogs = COGSItems.objects.filter(date__month=current_month)

    first_day = min(sales.first().date, cogs.first().date)
    
    def get_week_data(queryset, start_date, end_date, amount_field):
        week_data = queryset.filter(date__gte=start_date, date__lt=end_date).values(amount_field, 'date')
        logger.info(week_data)
        total = week_data.aggregate(total=Sum(amount_field))['total'] or 0
        return week_data, total

    data = {}
    for week in range(1, 5):
        week_start = first_day + timedelta(days=(week-1)*7)
        week_end = week_start + timedelta(days=7)

        logger.info(week_start)
        logger.info(week_end)

        sales_data, sales_total = get_week_data(sales, week_start, week_end, 'total_amount')
        cogs_data, cogs_total = get_week_data(cogs, week_start, week_end, 'product__cost')
        
        data[f'week {week}'] = {
            'sales': list(sales_data),
            'cogs': list(cogs_data),
            'total_sales': sales_total,
            'total_cogs': cogs_total
        }

    return JsonResponse(data)

@login_required
def income_json(request):
    current_month = get_current_month()
    today = datetime.date.today()
    
    month = request.GET.get('month', current_month)
    day = request.GET.get('day', today.day)

    sales = Sale.objects.filter(transaction__branch=request.user.branch)
    
    if request.GET.get('filter') == 'today':
        sales_total = sales.filter(date=today).aggregate(Sum('total_amount'))
    else:
        sales_total = sales.filter(date__month=month).aggregate(Sum('total_amount'))

    return JsonResponse({'sales_total': sales_total['total_amount__sum'] or 0})


@login_required
def expense_json(request):
    current_month = get_current_month()
    today = datetime.date.today()
    
    month = request.GET.get('month', current_month)
    day = request.GET.get('day', today.day)

    expenses = Expense.objects.filter(branch=request.user.branch)
    
    if request.GET.get('filter') == 'today':
        expense_total = expenses.filter(issue_date=today, status=False).aggregate(Sum('amount'))
    else:
        expense_total = expenses.filter(issue_date__month=month, status=False).aggregate(Sum('amount'))
    
    return JsonResponse({'expense_total': expense_total['amount__sum'] or 0})


@login_required
def pl_overview(request):
    filter_option = request.GET.get('filter')
    today = datetime.date.today()
    previous_month = get_previous_month()
    current_year = today.year
    current_month = today.month

    sales = Sale.objects.filter(transaction__branch=request.user.branch)
    expenses = Expense.objects.filter(branch=request.user.branch)
    cogs = COGSItems.objects.filter(invoice__branch=request.user.branch)

    if filter_option == 'today':
        date_filter = today
    elif filter_option == 'last_week':
        last_week_start = today - datetime.timedelta(days=today.weekday() + 7)
        last_week_end = last_week_start + datetime.timedelta(days=6)
        date_filter = (last_week_start, last_week_end)
    elif filter_option == 'this_month':
        date_filter = (datetime.date(current_year, current_month, 1), today)
    elif filter_option == 'year':
        year = int(request.GET.get('year', current_year))
        date_filter = (datetime.date(year, 1, 1), datetime.date(year, 12, 31))
    else:
        date_filter = (datetime.date(current_year, current_month, 1), today)

    if filter_option == 'today':
        current_month_sales = sales.filter(date=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(issue_date=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.objects.filter(date=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    elif filter_option == 'last_week':
        current_month_sales = sales.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(issue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    else:
        current_month_sales = sales.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = expenses.filter(dissue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = cogs.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0

    previous_month_sales = sales.filter(date__year=current_year, date__month=previous_month).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
    previous_month_expenses = expenses.filter(issue_date__year=current_year, issue_date__month=previous_month).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
    previous_cogs =  cogs.filter(date__year=current_year, date__month=previous_month).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    
    current_net_income = current_month_sales
    previous_net_income = previous_month_sales 
    current_expenses = current_month_expenses 
    
    current_gross_profit = current_month_sales - cogs_total
    previous_gross_profit = previous_month_sales - previous_cogs
    
    current_net_profit = current_gross_profit - current_month_expenses
    previous_net_profit = previous_gross_profit - previous_month_expenses

    current_gross_profit_margin = (current_gross_profit / current_month_sales * 100) if current_month_sales != 0 else 0
    previous_gross_profit_margin = (previous_gross_profit / previous_month_sales * 100) if previous_month_sales != 0 else 0
    
    # net_income_change = calculate_percentage_change(current_net_income, previous_net_income)
    # gross_profit_change = calculate_percentage_change(current_gross_profit, previous_gross_profit)
    # gross_profit_margin_change = calculate_percentage_change(current_gross_profit_margin, previous_gross_profit_margin)


    data = {
        'net_profit':current_net_profit,
        'cogs_total':cogs_total,
        'current_expenses':current_expenses,
        'current_net_profit': current_net_profit,
        'previous_net_profit':previous_net_profit,
        'current_net_income': current_net_income,
        'previous_net_income': previous_net_income,
        'current_gross_profit': current_gross_profit,
        'previous_gross_profit': previous_gross_profit,
        'current_gross_profit_margin': f'{current_gross_profit_margin:.2f}',
        'previous_gross_profit_margin': previous_gross_profit_margin,
    }
    
    return JsonResponse(data)

@login_required
def cash_deposit(request):
    if request.method == 'GET':
        deposits = CashDeposit.objects.all()
        return render(request, 'cash_deposit.html', 
            {
                'form':cashDepositForm(),
                'deposits':deposits
            }
        )
    
@login_required
def vat(request):
    if request.method == 'GET':
        
        filter_option = request.GET.get('filter', 'today')
        download = request.GET.get('download')
        
        now = datetime.datetime.now()
        end_date = now
        
        if filter_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_week':
            start_date = now - timedelta(days=now.weekday())
        elif filter_option == 'yesterday':
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_month':
            start_date = now.replace(day=1)
        elif filter_option == 'last_month':
            start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        elif filter_option == 'this_year':
            start_date = now.replace(month=1, day=1)
        elif filter_option == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now
            
        vat_transactions = VATTransaction.objects.filter(date__gte=start_date, date__lte=end_date).order_by('-date')
        
        if download:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="vat_report_{filter_option}.csv"'

            writer = csv.writer(response)
            writer.writerow(['Date', 'Description', 'Status', 'Input', 'Output'])

            balance = 0
            for transaction in vat_transactions:

                if transaction.vat_type == 'Input':
                    balance += transaction.tax_amount
                else:
                    balance -= transaction.tax_amount

                writer.writerow([
                    transaction.date,
                    transaction.invoice.invoice_number if transaction.invoice else transaction.purchase_order.order_number,
                    transaction.tax_amount if transaction.vat_type == 'Input' else  '',
                    transaction.tax_amount if transaction.vat_type == 'Output' else  ''
                ])

            writer.writerow(['Total', '', '', balance])
            
            return response
        return render(request, 'vat.html', 
            {
                'filter_option':filter_option,
                'vat_transactions':vat_transactions
            }
        )
    
    if request.method == 'POST':
        # payload 
        {
            'date_from':'date',
            'date_to':'date'
        }
        try:
            data = json.loads(request.body)
            
            date_to = data.get('date_to')
            date_from = data.get('date_from')

            vat_transactions = VATTransaction.objects.filter(
                date__gte=date_from, 
                date__lte=date_to
            )
            
            vat_transactions.update(paid=True)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status = 400)
        return JsonResponse({'success':False, 'message':'VAT successfully paid'}, status = 200)


@login_required
def cash_flow(request):
    """
    View to display a comprehensive financial overview including sales, income, and expenses.
    """
    today = datetime.datetime.today()
    income_form = IncomeCategoryForm()
    
    filter_type = request.GET.get('filter_type', 'today')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if filter_type == 'today':
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'weekly':
        start_date = (today - datetime.timedelta(days=today.weekday())).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'monthly':
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'yearly':
        start_date = today.replace(month=1, day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif filter_type == 'custom':
        if not start_date:
            start_date = today.strftime('%Y-%m-%d')
        if not end_date:
            end_date = today.strftime('%Y-%m-%d')
    
    start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d')
    end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d')

    end_date_query = end_date_obj + datetime.timedelta(days=1)
        
    # Query for invoice items in the date range
    invoice_items = InvoiceItem.objects.filter(
        invoice__issue_date__date__gte=start_date_obj,
        invoice__issue_date__date__lt=end_date_query
    )
    
    # Query for other income and expenses
    income = Income.objects.filter(
        created_at__date__gte=start_date_obj,
        created_at__date__lt=end_date_query
    )
    
    expenses = Expense.objects.filter(
        issue_date__date__gte=start_date_obj,
        issue_date__date__lt=end_date_query
    )
    
    logs = FinanceLog.objects.filter(
        date__gte=start_date_obj.date(),
        date__lt=end_date_query.date()
    )
    
    normalized_sales = invoice_items.annotate(
        type_label=Value('sale', output_field=CharField()),
        category_name=F('item__description'), 
        parent_category=Value('Sales', output_field=CharField()),
        datetime=F('invoice__issue_date'),
        source=Value('Invoice', output_field=CharField()),
        amount=F('total_amount')  
    ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')
    
    normalized_incomes = income.annotate(
        type_label=Value('income', output_field=CharField()),
        category_name=F('category__name'),
        parent_category=F('category__parent__name'),
        datetime=F('created_at'),
        source=Value('Income', output_field=CharField())
    ).values('datetime',  'sale__invoice_items__item__name', 'amount', 'type_label', 'category_name', 'parent_category', 'source', 'note')

    normalized_expenses = expenses.annotate(
        type_label=Value('expense', output_field=CharField()),
        category_name=F('category__name'),
        parent_category=F('category__parent__name'),
        datetime=F('issue_date'),
        source=Value('Expense', output_field=CharField())
    ).values('datetime', 'amount', 'description', 'type_label', 'category_name', 'parent_category', 'source')
    
    combined_cashflow = sorted(
        chain(normalized_incomes, normalized_expenses),
        key=lambda x: x['datetime']
    )
    
    product_sales = invoice_items.values(
        'item__id', 
        'item__name',
        'item__description'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('total_amount'),
        average_price=Avg('unit_price'),
        total_vat=Sum('vat_amount')
    ).order_by('-total_revenue')
    
    sales_total = invoice_items.aggregate(total=Sum('total_amount'))['total'] or 0
    income_total = income.filter(category__name="other").aggregate(total=Sum('amount'))['total'] or 0
    expenses_total = expenses.aggregate(total=Sum('amount'))['total'] or 0
    total_income = sales_total 
    balance = total_income - expenses_total
    
    expenses_by_category = expenses.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    income_by_category = income.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')


    for category in expenses_by_category:
        if expenses_total > 0:
            category['percentage'] = (category['total'] / expenses_total) * 100
        else:
            category['percentage'] = 0
    
    income_by_category = income.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    for category in income_by_category:
        if income_total > 0:
            category['percentage'] = (category['total'] / income_total) * 100
        else:
            category['percentage'] = 0

    expenses_categories = ExpenseCategory.objects.all()
    income_categories = IncomeCategory.objects.all()
    
    cash_ups = CashUp.objects.all().select_related(
        'branch', 'created_by'
    ).prefetch_related(
        'sales', 'expenses'
    ).values(
        'expected_cash',
        'branch__id',
        'branch__name',
        'received_amount',
        'sales',
        'expenses',
        'created_by__username',
        'sales_status',
        'expenses_status',
        'status',
        'date'
    ).order_by('-created_at')

    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'filter_type': filter_type,
        
        'sales': invoice_items,
        'income': income,
        'expenses': expenses,
        'logs': logs,
        
        'product_sales': product_sales,
        'expenses_by_category': expenses_by_category,
        'income_by_category': income_by_category,
        'combined_cashflow': combined_cashflow,
        
        'sales_total': sales_total,
        'income_total': income_total,
        'expenses_total': expenses_total,
        'total_income': total_income,
        'balance': balance,

        'expenses_categories':expenses_categories,
        'income_categories':income_categories,
        
        'cash_ups':cash_ups,
    }
    
    return render(request, 'cashflow.html', context)
    
# @login_required
# def cash_flow(request):
#     from itertools import chain
#     today = datetime.datetime.today()

#     # Income & Sales
#     sales = InvoiceItems.objects.filter(sale__issue_date__date=today)
#     income = Income.objects.filter(created_at__date=today)
#     logs = FinanceLog.objects.filter(date=today)

#     # Expenses
#     expenses = Expense.objects.filter(issue_date__date=today)

#     # Normalize income entries
#     normalized_incomes = income.annotate(
#         type_label=models.Value('income', output_field=models.CharField()),
#         category_name=models.F('category__name'),
#         parent_category=models.F('category__parent__name'),
#         datetime=models.F('created_at'),
#         source=models.Value('Income', output_field=models.CharField())
#     ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')

#     # Normalize expense entries
#     normalized_expenses = expenses.annotate(
#         type_label=models.Value('expense', output_field=models.CharField()),
#         category_name=models.F('category__name'),
#         parent_category=models.F('category__parent__name'),
#         datetime=models.F('issue_date'),
#         source=models.Value('Expense', output_field=models.CharField())
#     ).values('datetime', 'amount', 'type_label', 'category_name', 'parent_category', 'source')

#     # Combine and sort by datetime
#     combined_cashflow = sorted(
#         chain(normalized_incomes, normalized_expenses),
#         key=lambda x: x['datetime']
#     )

#     # Totals
#     sales_total = sales.aggregate(total=Sum('amount'))['total'] or 0
#     income_total = income.aggregate(total=Sum('amount'))['total'] or 0
#     expenses_total = expenses.aggregate(total=Sum('amount'))['total'] or 0
#     total_income = sales_total + income_total
#     balance = total_income - expenses_total

#     context = {
#         'sales': sales,
#         'sales_total': sales_total,
#         'income': income,
#         'income_total': income_total,
#         'expenses_total': expenses_total,
#         'total_income': total_income,
#         'combined_cashflow': combined_cashflow,  
#         'balance':balance,
#         'logs':logs
#         # 'grouped_expenses': grouped_expenses,
#         # 'grouped_income': grouped_income,
#     }

#     return render(request, 'cashflow.html', context)

@login_required
def cash_up_list(request):
    if request.method == 'GET':
        cashups = (
            CashUp.objects.filter() 
            .select_related('branch', 'created_by')
            .prefetch_related(
                'sales',
                'expenses'
            ).select_related(
                'branch',
                'created_by'
            ).values(
                'id',
                'branch__name',
                'expected_cash',
                'created_by__username',
                'created_at',
                'received_amount',
                'cashed_ammout',
                'short_fall'
            ).order_by('-created_at')
        )
        
        logger.info(f'cashup details:{cashups}')

        data = []
        for cashup in cashups:
            cashup_dict = dict(cashup)
            cashup_dict['created_at'] = cashup['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            data.append(cashup_dict)

        return JsonResponse({
            'success': True,
            'data': data
        })
   
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cash_up_type = data.get('type', '')
            branch_id = data.get('branch_id', '')
            
            cash_up=None

            if not branch_id:
                logger.info('here')
                cash_up = CashUp.objects.filter(
                    status=False,
                )   
                logger.info(f'cash up: {cash_up}')                     

            else:
                cash_up = CashUp.objects.filter(
                    branch__id=branch_id,
                    status=False
                )
            
            cash_up = cash_up.annotate(
                total_sales_amount=Coalesce(
                    Sum('sales__amount_paid', output_field=models.DecimalField(max_digits=10, decimal_places=2)), 
                    Value(0, output_field=models.DecimalField(max_digits=10, decimal_places=2))
                ),
                total_expenses_amount=Coalesce(
                    Sum('expenses__amount', output_field=models.DecimalField(max_digits=10, decimal_places=2)), 
                    Value(0, output_field=models.DecimalField(max_digits=10, decimal_places=2))
                )
                ).prefetch_related(
                    'sales',
                    'expenses'
                ).select_related(
                    'branch',
                    'created_by'
                ).order_by('-created_at__time')

            if not cash_up:
                return JsonResponse({'message': 'Cash up not found', 'success': False}, status=404)
            
            sales = []
            expenses = []

            logger.info(f'cash up: {cash_up}')
            for cash in cash_up:
                sales.append(
                    {
                        'cash_id': cash.id,
                        'sales': list(cash.sales.values(
                            'id',
                            'invoice_number',
                            'products_purchased',
                            'amount_paid',
                            'branch',
                            'branch__name',
                            'cash_up_status'
                        )),
                        'expenses': list(cash.expenses.values(
                            'id',
                            'amount',
                            'category__name',
                            'category__parent__name',
                            'issue_date'
                        ))
                    }
                )
            logger.info(f'cash up: {cash_up}')
            return JsonResponse({
                'success': True,
                'cash_up': list(cash_up.values(
                    'id',
                    'branch__name',
                    'expected_cash',
                    'created_by__username',
                    'created_at',
                    'received_amount',
                    'total_sales_amount',
                    'total_expenses_amount',
                    'cashed_amount',
                    'short_fall'
                )),
                'data': {
                    'sales': sales if sales else [],
                    'expenses': expenses if expenses else []
                },
            }, status=200)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@transaction.atomic
def cashflow_create(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            amount = data.get('amount')
            cash_up_id = data.get('cash_up_id')

            with transaction.atomic():
                cash_up = CashUp.objects.get(id=cash_up_id)
                cash_up.received_amount = Decimal(amount)
                cash_up.status = True
                cash_up.save()

                total_income = sum(sale.unit_price * sale.quantity for sale in cash_up.sales.all())
                total_expenses = sum(expense.amount for expense in cash_up.expenses.all())

                for sale in cash_up.sales.all():
                    logger.info(f'cash sale: {sale.unit_price * sale.quantity}', {sale})

                logger.info(f'total income: {total_income}')

                for expenses in cash_up.expenses.all():
                    logger.info(f'cash expense: {expenses}')
                
                logger.info(f'total expenses: {total_expenses}')

                expense_category, _ = MainExpenseCategory.objects.get_or_create(name='Expense')
                income_category, _ = MainIncomeCategory.objects.get_or_create(name='Income')
                
                # for income
                Cashflow.objects.create(
                    branch=cash_up.branch,
                    total=amount,
                    date=datetime.datetime.now(),
                    status=False,
                    cash_up=cash_up,
                    income = total_income,
                    category=income_category,
                    created_by=request.user
                )

                # for expense
                Cashflow.objects.create(
                    branch=cash_up.branch,
                    total=amount,
                    date=datetime.datetime.now(),
                    status=False,
                    cash_up=cash_up,
                    expense = total_expenses,
                    category = expense_category,
                    created_by=request.user
                )

                # if amount > total_income - total_expenses:
                #     user_account, _ = UserAccount.objects.get_or_create(
                #         user=request.user, 
                #         defaults={
                #             'balance': Decimal('0.00'),
                #             'total_credits': Decimal('0.00'),
                #             'total_debits': Decimal('0.00'),
                #             'last_transaction_date':datetime.datetime.now()
                #         }
                #     )

                #     transaction = UserTransaction.objects.create(
                #         account=user_account,
                #         branch=cash_up.branch,
                #         amount=amount - total_income + total_expenses,
                #         transaction_type=UserTransaction.TransactionType.CASH,
                #         description='Cashup deficit',
                #         created_by=request.user
                #     )

                #     user_account.balance += transaction.amount
                #     user_account.total_credits = 0
                #     user_account.total_debits = transaction.amount


                return JsonResponse({'success':True, 'message':'Cashflow successfully created'}, status=201)

        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        


@login_required
@transaction.atomic
def record_cashflow(request):
    try:
        data = json.loads(request.body)
        logger.info(data)
        type = data.get('type', '')
        id = data.get('id', '')
        branch = int(data.get('branch', ''))
        category = (data.get('category'))
        
        # get branch
        branch = get_object_or_404(Branch, id=branch)

        category = None
        if type == 'sale':
            sale = Invoice.objects.filter(id=id).first()
      
            category = IncomeCategory.objects.filter(name__iexact=category).first()

            if not category:
                new_main_category = IncomeCategory.objects.create(
                    name='sales',
                    parent=None
                )

                new_sub_category = IncomeCategory.objects.create(
                    name="sales",
                    parent=new_main_category
                )
                category = new_sub_category
        else:
            # expense = Expense.objects.filter(branch=branch, id=id)
            pass

        with transaction.atomic():

            Income.objects.create(
                amount = sale.amount_paid,
                currency = sale.currency,
                note = sale.products_purchased,
                branch = request.user.branch,
                status = False,
                sale = sale,
                user=request.user,
                category = category
            )

            sale.cash_up_status = True
            sale.save()
            
            return JsonResponse(
                {
                    'success':True, 
                    'message':'Sale recorded succesfully', 
                    'id':id,
                    'cash_up_status':True
                }, status=200)
            
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
    


@login_required
def check_cashup_status(request, cash_up_id):
    logger.info(cash_up_id)
    cash_up = CashUp.objects.filter(id=cash_up_id).first()
    
    if cash_up:
        if True:
            cash_up.status = True
            cash_up.save()
            return JsonResponse({
                'success':True,
                'status':True
            }, status=200)
        return JsonResponse({
                'success':True,
                'status':False
            }, status=200)
    return JsonResponse({
            'success':False,
            'message':'Error occured'
        }, status=400)
    
@login_required
def get_incomes(request):
    page_number = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('limit', 10))  

    incomes = Income.objects.filter(branch=request.user.branch).order_by('-created_at')
    paginator = Paginator(incomes, per_page)
    page_obj = paginator.get_page(page_number)

    income_data = [
        {
            'id': income.id,
            'created_at': income.created_at.strftime('%Y-%m-%d %H:%M'),
            'amount': str(income.amount),
            'category': str(income.category.name),
            'branch':income.branch.name,
            'note': income.note,
            'user': income.user.get_full_name() or income.user.username,
            'is_recurring': income.is_recurring,
            'recurrence': f"{income.recurrence_value} {income.recurrence_unit}" if income.is_recurring else '',
        }
        for income in page_obj.object_list
    ]

    return JsonResponse({
        'data': income_data,
        'has_next': page_obj.has_next()
    })
    
@transaction.atomic
@login_required
def record_income(request):
    try:
        data = json.loads(request.body)
        logger.info(data)

        name = data.get('name')
        amount = data.get('amount')
        category_name = data.get('category')
        branch_id = data.get('branch')
        r_value = data.get('r_value')
        r_unit = data.get('r_unit')

        if not all([name, amount, category_name, branch_id]):
            return JsonResponse({'success': False, 'message': 'Missing required fields'}, status=400)

        parent_category, _ = IncomeCategory.objects.get_or_create(name="Manual", parent=None)
        category, _ = IncomeCategory.objects.get_or_create(name=category_name, parent=parent_category)

        branch = Branch.objects.get(id=branch_id)
        currency = Currency.objects.filter(name__icontains="usd").first() #to be dynamic
        
        logger.info(currency)

        Income.objects.create(
            amount=amount,
            currency_id=currency.id,
            category=category,
            note=name,
            user=request.user,
            branch=branch,
            is_recurring=bool(r_value),
            recurrence_value=r_value if r_value else None,
            recurrence_unit=r_unit if r_unit else None
        )
        
        Cashbook.objects.create(
            amount=amount,
            description=f"Income: {name} -> {category.name}",
            debit=True,
            credit=False,
            branch=branch,
            created_by=request.user,
            updated_by=request.user,
            issue_date=timezone.now()
        )
        
        logger.info(f'Income recorded successfully: {Income.objects.last()}')

        return JsonResponse({'success': True, 'message': 'Income recorded successfully'}, status=200)
    except Exception as e:
        logger.error("Income record error: %s", e)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@transaction.atomic
def record_cashflow_transaction(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            income = float(data.get('IncomeAmount', 0))
            expense_amount = float(data.get('ExpenseAmount', 0))
            transaction_type = data.get('type', '')
            categories = data.get('categories', {})

            category = categories.get('category', {})
            subcategory = categories.get('subcategory', {})
            name = categories.get('name', {})

            if not all([category, subcategory, name]):
                return JsonResponse({
                    'success': False, 
                    'message': 'Missing required category information.'
                }, status=400)

            if not transaction_type or transaction_type not in ['income', 'expense']:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid transaction type.'
                }, status=400)
            
            with transaction.atomic():
                if transaction_type == 'income':
                    if not income or income <= 0:
                        return JsonResponse({
                            'success': False,
                            'message': 'Invalid amount.'
                        }, status=400)
                    
                    logger.info(f'Creating Income amount: {income}')
                    logger.info(f'Categories data: {categories}')

                    cash_flow_name, _ = CashFlowName.objects.get_or_create(
                        name=name.get('value')
                    )
                    sub_category, _ = IncomeSubCategory.objects.get_or_create(
                        name=subcategory.get('value')
                    )
                    main_category, _ = MainIncomeCategory.objects.get_or_create(
                        name=category.get('value'),
                        defaults={'sub_income_category': sub_category}
                    )

                    main_category.save()

                    logger.info(f'cash_flow name: {cash_flow_name}')
                    logger.info(f'sub category: {sub_category}')
                    logger.info(f'main category: {main_category.id} type: {type(main_category)}, main category sub: {main_category.sub_income_category}')

                    object = Cashflow.objects.create(
                        name=cash_flow_name,
                        branch=request.user.branch,
                        total=income,
                        date=datetime.datetime.now(),
                        status=False,
                        income=income,
                        expense=0,
                        income_category=main_category,
                        created_by=request.user
                    )

                    logger.info(f'Income created: {object}.')

                    return JsonResponse({
                        'success': True,
                        'message': 'Income cashflow successfully created'
                    }, status=201)
                
                else:  # expense
                    if not expense_amount or expense_amount <= 0:
                        return JsonResponse({
                            'success': False,
                            'message': 'Invalid amount.'
                        }, status=400)
                    
                    logger.info(f'Creating Expense amount: {expense_amount}')
                    
                    # Create or get the required objects
                    cash_flow_name, _ = CashFlowName.objects.get_or_create(
                        name=name.get('value')
                    )
                    sub_category, _ = ExpenseSubCategory.objects.get_or_create(
                        name=subcategory.get('value')
                    )
                    main_category, _ = MainExpenseCategory.objects.get_or_create(
                        name=category.get('value'),
                        defaults={'sub_expense': sub_category}
                    )

                    main_category.save()                    

                    logger.info(f'cash_flow name: {cash_flow_name}')
                    logger.info(f'sub category: {sub_category}')
                    logger.info(f'main category: {main_category.id} type: {type(main_category)}, main category sub: {main_category.sub_expense}')

                    object = Cashflow.objects.create(
                        name=cash_flow_name,
                        branch=request.user.branch,
                        total=expense_amount,
                        date=datetime.datetime.now(),
                        status=False,
                        expense=expense_amount,
                        income=0,
                        expense_category=main_category,
                        created_by=request.user
                    )

                    logger.info(f'Expense created: {object}.')

                    return JsonResponse({
                        'success': True,
                        'message': 'Expense cashflow successfully created'
                    }, status=201)

        except Exception as e:
            logger.error(f"Error recording transaction: {e}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
        
@login_required
def get_cashflow_categories(request):
    if request.method == 'GET':
        main_income_category = MainIncomeCategory.objects.all()
        main_expense_category = MainExpenseCategory.objects.all()
        return JsonResponse({'income': main_income_category, "expense":main_expense_category}, status == 200)
    return JsonResponse({'success': True, 'message': 'Invalid request'}, status == 500)

@login_required
def add_income_category(request):
    try:
        data = json.loads(request.body)
        name = data.get("name")
        parent_id = data.get("parent_id")
        new_parent_name = data.get("new_parent_name")

        if not name:
            return JsonResponse({"success": False, "error": "Name is required."})

        if new_parent_name:
            parent = IncomeCategory.objects.create(name=new_parent_name)
        elif parent_id:
            parent = IncomeCategory.objects.get(id=parent_id)
        else:
            parent = None

        category = IncomeCategory.objects.create(name=name, parent=parent)
        return JsonResponse({"success": True, "id": category.id, "name": str(category)})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

@login_required
def get_branch_data(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    
    cashup = CashUp.objects.filter(branch=branch).order_by('-date').first()
    
    if cashup:
        categories = {}
        logger.info(cashup.sales.all())
        for sale in cashup.sales.all():
            cat_name = sale.item.name
            logger.info(sale)
            if sale.invoice.amount_paid > 0:
                if cat_name not in categories:
                    categories[cat_name] = {
                        'product': cat_name,
                        'expense': 0,
                        'income': float(sale.quantity * sale.unit_price),
                        'total': float(sale.quantity * sale.unit_price)  
                    }
                else:
                    categories[cat_name]['income'] += float(sale.quantity * sale.unit_price)
                    categories[cat_name]['total'] += float(sale.quantity * sale.unit_price)
        
        for expense in cashup.expenses.all():
            logger.info(expense)
            cat_name = expense.category.name
            if cat_name not in categories:
                categories[cat_name] = {
                    'product': cat_name,
                    'expense': float(expense.amount),
                    'income': 0,
                    'total': -float(expense.amount)  
                }
            else:
                categories[cat_name]['expense'] += float(expense.amount)
                categories[cat_name]['total'] -= float(expense.amount)

        categories['SUMMARY'] = {
            'product': 'SUMMARY',
            'expense': float(cashup.get_total_expenses()),
            'income': float(cashup.get_total_sales()),
            'total': float(cashup.get_net_cash())
        }
        
        data = sorted(
            list(categories.values()),
            key=lambda x: (x['product'] == 'SUMMARY', x['product'])
        )
        
        response_data = {
            'categories': data,
            'cashup_details': {
                'status': cashup.status,
                'expected_cash': float(cashup.expected_cash),
                'received_amount': float(cashup.received_amount),
                'balance': float(cashup.balance),
                'date': datetime.datetime.today().strftime('%Y-%m-%d')
            }
        }

        logger.info(response_data)
        
        return JsonResponse(response_data, safe=False)
    
    return JsonResponse({
        'categories': [],
        'cashup_details': None
    })

@login_required
def daily_summary(request, date=None):
    """View to show daily summary of all cashflows and cashups"""
    from datetime import datetime
    
    if date is None:
        date = datetime.now().date()
    
    daily_cashflows = Cashflow.objects.filter(date=date)
    daily_cashups = CashUp.objects.filter(date=date)
    
    total_income = daily_cashflows.aggregate(Sum('income'))['income__sum'] or 0
    total_expense = daily_cashflows.aggregate(Sum('expense'))['expense__sum'] or 0
    net_total = total_income - total_expense
    
    context = {
        'date': date,
        'daily_cashflows': daily_cashflows,
        'daily_cashups': daily_cashups,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_total': net_total,
    }
    return render(request, 'cashflow/daily_summary.html', context)

@login_required
def branch_summary(request, branch_id):
    """View to show summary of all cashflows and cashups for a specific branch"""
    branch = get_object_or_404(Branch, id=branch_id)
    
    branch_cashflows = Cashflow.objects.filter(branch=branch)
    branch_cashups = CashUp.objects.filter(branch=branch)
    
    context = {
        'branch': branch,
        'cashflows': branch_cashflows,
        'cashups': branch_cashups,
        'total_income': branch_cashflows.aggregate(Sum('income'))['income__sum'] or 0,
        'total_expense': branch_cashflows.aggregate(Sum('expense'))['expense__sum'] or 0,
    }
    return render(request, 'cashflow/branch_summary.html', context)

@login_required
def user_accounts(request):
    users = User.objects.filter(is_active=True).prefetch_related('accounts')
    
    users_with_accounts = []
    for user in users:
        accounts = user.accounts.all()
        
        total_balance = accounts.aggregate(
            total=Coalesce(Sum('balance', output_field=DecimalField()), Decimal('0.00'))
        )['total']
        
        total_credits = accounts.aggregate(
            total=Coalesce(Sum('total_credits', output_field=DecimalField()), Decimal('0.00'))
        )['total']
        
        total_debits = accounts.aggregate(
            total=Coalesce(Sum('total_debits', output_field=DecimalField()), Decimal('0.00'))
        )['total']

        last_activity = accounts.aggregate(
            last_date=Max('last_transaction_date')
        )['last_date']
        
        last_activity = accounts.aggregate(
            last_date=Max('last_transaction_date')
        )['last_date']
        
        users_with_accounts.append({
            'user': user,
            'accounts': accounts,
            'total_balance': total_balance,
            'total_credits': total_credits,
            'total_debits': total_debits,
            'last_activity': last_activity
        })
    logger.info(users_with_accounts)
    context = {
        'users_with_accounts': users_with_accounts
    }
    return render(request, 'user_accounts/user_accounts.html', context)


@login_required
def tax(request):
    tax_receipts = OfflineReceipt.objects.all()
    return render(request, 'tax/tax.html', {
        'tax_receipts':tax_receipts,
        'receipts_count':tax_receipts.count(),
    })

@login_required
def get_config(request):
    try:
        get_config_response = zimra.get_config()
        logger.info(get_config_response)
        return JsonResponse({'success':True, 'data':get_config_response})
    except Exception as e:
        return JsonResponse({'success':False, 'message':f'{e}'})

@login_required
def open_fiscal_day(request):
    try:
        open_day_response = zimra.open_day()
        return JsonResponse({'success': True, 'data': open_day_response})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'{e}'})

@login_required
def check_fiscal_status(request):

    fiscal_day = FiscalDay.objects.filter(created_at__date=datetime.datetime.today(), is_open=True).first()

    if fiscal_day:
        return  JsonResponse({'success':True, 'message':False}, status=200)
    
    return JsonResponse({'success':False, 'message':True}, status=400)

@login_required
def close_fiscal_day(request):
    """
    Close the fiscal day and generate a report with counter values following the FDMS signature format.
    Handles two currencies (USD and ZIG) and three counter types
    (BalanceByMoneyType, SaleByTax, SaleTaxByTax) with currency information.
    """
    if request.method == 'GET':
        try:
            fiscal_day = FiscalDay.objects.filter(created_at__date=datetime.datetime.today(), is_open=True).first()
            if not fiscal_day:
                return JsonResponse({'success': False, 'message': 'No open fiscal day found for today'}, status=404)
                
            fiscal_day_counters = FiscalCounter.objects.filter(created_at__date=datetime.datetime.today())
            
            logger.debug(f'fiscal counters: {fiscal_day_counters.values("fiscal_counter_type", "fiscal_counter_money_type", "fiscal_counter_currency")}')

            sale_by_tax_string = ""  
            sale_tax_by_tax_string = ""
            balance_money_string = ""
            
            sale_by_tax_dict = {}
            sale_tax_by_tax_dict = {}
            balance_by_currency_and_type = {}
            

            for counter in fiscal_day_counters:
                counter_type = counter.fiscal_counter_type.upper()
                counter_currency = counter.fiscal_counter_currency.upper().replace("ZWL", "ZIG")
                counter_value = int(counter.fiscal_counter_value * 100)
                
                if counter_type == "BALANCEBYMONEYTYPE":
                    money_type = counter.fiscal_counter_money_type.upper()
                    
                    key = f"{counter_currency}_{money_type}"
                    
                    if key not in balance_by_currency_and_type:
                        balance_by_currency_and_type[key] = {
                            "type": counter_type,
                            "currency": counter_currency,
                            "money_type": money_type,
                            "value": 0
                        }
                    
                    balance_by_currency_and_type[key]["value"] += counter_value
                
                elif counter_type == "SALEBYTAX":
                    tax_percent = float(counter.fiscal_counter_tax_percent)
                    key = f"{counter_currency}_{tax_percent}"
                    
                    if key not in sale_by_tax_dict:
                        sale_by_tax_dict[key] = {
                            "type": counter_type,
                            "currency": counter_currency,
                            "tax_percent": tax_percent,
                            "value": 0
                        }
                    
                    sale_by_tax_dict[key]["value"] += counter_value
                
                elif counter_type == "SALETAXBYTAX":
                    tax_percent = float(counter.fiscal_counter_tax_percent)
                    key = f"{counter_currency}_{tax_percent}"
                    
                    if key not in sale_tax_by_tax_dict:
                        sale_tax_by_tax_dict[key] = {
                            "type": counter_type,
                            "currency": counter_currency,
                            "tax_percent": tax_percent,
                            "value": 0
                        }
                    
                    sale_tax_by_tax_dict[key]["value"] += counter_value
            
            for key, data in sale_by_tax_dict.items():
                tax_percent = format_tax_percent(data["tax_percent"])
                sale_by_tax_string += f"{data['type']}{data['currency']}{tax_percent}{data['value']}"
            
            for key, data in sale_tax_by_tax_dict.items():
                tax_percent = format_tax_percent(data["tax_percent"])
                sale_tax_by_tax_string += f"{data['type']}{data['currency']}{tax_percent}{data['value']}"

            for key, data in balance_by_currency_and_type.items():
                balance_money_string += f"{data['type']}{data['currency']}{data['money_type']}{data['value']}"

            fiscal_day_counters_string = sale_by_tax_string + sale_tax_by_tax_string + balance_money_string

            hash_input = f"{ZIMRA.device_identification}{fiscal_day.day_no}{datetime.datetime.today().date()}{fiscal_day_counters_string}"
            
            logger.info(f'Hash input for fiscal day signature: {hash_input}')
           
            return JsonResponse({
                'success': True, 
                'data': hash_input
            }, status=200)
        
        except Exception as e:
            logger.error(f"Error closing fiscal day: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': f'{str(e)}'}, status=400)
        
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            signature_string = data.get('sig_string')

            if not signature_string:
                return JsonResponse({'message':'Hash data missing.', 'success':False})
            
            fiscal_day_counters = FiscalCounter.objects.filter(created_at__date=datetime.datetime.today())
            
            day_signature = run(signature_string)

            close_day_response = zimra.close_day(day_signature['hash'], day_signature['signature'], fiscal_day_counters)

            return JsonResponse({'message':close_day_response, 'success':True})

        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
def format_tax_percent(tax_percent):
        """Format tax percent to have 2 decimal places as required by the documentation."""
        if tax_percent == int(tax_percent):
            return f"{int(tax_percent)}.00"
        else:
            return f"{tax_percent:.2f}"
@login_required
def submit_z_report(request):
    try:
        z_report_response = zimra.z_report()
        return JsonResponse({'success': True, 'data': z_report_response})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'{e}'})

@login_required
def send_quote_email(request, quote_id):
    if request.method == 'POST':
        return send_quotation_email(request, quote_id)
    return JsonResponse({'success': False, 'error': 'Only POST method is allowed'})


#reporting
@login_required
def generate_financial_report(request):
    report_type = request.POST.get('reportType')
    time_frame = request.POST.get('timeFrame')
    report_branch = request.POST.get('reportBranch')
    print(report_branch)
    start_date, end_date = get_date_range_from_time_frame(time_frame, request)
    
    
    if report_branch == "all":
        branch = None
    else:
        branch = Branch.objects.filter(id=report_branch).first()
        
    logger.info(branch)
    
    invoices = Invoice.objects.filter(
        cancelled=False,
        invoice_return=False,
        status=True,
        issue_date__date__range=[start_date, end_date]
    )
    
    if branch:
        invoices = invoices.filter(branch=branch)
    
    logger.info(invoices)
    
    expenses = Expense.objects.filter(
        status=False,  # to be changed if confirm feature is ready
        # branch=branch,
        issue_date__date__range=[start_date, end_date]
    ).select_related('category', 'category__parent', 'currency')
    
    total_sales = invoices.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0

    if report_type == 'profit_loss':
        invoice_items = InvoiceItem.objects.filter(invoice__in=invoices)
        total_cost = invoice_items.aggregate(
            cost=Sum(
                ExpressionWrapper(
                    F('quantity') * F('item__price'),
                    output_field=DecimalField(max_digits=15, decimal_places=2)
                )
            )
        )['cost'] or 0
        
        gross_profit = total_sales - total_cost
        net_profit = gross_profit - total_expenses
        
        expenses_by_category = {}
        for expense in expenses:
            parent_category = expense.category.parent or expense.category
            parent_id = parent_category.id
            
            if parent_id not in expenses_by_category:
                expenses_by_category[parent_id] = {
                    'category': parent_category,
                    'items': [],
                    'total': 0
                }
            
            if expense.category.parent is not None:
                expenses_by_category[parent_id]['items'].append(expense)
            
            expenses_by_category[parent_id]['total'] += expense.amount

        html = render_to_string('reports/partials/p_l_report.html', {
            'sales': total_sales,
            'expenses_by_category': expenses_by_category.values(),
            'total_expenses': total_expenses,
            'cost_of_sales': total_cost,
            'gross_profit': gross_profit,
            'net_profit': net_profit,
            'start_date': start_date,
            'end_date': end_date,
            'branch':branch
        })
        
        return JsonResponse({
            'success': True,
            'html': html,
        })
        
    elif report_type == 'sales':
        total_paid = invoices.aggregate(total=Sum('amount_paid'))['total'] or 0
        total_unpaid = total_sales - total_paid
        invoice_count = invoices.count()

        invoice_items = InvoiceItem.objects.filter(invoice__in=invoices)
        
        total_cost = invoice_items.aggregate(
            cost=Sum(
                ExpressionWrapper(
                    F('quantity') * F('item__price'),
                    output_field=DecimalField(max_digits=15, decimal_places=2)
                )
            )
        )['cost'] or 0

        gross_profit = total_sales - total_cost

        product_summary = invoice_items.values(
            name=F('item__name')
        ).annotate(
            quantity_sold=Sum('quantity'),
            total_sales=Sum('total_amount'),
        ).order_by('-total_sales')

        html = render_to_string('reports/partials/sales_report_products.html', {
            'products': product_summary,
            'total_sales': round(total_sales, 2),
            'total_paid': round(total_paid, 2),
            'total_unpaid': round(total_unpaid, 2),
            'invoice_count': invoice_count,
            'branch': branch
        })

        return JsonResponse({
            'success': True,
            'html': html,
            'start_date': start_date,
            'end_date': end_date,
        })
        
    elif report_type == 'expenses':
        html = render_to_string('reports/partials/expense_report_table.html', {
            'expenses': expenses,
            'total_expenses': total_expenses,
            'start_date': start_date,
            'end_date': end_date,
            'branch': branch
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'report_type': report_type,
            'time_frame': time_frame,
            'start_date': str(start_date),
            'end_date': str(end_date),
            'total_expenses': float(total_expenses),
            'html': html,
        })

    return JsonResponse({'status': 'error', 'message': 'Unsupported report type'}, status=400)

def get_date_range_from_time_frame(time_frame, request):

    if time_frame == 'today':
        return today, today

    elif time_frame == 'weekly':
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end

    elif time_frame == 'monthly':
        return today.replace(day=1), today

    elif time_frame == 'yearly':
        return today.replace(month=1, day=1), today

    elif time_frame == 'custom':
        start_date = parse_date(request.POST.get('startDate'))
        end_date = parse_date(request.POST.get('endDate'))
        return start_date, end_date

    return today, today

    

#API 
#############################################################################################################
from rest_framework import status, views, viewsets
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.viewsets import ModelViewSet
from .serializers import *

# class CustomerCrud(viewsets.ModelViewSet):
#     permission_classes = [IsAuthenticated]
#     queryset = Customer.objects.all()
#     serializer_class = CustomerSerializer

#     def create(self, request):
#         data = request.data
        
#         validation_errors = validate_customer_data(data)
#         if validation_errors:
#             return Response(validation_errors, status= status.HTTP_406_NOT_ACCEPTABLE)
#         if Customer.objects.filter(phone_number=data['phonenumber']).exists():
#             return Response('Customer exists', status= status.HTTP_409_CONFLICT)
#         else:
#             customer = Customer.objects.create(
#                 name=data['name'],
#                 email=data['email'],
#                 address=data['address'],
#                 phone_number=data['phonenumber'],
#                 branch=request.user.branch
#             )
#             account = CustomerAccount.objects.create(customer=customer)

#         balances_to_create = [
#             CustomerAccountBalances(account=account, currency=currency, balance=0) 
#             for currency in Currency.objects.all()
#         ]
#         CustomerAccountBalances.objects.bulk_create(balances_to_create)
#         return JsonResponse(status.HTTP_201_CREATED)

class CustomersViewset(ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request):
        data = request.data
        branch = request.user.branch

        if Customer.objects.filter(Q(phone_number=data.get('phonenumber')) | Q(email=data.get('email'))).exists():
            return Response({'error': 'Customer with this phone number or email already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        customer = Customer.objects.create(
            name=data.get('name'),
            email=data.get('email'),
            address=data.get('address'),
            phone_number=data.get('phonenumber'),
            branch=branch
        )
        account = CustomerAccount.objects.create(customer=customer)

        balances_to_create = [
            CustomerAccountBalances(account=account, currency=currency, balance=0) 
            for currency in Currency.objects.all()
        ]
        CustomerAccountBalances.objects.bulk_create(balances_to_create)

        customer_serializer = self.get_serializer(customer)
        return Response(customer_serializer.data, status=status.HTTP_201_CREATED)
    
class AllCustomerAccounts(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        customer_infomation = Customer.objects.all()
        customer_balances = CustomerAccountBalances.objects.all()

        customer_infomation_list = {}
        # balance = []
        for items in customer_balances:
            balance = []
            customer_infomation_list[items.account.customer.id] = {
                'name':items.account.customer.name,
                'phone_number': items.account.customer.phone_number,
                'email': items.account.customer.email,
                'accounts':
                {   
                
                }
            }
            id = items.account.customer.id
            for item in customer_balances:
                if id == item.account.customer.id:
                    balance.append(
                        {
                            'customer_id': item.account.customer.id,
                            'currency': item.currency.name,
                            'balance': item.balance
                            
                        }
                    )
            customer_infomation_list[items.account.customer.id]['accounts'] = balance

        logger.info(customer_infomation_list)
        return Response(customer_infomation_list, status.HTTP_200_OK)

class CustomerCurrenciesTotal(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        currencies = Currency.objects.all()
        balances  = CustomerAccountBalances.objects.all()

        currencies_information = {}

        for item in currencies:
            total_balance = 0
            currencies_information[item.name] = {
                'balance': 0
            }
            currency_name = item.name
            for items in balances:
                if items.currency.name == currency_name:
                    total_balance += items.balance
            currencies_information[item.name]['balance'] = total_balance

        return Response(currencies_information, status.HTTP_200_OK)

class CustomerAccountView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, customer_id):
        customer = get_object_or_404(Customer, id=customer_id)
        customer_serializer = CustomerSerializer(customer)

        account = CustomerAccountBalances.objects.filter(account__customer=customer).values()

        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        ).values()

        invoice_payments = Payment.objects.filter(
            invoice__branch=request.user.branch, 
            invoice__customer=customer
        ).order_by('-payment_date').values()

        filters = Q()
        if request.GET.get('q'):
            filters &= Q(payment_status=request.GET['q'])
        if request.GET.get('search_query'):
            search_query = request.GET['search_query']
            filters &= (Q(invoice_number__icontains=search_query) | Q(issue_date__icontains=search_query))

        invoices = invoices.filter(filters)

        if request.GET.get('email_bool'):
            send_account_statement_email(customer.id, request.user.branch.id, request.user.id)
            return Response({'message': 'Email sent'},status.HTTP_200_OK)

        paid_invoice = invoices.filter(payment_status='Paid').count()
        due_invoice = invoices.filter(payment_status='Partial').count()
        return Response({
            'account': account,
            'invoices': invoices,
            'customer': customer_serializer.data,
            'invoice_count': invoices.count(),
            'invoice_payments': invoice_payments,
            'paid': paid_invoice,  
            'due': due_invoice, 
        },status.HTTP_200_OK)
    
class CustomerPaymentsJsonView(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, customer_id):
        customer_id = customer_id
        transaction_type = request.data.get('type')

        customer = get_object_or_404(Customer, id=customer_id)

        if transaction_type == 'invoice_payments':
            invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
                invoice__branch=request.user.branch, 
                invoice__customer=customer
            ).order_by('-payment_date').values(
                'invoice__products_purchased',
                'payment_date',
                'invoice__invoice_number',
                'invoice__currency__symbol', 
                'invoice__payment_status',
                'invoice__amount_due',
                'invoice__amount', 
                'user__username',
                'amount_paid', 
                'amount_due'
            )
            return Response(invoice_payments,status.HTTP_200_OK)
        else:
            return Response(status.HTTP_400_BAD_REQUEST)

class EditCustomerDeposit(views.APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, deposit_id):
        try:
            deposit = CustomerDeposits.objects.get(id=deposit_id)
        except CustomerDeposits.DoesNotExist:
            # messages.warning(request, 'Deposit not found')
            return Response(status.HTTP_404_NOT_FOUND)
        
        form = CustomerDepositSerializer(data = request.data)
        if not form.is_valid():
            #messages.warning(request, 'Invalid form submission')
            return Response({'message':'Invalid form submission, redirect to finance:edit_customer_deposit' , 'Deposti Id':f'{deposit_id}'}, status.HTTP_400_BAD_REQUEST)

        amount = Decimal(request.data.get('amount'))
        if amount <= 0:
            #messages.warning(request, 'Amount cannot be zero or negative')
            return Response({'message':'Amount cannot be zero or negative, redirect to finance:edit_customer_deposit' , 'Deposti Id':f'{deposit_id}'}, status.HTTP_400_BAD_REQUEST)

        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"
        
        try:
            account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
            account_balance = AccountBalance.objects.get(
                account=account,
                currency=deposit.currency,
                branch=request.user.branch,
            )
        except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
            #messages.warning(request, str(e))
            return Response({'message':f'{e}' , 'Deposti Id':f'{deposit_id}'}, status.HTTP_400_BAD_REQUEST)
        
        adj_amount = amount - deposit.amount

        if adj_amount != 0:
            if adj_amount > 0:
                account_balance.balance += adj_amount
                debit, credit = True, False
            else:
                account_balance.balance += adj_amount 
                debit, credit = False, True

        Cashbook.objects.create(
            issue_date=datetime.date.today(),
            description=f'{deposit.payment_method.upper()} deposit adjustment ({deposit.customer_account.account.customer.name})',
            debit=debit,
            credit=credit,
            amount=abs(adj_amount),
            currency=deposit.currency,
            branch=deposit.branch
        )

        account_balance.save()
        deposit.amount = amount
        deposit.save()
        #messages.success(request, 'Customer deposit successfully updated')
        return Response({deposit.customer_account.account.customer.id}, status.HTTP_200_OK)

class CustomerAccountJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, customer_id):
        account = CustomerAccountBalances.objects.filter(account__customer__id=customer_id).values(
            'currency__symbol', 'balance'
        )  
        return Response(account, status.HTTP_200_OK)

class CustomerAccountTransactionsJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, id):
        customer_id = id
        transaction_type = request.data.get('type')

        customer = get_object_or_404(Customer, id=customer_id)  

        if transaction_type == 'invoices':
            invoices = Invoice.objects.filter(
                customer=customer, 
                branch=request.user.branch, 
                status=True
            ).order_by('-issue_date').values(
                'issue_date',
                'invoice_number',
                'products_purchased', 
                'amount_paid', 
                'amount_due', 
                'amount', 
                'user__username',
                'payment_status'
            )
            return Response(invoices, status.HTTP_200_OK)
        else:
            return Response({'message': 'Invalid transaction type.'}, status.HTTP_400_BAD_REQUEST)

class RefundCustomerDeposit(views.APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, deposit_id):
        try:
            deposit = CustomerDeposits.objects.get(id=deposit_id)
        except CustomerDeposits.DoesNotExist:
            return Response({'message': 'Deposit not found'}, status.HTTP_404_NOT_FOUND)
        
        try:
            data = request.data
            amount = Decimal(data.get('amount', 0))
            if amount <= 0:
                return Response({'message': 'Invalid amount'}, status.HTTP_400_BAD_REQUEST)
        except (json.JSONDecodeError, TypeError, ValueError):
            return Response({'message': 'Invalid input data'}, status.HTTP_400_BAD_REQUEST)

        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"

        try:
            account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
            account_balance = AccountBalance.objects.get(
                account=account,
                currency=deposit.currency,
                branch=request.user.branch,
            )
        except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
            return Response({'message': str(e)}, status.HTTP_400_BAD_REQUEST)
        
        if amount > deposit.amount:
            return Response({'message': 'Refund amount exceeds deposit amount'}, status.HTTP_400_BAD_REQUEST)
        
        account_balance.balance -= amount
        diff_amount = deposit.amount - amount

        if diff_amount == 0:
            deposit.delete()
        else:
            deposit.amount = diff_amount
            deposit.save()

        Cashbook.objects.create(
            issue_date=datetime.date.today(),
            description=f'{deposit.payment_method.upper()} deposit refund ({deposit.customer_account.account.customer.name})',
            debit=False,
            credit=True,
            amount=amount,
            currency=deposit.currency,
            branch=deposit.branch
        )

        account_balance.save()
        return Response(status.HTTP_200_OK)

class PrintAccountStatement(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, customer_id):
        try:
            customer = get_object_or_404(Customer, id=customer_id)
            
            account = CustomerAccountBalances.objects.filter(account__customer=customer).values()
            
            invoices = Invoice.objects.filter(
                customer=customer, 
                branch=request.user.branch, 
                status=True
            ).values()
        except:
            #messages.warning(request, 'Error in processing the request')
            return Response({'message':'Error in processing the request'}, status.HTTP_400_BAD_REQUEST)

        invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
            invoice__branch=request.user.branch, 
            invoice__customer=customer
        ).order_by('-payment_date').values()
        
        customer_serializer = CustomerSerializer(customer)

        return Response({
            'customer':customer_serializer.data,
            'account':account,
            'invoices':invoices, 
            'invoice_payments':invoice_payments
        }, status.HTTP_200_OK)

class CustomerDepositsView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, id): 
        customer_id = id
        
        if customer_id:
            deposits = CustomerDeposits.objects.filter(branch=request.user.branch).values(
                'customer_account__account__customer_id',
                'date_created',
                'amount', 
                'reason',
                'currency__name', 
                'currency__symbol', 
                'payment_method',
                'payment_reference',
                'cashier__username', 
                'id'
            ).order_by('-date_created')

            return Response(deposits, status.HTTP_200_OK)
        else:
            return Response({
                'message':f'{customer_id} was not provided'
            }, status.HTTP_400_BAD_REQUEST)

class DepositList(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        deposits = CustomerDeposits.objects.filter(branch=request.user.branch).order_by('-date_created').values()
        return Response({
            'deposits':deposits,
            'total_deposits': deposits.aggregate(Sum('amount'))['amount__sum'] or 0,
        }, status.HTTP_200_OK)

class CashTransfer(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        transfers = CashTransfers.objects.filter(branch=request.user.branch)
        
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        form = TransferSerializer(data = request.data)
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.user = request.user
            transfer.notification_type = 'Expense'
            transfer.from_branch = request.user.branch
            transfer.branch = request.user.branch
            transfer.received_status = False
            
            account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"

            try:
                account = Account.objects.get(name=account_name, type=account_types[transfer.transfer_method.lower()])
            except Account.DoesNotExist:
                #messages.error(request, f"Account '{account_name}' not found.")
                return Response({'message': f'Account {account_name} not found.'}, status.HTTP_400_BAD_REQUEST)  

            try:
                account_balance = AccountBalance.objects.select_for_update().get(
                    account=account,
                    currency=transfer.currency,
                    branch=request.user.branch
                )
            except AccountBalance.DoesNotExist:
                #messages.error(request, "Account balance record not found.")
                return Response({'message':'Account balance record not found.'}, status.HTTP_400_BAD_REQUEST)

            if account_balance.balance < transfer.amount:
                #messages.error(request, "Insufficient funds in the account.")
                return Response({'message':'Insufficient funds in the account.'}, status.HTTP_400_BAD_REQUEST) 

            account_balance.balance -= transfer.amount
            account_balance.save()
            transfer.save()  
            
            #messages.success(request, 'Money successfully transferred.')
            return Response({'message':'Money successfully transferred.'}, status.HTTP_200_OK)  
        else:
            #messages.error(request, "Invalid form data. Please correct the errors.")
            return Response({'message':'Invalid form data. Please correct the errors.'}, status.HTTP_400_BAD_REQUEST)

class CashTransferList(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        search_query = request.data.get('q', '')
        transfers = CashTransfers.objects.filter(to=request.user.branch.id).values()

        if search_query:
            transfers = transfers.filter(Q(date__icontains=search_query))   
        return Response({'transfers':transfers, 'search_query':search_query}, status.HTTP_200_OK)

class ReceiveMoneyTransfer(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, transfer_id):
        if transfer_id:
            transfer = get_object_or_404(CashTransfers, id=transfer_id)
            account_types = {
                'cash': Account.AccountType.CASH,
                'bank': Account.AccountType.BANK,
                'ecocash': Account.AccountType.ECOCASH,
            }
            
            account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"

            try:
                account, _ = Account.objects.get_or_create(name=account_name, type=account_types[transfer.transfer_method.lower()])
            except Account.DoesNotExist:
                return Response({'message':f"Account '{account_name}' not found."}, status.HTTP_400_BAD_REQUEST) 

            try:
                account_balance, _ = AccountBalance.objects.get_or_create(
                    account=account,
                    currency=transfer.currency,
                    branch=request.user.branch
                )
            except AccountBalance.DoesNotExist:
                #messages.error(request, )
                return Response({'message':"Account balance record not found."}, status.HTTP_400_BAD_REQUEST)  

            account_balance.balance += transfer.amount
            account_balance.save()
            
            transfer.received_status = True
            transfer.save() 
            return Response(status.HTTP_200_OK)  
        return Response({'message':"Transfer ID is needed"}, status.HTTP_400_BAD_REQUEST)  

class FinanceNotification(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        notifications = FinanceNotifications.objects.filter(status=True).values(
            'transfer__id', 
            'transfer__to',
            'expense__id',
            'expense__branch',
            'invoice__id',
            'invoice__branch',
            'notification',
            'notification_type',
            'id'
        )
        return Response(notifications, status.HTTP_200_OK)

class CurrencyViewset(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer
    def retrieve(self, request, pk):
        logger.info(pk)
        data = Currency.objects.get(id = pk)
        info = {
            'code': data.code,
            'name': data.name,
            'symbol': data.symbol,
            'exchange rate': data.exchange_rate
        }
        return Response(info, status.HTTP_200_OK)
    def list(self, request):
        data = self.queryset.values()
        logger.info(data)
        return Response(data, status.HTTP_200_OK)
    def create(self, request):
        con_data = request.data
        data = {
            'code': con_data.get('code'),
            'name': con_data.get('name'),
            'symbol': con_data.get('symbol'),
            'exchange_rate': con_data.get('exchange_rate')
        }
        if not Currency.objects.filter(name = con_data.get('name')):
            Currency.objects.create(
                code = data.get('code'),
                name = data.get('name'),
                symbol = data.get('symbol'),
                exchange_rate = data.get('exchange_rate')
            )
            data_saved = Currency.objects.get(name = data.get('name'))
            logger.info(data_saved)
            data_returned = {
                'code': data_saved.code,
                'name': data_saved.name,
                'symbol': data_saved.symbol,
                'exchange_rate': data_saved.exchange_rate
            }
            return Response(data_returned,status.HTTP_201_CREATED)
        return Response(status.HTTP_400_BAD_REQUEST)
    def update(self, request, pk):
        logger.info(pk)
        data = request.data
        change_data = Currency.objects.get(id = pk)
        change_data.code = data.get('code')
        change_data.name = data.get('name')
        change_data.symbol = data.get('symbol')
        change_data.exchange_rate = data.get('exchange rate')

        change_data.save()
        updated_data = Currency.objects.get(id = pk)

        data_returned = {
                'code': updated_data.code,
                'name': updated_data.name,
                'symbol': updated_data.symbol,
                'exchange_rate': updated_data.exchange_rate
        }
        return Response(data_returned, status.HTTP_200_OK)

class CashWithdrawalsViewset(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = CashWithdrawals.objects.all()
    serializer_class = CashWithdrawalSerializer

class EndOfDay(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        today = timezone.now().date()
        user_timezone_str = request.user.timezone if hasattr(request.user, 'timezone') else 'UTC'
        user_timezone = pytz_timezone(user_timezone_str)  

        # make a utility
        def filter_by_date_range(start_date, end_date):
            start_datetime = user_timezone.localize(
                timezone.datetime.combine(start_date, timezone.datetime.min.time())
            )
            end_datetime = user_timezone.localize(
                timezone.datetime.combine(end_date, timezone.datetime.max.time())
            )
            return Invoice.objects.filter(branch=request.user.branch, issue_date__range=[start_datetime, end_datetime])

        now = timezone.now().astimezone(user_timezone)
        today = now.date()

        now = timezone.now() 
        today = now.date()  
        
        # invoices = filter_by_date_range(today, today)
        invoices = filter_by_date_range(today, today)

        logger.info(f'Invoices {invoices}')
        withdrawals = CashWithdraw.objects.filter(user__branch=request.user.branch, date=today, status=False)
        
        total_cash_amounts = [
            {
                'total_invoices_amount' : invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'total_withdrawals_amount' : withdrawals.aggregate(Sum('amount'))['amount__sum'] or 0
            }
        ]

        logger.info(f'cash amounts: {total_cash_amounts}')

        sold_inventory = (
            ActivityLog.objects
            .filter(invoice__branch=request.user.branch, timestamp__date=today, action='Sale')
            .values('inventory__id', 'inventory__name')
            .annotate(quantity_sold=Sum('quantity'))
        )

        logger.info(f'Sold Inventory: {sold_inventory}')
        
        all_inventory = Inventory.objects.filter(branch=request.user.branch, status=True).values(
            'id', 'name', 'quantity'
        )

        inventory_data = []
        for item in sold_inventory:
            logger.info(item)
            sold_info = next((inv for inv in all_inventory if item['inventory__id'] == inv['id']), None)
            
            if sold_info:
                inventory_data.append({
                    'id': item['inventory__id'],
                    'name': item['inventory__name'],
                    'initial_quantity': item['quantity_sold'] + sold_info['quantity'] if sold_info else 0,
                    'quantity_sold':  item['quantity_sold'],
                    'remaining_quantity': sold_info['quantity'] if sold_info else 0,
                    'physical_count': None
                })
        logger.info(inventory_data)
        logger.info(total_cash_amounts)
        return Response({'inventory': inventory_data, 'total_cash_amounts':total_cash_amounts}, status.HTTP_200_OK)    
    def post(self, request):
        try:
            today = timezone.now().date()
            data = request.data
            inventory_data = []
            
            sold_inventory = (
            ActivityLog.objects
            .filter(invoice__branch=request.user.branch, timestamp__date= today, action='Sale')
            .values('inventory__id', 'inventory__name')
            .annotate(quantity_sold=Sum('quantity'))
            )

            logger.info(f'Sold Inventory: {sold_inventory}')

            for item in data:
                try:
                    inventory = Inventory.objects.get(id=item.get('item_id'), branch=request.user.branch, status=True)
                    inventory.physical_count = item.get('physical_count')
                    inventory.save()

                    sold_info = next((i for i in sold_inventory if i['inventory__id'] == inventory.id), None)
                    inventory_data.append({
                        'id': inventory.id,
                        'name': inventory.name,
                        'initial_quantity': inventory.quantity,
                        'quantity_sold': sold_info['quantity_sold'] if sold_info else 0,
                        'remaining_quantity': inventory.quantity - (sold_info['quantity_sold'] if sold_info else 0),
                        'physical_count': inventory.physical_count,
                        'difference': inventory.physical_count - (inventory.quantity - (sold_info['quantity_sold'] if sold_info else 0))
                    })
                except Inventory.DoesNotExist:
                    return Response({'error': f'Inventory item with id {item["inventory_id"]} does not exist.'}, status.HTTP_404_NOT_FOUND)

            today_min = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
            today_max = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
            
            # Invoice data
            invoices = Invoice.objects.filter(branch=request.user.branch, issue_date__range=(today_min, today_max))
            partial_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PARTIAL)
            paid_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PAID, branch=request.user.branch)
        
            # Expenses
            expenses = Expense.objects.filter(branch=request.user.branch, date=today)
            confirmed_expenses = expenses.filter(status=True)
            unconfirmed_expenses = expenses.filter(status=False)
            
            # Accounts
            account_balances = AccountBalance.objects.filter(branch=request.user.branch)

            html_string = render_to_string('day_report.html', {
                'request':request,
                'invoices':invoices,
                'expenses':expenses,
                'date': today,
                'inventory_data': inventory_data,
                'total_sales': paid_invoices.aggregatea,
                'partial_payments': partial_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'total_paid_invoices': paid_invoices.count(),
                'total_partial_invoices': partial_invoices.count(),
                'total_expenses': confirmed_expenses.aggregate(Sum('amount'))['amount__sum'] or 0,
                'confirmed_expenses': confirmed_expenses,
                'unconfirmed_expenses': unconfirmed_expenses,
                'account_balances': account_balances,
            })
            
            pdf_buffer = BytesIO()
            pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
            if not pisa_status.err:
                filename = f"{request.user.branch.name}_today_report_{today}.pdf"
                return Response(status.HTTP_200_OK)
            else:
                return Response({"error": "Error generating PDF."})
        except json.JSONDecodeError:
            return Response({'error': 'Invalid JSON data.'}, status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception(f"Error processing request: {e}")
            return Response({'error': str(e)}, status.HTTP_400_BAD_REQUEST)
        

class QuatationCrud(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Qoutation.objects.all()
    serializer_class = QuotationSerializer
    def create(self, request):
        data = request.data
        qoute_id = int(data.get('id'))
        qoute_subtotal = data.get('subtotal')
        qoute_currency_id = data.get('currency_id')
        items_data = data.get('items')
        
        customer = Customer.objects.get(id=qoute_id)
        currency = Currency.objects.get(id=qoute_currency_id)
        
        qoute = Qoutation.objects.create(
            customer = customer,
            amount =  Decimal(qoute_subtotal),
            branch = request.user.branch,
            currency = currency,
            qoute_reference = Qoutation.generate_qoute_number(request.user.branch.name),
            products = ', '.join([f'{item['product_name']} x {item['quantity']}' for item in items_data])
        )
        logger.info(data)
        logger.info(items_data)
        for item_data in items_data:
            pk_id = item_data['inventory_id']
            logger.info(pk_id)
            item = Inventory.objects.get(pk=pk_id)
            
            QoutationItems.objects.create(
                qoute=qoute,
                product=item,
                unit_price=item.price,
                quantity=item_data['quantity'],
                total_amount= item.price * item_data['quantity'],
            )
        return Response({'qoute_id': qoute.id}, status.HTTP_201_CREATED)
    def list(self, request, *args, **kwargs):
        search_query = request.GET.get('q', '')
        qoutations = Qoutation.objects.filter(branch=request.user.branch).order_by('-date').values()
    
        if search_query:
            qoutations = qoutations.filter(
                Q(customer__name__icontains=search_query)|
                Q(products__icontains=search_query)|
                Q(date__icontains=search_query)|
                Q(qoute_reference__icontains=search_query)
            )
        return Response(qoutations, status.HTTP_200_OK)
    def retrieve(self, request, pk):
        search_query = request.GET.get('q', '')
        qoutations = Qoutation.objects.filter(id = pk, branch=request.user.branch).order_by('-date').values()
    
        if search_query:
            qoutations = qoutations.filter(
                Q(customer__name__icontains=search_query)|
                Q(products__icontains=search_query)|
                Q(date__icontains=search_query)|
                Q(qoute_reference__icontains=search_query)
            )
        return Response(qoutations, status.HTTP_200_OK)
    
# class QuotationList(views.APIView):
#     permission_classes = [IsAuthenticated]
#     def get(self, request):
#         search_query = request.GET.get('q', '')
#         qoutations = Qoutation.objects.filter(branch=request.user.branch).order_by('-date').values()
    
#         if search_query:
#             qoutations = qoutations.filter(
#                 Q(customer__name__icontains=search_query)|
#                 Q(products__icontains=search_query)|
#                 Q(date__icontains=search_query)|
#                 Q(qoute_reference__icontains=search_query)
#             )
#         return Response(qoutations, status.HTTP_200_OK)

class QuotationDelete(views.APIView):
    permission_classes = [IsAuthenticated]
    def delete(request, qoutation_id):
        qoute = get_object_or_404(Qoutation, id=qoutation_id)
        qoute.delete()
        return Response(status.HTTP_202_ACCEPTED)

class QuotationView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(request, qoutation_id):
        qoute = Qoutation.objects.get(id=qoutation_id)
        quote_serializer = QuotationSerializer(qoute)
        qoute_items = QoutationItems.objects.filter(qoute=qoute).values()
        return Response({quote_serializer.data, qoute_items}, status.HTTP_200_OK)

class InvoiceList(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        invoices = Invoice.objects.filter(branch=request.user.branch, status=True).order_by('-invoice_number').values()

        query_params = request.GET
        if query_params.get('q'):
            search_query = query_params['q']
            invoices = invoices.filter(
                Q(customer__name__icontains=search_query) |
                Q(invoice_number__icontains=search_query) |
                Q(issue_date__icontains=search_query)
            )

        user_timezone_str = request.user.timezone if hasattr(request.user, 'timezone') else 'UTC'
        user_timezone = pytz_timezone(user_timezone_str)  

        def filter_by_date_range(start_date, end_date):
            start_datetime = user_timezone.localize(
                timezone.datetime.combine(start_date, timezone.datetime.min.time())
            )
            end_datetime = user_timezone.localize(
                timezone.datetime.combine(end_date, timezone.datetime.max.time())
            )
            return invoices.filter(issue_date__range=[start_datetime, end_datetime])

        now = timezone.now().astimezone(user_timezone)
        today = now.date()

        now = timezone.now() 
        today = now.date()  
        
        date_filters = {
            'today': lambda: filter_by_date_range(today, today),
            'yesterday': lambda: filter_by_date_range(today - timedelta(days=1), today - timedelta(days=1)),
            't_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday()), today),
            'l_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday() + 7), today - timedelta(days=today.weekday() + 1)),
            't_month': lambda: invoices.filter(issue_date__month=today.month, issue_date__year=today.year),
            'l_month': lambda: invoices.filter(issue_date__month=today.month - 1 if today.month > 1 else 12, issue_date__year=today.year if today.month > 1 else today.year - 1),
            't_year': lambda: invoices.filter(issue_date__year=today.year),
        }

        if query_params.get('day') in date_filters:
            invoices = date_filters[query_params['day']]()

        total_partial = invoices.filter(payment_status='Partial').aggregate(Sum('amount'))['amount__sum'] or 0
        total_paid = invoices.filter(payment_status='Paid').aggregate(Sum('amount'))['amount__sum'] or 0
        total_amount = invoices.aggregate(Sum('amount'))['amount__sum'] or 0

        logger.info(f'Invoices: {invoices.values}')

        return Response({
            'invoices': invoices,
            'total_paid': total_paid,
            'total_due': total_partial,
            'total_amount': total_amount,
        },status.HTTP_200_OK)

class ExpenseView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        filter_option = request.GET.get('filter', 'today')
        download = request.GET.get('download')
        
        now = datetime.datetime.now()
        end_date = now
        
        if filter_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_week':
            start_date = now - timedelta(days=now.weekday())
        elif filter_option == 'yesterday':
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_month':
            start_date = now.replace(day=1)
        elif filter_option == 'last_month':
            start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        elif filter_option == 'this_year':
            start_date = now.replace(month=1, day=1)
        elif filter_option == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now
            
        expenses = Expense.objects.filter(issue_date__gte=start_date, issue_date__lte=end_date, branch=request.user.branch).order_by('issue_date').values()
        
        # if user role is sales filter expense by user
        if request.user.role == 'sale':
            expenses = expenses.filter(user=request.user)

        if download:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="expenses_report_{filter_option}.csv"'

            writer = csv.writer(response)
            writer.writerow(['Date', 'Description', 'Done By', 'Amount'])

            total_expense = 0  
            for expense in expenses:
                total_expense += expense.amount

                writer.writerow([
                    expense.issue_date,
                    expense.description,
                    expense.user.first_name,
                    expense.amount,
                ])

            writer.writerow(['Total', '', '', total_expense])
            
            return response
        
        return Response( 
            {
                'expenses':expenses,
                'filter_option': filter_option,
            }
        )

class ExpenseDetail(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, expense_id):
        expense = get_object_or_404(Expense, id=expense_id)
        return Response(
        {
            'id': expense.id,
            'amount': expense.amount,
            'description': expense.description,
            'category': expense.category.id
        }, status.HTTP_200_OK)

class AddExpenseCategory(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        categories = ExpenseCategory.objects.all().values()
        data = request.data
        category = data.get('name')
        logger.info(data)
        
        if ExpenseCategory.objects.filter(name=category).exists():
            return Response({'message':f'Category with ID {category} Exists.'}, status.HTTP_400_BAD_REQUEST)
        
        ExpenseCategory.objects.create(
            name=category
        )
        return Response(categories, status.HTTP_201_CREATED)
    
class EditExpense(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, id):
        try:
            data = request.data
            amount = data.get('amount')
            description = data.get('description')
            category_id = data.get('category')
            expense_id = id

            if not amount or not description or not category_id:
                return Response({'message': 'Missing fields: amount, description, category.'}, status.HTTP_400_BAD_REQUEST)
            
            category = get_object_or_404(ExpenseCategory, id=category_id)
            
            if expense_id:  
                expense = get_object_or_404(Expense, id=expense_id)
                before_amount = expense.amount
                
                expense.amount = amount
                expense.description = description
                expense.category = category
                expense.save()
                message = 'Expense successfully updated'
                
                try:
                    cashbook_expense = Cashbook.objects.get(expense=expense)
                    expense_amount = Decimal(expense.amount)
                    if cashbook_expense.amount < expense_amount:
                        cashbook_expense.amount = expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'Expense (update from {before_amount} to {cashbook_expense.amount})'
                    else:
                        cashbook_expense.amount -= cashbook_expense.amount - expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'(update from {before_amount} to {cashbook_expense.amount})'
                    cashbook_expense.save()
                except Exception as e:
                    return Response({str(e)}, status.HTTP_400_BAD_REQUEST)
            return Response({'message': message}, status.HTTP_201_CREATED)
        except Exception as e:
            return Response({str(e)}, status.HTTP_400_BAD_REQUEST)

class DeleteExpense(views.APIView):
    permission_classes = [IsAuthenticated]
    def delete(self, request, expense_id):
        try:
            expense = get_object_or_404(Expense, id=expense_id)
            expense.cancel = True
            expense.save()
            
            Cashbook.objects.create(
                amount=expense.amount,
            debit=True,
            credit=False,
                description=f'Expense ({expense.description}): cancelled'
            )
            return Response({'message': 'Expense successfully deleted'}, status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': str(e)}, status.HTTP_400_BAD_REQUEST)

class UpdateExpenseStatus(views.APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, id):
        try:
            data = request.data
            expense_id = id
            expense_status = data.get('status')

            expense = Expense.objects.get(id=expense_id)
            expense.status = expense_status
            expense.save()

            return Response(status.HTTP_202_ACCEPTED)
        except Exception as e:
                return Response({'message': str(e)}, status.HTTP_400_BAD_REQUEST)

class InvoicePDF(views.APIView):  
    permission_classes = [IsAuthenticated]   
    def get(request, id):
        invoice_id = id
        # if invoice_id:
        #     try:
        #         invoice = get_object_or_404(Invoice, pk=invoice_id)

        #         invoice_items = InvoiceItem.objects.filter(invoice=invoice).values()
                
        #     except Invoice.DoesNotExist:
        #         return Response("Invoice not found",status.HTTP_404_NOT_FOUND)
        # else:
        #     return Response("Invoice ID is required",status.HTTP_404_NOT_FOUND)
        
        # invoice_serializer = InvoiceSerializer(invoice)
        return Response({'https://web-production-86a7.up.railway.app/finance/invoice/pdf/'}, status.HTTP_200_OK)
        # return generate_pdf(
        #     None,
        #     {
        #         'title': 'Invoice', 
        #         'report_date': datetime.date.today(),
        #         'invoice':invoice,
        #         'invoice_items':invoice_items
        #     }
        # )


class CreateInvoice(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        try:
            data = request.data
            invoice_data = data['data'][0]  
            items_data = data.get('items')
            layby_dates = data.get('layby_dates')
           
            # get currency
            currency = Currency.objects.get(id=invoice_data['currency'])
            
            # create or get accounts
            account_types = {
                'cash': Account.AccountType.CASH,
                'bank': Account.AccountType.BANK,
                'ecocash': Account.AccountType.ECOCASH,
            }

            account_name = f"{request.user.branch} {currency.name} {invoice_data['payment_method'].capitalize()} Account"
            
            account, _ = Account.objects.get_or_create(name=account_name, type=account_types[invoice_data['payment_method']])
            
            account_balance, _ = AccountBalance.objects.get_or_create(
                account=account,
                currency=currency,
                branch=request.user.branch,
                defaults={'balance': 0}  
            )
            logger.info(f"[Create Invoice]: {account_balance}")

            
            # accountts_receivable
            accounts_receivable, _ = ChartOfAccounts.objects.get_or_create(name="Accounts Receivable")

            # VAT rate
            vat_rate = VATRate.objects.get(status=True)

            # customer
            customer = Customer.objects.get(id=int(invoice_data['client_id'])) 
            
            # customer account
            customer_account = CustomerAccount.objects.get(customer=customer)

            # customer Account + Balances
            customer_account_balance, _ = CustomerAccountBalances.objects.get_or_create(
                account=customer_account,
                currency=currency, 
                defaults={'balance': 0}
            )
            
            amount_paid = update_latest_due(customer, Decimal(invoice_data['amount_paid']), request, invoice_data['paymentTerms'], customer_account_balance)
            invoice_total_amount = Decimal(invoice_data['payable'])


            logger.info(f'amount paid: {amount_paid}')
            if amount_paid > invoice_total_amount:
                amount_paid = invoice_total_amount
                amount_due = 0
            else:
                amount_paid = amount_paid
                amount_due = invoice_total_amount - amount_paid  
                
            logger.info(f'amount due: {amount_due}')
                    

            cogs = COGS.objects.create(amount=Decimal(0))
            
            with transaction.atomic():
                
                invoice = Invoice.objects.create(
                    invoice_number=Invoice.generate_invoice_number(request.user.branch.name),  
                    customer=customer,
                    issue_date=timezone.now(),
                    amount=invoice_total_amount,
                    amount_paid=amount_paid,
                    amount_due=amount_due,
                    vat=Decimal(invoice_data['vat_amount']),
                    payment_status = Invoice.PaymentStatus.PARTIAL if amount_due > 0 else Invoice.PaymentStatus.PAID,
                    branch = request.user.branch,
                    user=request.user,
                    currency=currency,
                    subtotal=invoice_data['subtotal'],
                    reocurring = invoice_data['recourring'],
                    products_purchased = ', '.join([f'{item['product_name']} x {item['quantity']} ' for item in items_data]),
                    payment_terms = invoice_data['paymentTerms'],
                    hold_status = invoice_data['hold_status'],
                    amount_received = invoice_data['amount_paid']
                )

                logger.info(invoice.hold_status)

                logger.info(f'Invoice created for customer: {invoice}')

                # check if invoice status is hold
                if invoice.hold_status == True:
                    held_invoice(items_data, invoice, request, vat_rate)
                    return Response({'message':'Invoice succesfully on hold'}, status.HTTP_200_OK)

                # create layby object
                if invoice.payment_terms == 'layby':
                    layby_obj = layby.objects.create(invoice=invoice)

                    logger.info(layby_obj)

                    layby_dates_list = []
                    for date in layby_dates:
                        laybyDates.objects.create(
                            layby=layby_obj,
                            due_date=date
                        )
                    
                    # laybyDates.objects.bulk_create(layby_dates)
                
                # create monthly installment object
                if invoice.payment_terms == 'installment':
                    recurringInvoices.objects.create(
                        invoice = invoice,
                        status = False
                    )

                # #create transaction
                transaction_obj = Transaction.objects.create(
                    date=timezone.now(),
                    description=invoice.products_purchased,
                    account=accounts_receivable,
                    debit=Decimal(invoice_data['payable']),
                    credit=Decimal('0.00'),
                    customer=customer
                )

                logger.info(f'Creating transaction obj for invoice: {invoice}')
                
                # Cost of sales parent object
                
                
                # Create InvoiceItem objects
                for item_data in items_data:
                    item = Inventory.objects.get(pk=item_data['inventory_id'])
                    # product = Product.objects.get(pk=item.product.id)
                    
                    item.quantity -= item_data['quantity']
                    item.save()
                  
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        item=item,
                        quantity=item_data['quantity'],
                        unit_price=item_data['price'],
                        vat_rate = vat_rate
                    )
                    
                    # Create StockTransaction for each sold item
                    # stock_transaction = StockTransaction.objects.create(
                    #     item=product,
                    #     transaction_type=StockTransaction.TransactionType.SALE,
                    #     quantity=item_data['quantity'],
                    #     unit_price=item.price,
                    #     invoice=invoice,
                    #     date=timezone.now()
                    # )
                    # stock_transaction.save()
                    
                    # cost of sales item
                    COGSItems.objects.get_or_create(
                        invoice=invoice,
                        defaults={'cogs': cogs, 'product': Inventory.objects.get(id=item.id, branch=request.user.branch)}
                    )
                
                    # stock log  
                    ActivityLog.objects.create(
                        branch=request.user.branch,
                        inventory=item,
                        user=request.user,
                        quantity = -item_data['quantity'],
                        total_quantity = item.quantity,
                        action='Sale',
                        invoice=invoice
                    )

                    accessories = Accessory.objects.filter(main_product=item).values('accessory_product', 'accessory_product__quantity')

                    logger.info(f'Accessories for this product: {accessories}')

                    for acc in accessories:
                        COGSItems.objects.get_or_create(
                            invoice=invoice,
                            defaults={'cogs': cogs, 'product': Inventory.objects.get(id=acc['accessory_product'], branch=request.user.branch)}
                        )
                        prod_acc = Inventory.objects.get(id = acc['accessory_product'] )
                        prod_acc.quantity -= 1

                        logger.info(f'accessory quantity: {acc['accessory_product__quantity']}')

                        ActivityLog.objects.create(
                            branch=request.user.branch,
                            inventory=prod_acc,
                            user=request.user,
                            quantity=1,
                            total_quantity = acc['accessory_product__quantity'],
                            action='Sale',
                            invoice=invoice
                        )
                        prod_acc.save()
                        
                # # Create VATTransaction
                VATTransaction.objects.create(
                    invoice=invoice,
                    vat_type=VATTransaction.VATType.OUTPUT,
                    vat_rate=VATRate.objects.get(status=True).rate,
                    tax_amount=invoice_data['vat_amount']
                )                                                          
                # Create Sale object
                sale = Sale.objects.create(
                    date=timezone.now(),
                    transaction=invoice,
                    total_amount=invoice_total_amount
                )
                sale.save()
                
                #payment
                Payment.objects.create(
                    invoice=invoice,
                    amount_paid=amount_paid,
                    payment_method=invoice_data['payment_method'],
                    amount_due=invoice_total_amount - amount_paid,
                    user=request.user
                )

                # calculate total cogs amount
                cogs.amount = COGSItems.objects.filter(cogs=cogs, cogs__date=datetime.datetime.today())\
                                               .aggregate(total=Sum('product__cost'))['total'] or 0
                logger.info(f'COGS amount: {cogs.amount}')
                cogs.save()
                
                # updae account balance
                if invoice.payment_status == 'Partial':
                    customer_account_balance.balance += -amount_due
                    customer_account_balance.save()
                    
                # Update customer balance
                account_balance.balance = Decimal(invoice_data['payable']) + Decimal(account_balance.balance)
                account_balance.save()

                # try:
                #     return create_invoice_pdf(invoice)
                # except Exception as e:
                #     logger.info(e)
                
                # return redirect('finance:invoice_preview', invoice.id)
                return Response({'invoice_id': invoice.id}, status.HTTP_201_CREATED)
        except (KeyError, json.JSONDecodeError, Customer.DoesNotExist, Inventory.DoesNotExist) as e:
            return Response({'error': str(e)}, status.HTTP_400_BAD_REQUEST)
        
class InvoicePaymentTrack(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, invoice_id):
        invoice_id = invoice_id
        
        if invoice_id:
            payments = Payment.objects.filter(invoice__id=invoice_id).order_by('-payment_date').values(
                'payment_date', 'amount_paid', 'payment_method', 'user__username'
            )
        return Response(payments, status.HTTP_200_OK)
    
class InvoiceDelete(views.APIView):
    permission_classes = [IsAuthenticated]
    def delete(self, request, invoice_id):
        try:
            invoice = get_object_or_404(Invoice, id=invoice_id)
            account = get_object_or_404(CustomerAccount, customer=invoice.customer)
            customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

            sale = get_object_or_404(Sale, transaction=invoice)
            invoice_payment = get_object_or_404(Payment, invoice=invoice)
            stock_transactions = invoice.stocktransaction_set.all()  
            vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)

            with transaction.atomic():
                if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
                    customer_account_balance.balance -= invoice.amount_due

                account_types = {
                    'cash': Account.AccountType.CASH,
                    'bank': Account.AccountType.BANK,
                    'ecocash': Account.AccountType.ECOCASH,
                }

                account = get_object_or_404(
                    Account, 
                    name=f"{request.user.branch} {invoice.currency.name} {invoice_payment.payment_method.capitalize()} Account", 
                    type=account_types.get(invoice_payment.payment_method, None)  
                )
                account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
                account_balance.balance -= invoice.amount_paid

                for stock_transaction in stock_transactions:
                    product = Inventory.objects.get(product=stock_transaction.item, branch=request.user.branch)
                    product.quantity += stock_transaction.quantity
                    product.save()

                    ActivityLog.objects.create(
                        invoice=invoice,
                        product_transfer=None,
                        branch=request.user.branch,
                        user=request.user,
                        action='sale return',
                        inventory=product,
                        quantity=stock_transaction.quantity,
                        total_quantity=product.quantity
                    )

                InvoiceItem.objects.filter(invoice=invoice).delete() 
                StockTransaction.objects.filter(invoice=invoice).delete()
                Payment.objects.filter(invoice=invoice).delete()

                account_balance.save()
                customer_account_balance.save()
                sale.delete()
                vat_transaction.delete()
                invoice.cancelled=True
                invoice.save()

            return Response({'message': f'Invoice {invoice.invoice_number} successfully deleted'}, status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': f"{e}"}, status.HTTP_400_BAD_REQUEST)
        
class InvoiceUpdate(views.APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, invoice_id):
        invoice = get_object_or_404(Invoice, id=invoice_id)
        customer_account = get_object_or_404(CustomerAccount, customer=invoice.customer)
        customer_account_balance = get_object_or_404(
            CustomerAccountBalances, account=customer_account, currency=invoice.currency
        )

        data = request.data
        amount_paid = Decimal(data.get('amount_paid'))

        invoice = Invoice.objects.select_for_update().get(pk=invoice.pk)
        customer_account_balance = CustomerAccountBalances.objects.select_for_update().get(pk=customer_account_balance.pk)

        if amount_paid <= 0:
            return Response({'message': 'Invalid amount paid.'}, status.HTTP_400_BAD_REQUEST)

        if amount_paid >= invoice.amount_due:
            invoice.payment_status = Invoice.PaymentStatus.PAID
            invoice.amount_due = 0
        else:
            invoice.amount_due -= amount_paid

        invoice.amount_paid += amount_paid
        
        # get the latest payment for the invoice
        latest_payment = Payment.objects.filter(invoice=invoice).order_by('-payment_date').first()
        if latest_payment:
            amount_due = latest_payment.amount_due - amount_paid 
        else:
            amount_due = invoice.amount - invoice.amount_paid 

        payment = Payment.objects.create(
            invoice=invoice,
            amount_paid=amount_paid,
            amount_due=amount_due, 
            payment_method=data['payment_method'],
            user=request.user
        )

        account, _ = Account.objects.get_or_create(
            name=f"{request.user.branch} {invoice.currency.name} {payment.payment_method.capitalize()} Account",
            type=Account.AccountType[payment.payment_method.upper()] 
        )
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=invoice.currency,
            branch=request.user.branch,
            defaults={'balance': 0}
        )

        account_balance.balance += amount_paid
        if customer_account_balance.balance < 0:
            customer_account_balance.balance += amount_paid
        else:
            customer_account_balance.balance -= amount_paid

        description = ''
        if invoice.hold_status:
            description = 'Held invoice payment'
            sale = Sale.objects.create(
                date=timezone.now(),
                transaction=invoice,
                total_amount=invoice.amount # invoice delivery amount
            )
            
            VATTransaction.objects.create(
                invoice=invoice,
                vat_type=VATTransaction.VATType.OUTPUT,
                vat_rate=VATRate.objects.get(status=True).rate,
                tax_amount=invoice.vat
            ) 

        else:
            description = 'Invoice payment update'
        
        Cashbook.objects.create(
            issue_date=invoice.issue_date,
            description=f'({description} {invoice.invoice_number})',
            debit=True,
            credit=False,
            amount=invoice.amount_paid,
            currency=invoice.currency,
            branch=invoice.branch
        )

        invoice.hold_status = False
        account_balance.save()
        customer_account_balance.save()
        invoice.save()
        payment.save()
        
        return Response(status.HTTP_202_ACCEPTED)
    
class InvoiceDetails(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, invoice_id):
        invoice = Invoice.objects.filter(id=invoice_id, branch=request.user.branch).values(
            'invoice_number',
            'customer__id', 
            'customer__name', 
            'products_purchased', 
            'payment_status', 
            'amount'
        )
        return Response(invoice, status.HTTP_200_OK)
    
class InvoicePreview(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, invoice_id):
        invoice = Invoice.objects.get(id=invoice_id)
        invoice_serializer = InvoiceSerializer(invoice)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice).values()
        return Response({'invoice_id':invoice_id, 'invoice':invoice_serializer.data, 'invoice_items':invoice_items}, status.HTTP_200_OK)

class InvoicePreviewJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, invoice_id):
        try:
            invoice = Invoice.objects.get(id=invoice_id)

        except Invoice.DoesNotExist:
            return Response({"error": "Invoice not found"}, status.HTTP_400_BAD_REQUEST) 
        
        dates = {}
        if invoice.payment_terms == 'layby':
            dates = laybyDates.objects.filter(layby__invoice=invoice).values('due_date')
        
        invoice_items = InvoiceItem.objects.filter(invoice=invoice).values(
            'item__name', 
            'quantity',
            'item__description',
            'total_amount',
            'unit_price'
        )

        invoice_dict = {
            field.name: getattr(invoice, field.name)
            for field in invoice._meta.fields
            if field.name not in ['customer', 'currency', 'branch', 'user']
        }

        invoice_dict['customer_name'] = invoice.customer.name
        invoice_dict['customer_email'] = invoice.customer.email
        invoice_dict['customer_cell'] = invoice.customer.phone_number
        invoice_dict['customer_address'] = invoice.customer.address
        invoice_dict['currency_symbol'] = invoice.currency.symbol
        invoice_dict['amount_paid'] = invoice.amount_paid
        invoice_dict['payment_terms'] = invoice.payment_terms
        
        if invoice.branch:
            invoice_dict['branch_name'] = invoice.branch.name
            invoice_dict['branch_phone'] = invoice.branch.phonenumber
            invoice_dict['branch_email'] = invoice.branch.email
            
        invoice_dict['user_username'] = invoice.user.username  
        
        invoice_data = {
            'invoice': invoice_dict,
            'invoice_items': invoice_items,
            'dates':dates
        }
        return Response(invoice_data, status.HTTP_200_OK)

class HeldInvoiceView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        invoices = Invoice.objects.filter(branch=request.user.branch, status=True, hold_status =True).order_by('-invoice_number').values()
        logger.info(f'Held invoices: {invoices}')
        return Response(invoices, status.HTTP_200_OK)
    
class ExpenseReport(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        search = request.data.get('search', '')
        start_date_str = request.data.get('startDate', '')
        end_date_str = request.data.get('endDate', '')
        category_id = request.data.get('category', '')
    
        if start_date_str and end_date_str:
            try:
                end_date = datetime.date.fromisoformat(end_date_str)
                start_date = datetime.date.fromisoformat(start_date_str)
            except ValueError:
                return Response({'messgae':'Invalid date format. Please use YYYY-MM-DD.'}, status.HTTP_400_BAD_REQUEST)
        else:
            start_date = ''
            end_date= ''
            
        try:
            category_id = int(category_id) if category_id else None
        except ValueError:
            return Response({'messgae':'Invalid category or search ID.'}, status.HTTP_400_BAD_REQUEST)

        expenses = Expense.objects.all()  
        
        if search:
            expenses = expenses.filter(Q('amount=search'))
        if start_date:
            start_date = parse_date(start_date_str)
            expenses = expenses.filter(date__gte=start_date)
        if end_date:
            end_date = parse_date(end_date_str)
            expenses = expenses.filter(date__lte=end_date)
        if category_id:
            expenses = expenses.filter(category__id=category_id)
        
        return Response(
            {
                'title': 'Expenses', 
                'date_range': f"{start_date} to {end_date}", 
                'report_date': datetime.date.today(),
                'total_expenses':calculate_expenses_totals(expenses),
                'expenses':expenses
            },status.HTTP_200_OK
        )

class SendEmails(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        # data = request.data
        # invoice_id = data['invoice_id']
        # invoice = Invoice.objects.get(id=invoice_id)
        # invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        # account = CustomerAccount.objects.get(customer__id = invoice.customer.id)
        
        # html_string = render_to_string('Pos/receipt.html', {'invoice': invoice, 'invoice_items':invoice_items, 'account':account})
        # buffer = BytesIO()

        # pisa.CreatePDF(html_string, dest=buffer) 

        # email = EmailMessage(
        #     'Your Invoice',
        #     'Please find your invoice attached.',
        #     'your_email@example.com',
        #     ['recipient_email@example.com'],
        # )
        
        # buffer.seek(0)
        # email.attach(f'invoice_{invoice.invoice_number}.pdf', buffer.getvalue(), 'application/pdf')

        # # Send the email
        # email.send()

        # task = send_invoice_email_task.delay(data['invoice_id']) 
        # task_id = task.id 
        # buffer.seek(0)
        # response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        # response['Content-Disposition'] = f'attachment; filename=invoice_{invoice.invoice_number}.pdf'
        
        return Response({'https://web-production-86a7.up.railway.app/finance/invoice/send/email/'}, status.HTTP_200_OK)

class SendWhatsapp(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, invoice_id):
        try:
            
        #     invoice = Invoice.objects.get(pk=invoice_id)
        #     invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        #     img = settings.STATIC_URL + "/assets/logo.png"
        
        #     html_string = render_to_string('Pos/invoice_template.html', {'invoice': invoice, 'request':request, 'invoice_items':invoice_items, 'img':img})
        #     pdf_buffer = BytesIO()
        #     pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        #     if not pisa_status.err:
            
        #         s3 = boto3.client(
        #             "s3",
        #             aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        #             aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        #             region_name=settings.AWS_S3_REGION_NAME,
        #         )
        #         invoice_filename = f"invoice_{invoice.invoice_number}.pdf"
        #         s3.put_object(
        #             Bucket=settings.AWS_STORAGE_BUCKET_NAME,
        #             Key=f"invoices/{invoice_filename}",
        #             Body=pdf_buffer.getvalue(),
        #             ContentType="application/pdf",
        #             ACL="public-read",
        #         )
        #         s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

        #         account_sid = 'AC6890aa7c095ce1315c4a3a86f13bb403'
        #         auth_token = '897e02139a624574c5bd175aa7aaf628'
        #         client = Client(account_sid, auth_token)
        #         from_whatsapp_number = 'whatsapp:' + '+14155238886'
        #         to_whatsapp_number = 'whatsapp:' + '+263778587612'

        #         message = client.messages.create(
        #             from_=from_whatsapp_number,
        #             body="Your invoice is attached.",
        #             to=to_whatsapp_number,
        #             media_url=s3_url
        #         )
        #         logger.info(f"WhatsApp message SID: {message.sid}")
        #         return Response({"message_sid": message.sid}, status.HTTP_200_OK)
        #     else:
        #         logger.error(f"PDF generation error for Invoice ID: {invoice_id}")
        #         return Response({"error": "PDF generation failed"}, status.HTTP_400_BAD_REQUEST)
        # except Invoice.DoesNotExist:
        #     logger.error(f"Invoice not found with ID: {invoice_id}")
        #     return Response({"error": "Invoice not found"}, status.HTTP_400_BAD_REQUEST)
            return Response({f'https://web-production-86a7.up.railway.app/finance/send_invoice_whatsapp/{invoice_id}/'}, status.HTTP_200_OK)
        except Exception as e:
            logger.exception(f"Error sending invoice via WhatsApp: {e}")
            return Response({"error": "Error sending invoice via WhatsApp"}, status.HTTP_400_BAD_REQUEST)
        

class CashbookView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        filter_option = request.GET.get('filter', 'today')
        now = datetime.datetime.now()
        end_date = now
        
        if filter_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_week':
            start_date = now - timedelta(days=now.weekday())
        elif filter_option == 'yesterday':
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_month':
            start_date = now.replace(day=1)
        elif filter_option == 'last_month':
            start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        elif filter_option == 'this_year':
            start_date = now.replace(month=1, day=1)
        elif filter_option == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now

        entries = Cashbook.objects.filter(issue_date__gte=start_date, issue_date__lte=end_date, branch=request.user.branch).order_by('issue_date').values()
        
        total_debit = entries.filter(debit=True, cancelled=False).aggregate(Sum('amount'))['amount__sum'] or 0
        total_credit = entries.filter(credit=True, cancelled=False).aggregate(Sum('amount'))['amount__sum'] or 0
        
        balance_bf = 0 
        
        previous_entries = Cashbook.objects.filter(issue_date__lt=start_date, branch=request.user.branch)

        previous_debit = previous_entries.filter(debit=True).aggregate(Sum('amount'))['amount__sum'] or 0
        previous_credit = previous_entries.filter(credit=True).aggregate(Sum('amount'))['amount__sum'] or 0
        balance_bf = previous_debit - previous_credit

        total_balance = total_debit - total_credit
        logger.info(total_balance)
        invoice_items = InvoiceItem.objects.all().values()

        return Response({
            'filter_option': filter_option,
            'entries': entries,
            'balance_bf': balance_bf,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'total_balance': total_balance,
            'end_date': end_date,
            'start_date': start_date,
            'invoice_items': invoice_items
        })
    def post(self, request):
        filter_option = request.data.get('filter', 'today')
        now = datetime.datetime.now()
        end_date = now
        
        if filter_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_week':
            start_date = now - timedelta(days=now.weekday())
        elif filter_option == 'yesterday':
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_month':
            start_date = now.replace(day=1)
        elif filter_option == 'last_month':
            start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        elif filter_option == 'this_year':
            start_date = now.replace(month=1, day=1)
        elif filter_option == 'custom':
            start_date = request.data.get('start_date')
            end_date = request.data.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now

        entries = Cashbook.objects.filter(issue_date__gte=start_date, issue_date__lte=end_date, branch=request.user.branch).order_by('issue_date').values()
        
        total_debit = entries.filter(debit=True, cancelled=False).aggregate(Sum('amount'))['amount__sum'] or 0
        total_credit = entries.filter(credit=True, cancelled=False).aggregate(Sum('amount'))['amount__sum'] or 0
        
        balance_bf = 0 
        
        previous_entries = Cashbook.objects.filter(issue_date__lt=start_date, branch=request.user.branch)

        previous_debit = previous_entries.filter(debit=True).aggregate(Sum('amount'))['amount__sum'] or 0
        previous_credit = previous_entries.filter(credit=True).aggregate(Sum('amount'))['amount__sum'] or 0
        balance_bf = previous_debit - previous_credit

        total_balance = total_debit - total_credit
        logger.info(total_balance)
        invoice_items = InvoiceItem.objects.all().values()

        return Response({
            'filter_option': filter_option,
            'entries': entries,
            'balance_bf': balance_bf,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'total_balance': total_balance,
            'end_date': end_date,
            'start_date': start_date,
            'invoice_items': invoice_items
        }, status.HTTP_200_OK)
    
class CashbookNote(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        #payload
        """
            entry_id:id,
            note:str
        """
        try:
            data = request.data
            entry_id = data.get('entry_id')
            note = data.get('note')
            
            entry = Cashbook.objects.get(id=entry_id)
            entry.note = note
            
            entry.save()
        except Exception as e:
            return Response({'message':f'{e}.'}, status.HTTP_400_BAD_REQUEST)
        return Response({'message':'Note successfully saved.'}, status.HTTP_201_CREATED)
    
class CashbookReport(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        # filter_option = request.data.get('filter', 'this_week')
        # now = datetime.datetime.now()
        # end_date = now
        
        # if filter_option == 'today':
        #     start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        # elif filter_option == 'this_week':
        #     start_date = now - timedelta(days=now.weekday())
        # elif filter_option == 'yesterday':
        #     start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        # elif filter_option == 'this_month':
        #     start_date = now.replace(day=1)
        # elif filter_option == 'last_month':
        #     start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        # elif filter_option == 'this_year':
        #     start_date = now.replace(month=1, day=1)
        # elif filter_option == 'custom':
        #     start_date = request.data.get('start_date')
        #     end_date = request.data.get('end_date')
        #     start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        #     end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        # else:
        #     start_date = now - timedelta(days=now.weekday())
        #     end_date = now

        # entries = Cashbook.objects.filter(issue_date__gte=start_date, issue_date__lte=end_date, branch=request.user.branch).order_by('issue_date')

        # # Create a CSV response
        # response = HttpResponse(content_type='text/csv')
        # response['Content-Disposition'] = f'attachment; filename="cashbook_report_{filter_option}.csv"'

        # writer = csv.writer(response)
        # writer.writerow(['Date', 'Description', 'Expenses', 'Income', 'Balance'])

        # balance = 0  
        # for entry in entries:
        #     if entry.debit:
        #         balance += entry.amount
        #     elif entry.credit:
        #         balance -= entry.amount

        #     writer.writerow([
        #         entry.issue_date,
        #         entry.description,
        #         entry.amount if entry.debit else '',
        #         entry.amount if entry.credit else '',
        #         balance,
        #         entry.accountant,
        #         entry.manager,
        #         entry.director
        #     ])
        return Response({'https://web-production-86a7.up.railway.app/finance/report/'}, status.HTTP_200_OK)

class CancelTransaction(views.APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        #payload
        """
            entry_id:id,
        """
        try:
            data = request.data
            entry_id = int(data.get('entry_id'))
            
            logger.info(entry_id)
            
            entry = Cashbook.objects.get(id=entry_id)
            
            entry.cancelled = True
            
            if entry.director:
                entry.director = False
            elif entry.manager:
                entry.manager = False
            elif entry.accountant:
                entry.accountant = False
                
            entry.save()
            logger.info(entry)
            return Response(status.HTTP_201_CREATED)
        except Exception as e:
            logger.info(e)
            return Response({'message': str(e)}, status.HTTP_400_BAD_REQUEST)
        
class CashbookNoteView(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, entry_id):
        entry = get_object_or_404(Cashbook, id=entry_id)
    
        notes = entry.notes.all().order_by('timestamp')
        notes_data = [
            {'user': note.user.username, 'note': note.note, 'timestamp': note.timestamp.strftime("%Y-%m-%d %H:%M:%S")}
            for note in notes
        ]
        return Response({'notes': notes_data}, status.HTTP_200_OK)
    
    def post(self, request, entry_id):
        entry = get_object_or_404(Cashbook, id=entry_id)
        try:
            data = json.loads(request.body)
            note_text = data.get('note')
            CashBookNote.objects.create(entry=entry, user=request.user, note=note_text)
            return Response({'message': 'Note successfully added.'}, status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'message': str(e)}, status.HTTP_400_BAD_REQUEST)

class UpdateTransactionStatus(views.APIView):    
    permission_classes = [IsAuthenticated]
    def put(self, request, pk):
        entry = get_object_or_404(Cashbook, pk=pk)
        data = request.data
        
        status_t = data.get('status')
        field = data.get('field')  

        if field in ['manager', 'accountant', 'director']:
            setattr(entry, field, status_t)

            if entry.cancelled:
                entry.cancelled = False
            entry.save()
            return Response({'status': getattr(entry, field)}, status.HTTP_200_OK)
            
        return Response(status.HTTP_400_BAD_REQUEST)   
    
class CashFlowView(views.APIView):
    def get(self, request):
        cashflows = Cashflow.objects.all().order_by('-date').values()
        cashups = CashUp.objects.select_related('branch').filter(status=False).values()

        cashFlows_total = cashflows.aggregate(total=Sum('total'))['total'] or 0
        cash_flow_items = CashUp.objects.filter(status=False).aggregate(total=Sum('expected_cash'))['total'] or 0

        # categories
        main_income_category = MainIncomeCategory.objects.all().select_related('sub_income_category').values()
        main_expense_category = MainExpenseCategory.objects.all().select_related('sub_expense').values()
        cash_flow_names = CashFlowName.objects.all().values()
        
        return Response(
        {
            'cashflows': cashflows,
            'cashups': cashups,
            'cashFlows_total': cashFlows_total,
            'cash_flow_items': cash_flow_items,
            'cash_flow_names': cash_flow_names,
            'income_categories': main_income_category,
            'expense_categories': main_expense_category,
        }, status.HTTP_200_OK)
    def post(self, request):
        try:
            data = request.data
            income = float(data.get('IncomeAmount', 0))
            expense_amount = float(data.get('ExpenseAmount', 0))
            transaction_type = data.get('type', '')
            income_category = data.get('incomeCategory', '')
            expense_category = data.get('expenseCategory', '')
            #cashflow name
            name = data.get('name')
            #adding subcategories
            income_sub_category = data.get('incomeSubCategory', '')
            expense_sub_category = data.get('expenseSubCategory', '')
            #ends here
            income_branch = data.get('incomeBranch', '')
            expense_branch = data.get('expenseBranch', '')

            logger.info(expense_category)

            # validations 
            
            if not transaction_type or transaction_type not in ['income', 'expense']:
                return JsonResponse({'success': False, 'message': 'Invalid transaction type.'}, status=400)
            
            with transaction.atomic():
                if transaction_type == 'income':

                    if not income or income <= 0:
                        return JsonResponse({'success': False, 'message': 'Invalid amount.'}, status=400)
                    
                    if transaction_type == 'income' and not income_category:
                        return Response({'message': 'Income category is required.'}, status.HTTP_400_BAD_REQUEST)

                    # if transaction_type == 'income' and not income_branch:
                    #     return JsonResponse({'success': False, 'message': 'Income branch is required.'}, status=400)
            
                    logger.info(f'Creating Income amount: {income}')
                    cash_flow_name, _ = CashFlowName.objects.get_or_create(name=name)
                    income_category, _ = MainIncomeCategory.objects.get_or_create(id=income_category, defaults={'name': 'Income'})

                    object = Cashflow.objects.create(
                        name=cash_flow_name,
                        branch=request.user.branch,
                        total=income,
                        date=datetime.datetime.now(),
                        status=False,
                        income=income,
                        income_category=income_category,
                        created_by=request.user
                    )

                    logger.info(f'Income created: {object}.')

                    return Response({'message': 'Income cashflow successfully created'}, status.HTTP_201_CREATED)
                
                else:

                    if not expense_amount or expense_amount <= 0:
                        return JsonResponse({'success': False, 'message': 'Invalid amount.'}, status=400)

                    if transaction_type == 'expense' and not expense_category:
                        return JsonResponse({'success': False, 'message': 'Expense category is required.'}, status=400)
                    
                    # if transaction_type == 'expense' and not expense_branch:
                    #     return JsonResponse({'success': False, 'message': 'Expense branch is required.'}, status=400)
                    
                    logger.info(f'Creating Expense amount: {expense_amount}')
                    cash_flow_name, _ = CashFlowName.objects.get_or_create(name=name)
                    expense_category, _ = MainExpenseCategory.objects.get_or_create(id=expense_category, defaults={'name': 'Expense'})

                    object = Cashflow.objects.create(
                        name=cash_flow_name,
                        branch=request.user.branch,
                        total=expense_amount,
                        date=datetime.datetime.now(),
                        status=False,
                        expense=expense_amount,
                        income=0,
                        expense_category=expense_category,
                        created_by=request.user
                    )

                    logger.info(f'Expense created: {object}.')

                    return JsonResponse({'success': True, 'message': 'Expense cashflow successfully created'}, status=201)

        except Exception as e:
            logger.error(f"Error recording transaction {e}.")
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

class CashUpList(views.APIView):
    def get(self, request):
        cashups = (
            CashUp.objects
            .select_related('branch', 'created_by')
            .filter(status=False)
            .values(
                'id',
                'branch__name',
                'expected_cash',
                'created_by__username',
                'created_at',
                'received_amount'
            )
            .order_by('-created_at')
        )

        data = []
        for cashup in cashups:
            cashup_dict = dict(cashup)
            cashup_dict['created_at'] = cashup['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            data.append(cashup_dict)

        return Response({
            'success': True,
            'data': data
        }, status.HTTP_200_OK)

class DaysData(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        current_month = get_current_month()

        sales = Sale.objects.filter(date__month=current_month).values()
        cogs = COGSItems.objects.filter(date__month=current_month).values()
        if sales or cogs:
            first_day = min(sales.first().date, cogs.first().date)
        else:
            return Response({'Empty':{ 'sales':sales, 'COGS':cogs}}, status.HTTP_200_OK)
        logger.info(first_day)
        def get_week_data(queryset, start_date, end_date, amount_field):
            week_data = queryset.filter(date__gte=start_date, date__lt=end_date).values(amount_field, 'date')
            logger.info(week_data)
            total = week_data.aggregate(total=Sum(amount_field))['total'] or 0
            return week_data, total

        data = {}
        for week in range(1, 5):
            week_start = first_day + timedelta(days=(week-1)*7)
            week_end = week_start + timedelta(days=7)

            logger.info(week_start)
            logger.info(week_end)

            sales_data, sales_total = get_week_data(sales, week_start, week_end, 'total_amount')
            cogs_data, cogs_total = get_week_data(cogs, week_start, week_end, 'product__cost')
            
            data[f'week {week}'] = {
                'sales': list(sales_data),
                'cogs': list(cogs_data),
                'total_sales': sales_total,
                'total_cogs': cogs_total
            }

        return Response(data, status.HTTP_200_OK)

class VAT(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        filter_option = request.GET.get('filter', 'this_week')
        download = request.GET.get('download')
        
        now = datetime.datetime.now()
        end_date = now
        
        if filter_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_week':
            start_date = now - timedelta(days=now.weekday())
        elif filter_option == 'yesterday':
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_month':
            start_date = now.replace(day=1)
        elif filter_option == 'last_month':
            start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        elif filter_option == 'this_year':
            start_date = now.replace(month=1, day=1)
        elif filter_option == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now
            
        vat_transactions = VATTransaction.objects.filter(date__gte=start_date, date__lte=end_date).values().order_by('-date')
        
        if download:
            # response = HttpResponse(content_type='text/csv')
            # response['Content-Disposition'] = f'attachment; filename="vat_report_{filter_option}.csv"'

            # writer = csv.writer(response)
            # writer.writerow(['Date', 'Description', 'Status', 'Input', 'Output'])

            # balance = 0
            # for transaction in vat_transactions:

            #     if transaction.vat_type == 'Input':
            #         balance += transaction.tax_amount
            #     else:
            #         balance -= transaction.tax_amount

            #     writer.writerow([
            #         transaction.date,
            #         transaction.invoice.invoice_number if transaction.invoice else transaction.purchase_order.order_number,
            #         transaction.tax_amount if transaction.vat_type == 'Input' else  '',
            #         transaction.tax_amount if transaction.vat_type == 'Output' else  ''
            #     ])

            # writer.writerow(['Total', '', '', balance])
            return Response({'https://web-production-86a7.up.railway.app/finance/vat/'}, status.HTTP_200_OK)
        return Response( 
            {
                'filter_option':filter_option,
                'vat_transactions':vat_transactions
            },
            status.HTTP_200_OK
        )
    
    def post(self, request):
        # payload 
        {
            'date_from':'date',
            'date_to':'date'
        }
        try:
            data = request.data
            
            date_to = data.get('date_to')
            date_from = data.get('date_from')

            vat_transactions = VATTransaction.objects.filter(
                date__gte=date_from, 
                date__lte=date_to
            )
            
            vat_transactions.update(paid=True)
        except Exception as e:
            return Response({'message':f'{e}'}, status.HTTP_400_BAD_REQUEST)
        return Response({'message':'VAT successfully paid'}, status.HTTP_200_OK)

class PLOverview(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        filter_option = request.data.get('filter', 'today')
        today = datetime.date.today()
        previous_month = get_previous_month()
        current_year = today.year
        current_month = today.month

        sales = Sale.objects.filter(transaction__branch=request.user.branch)
        expenses = Expense.objects.filter(branch=request.user.branch)
        cogs = COGSItems.objects.filter(invoice__branch=request.user.branch)

        if filter_option == 'today':
            date_filter = today
        elif filter_option == 'last_week':
            last_week_start = today - datetime.timedelta(days=today.weekday() + 7)
            last_week_end = last_week_start + datetime.timedelta(days=6)
            date_filter = (last_week_start, last_week_end)
        elif filter_option == 'this_month':
            date_filter = (datetime.date(current_year, current_month, 1), today)
        elif filter_option == 'year':
            year = int(request.GET.get('year', current_year))
            date_filter = (datetime.date(year, 1, 1), datetime.date(year, 12, 31))
        else:
            date_filter = (datetime.date(current_year, current_month, 1), today)

        if filter_option == 'today':
            current_month_sales = sales.filter(date=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
            current_month_expenses = expenses.filter(issue_date=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
            cogs_total = cogs.filter(date=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
        elif filter_option == 'last_week':
            current_month_sales = sales.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
            current_month_expenses = expenses.filter(issue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
            cogs_total = cogs.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
        else:
            current_month_sales = sales.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
            current_month_expenses = expenses.filter(dissue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
            cogs_total = cogs.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0

        previous_month_sales = sales.filter(date__year=current_year, date__month=previous_month).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        previous_month_expenses = expenses.filter(issue_date__year=current_year, issue_date__month=previous_month).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        previous_cogs =  cogs.filter(date__year=current_year, date__month=previous_month).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
        
        current_net_income = current_month_sales
        previous_net_income = previous_month_sales 
        current_expenses = current_month_expenses 
        
        current_gross_profit = current_month_sales - cogs_total
        previous_gross_profit = previous_month_sales - previous_cogs
        
        current_net_profit = current_gross_profit - current_month_expenses
        previous_net_profit = previous_gross_profit - previous_month_expenses

        current_gross_profit_margin = (current_gross_profit / current_month_sales * 100) if current_month_sales != 0 else 0
        previous_gross_profit_margin = (previous_gross_profit / previous_month_sales * 100) if previous_month_sales != 0 else 0
        
        # net_income_change = calculate_percentage_change(current_net_income, previous_net_income)
        # gross_profit_change = calculate_percentage_change(current_gross_profit, previous_gross_profit)
        # gross_profit_margin_change = calculate_percentage_change(current_gross_profit_margin, previous_gross_profit_margin)


        data = {
            'net_profit':current_net_profit,
            'cogs_total':cogs_total,
            'current_expenses':current_expenses,
            'current_net_profit': current_net_profit,
            'previous_net_profit':previous_net_profit,
            'current_net_income': current_net_income,
            'previous_net_income': previous_net_income,
            'current_gross_profit': current_gross_profit,
            'previous_gross_profit': previous_gross_profit,
            'current_gross_profit_margin': f'{current_gross_profit_margin:.2f}',
            'previous_gross_profit_margin': previous_gross_profit_margin,
        }
        
        return Response(data, status.HTTP_200_OK)

class IncomeJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        current_month = get_current_month()
        today = datetime.date.today()
        
        month = request.data.get('month', current_month)
        day = request.data.get('day', today.day)

        sales = Sale.objects.filter(transaction__branch=request.user.branch).values()
        
        if request.data.get('filter') == 'today':
            sales_total = sales.filter(date=today).aggregate(Sum('total_amount'))
        else:
            sales_total = sales.filter(date__month=month).aggregate(Sum('total_amount'))

        return Response({'sales_total': sales_total['total_amount__sum'] or 0}, status.HTTP_200_OK)
    

class ExpenseJson(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        current_month = get_current_month()
        today = datetime.date.today()
        
        month = request.data.get('month', current_month)
        day = request.data.get('day', today.day)

        expenses = Expense.objects.filter(branch=request.user.branch).values()
        
        if request.data.get('filter') == 'today':
            expense_total = expenses.filter(issue_date=today, status=False).aggregate(Sum('amount'))
        else:
            expense_total = expenses.filter(issue_date__month=month, status=False).aggregate(Sum('amount'))
        
        return Response({'expense_total': expense_total['amount__sum'] or 0}, status.HTTP_200_OK)

class AccountType(views.APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        payment_options = Account.objects.all().values('name', 'type')
        return Response({'payment options': payment_options}, status.HTTP_200_OK)

class FinanceApi(views.APIView):
    def get(self, request, *args, **kwargs):

        balances = AccountBalance.objects.filter(branch=request.user.branch).values()
    
        recent_sales = Sale.objects.filter(transaction__branch=request.user.branch).order_by('-date')[:5].values()

        expenses_by_category = Expense.objects.values('category__name').annotate(
            total_amount=Sum('amount', output_field=DecimalField())
        ).values()
        
        return Response({
            'balances': balances,
            'recent_transactions': recent_sales,
            'expenses_by_category': expenses_by_category,
        })

class UserAccountsView(views.APIView):
    def get(self, request):
        users = User.objects.filter(is_active=True).prefetch_related('accounts')
    
        users_with_accounts = []
        for user in users:
            accounts = user.accounts.all()
            
            total_balance = accounts.aggregate(
                total=Coalesce(Sum('balance', output_field=DecimalField()), Decimal('0.00'))
            )['total']
            
            total_credits = accounts.aggregate(
                total=Coalesce(Sum('total_credits', output_field=DecimalField()), Decimal('0.00'))
            )['total']
            
            total_debits = accounts.aggregate(
                total=Coalesce(Sum('total_debits', output_field=DecimalField()), Decimal('0.00'))
            )['total']

            last_activity = accounts.aggregate(
                last_date=Max('last_transaction_date')
            )['last_date']
            
            last_activity = accounts.aggregate(
                last_date=Max('last_transaction_date')
            )['last_date']
            
            users_with_accounts.append({
                'user': user.get_full_name(),
                # 'accounts': accounts,
                'total_balance': total_balance,
                'total_credits': total_credits,
                'total_debits': total_debits,
                'last_activity': last_activity
            })

        return Response({'Account Data':users_with_accounts}, status.HTTP_200_OK)